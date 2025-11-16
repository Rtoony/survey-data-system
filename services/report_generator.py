"""
Report Generation Service
Handles PDF and Excel report generation for the Report Builder feature
"""

from weasyprint import HTML, CSS
from jinja2 import Environment, FileSystemLoader, select_autoescape
import openpyxl
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import io
import os
from datetime import datetime
from decimal import Decimal
import json

class ReportGenerator:
    """Main class for generating reports in PDF and Excel formats"""

    def __init__(self, templates_dir='templates'):
        """Initialize the report generator with Jinja2 environment"""
        self.templates_dir = templates_dir
        self.env = Environment(
            loader=FileSystemLoader(templates_dir),
            autoescape=select_autoescape(['html', 'xml'])
        )

        # Add custom filters for Jinja2
        self.env.filters['format_number'] = self._format_number
        self.env.filters['format_date'] = self._format_date

    def _format_number(self, value, decimals=2):
        """Format number with specified decimal places"""
        if value is None:
            return 'N/A'
        if isinstance(value, (int, float, Decimal)):
            return f"{float(value):,.{decimals}f}"
        return value

    def _format_date(self, value, format='%Y-%m-%d'):
        """Format datetime object"""
        if value is None:
            return 'N/A'
        if isinstance(value, datetime):
            return value.strftime(format)
        return str(value)

    def generate_pdf_report(self, template_path, data, output_path, css_string=None):
        """
        Generate a PDF report using WeasyPrint

        Args:
            template_path: Path to Jinja2 HTML template (relative to templates_dir)
            data: Dictionary of data to pass to template
            output_path: Path where PDF should be saved
            css_string: Optional CSS string for styling

        Returns:
            output_path if successful, None if failed
        """
        try:
            # Load and render template
            template = self.env.get_template(template_path)
            html_content = template.render(**data, generated_at=datetime.now())

            # Generate PDF
            html = HTML(string=html_content)

            if css_string:
                css = CSS(string=css_string)
                html.write_pdf(output_path, stylesheets=[css])
            else:
                html.write_pdf(output_path)

            return output_path

        except Exception as e:
            print(f"Error generating PDF: {str(e)}")
            return None

    def generate_excel_report(self, config, data, output_path):
        """
        Generate an Excel report with data and charts

        Args:
            config: Configuration dict with sheets, columns, chart specs
            data: Dictionary with data for each sheet
            output_path: Path where Excel file should be saved

        Returns:
            output_path if successful, None if failed
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Process each sheet in config
            for sheet_config in config.get('sheets', []):
                sheet_name = sheet_config.get('name', 'Sheet1')
                sheet_data = data.get(sheet_config.get('data_key', 'data'), [])
                columns = sheet_config.get('columns', [])

                # Create worksheet
                ws = wb.create_sheet(sheet_name)

                # Write headers
                header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
                header_font = Font(bold=True, color="000000")

                for col_idx, column in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = column.get('label', column.get('field', ''))
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Write data rows
                for row_idx, row_data in enumerate(sheet_data, 2):
                    for col_idx, column in enumerate(columns, 1):
                        field = column.get('field')
                        value = row_data.get(field, '')

                        cell = ws.cell(row=row_idx, column=col_idx)

                        # Handle different data types
                        if isinstance(value, (int, float, Decimal)):
                            cell.value = float(value)
                            if column.get('format') == 'number':
                                cell.number_format = '#,##0.00'
                        elif isinstance(value, datetime):
                            cell.value = value
                            cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                        else:
                            cell.value = str(value) if value is not None else ''

                        # Apply alignment
                        align = column.get('align', 'left')
                        cell.alignment = Alignment(horizontal=align)

                # Auto-size columns
                for col_idx in range(1, len(columns) + 1):
                    column_letter = get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = 15

                # Add chart if specified
                if sheet_config.get('chart'):
                    self._add_excel_chart(ws, sheet_config['chart'], len(sheet_data))

            # Save workbook
            wb.save(output_path)
            return output_path

        except Exception as e:
            print(f"Error generating Excel: {str(e)}")
            return None

    def _add_excel_chart(self, worksheet, chart_config, row_count):
        """Add a chart to an Excel worksheet"""
        try:
            chart_type = chart_config.get('type', 'bar')

            # Create appropriate chart
            if chart_type == 'bar':
                chart = BarChart()
            elif chart_type == 'pie':
                chart = PieChart()
            elif chart_type == 'line':
                chart = LineChart()
            else:
                return

            # Set chart properties
            chart.title = chart_config.get('title', 'Chart')
            chart.style = 10
            chart.height = 10
            chart.width = 20

            # Add data
            data_col = chart_config.get('data_column', 2)
            label_col = chart_config.get('label_column', 1)

            data = Reference(worksheet, min_col=data_col, min_row=1, max_row=row_count + 1)
            labels = Reference(worksheet, min_col=label_col, min_row=2, max_row=row_count + 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)

            # Position chart
            chart_position = chart_config.get('position', 'E2')
            worksheet.add_chart(chart, chart_position)

        except Exception as e:
            print(f"Error adding chart: {str(e)}")

    def generate_chart_svg(self, chart_config, data):
        """
        Generate a chart as SVG using matplotlib

        Args:
            chart_config: Dict with chart type, labels, colors
            data: List of values or dict with labels and values

        Returns:
            SVG string or None if failed
        """
        try:
            fig, ax = plt.subplots(figsize=(10, 6))

            chart_type = chart_config.get('type', 'bar')

            if isinstance(data, dict):
                labels = list(data.keys())
                values = list(data.values())
            else:
                labels = [f"Item {i+1}" for i in range(len(data))]
                values = data

            # Create chart based on type
            if chart_type == 'bar':
                ax.bar(labels, values, color='#00ffff')
                ax.set_ylabel(chart_config.get('y_label', 'Values'))
            elif chart_type == 'pie':
                ax.pie(values, labels=labels, autopct='%1.1f%%', colors=['#00ffff', '#ff00ff', '#00ff88', '#ffff00'])
            elif chart_type == 'line':
                ax.plot(labels, values, marker='o', color='#00ffff', linewidth=2)
                ax.set_ylabel(chart_config.get('y_label', 'Values'))
                ax.grid(True, alpha=0.3)

            ax.set_title(chart_config.get('title', 'Chart'), color='#00ffff', fontsize=14, fontweight='bold')

            # Style the chart
            fig.patch.set_facecolor('#0a0e27')
            ax.set_facecolor('#1a1f3a')
            ax.tick_params(colors='#e0e6ed')
            ax.spines['bottom'].set_color('#00ffff')
            ax.spines['left'].set_color('#00ffff')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            # Save to SVG string
            svg_buffer = io.StringIO()
            plt.savefig(svg_buffer, format='svg', bbox_inches='tight', facecolor=fig.get_facecolor())
            plt.close(fig)

            return svg_buffer.getvalue()

        except Exception as e:
            print(f"Error generating chart: {str(e)}")
            return None


# Convenience functions for direct use
def generate_pdf(template_path, data, output_path, css_string=None):
    """Generate PDF report"""
    generator = ReportGenerator()
    return generator.generate_pdf_report(template_path, data, output_path, css_string)


def generate_excel(config, data, output_path):
    """Generate Excel report"""
    generator = ReportGenerator()
    return generator.generate_excel_report(config, data, output_path)


def generate_chart(chart_config, data):
    """Generate chart as SVG"""
    generator = ReportGenerator()
    return generator.generate_chart_svg(chart_config, data)
