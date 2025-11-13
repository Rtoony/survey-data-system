"""
ACAD-GIS Schema Explorer & Data Manager
A companion tool for viewing and managing your Supabase database
"""

from flask import Flask, render_template, jsonify, request, send_file, make_response, redirect, url_for
from flask_cors import CORS
from flask_caching import Cache
import psycopg2
from psycopg2.extras import RealDictCursor
import os
import csv
import io
import uuid
from dotenv import load_dotenv
from contextlib import contextmanager
from werkzeug.utils import secure_filename
from dxf_importer import DXFImporter
from dxf_exporter import DXFExporter
from map_export_service import MapExportService
import threading
import json
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill
import zipfile
import tempfile
from weasyprint import HTML, CSS
from services.project_mapping_service import ProjectMappingService
from project_mapping_registry import get_supported_entity_types
from database import get_db, execute_query, DB_CONFIG

# Load environment variables (works with both .env file and Replit secrets)
load_dotenv()

# Custom JSON encoder for datetime, date, and Decimal objects
class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, uuid.UUID):
            return str(obj)
        return super().default(obj)

app = Flask(__name__)
app.json_encoder = CustomJSONEncoder
CORS(app)

# Configure caching
app.config['CACHE_TYPE'] = 'SimpleCache'  # In-memory cache
app.config['CACHE_DEFAULT_TIMEOUT'] = 300  # 5 minutes default
cache = Cache(app)

# Debug: Check if DB credentials are available
print("=" * 50)
print("Database Configuration Status:")
print(f"DB_HOST: {'SET' if DB_CONFIG['host'] else 'MISSING'}")
print(f"DB_USER: {'SET' if DB_CONFIG['user'] else 'MISSING'}")
print(f"DB_NAME: {'SET' if DB_CONFIG['database'] else 'MISSING'}")
print(f"DB_PASSWORD: {'SET' if DB_CONFIG['password'] else 'MISSING'}")
print("=" * 50)

# ============================================
# PAGES
# ============================================

@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@app.route('/schema')
def schema_page():
    """Schema viewer page"""
    return render_template('schema.html')

@app.route('/schema/graph')
def schema_graph_page():
    """Schema knowledge graph visualization page"""
    return render_template('schema_graph.html')

@app.route('/projects')
def projects_page():
    """Projects manager page"""
    return render_template('projects.html')

@app.route('/projects/<project_id>')
def project_overview(project_id):
    """Project Overview dashboard page"""
    return render_template('project_overview.html', project_id=project_id)

@app.route('/projects/<project_id>/survey-points')
def project_survey_points(project_id):
    """Project Survey Point Manager page"""
    return render_template('project_survey_points.html', project_id=project_id)

@app.route('/standards-library')
def standards_library():
    """Standards Library landing page"""
    return render_template('standards_library.html')

@app.route('/project-operations')
def project_operations():
    """Project Operations landing page"""
    return render_template('project_operations.html')

# ============================================
# DATA MANAGER PAGES
# ============================================

@app.route('/data-manager')
def data_manager_home():
    """Data Manager home page"""
    return render_template('data_manager/index.html')

@app.route('/data-manager/abbreviations')
def data_manager_abbreviations():
    """Abbreviations data manager page"""
    return render_template('data_manager/abbreviations.html')

@app.route('/data-manager/layers')
def data_manager_layers():
    """Layers data manager page"""
    return render_template('data_manager/layers.html')

@app.route('/data-manager/blocks')
def data_manager_blocks():
    """Blocks data manager page"""
    return render_template('data_manager/blocks.html')

@app.route('/data-manager/details')
def data_manager_details():
    """Details data manager page"""
    return render_template('data_manager/details.html')

@app.route('/data-manager/categories')
def data_manager_categories():
    """Categories data manager page"""
    return render_template('data_manager/categories.html')

@app.route('/data-manager/disciplines')
def data_manager_disciplines():
    """Disciplines data manager page"""
    return render_template('data_manager/disciplines.html')

@app.route('/tools/batch-cad-import')
def batch_cad_import_tool():
    """Unified Batch CAD Import Tool (Blocks, Details, Hatches, Linetypes)"""
    return render_template('tools/batch_block_import.html')

@app.route('/tools/batch-block-import')
def batch_block_import_tool_redirect():
    """Legacy redirect - Batch Block Import Tool"""
    return redirect(url_for('batch_cad_import_tool'))

@app.route('/tools/batch-point-import')
def batch_point_import_tool():
    """Batch Point Import Tool - Import survey points from PNEZD text files (includes Survey Code Import mode)"""
    return render_template('tools/batch_point_import.html')

@app.route('/tools/survey-point-manager')
def survey_point_manager_tool():
    """Survey Point Manager - Manage imported survey points with filtering and soft-delete"""
    return render_template('tools/survey_point_manager.html')

@app.route('/tools/specialized-tools-directory')
def specialized_tools_directory():
    """Specialized Tools Directory - comprehensive list of interactive management tools"""
    return render_template('tools/specialized_tools_directory.html')

@app.route('/tools/z-stress-test')
def z_stress_test_tool():
    """Z-Value Elevation Preservation Stress Test - interactive testing tool"""
    return render_template('tools/z_stress_test.html')

@app.route('/tools/survey-codes')
def survey_codes_manager():
    """Survey Code Library Manager - CRUD interface for field survey codes"""
    return render_template('tools/survey_codes.html')

@app.route('/tools/survey-code-tester')
def survey_code_tester():
    """Survey Code Testing Interface - Test parsing, preview CAD output, simulate field shots"""
    return render_template('tools/survey_code_tester.html')

@app.route('/project-standards-assignment')
def project_standards_assignment():
    """Project Standards Assignment page"""
    return render_template('project_standards_assignment.html')

@app.route('/project-compliance')
def project_compliance():
    """Project Compliance Dashboard page"""
    return render_template('project_compliance.html')

# ============================================
# CAD STANDARDS PORTAL PAGES
# ============================================

@app.route('/standards')
def standards_home():
    """CAD Standards Portal home page"""
    return render_template('standards/index.html')

@app.route('/standards/layers')
def standards_layers():
    """Layer standards page"""
    return render_template('standards/layers.html')

@app.route('/standards/blocks')
def standards_blocks():
    """Block/symbol standards page"""
    return render_template('standards/blocks.html')

@app.route('/standards/colors')
def standards_colors():
    """Color standards page"""
    return render_template('standards/colors.html')

@app.route('/standards/linetypes')
def standards_linetypes():
    """Linetype standards page"""
    return render_template('standards/linetypes.html')

@app.route('/standards/text')
def standards_text():
    """Text style standards page"""
    return render_template('standards/text.html')

@app.route('/standards/hatches')
def standards_hatches():
    """Hatch pattern standards page"""
    return render_template('standards/hatches.html')

@app.route('/standards/details')
def standards_details():
    """Detail standards page"""
    return render_template('standards/details.html')

@app.route('/standards/abbreviations')
def standards_abbreviations():
    """Abbreviation standards page"""
    return render_template('standards/abbreviations.html')

@app.route('/standards/materials')
def standards_materials():
    """Material standards page"""
    return render_template('standards/materials.html')

@app.route('/standards/vocabulary')
def standards_vocabulary():
    """CAD Standards Vocabulary browser page"""
    return render_template('standards/vocabulary.html')

@app.route('/standards/layer-vocabulary')
def standards_layer_vocabulary():
    """CAD Layer Vocabulary - layer naming classification system"""
    return render_template('standards/layer-vocabulary.html')

@app.route('/standards/reference-data')
def standards_reference_data():
    """Reference Data Hub - system configuration and reference tables"""
    return render_template('standards/reference-data.html')

@app.route('/standards/reference')
def standards_reference():
    """CAD Standards Layer Reference - visual layer examples"""
    return render_template('standards/reference.html')

@app.route('/standards/sheets')
def standards_sheets():
    """Sheet template standards page"""
    return render_template('standards/sheets.html')

@app.route('/standards/plotstyles')
def standards_plotstyles():
    """Plot style standards page"""
    return render_template('standards/plotstyles.html')

@app.route('/standards/viewports')
def standards_viewports():
    """Viewport standards page"""
    return render_template('standards/viewports.html')

@app.route('/standards/annotations')
def standards_annotations():
    """Annotation standards page"""
    return render_template('standards/annotations.html')

@app.route('/standards/categories')
def standards_categories():
    """Symbol categories page"""
    return render_template('standards/categories.html')

@app.route('/standards/codes')
def standards_codes():
    """Code references page"""
    return render_template('standards/codes.html')

@app.route('/standards/notes')
def standards_notes():
    """Standard notes page"""
    return render_template('standards/notes.html')

@app.route('/standards/scales')
def standards_scales():
    """Drawing scale standards page"""
    return render_template('standards/scales.html')

@app.route('/standards/import-manager')
def standards_import_manager():
    """Import Template Manager page"""
    return render_template('standards/import_manager.html')

@app.route('/standards/bulk-editor')
def standards_bulk_editor():
    """Bulk Standards Editor page"""
    return render_template('standards/bulk_editor.html')

# ============================================
# API ENDPOINTS
# ============================================

@app.route('/api/health')
def health():
    """Check database connection"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT version()')
                version = cur.fetchone()[0]
                return jsonify({
                    'status': 'connected',
                    'database': version
                })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

@app.route('/api/schema')
def get_schema():
    """Get database schema information"""
    try:
        query = """
            SELECT 
                table_name,
                column_name,
                data_type,
                is_nullable,
                column_default
            FROM information_schema.columns
            WHERE table_schema = 'public'
            ORDER BY table_name, ordinal_position
        """
        columns = execute_query(query)
        
        # Group by table
        schema = {}
        for col in columns:
            table = col['table_name']
            if table not in schema:
                schema[table] = []
            schema[table].append({
                'name': col['column_name'],
                'type': col['data_type'],
                'nullable': col['is_nullable'] == 'YES',
                'default': col['column_default']
            })
        
        # Get row counts
        counts = {}
        for table in schema.keys():
            try:
                result = execute_query(f'SELECT COUNT(*) as count FROM {table}')
                counts[table] = result[0]['count'] if result else 0
            except:
                counts[table] = 0
        
        return jsonify({
            'tables': schema,
            'counts': counts
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects')
def get_projects():
    """Get all projects"""
    try:
        query = """
            SELECT p.*
            FROM projects p
            ORDER BY p.created_at DESC
        """
        projects = execute_query(query)
        return jsonify({'projects': projects})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects', methods=['POST'])
def create_project():
    """Create a new project"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
            
        project_name = data.get('project_name')
        client_name = data.get('client_name')
        project_number = data.get('project_number')
        description = data.get('description')

        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400

        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO projects (
                        project_name, client_name, project_number, description,
                        quality_score, tags, attributes
                    )
                    VALUES (%s, %s, %s, %s, 0.5, '{}', '{}')
                    RETURNING project_id, project_name, client_name, project_number, created_at
                    """,
                    (project_name, client_name, project_number, description)
                )
                result = cur.fetchone()
                conn.commit()
                
                return jsonify({
                    'project_id': str(result[0]),
                    'project_name': result[1],
                    'client_name': result[2],
                    'project_number': result[3],
                    'created_at': result[4].isoformat() if result[4] else None
                })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>', methods=['DELETE'])
def delete_project(project_id):
    """Delete a project"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                # Delete project (cascades will handle related entities)
                cur.execute('DELETE FROM projects WHERE project_id = %s', (project_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>')
def get_project(project_id):
    """Get a single project by ID"""
    try:
        query = """
            SELECT 
                p.*,
                c.client_name as client_name_from_ref
            FROM projects p
            LEFT JOIN clients c ON p.client_id = c.client_id
            WHERE p.project_id = %s
        """
        result = execute_query(query, (project_id,))
        
        if not result:
            return jsonify({'error': 'Project not found'}), 404
        
        project_data = result[0]
        if project_data.get('client_name_from_ref'):
            project_data['client_name'] = project_data['client_name_from_ref']
        
        return jsonify(project_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/survey-points')
def get_project_survey_points(project_id):
    """Get all survey points for a specific project with optional filtering"""
    try:
        # Get query parameters for filtering
        point_type = request.args.get('point_type')
        is_control = request.args.get('is_control')
        search = request.args.get('search')
        
        # Build dynamic WHERE clause
        where_conditions = ["sp.project_id = %s", "sp.is_active = true"]
        params = [project_id]
        
        if point_type:
            where_conditions.append("sp.point_type = %s")
            params.append(point_type)
        
        if is_control is not None:
            is_control_bool = is_control.lower() == 'true'
            where_conditions.append("sp.is_control_point = %s")
            params.append(is_control_bool)
        
        if search:
            where_conditions.append("LOWER(sp.point_number::text) LIKE %s")
            params.append(f"%{search.lower()}%")
        
        where_clause = " AND ".join(where_conditions)
        
        query = f"""
            SELECT 
                sp.point_id,
                sp.point_number,
                sp.point_description,
                sp.point_type,
                sp.elevation,
                sp.northing,
                sp.easting,
                sp.is_control_point,
                sp.survey_date,
                sp.surveyed_by,
                sp.survey_method,
                sp.horizontal_accuracy,
                sp.vertical_accuracy,
                sp.quality_score,
                sp.notes,
                sp.created_at
            FROM survey_points sp
            WHERE {where_clause}
            ORDER BY sp.point_number
        """
        points = execute_query(query, tuple(params))
        return jsonify({
            'project_id': project_id,
            'count': len(points),
            'points': points
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/survey-points', methods=['DELETE'])
def delete_survey_points(project_id):
    """Soft-delete survey points by setting is_active = false"""
    try:
        data = request.get_json()
        if not data or 'point_ids' not in data:
            return jsonify({'error': 'point_ids array is required'}), 400
        
        point_ids = data['point_ids']
        if not isinstance(point_ids, list) or len(point_ids) == 0:
            return jsonify({'error': 'point_ids must be a non-empty array'}), 400
        
        # Soft delete points (set is_active = false)
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE survey_points 
                    SET is_active = false, updated_at = CURRENT_TIMESTAMP
                    WHERE project_id = %s 
                      AND point_id = ANY(%s::uuid[])
                      AND is_active = true
                    """,
                    (project_id, point_ids)
                )
                deleted_count = cur.rowcount
                conn.commit()
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'{deleted_count} survey point(s) deleted successfully'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>', methods=['PUT'])
def update_project(project_id):
    """Update an existing project"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
            
        project_name = data.get('project_name')
        client_id = data.get('client_id')
        client_name = data.get('client_name')
        project_number = data.get('project_number')
        description = data.get('description')

        if not project_name:
            return jsonify({'error': 'project_name is required'}), 400

        with get_db() as conn:
            with conn.cursor() as cur:
                if client_id:
                    cur.execute(
                        "SELECT client_name FROM clients WHERE client_id = %s",
                        (client_id,)
                    )
                    client_result = cur.fetchone()
                    if client_result:
                        client_name = client_result[0]
                
                cur.execute(
                    """
                    UPDATE projects 
                    SET project_name = %s, 
                        client_id = %s,
                        client_name = %s, 
                        project_number = %s, 
                        description = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE project_id = %s
                    RETURNING project_id, project_name, client_id, client_name, project_number, updated_at
                    """,
                    (project_name, client_id, client_name, project_number, description, project_id)
                )
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Project not found'}), 404
                conn.commit()
                
                return jsonify({
                    'project_id': str(result[0]),
                    'project_name': result[1],
                    'client_id': result[2],
                    'client_name': result[3],
                    'project_number': result[4],
                    'updated_at': result[5].isoformat() if result[5] else None
                })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/map-summary')
def get_project_map_summary(project_id):
    """Get project spatial extent and simplified geometries for map display"""
    try:
        # First, check if project exists
        project_check = execute_query(
            "SELECT project_id, project_name FROM projects WHERE project_id = %s",
            (project_id,)
        )
        if not project_check:
            return jsonify({'error': 'Project not found'}), 404
        
        # Calculate bounding box from all drawing entities associated with this project
        # Handle both SRID 2226 and SRID 0 geometries
        extent_query = """
            SELECT 
                ST_XMin(extent_geom) as min_x,
                ST_YMin(extent_geom) as min_y,
                ST_XMax(extent_geom) as max_x,
                ST_YMax(extent_geom) as max_y,
                ST_AsGeoJSON(extent_geom)::json as bbox_geojson
            FROM (
                SELECT ST_Extent(
                    CASE 
                        WHEN ST_SRID(de.geometry) = 0 THEN ST_Transform(ST_SetSRID(de.geometry, 2226), 4326)
                        WHEN ST_SRID(de.geometry) = 2226 THEN ST_Transform(de.geometry, 4326)
                        ELSE ST_Transform(de.geometry, 4326)
                    END
                ) as extent_geom
                FROM drawing_entities de
                WHERE de.project_id = %s
                  AND de.geometry IS NOT NULL
            ) sub
        """
        extent_result = execute_query(extent_query, (project_id,))
        
        # If no geometries found, return placeholder response
        if not extent_result or not extent_result[0]['min_x']:
            return jsonify({
                'project_id': project_id,
                'project_name': project_check[0]['project_name'],
                'has_spatial_data': False,
                'message': 'No spatial data available for this project yet',
                'bbox': None,
                'features': []
            })
        
        extent = extent_result[0]
        
        # Get simplified geometries for map display (limit to 100 features for performance)
        # Handle SRID 0 by setting it to 2226
        features_query = """
            SELECT 
                de.entity_id,
                COALESCE(l.layer_name, 'Unknown') as layer_name,
                de.entity_type,
                ST_AsGeoJSON(
                    ST_Transform(
                        ST_Simplify(
                            CASE 
                                WHEN ST_SRID(de.geometry) = 0 THEN ST_SetSRID(de.geometry, 2226)
                                ELSE de.geometry
                            END,
                            1.0
                        ),
                        4326
                    )
                )::json as geometry
            FROM drawing_entities de
            LEFT JOIN layers l ON de.layer_id = l.layer_id
            WHERE de.project_id = %s
              AND de.geometry IS NOT NULL
            ORDER BY ST_Area(
                CASE 
                    WHEN ST_SRID(de.geometry) = 0 THEN ST_SetSRID(de.geometry, 2226)
                    ELSE de.geometry
                END
            ) DESC
            LIMIT 100
        """
        features = execute_query(features_query, (project_id,))
        
        # Build GeoJSON feature collection
        geojson_features = []
        for feat in features:
            geojson_features.append({
                'type': 'Feature',
                'properties': {
                    'entity_id': str(feat['entity_id']),
                    'layer_name': feat['layer_name'],
                    'entity_type': feat['entity_type']
                },
                'geometry': feat['geometry']
            })
        
        return jsonify({
            'project_id': project_id,
            'project_name': project_check[0]['project_name'],
            'has_spatial_data': True,
            'bbox': {
                'min_x': float(extent['min_x']),
                'min_y': float(extent['min_y']),
                'max_x': float(extent['max_x']),
                'max_y': float(extent['max_y'])
            },
            'features': {
                'type': 'FeatureCollection',
                'features': geojson_features
            },
            'feature_count': len(features)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/statistics')
def get_project_statistics(project_id):
    """Get project statistics and KPIs"""
    try:
        # Verify project exists
        project_check = execute_query(
            "SELECT project_id, project_name FROM projects WHERE project_id = %s",
            (project_id,)
        )
        if not project_check:
            return jsonify({'error': 'Project not found'}), 404
        
        # Get total entity count across all drawing entities
        entity_count_query = """
            SELECT COUNT(*) as count 
            FROM drawing_entities de
            WHERE de.project_id = %s
        """
        entity_count = execute_query(entity_count_query, (project_id,))[0]['count']
        
        # Get survey point count (project-level entities, active only)
        survey_point_count_query = """
            SELECT COUNT(*) as count 
            FROM survey_points 
            WHERE project_id = %s
              AND is_active = true
        """
        survey_point_count = execute_query(survey_point_count_query, (project_id,))[0]['count']
        
        # Get reference data counts using efficient COUNT queries (not full entity loads)
        # Single query with UNION ALL for all mapping tables
        # NOTE: Only counts ACTIVE attachments (is_active = true) to reflect current project state.
        # Inactive/soft-deleted attachments are excluded from dashboard KPIs.
        reference_counts_query = """
            SELECT 'clients' as entity_type, COUNT(*) as count
            FROM project_clients WHERE project_id = %s AND is_active = true
            UNION ALL
            SELECT 'vendors', COUNT(*)
            FROM project_vendors WHERE project_id = %s AND is_active = true
            UNION ALL
            SELECT 'municipalities', COUNT(*)
            FROM project_municipalities WHERE project_id = %s AND is_active = true
            UNION ALL
            SELECT 'coordinate_systems', COUNT(*)
            FROM project_coordinate_systems WHERE project_id = %s AND is_active = true
            UNION ALL
            SELECT 'survey_point_descriptions', COUNT(*)
            FROM project_survey_point_descriptions WHERE project_id = %s AND is_active = true
            UNION ALL
            SELECT 'gis_layers', COUNT(*)
            FROM project_gis_layers WHERE project_id = %s AND is_active = true
        """
        
        counts_result = execute_query(reference_counts_query, 
                                     (project_id, project_id, project_id, 
                                      project_id, project_id, project_id))
        
        reference_data_counts = {row['entity_type']: row['count'] for row in counts_result}
        
        # Get spatial extent
        has_spatial_data = False
        bbox = None
        extent_query = """
            SELECT 
                ST_XMin(extent_geom) as min_x,
                ST_YMin(extent_geom) as min_y,
                ST_XMax(extent_geom) as max_x,
                ST_YMax(extent_geom) as max_y
            FROM (
                SELECT ST_Extent(
                    CASE 
                        WHEN ST_SRID(de.geometry) = 0 THEN ST_Transform(ST_SetSRID(de.geometry, 2226), 4326)
                        WHEN ST_SRID(de.geometry) = 2226 THEN ST_Transform(de.geometry, 4326)
                        ELSE ST_Transform(de.geometry, 4326)
                    END
                ) as extent_geom
                FROM drawing_entities de
                WHERE de.project_id = %s
                  AND de.geometry IS NOT NULL
            ) sub
        """
        extent_result = execute_query(extent_query, (project_id,))
        
        if extent_result and extent_result[0]['min_x']:
            has_spatial_data = True
            bbox = {
                'min_x': float(extent_result[0]['min_x']),
                'min_y': float(extent_result[0]['min_y']),
                'max_x': float(extent_result[0]['max_x']),
                'max_y': float(extent_result[0]['max_y'])
            }
        
        return jsonify({
            'project_id': project_id,
            'project_name': project_check[0]['project_name'],
            'drawing_count': drawing_count,
            'entity_count': entity_count,
            'survey_point_count': survey_point_count,
            'reference_data_counts': reference_data_counts,
            'total_reference_data': sum(reference_data_counts.values()),
            'has_spatial_data': has_spatial_data,
            'bbox': bbox
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# PROJECT MAPPING API ENDPOINTS (Generic Entity Attachments)
# ============================================================================

@app.route('/api/projects/<project_id>/<entity_type>', methods=['GET'])
def list_project_entities(project_id, entity_type):
    """List all entities of a type attached to a project"""
    try:
        service = ProjectMappingService(entity_type)
        entities = service.list_attached(project_id)
        return jsonify({
            'entity_type': entity_type,
            'count': len(entities),
            'entities': entities
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/<entity_type>/available', methods=['GET'])
def list_available_entities(project_id, entity_type):
    """List all entities of a type available to attach (not already attached)"""
    try:
        service = ProjectMappingService(entity_type)
        entities = service.list_available(project_id)
        return jsonify({
            'entity_type': entity_type,
            'count': len(entities),
            'entities': entities
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/<entity_type>', methods=['POST'])
def attach_entity_to_project(project_id, entity_type):
    """Attach an entity to a project"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
        
        entity_id = data.get('entity_id')
        if not entity_id:
            return jsonify({'error': 'entity_id is required'}), 400
        
        service = ProjectMappingService(entity_type)
        result = service.attach(
            project_id=project_id,
            entity_id=entity_id,
            is_primary=data.get('is_primary', False),
            relationship_notes=data.get('relationship_notes'),
            display_order=data.get('display_order')
        )
        
        return jsonify(result), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/<entity_type>/<int:mapping_id>', methods=['PUT'])
def update_project_entity_mapping(project_id, entity_type, mapping_id):
    """Update a project-entity mapping (primary status, notes, order)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
        
        service = ProjectMappingService(entity_type)
        result = service.update(
            project_id=project_id,
            mapping_id=mapping_id,
            is_primary=data.get('is_primary'),
            relationship_notes=data.get('relationship_notes'),
            display_order=data.get('display_order')
        )
        
        return jsonify(result)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<project_id>/<entity_type>/<int:mapping_id>', methods=['DELETE'])
def detach_entity_from_project(project_id, entity_type, mapping_id):
    """Detach an entity from a project (soft delete)"""
    try:
        service = ProjectMappingService(entity_type)
        result = service.detach(project_id, mapping_id)
        return jsonify(result)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recent-activity')
@cache.cached(timeout=60)  # Cache for 1 minute
def get_recent_activity():
    """Get recent projects for dashboard"""
    try:
        # Get 5 most recent projects
        projects_query = """
            SELECT 
                p.project_id,
                p.project_name,
                p.project_number,
                p.client_name,
                p.created_at
            FROM projects p
            ORDER BY p.created_at DESC
            LIMIT 5
        """
        recent_projects = execute_query(projects_query)
        
        # Get database stats
        stats_query = """
            SELECT 
                (SELECT COUNT(*) FROM projects) as total_projects,
                (SELECT COUNT(*) FROM layer_standards) as total_layers,
                (SELECT COUNT(*) FROM block_definitions) as total_blocks
        """
        stats = execute_query(stats_query)
        
        return jsonify({
            'recent_projects': recent_projects,
            'stats': stats[0] if stats else {}
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# CAD STANDARDS API ENDPOINTS
# ============================================

@app.route('/api/standards/overview')
@cache.cached(timeout=600)  # Cache for 10 minutes
def standards_overview():
    """Get overview of all standards"""
    try:
        overview = {}
        tables = {
            'layers': 'layer_standards',
            'blocks': 'block_definitions',
            'colors': 'color_standards',
            'linetypes': 'linetypes',
            'text_styles': 'text_styles',
            'hatches': 'hatch_patterns',
            'details': 'detail_standards',
            'dimensions': 'dimension_styles',
            'abbreviations': 'abbreviation_standards',
            'materials': 'material_standards',
            'sheets': 'sheet_templates',
            'plotstyles': 'plot_style_standards',
            'viewports': 'viewport_standards',
            'annotations': 'annotation_standards',
            'categories': 'symbol_categories',
            'codes': 'code_references',
            'notes': 'standard_notes',
            'scales': 'drawing_scale_standards'
        }
        
        for key, table in tables.items():
            try:
                result = execute_query(f'SELECT COUNT(*) as count FROM {table}')
                overview[key] = result[0]['count'] if result else 0
            except:
                overview[key] = 0
        
        return jsonify(overview)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/layers')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_layer_standards():
    """Get all layer standards"""
    try:
        query = """
            SELECT 
                layer_id,
                layer_name,
                discipline,
                category,
                color,
                color_rgb,
                aci_color,
                linetype,
                lineweight,
                description,
                usage_notes,
                tags,
                is_active,
                is_deprecated
            FROM layer_standards
            WHERE is_active = true OR is_active IS NULL
            ORDER BY discipline, category, layer_name
        """
        layers = execute_query(query)
        return jsonify(layers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/blocks')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_block_standards():
    """Get all block/symbol standards"""
    try:
        query = """
            SELECT 
                block_id,
                block_name,
                block_type,
                category,
                description,
                has_attributes,
                is_dynamic,
                tags,
                dxf_file_path,
                preview_image_path,
                is_active
            FROM block_definitions
            WHERE is_active = true OR is_active IS NULL
            ORDER BY category, block_name
        """
        blocks = execute_query(query)
        return jsonify(blocks)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/colors')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_color_standards():
    """Get all color standards"""
    try:
        query = """
            SELECT 
                color_id,
                color_name,
                aci_number,
                rgb_value,
                hex_value,
                cmyk_value,
                usage_context,
                description,
                discipline,
                is_active
            FROM color_standards
            WHERE is_active = true OR is_active IS NULL
            ORDER BY aci_number
        """
        colors = execute_query(query)
        return jsonify(colors)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/linetypes')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_linetype_standards():
    """Get all linetype standards"""
    try:
        query = """
            SELECT 
                linetype_id,
                linetype_name,
                description,
                pattern_definition,
                usage_context
            FROM linetypes
            WHERE is_active = true
            ORDER BY linetype_name
        """
        linetypes = execute_query(query)
        return jsonify(linetypes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/text')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_text_standards():
    """Get all text style standards"""
    try:
        query = """
            SELECT 
                style_id,
                style_name,
                font_name,
                font_file,
                height,
                width_factor,
                oblique_angle,
                description,
                usage_context,
                discipline,
                is_active
            FROM text_styles
            WHERE is_active = true OR is_active IS NULL
            ORDER BY style_name
        """
        text_styles = execute_query(query)
        return jsonify(text_styles)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/hatches')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_hatch_standards():
    """Get all hatch pattern standards"""
    try:
        query = """
            SELECT 
                hatch_id,
                pattern_name,
                pattern_type,
                pattern_definition,
                description,
                usage_context,
                pat_file_path,
                preview_image_path,
                is_active
            FROM hatch_patterns
            WHERE is_active = true OR is_active IS NULL
            ORDER BY pattern_name
        """
        hatches = execute_query(query)
        return jsonify(hatches)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/details')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_detail_standards():
    """Get all detail standards"""
    try:
        query = """
            SELECT 
                ds.detail_id,
                ds.detail_number,
                ds.detail_title,
                ds.detail_category,
                ds.applicable_layers,
                ds.applicable_symbols,
                ds.scale,
                ds.description,
                ds.usage_context,
                ds.typical_application,
                ds.code_references,
                bd.block_name,
                bd.svg_content
            FROM detail_standards ds
            LEFT JOIN block_definitions bd ON ds.block_id = bd.block_id
            ORDER BY ds.detail_category, ds.detail_number
        """
        details = execute_query(query)
        return jsonify(details)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/abbreviations')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_abbreviation_standards():
    """Get all abbreviation standards"""
    try:
        query = """
            SELECT * FROM abbreviation_standards
            ORDER BY discipline, abbreviation
        """
        abbreviations = execute_query(query)
        return jsonify(abbreviations)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/materials')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_material_standards():
    """Get all material standards"""
    try:
        query = """
            SELECT * FROM material_standards
            ORDER BY category, material_name
        """
        materials = execute_query(query)
        return jsonify(materials)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/sheets')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_sheet_templates():
    """Get all sheet template standards"""
    try:
        query = """
            SELECT * FROM sheet_templates
            ORDER BY discipline, sheet_size, template_name
        """
        sheets = execute_query(query)
        return jsonify(sheets)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/plotstyles')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_plotstyle_standards():
    """Get all plot style standards"""
    try:
        query = """
            SELECT * FROM plot_style_standards
            ORDER BY style_name
        """
        plotstyles = execute_query(query)
        return jsonify(plotstyles)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/viewports')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_viewport_standards():
    """Get all viewport standards"""
    try:
        query = """
            SELECT * FROM viewport_standards
            ORDER BY discipline, scale_factor
        """
        viewports = execute_query(query)
        return jsonify(viewports)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/annotations')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_annotation_standards():
    """Get all annotation standards"""
    try:
        query = """
            SELECT * FROM annotation_standards
            ORDER BY discipline, annotation_type, annotation_name
        """
        annotations = execute_query(query)
        return jsonify(annotations)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/categories')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_symbol_categories():
    """Get all symbol categories"""
    try:
        query = """
            SELECT * FROM symbol_categories
            ORDER BY sort_order, category_name
        """
        categories = execute_query(query)
        return jsonify(categories)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/codes')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_code_references():
    """Get all code references"""
    try:
        query = """
            SELECT * FROM code_references
            ORDER BY code_type, code_name
        """
        codes = execute_query(query)
        return jsonify(codes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/notes')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_standard_notes():
    """Get all standard notes"""
    try:
        query = """
            SELECT * FROM standard_notes
            ORDER BY discipline, note_category, sort_order
        """
        notes = execute_query(query)
        return jsonify(notes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards/scales')
@cache.cached(timeout=600)  # Cache for 10 minutes
def get_drawing_scales():
    """Get all drawing scale standards"""
    try:
        query = """
            SELECT * FROM drawing_scale_standards
            ORDER BY scale_type, scale_factor
        """
        scales = execute_query(query)
        return jsonify(scales)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# DATA MANAGER API ENDPOINTS
# ============================================

# ============================================================================
# CATEGORIES MANAGEMENT API ENDPOINTS
# ============================================================================

@app.route('/api/categories', methods=['GET'])
def get_standard_categories():
    """Get all categories, optionally filtered by standard_type"""
    try:
        standard_type = request.args.get('standard_type')
        
        if standard_type:
            # Get categories filtered by standard type
            query = """
                SELECT DISTINCT sc.category_id, sc.category_code, sc.category_name, 
                       sc.description, sc.sort_order, sc.is_active
                FROM standard_categories sc
                JOIN standard_category_applications sca ON sc.category_id = sca.category_id
                WHERE sc.is_active = TRUE AND sca.standard_type = %s
                ORDER BY sc.sort_order, sc.category_name
            """
            categories = execute_query(query, (standard_type,))
        else:
            # Get all categories
            query = """
                SELECT category_id, category_code, category_name, description, 
                       sort_order, is_active, parent_category_id
                FROM standard_categories
                WHERE is_active = TRUE
                ORDER BY sort_order, category_name
            """
            categories = execute_query(query)
        
        return jsonify({'categories': categories})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories', methods=['POST'])
def create_category():
    """Create a new category"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO standard_categories 
                    (category_code, category_name, description, sort_order)
                    VALUES (%s, %s, %s, %s)
                    RETURNING category_id
                """, (
                    data.get('category_code').upper(),
                    data.get('category_name'),
                    data.get('description'),
                    data.get('sort_order', 0)
                ))
                category_id = cur.fetchone()[0]
                
                # Add standard type applications if provided
                if data.get('standard_types'):
                    for std_type in data.get('standard_types'):
                        cur.execute("""
                            INSERT INTO standard_category_applications (category_id, standard_type)
                            VALUES (%s, %s)
                        """, (category_id, std_type))
                
                conn.commit()
        
        cache.clear()
        return jsonify({'category_id': category_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories/<category_id>', methods=['PUT'])
def update_category(category_id):
    """Update an existing category"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE standard_categories
                    SET category_code = %s, category_name = %s, description = %s, 
                        sort_order = %s, is_active = %s
                    WHERE category_id = %s
                """, (
                    data.get('category_code').upper(),
                    data.get('category_name'),
                    data.get('description'),
                    data.get('sort_order', 0),
                    data.get('is_active', True),
                    category_id
                ))
                
                # Update standard type applications if provided
                if 'standard_types' in data:
                    # Delete existing applications
                    cur.execute("DELETE FROM standard_category_applications WHERE category_id = %s", (category_id,))
                    # Add new applications
                    for std_type in data.get('standard_types'):
                        cur.execute("""
                            INSERT INTO standard_category_applications (category_id, standard_type)
                            VALUES (%s, %s)
                        """, (category_id, std_type))
                
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories/<category_id>', methods=['DELETE'])
def delete_category(category_id):
    """Delete a category (soft delete by setting is_active=false)"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE standard_categories 
                    SET is_active = FALSE 
                    WHERE category_id = %s
                """, (category_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# DISCIPLINES MANAGEMENT API ENDPOINTS
# ============================================================================

@app.route('/api/disciplines', methods=['GET'])
def get_standard_disciplines():
    """Get all disciplines, optionally filtered by standard_type"""
    try:
        standard_type = request.args.get('standard_type')
        
        if standard_type:
            # Get disciplines filtered by standard type
            query = """
                SELECT DISTINCT sd.discipline_id, sd.discipline_code, sd.discipline_name, 
                       sd.description, sd.sort_order, sd.is_active
                FROM standard_disciplines sd
                JOIN standard_discipline_applications sda ON sd.discipline_id = sda.discipline_id
                WHERE sd.is_active = TRUE AND sda.standard_type = %s
                ORDER BY sd.sort_order, sd.discipline_name
            """
            disciplines = execute_query(query, (standard_type,))
        else:
            # Get all disciplines
            query = """
                SELECT discipline_id, discipline_code, discipline_name, description, 
                       sort_order, is_active
                FROM standard_disciplines
                WHERE is_active = TRUE
                ORDER BY sort_order, discipline_name
            """
            disciplines = execute_query(query)
        
        return jsonify({'disciplines': disciplines})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/disciplines', methods=['POST'])
def create_discipline():
    """Create a new discipline"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO standard_disciplines 
                    (discipline_code, discipline_name, description, sort_order)
                    VALUES (%s, %s, %s, %s)
                    RETURNING discipline_id
                """, (
                    data.get('discipline_code').upper(),
                    data.get('discipline_name'),
                    data.get('description'),
                    data.get('sort_order', 0)
                ))
                discipline_id = cur.fetchone()[0]
                
                # Add standard type applications if provided
                if data.get('standard_types'):
                    for std_type in data.get('standard_types'):
                        cur.execute("""
                            INSERT INTO standard_discipline_applications (discipline_id, standard_type)
                            VALUES (%s, %s)
                        """, (discipline_id, std_type))
                
                conn.commit()
        
        cache.clear()
        return jsonify({'discipline_id': discipline_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/disciplines/<discipline_id>', methods=['PUT'])
def update_discipline(discipline_id):
    """Update an existing discipline"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE standard_disciplines
                    SET discipline_code = %s, discipline_name = %s, description = %s, 
                        sort_order = %s, is_active = %s
                    WHERE discipline_id = %s
                """, (
                    data.get('discipline_code').upper(),
                    data.get('discipline_name'),
                    data.get('description'),
                    data.get('sort_order', 0),
                    data.get('is_active', True),
                    discipline_id
                ))
                
                # Update standard type applications if provided
                if 'standard_types' in data:
                    # Delete existing applications
                    cur.execute("DELETE FROM standard_discipline_applications WHERE discipline_id = %s", (discipline_id,))
                    # Add new applications
                    for std_type in data.get('standard_types'):
                        cur.execute("""
                            INSERT INTO standard_discipline_applications (discipline_id, standard_type)
                            VALUES (%s, %s)
                        """, (discipline_id, std_type))
                
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/disciplines/<discipline_id>', methods=['DELETE'])
def delete_discipline(discipline_id):
    """Delete a discipline (soft delete by setting is_active=false)"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE standard_disciplines 
                    SET is_active = FALSE 
                    WHERE discipline_id = %s
                """, (discipline_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ABBREVIATIONS CRUD

@app.route('/api/data-manager/abbreviations', methods=['POST'])
def create_abbreviation():
    """Create a new abbreviation"""
    try:
        data = request.json
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO abbreviation_standards (category, discipline, abbreviation, full_text, context_usage)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING abbreviation_id
                """, (data.get('category'), data.get('discipline', 'CIVIL'), data.get('abbreviation'), 
                      data.get('full_text'), data.get('description')))
                abbreviation_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True, 'abbreviation_id': abbreviation_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/abbreviations/<int:abbreviation_id>', methods=['PUT'])
def update_abbreviation(abbreviation_id):
    """Update an existing abbreviation"""
    try:
        data = request.json
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE abbreviation_standards
                    SET category = %s, abbreviation = %s, full_text = %s, context_usage = %s
                    WHERE abbreviation_id = %s
                """, (data.get('category'), data.get('abbreviation'), data.get('full_text'), 
                      data.get('description'), abbreviation_id))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/abbreviations/<int:abbreviation_id>', methods=['DELETE'])
def delete_abbreviation(abbreviation_id):
    """Delete an abbreviation"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM abbreviation_standards WHERE abbreviation_id = %s", (abbreviation_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/abbreviations/import-csv', methods=['POST'])
def import_abbreviations_csv():
    """Import abbreviations from CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Read CSV file
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported_count = 0
        updated_count = 0
        
        with get_db() as conn:
            with conn.cursor() as cur:
                for row in csv_reader:
                    # Map CSV columns to database columns
                    # CSV: Category, Full_Term, Abbreviation, Context_Usage_Notes
                    # DB: category, discipline, abbreviation, full_text, context_usage_notes
                    
                    category = row.get('Category', '').strip()
                    abbreviation_text = row.get('Abbreviation', '').strip()
                    full_text = row.get('Full_Term', '').strip()
                    context_notes = row.get('Context_Usage_Notes', '').strip()
                    
                    if not abbreviation_text:
                        continue
                    
                    # Check if abbreviation already exists
                    cur.execute("""
                        SELECT abbreviation_id FROM abbreviation_standards 
                        WHERE LOWER(abbreviation) = LOWER(%s) AND LOWER(category) = LOWER(%s)
                    """, (abbreviation_text, category))
                    
                    existing = cur.fetchone()
                    
                    if existing:
                        # Update existing record
                        cur.execute("""
                            UPDATE abbreviation_standards
                            SET full_text = %s, context_usage = %s
                            WHERE abbreviation_id = %s
                        """, (full_text, context_notes, existing[0]))
                        updated_count += 1
                    else:
                        # Insert new record (discipline defaults to 'civil')
                        cur.execute("""
                            INSERT INTO abbreviation_standards (category, discipline, abbreviation, full_text, context_usage)
                            VALUES (%s, %s, %s, %s, %s)
                        """, (category, 'CIVIL', abbreviation_text, full_text, context_notes))
                        imported_count += 1
                
                conn.commit()
        
        cache.clear()
        return jsonify({
            'success': True, 
            'imported': imported_count, 
            'updated': updated_count,
            'total': imported_count + updated_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/abbreviations/export-csv', methods=['GET'])
def export_abbreviations_csv():
    """Export abbreviations to CSV file"""
    try:
        query = """
            SELECT category, abbreviation, full_text, context_usage_notes
            FROM abbreviation_standards
            ORDER BY category, abbreviation
        """
        data = execute_query(query)
        
        # Create CSV in memory
        output = io.StringIO()
        if data:
            fieldnames = ['Category', 'Abbreviation', 'Full_Term', 'Context_Usage_Notes']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for row in data:
                writer.writerow({
                    'Category': row.get('category', ''),
                    'Abbreviation': row.get('abbreviation', ''),
                    'Full_Term': row.get('full_text', ''),
                    'Context_Usage_Notes': row.get('context_usage_notes', '')
                })
        
        # Convert to bytes
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='abbreviations.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# LAYERS MANAGER ROUTES
# ============================================================================

@app.route('/data-manager/layers')
def layers_manager():
    """Render the Layers Manager page"""
    return render_template('data_manager/layers.html')

@app.route('/api/data-manager/layers', methods=['GET'])
def get_layers():
    """Get all layers"""
    try:
        query = """
            SELECT layer_id, category, discipline, layer_name, color_rgb, 
                   description, linetype, lineweight
            FROM layer_standards
            ORDER BY category, layer_name
        """
        layers = execute_query(query)
        return jsonify({'layers': layers})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/layers', methods=['POST'])
def create_layer():
    """Create a new layer"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO layer_standards 
                    (category, discipline, layer_name, color_rgb, description)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING layer_id
                """, (
                    data.get('category'),
                    data.get('discipline', 'CIVIL'),
                    data.get('layer_name'),
                    data.get('color_rgb'),
                    data.get('description')
                ))
                layer_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'layer_id': layer_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/layers/<layer_id>', methods=['PUT'])
def update_layer(layer_id):
    """Update an existing layer"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE layer_standards
                    SET category = %s, layer_name = %s, color_rgb = %s, description = %s
                    WHERE layer_id = %s
                """, (
                    data.get('category'),
                    data.get('layer_name'),
                    data.get('color_rgb'),
                    data.get('description'),
                    layer_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/layers/<layer_id>', methods=['DELETE'])
def delete_layer(layer_id):
    """Delete a layer"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM layer_standards WHERE layer_id = %s", (layer_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/layers/import-csv', methods=['POST'])
def import_layers_csv():
    """Import layers from CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Read CSV file
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported_count = 0
        updated_count = 0
        
        with get_db() as conn:
            with conn.cursor() as cur:
                for row in csv_reader:
                    # Map CSV columns to database columns
                    # CSV: Category, Layer_Name, Color_RGB, Color_Name, Description
                    # DB: category, discipline, layer_name, color_rgb, color_name, description
                    
                    category = row.get('Category', '').strip()
                    layer_name = row.get('Layer_Name', '').strip()
                    color_rgb = row.get('Color_RGB', '').strip()
                    color_name = row.get('Color_Name', '').strip()
                    description = row.get('Description', '').strip()
                    
                    if not layer_name:
                        continue
                    
                    # Check if layer already exists
                    cur.execute("""
                        SELECT layer_id FROM layer_standards 
                        WHERE LOWER(layer_name) = LOWER(%s)
                    """, (layer_name,))
                    
                    existing = cur.fetchone()
                    
                    if existing:
                        # Update existing record
                        cur.execute("""
                            UPDATE layer_standards
                            SET category = %s, color_rgb = %s, description = %s
                            WHERE layer_id = %s
                        """, (category, color_rgb, description, existing[0]))
                        updated_count += 1
                    else:
                        # Insert new record (discipline defaults to 'CIVIL')
                        cur.execute("""
                            INSERT INTO layer_standards (category, discipline, layer_name, color_rgb, description)
                            VALUES (%s, %s, %s, %s, %s)
                        """, (category, 'CIVIL', layer_name, color_rgb, description))
                        imported_count += 1
                
                conn.commit()
        
        cache.clear()
        return jsonify({
            'imported': imported_count,
            'updated': updated_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/layers/export-csv', methods=['GET'])
def export_layers_csv():
    """Export layers to CSV file"""
    try:
        query = """
            SELECT category, layer_name, color_rgb, description
            FROM layer_standards
            ORDER BY category, layer_name
        """
        data = execute_query(query)
        
        # Create CSV in memory
        output = io.StringIO()
        if data:
            fieldnames = ['Category', 'Layer_Name', 'Color_RGB', 'Description']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for row in data:
                writer.writerow({
                    'Category': row.get('category', ''),
                    'Layer_Name': row.get('layer_name', ''),
                    'Color_RGB': row.get('color_rgb', ''),
                    'Description': row.get('description', '')
                })
        
        # Convert to bytes
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='layers.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# BLOCKS MANAGER ROUTES
# ============================================================================

@app.route('/data-manager/blocks')
def blocks_manager():
    """Render the Blocks Manager page"""
    return render_template('data_manager/blocks.html')

@app.route('/api/data-manager/blocks', methods=['GET'])
def get_blocks():
    """Get all blocks"""
    try:
        query = """
            SELECT block_id, block_name, block_type, category, description,
                   svg_content, preview_image_path
            FROM block_definitions
            ORDER BY category, block_name
        """
        blocks = execute_query(query)
        return jsonify({'blocks': blocks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/blocks', methods=['POST'])
def create_block():
    """Create a new block"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO block_definitions 
                    (block_name, category, block_type, description)
                    VALUES (%s, %s, %s, %s)
                    RETURNING block_id
                """, (
                    data.get('block_name'),
                    data.get('category'),
                    data.get('block_type'),
                    data.get('description')
                ))
                block_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'block_id': block_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/blocks/<block_id>', methods=['PUT'])
def update_block(block_id):
    """Update an existing block"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE block_definitions
                    SET block_name = %s, category = %s, block_type = %s, description = %s
                    WHERE block_id = %s
                """, (
                    data.get('block_name'),
                    data.get('category'),
                    data.get('block_type'),
                    data.get('description'),
                    block_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/blocks/<block_id>', methods=['DELETE'])
def delete_block(block_id):
    """Delete a block"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM block_definitions WHERE block_id = %s", (block_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/blocks/import-csv', methods=['POST'])
def import_blocks_csv():
    """Import blocks from CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Read CSV file
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported_count = 0
        updated_count = 0
        
        with get_db() as conn:
            with conn.cursor() as cur:
                for row in csv_reader:
                    # Map CSV columns to database columns
                    # CSV: Block_Name, Category, Block_Type, Description
                    # DB: block_name, category, block_type, description
                    
                    block_name = row.get('Block_Name', '').strip()
                    category = row.get('Category', '').strip()
                    block_type = row.get('Block_Type', '').strip()
                    description = row.get('Description', '').strip()
                    
                    if not block_name:
                        continue
                    
                    # Check if block already exists
                    cur.execute("""
                        SELECT block_id FROM block_definitions 
                        WHERE LOWER(block_name) = LOWER(%s)
                    """, (block_name,))
                    
                    existing = cur.fetchone()
                    
                    if existing:
                        # Update existing record
                        cur.execute("""
                            UPDATE block_definitions
                            SET category = %s, block_type = %s, description = %s
                            WHERE block_id = %s
                        """, (category, block_type, description, existing[0]))
                        updated_count += 1
                    else:
                        # Insert new record
                        cur.execute("""
                            INSERT INTO block_definitions (block_name, category, block_type, description)
                            VALUES (%s, %s, %s, %s)
                        """, (block_name, category, block_type, description))
                        imported_count += 1
                
                conn.commit()
        
        cache.clear()
        return jsonify({
            'imported': imported_count,
            'updated': updated_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/blocks/export-csv', methods=['GET'])
def export_blocks_csv():
    """Export blocks to CSV file"""
    try:
        query = """
            SELECT block_name, category, block_type, description
            FROM block_definitions
            ORDER BY category, block_name
        """
        data = execute_query(query)
        
        # Create CSV in memory
        output = io.StringIO()
        if data:
            fieldnames = ['Block_Name', 'Category', 'Block_Type', 'Description']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for row in data:
                writer.writerow({
                    'Block_Name': row.get('block_name', ''),
                    'Category': row.get('category', ''),
                    'Block_Type': row.get('block_type', ''),
                    'Description': row.get('description', '')
                })
        
        # Convert to bytes
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='blocks.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-block-import/extract-from-dxf', methods=['POST'])
def extract_blocks_from_dxf():
    """Extract block definitions from uploaded DXF files"""
    try:
        if 'files[]' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files[]')
        if not files or len(files) == 0:
            return jsonify({'error': 'No files selected'}), 400
        
        import ezdxf
        from batch_block_extractor import BatchBlockExtractor
        
        extractor = BatchBlockExtractor(DB_CONFIG)
        extracted_blocks = []
        errors = []
        
        for file in files:
            if file.filename == '':
                continue
                
            if not file.filename.lower().endswith('.dxf'):
                errors.append(f"{file.filename}: Not a DXF file")
                continue
            
            try:
                temp_path = os.path.join(tempfile.gettempdir(), secure_filename(file.filename))
                file.save(temp_path)
                
                blocks = extractor.extract_blocks_from_file(temp_path, file.filename)
                extracted_blocks.extend(blocks)
                
                os.remove(temp_path)
                
            except Exception as e:
                errors.append(f"{file.filename}: {str(e)}")
        
        return jsonify({
            'blocks': extracted_blocks,
            'total_files': len(files),
            'total_blocks': len(extracted_blocks),
            'errors': errors,
            'valid_categories': extractor.get_valid_categories()
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-block-import/save-blocks', methods=['POST'])
def save_extracted_blocks():
    """Save extracted blocks to the database"""
    try:
        data = request.get_json()
        blocks = data.get('blocks', [])
        
        if not blocks:
            return jsonify({'error': 'No blocks provided'}), 400
        
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        with get_db() as conn:
            with conn.cursor() as cur:
                for block in blocks:
                    action = block.get('action', 'skip')
                    if action == 'skip':
                        skipped_count += 1
                        continue
                    
                    block_name = block.get('name')
                    category = block.get('category', '')
                    description = block.get('description', '')
                    svg_content = block.get('svg_preview', '')
                    
                    cur.execute("""
                        SELECT block_id FROM block_definitions
                        WHERE LOWER(block_name) = LOWER(%s)
                    """, (block_name,))
                    
                    existing = cur.fetchone()
                    
                    if action == 'update' and existing:
                        cur.execute("""
                            UPDATE block_definitions
                            SET category = %s, description = %s, svg_content = %s,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE block_id = %s
                        """, (category, description, svg_content, existing[0]))
                        updated_count += 1
                        
                    elif action == 'import' and not existing:
                        cur.execute("""
                            INSERT INTO block_definitions 
                            (block_name, category, description, svg_content, is_active)
                            VALUES (%s, %s, %s, %s, TRUE)
                        """, (block_name, category, description, svg_content))
                        imported_count += 1
                    else:
                        skipped_count += 1
                
                conn.commit()
        
        cache.clear()
        return jsonify({
            'imported': imported_count,
            'updated': updated_count,
            'skipped': skipped_count
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# UNIFIED BATCH CAD IMPORT ROUTES (Blocks, Details, Hatches, Linetypes)
# ============================================================================

@app.route('/api/batch-cad-import/extract', methods=['POST'])
def extract_cad_elements():
    """Extract CAD elements (blocks, details, hatches, linetypes) from uploaded DXF files"""
    try:
        import_type = request.form.get('import_type', 'blocks')
        
        if 'files[]' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files[]')
        if not files or len(files) == 0:
            return jsonify({'error': 'No files selected'}), 400
        
        from batch_cad_extractor import BatchCADExtractor
        
        extractor = BatchCADExtractor(DB_CONFIG)
        extracted_elements = []
        errors = []
        
        for file in files:
            if file.filename == '':
                continue
                
            if not file.filename.lower().endswith('.dxf'):
                errors.append(f"{file.filename}: Not a DXF file")
                continue
            
            try:
                temp_path = os.path.join(tempfile.gettempdir(), secure_filename(file.filename))
                file.save(temp_path)
                
                elements = extractor.extract_from_file(temp_path, file.filename, import_type)
                extracted_elements.extend(elements)
                
                os.remove(temp_path)
                
            except Exception as e:
                errors.append(f"{file.filename}: {str(e)}")
        
        metadata = extractor.get_metadata(import_type)
        
        return jsonify({
            'elements': extracted_elements,
            'total_files': len(files),
            'total_elements': len(extracted_elements),
            'errors': errors,
            'import_type': import_type,
            'metadata': metadata
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-cad-import/save', methods=['POST'])
def save_cad_elements():
    """Save extracted CAD elements to appropriate database tables"""
    try:
        data = request.get_json()
        import_type = data.get('import_type', 'blocks')
        elements = data.get('elements', [])
        
        if not elements:
            return jsonify({'error': 'No elements provided'}), 400
        
        # Dispatch to appropriate save handler
        save_handlers = {
            'blocks': _save_blocks,
            'details': _save_details,
            'hatches': _save_hatches,
            'linetypes': _save_linetypes
        }
        
        if import_type not in save_handlers:
            return jsonify({'error': f'Invalid import type: {import_type}'}), 400
        
        result = save_handlers[import_type](elements)
        cache.clear()
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _save_blocks(elements):
    """Save blocks to block_definitions table"""
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    
    with get_db() as conn:
        with conn.cursor() as cur:
            for elem in elements:
                action = elem.get('action', 'skip')
                if action == 'skip':
                    skipped_count += 1
                    continue
                
                name = elem.get('name')
                category = elem.get('category', '')
                description = elem.get('description', '')
                svg_content = elem.get('svg_preview', '')
                
                cur.execute("""
                    SELECT block_id FROM block_definitions
                    WHERE LOWER(block_name) = LOWER(%s)
                """, (name,))
                
                existing = cur.fetchone()
                
                if action == 'update' and existing:
                    cur.execute("""
                        UPDATE block_definitions
                        SET category = %s, description = %s, svg_content = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE block_id = %s
                    """, (category, description, svg_content, existing[0]))
                    updated_count += 1
                    
                elif action == 'import' and not existing:
                    cur.execute("""
                        INSERT INTO block_definitions 
                        (block_name, category, description, svg_content, is_active)
                        VALUES (%s, %s, %s, %s, TRUE)
                    """, (name, category, description, svg_content))
                    imported_count += 1
                else:
                    skipped_count += 1
            
            conn.commit()
    
    return {
        'imported': imported_count,
        'updated': updated_count,
        'skipped': skipped_count
    }

def _save_details(elements):
    """Save details to detail_standards table"""
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    
    with get_db() as conn:
        with conn.cursor() as cur:
            for elem in elements:
                action = elem.get('action', 'skip')
                if action == 'skip':
                    skipped_count += 1
                    continue
                
                name = elem.get('name')
                detail_category = elem.get('detail_category', '')
                discipline = elem.get('discipline', '')
                description = elem.get('description', '')
                svg_content = elem.get('svg_preview', '')
                
                cur.execute("""
                    SELECT detail_id FROM detail_standards
                    WHERE LOWER(detail_number) = LOWER(%s)
                """, (name,))
                
                existing = cur.fetchone()
                
                if action == 'update' and existing:
                    cur.execute("""
                        UPDATE detail_standards
                        SET detail_category = %s, discipline = %s, 
                            description = %s, svg_content = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE detail_id = %s
                    """, (detail_category, discipline, description, svg_content, existing[0]))
                    updated_count += 1
                    
                elif action == 'import' and not existing:
                    cur.execute("""
                        INSERT INTO detail_standards 
                        (detail_number, detail_category, discipline, description, svg_content)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (name, detail_category, discipline, description, svg_content))
                    imported_count += 1
                else:
                    skipped_count += 1
            
            conn.commit()
    
    return {
        'imported': imported_count,
        'updated': updated_count,
        'skipped': skipped_count
    }

def _save_hatches(elements):
    """Save hatches to hatch_patterns table"""
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    
    with get_db() as conn:
        with conn.cursor() as cur:
            for elem in elements:
                action = elem.get('action', 'skip')
                if action == 'skip':
                    skipped_count += 1
                    continue
                
                name = elem.get('name')
                pattern_type = elem.get('pattern_type', 'User-defined')
                description = elem.get('description', '')
                
                cur.execute("""
                    SELECT hatch_id FROM hatch_patterns
                    WHERE LOWER(pattern_name) = LOWER(%s)
                """, (name,))
                
                existing = cur.fetchone()
                
                if action == 'update' and existing:
                    cur.execute("""
                        UPDATE hatch_patterns
                        SET pattern_type = %s, description = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE hatch_id = %s
                    """, (pattern_type, description, existing[0]))
                    updated_count += 1
                    
                elif action == 'import' and not existing:
                    cur.execute("""
                        INSERT INTO hatch_patterns 
                        (pattern_name, pattern_type, description, is_active)
                        VALUES (%s, %s, %s, TRUE)
                    """, (name, pattern_type, description))
                    imported_count += 1
                else:
                    skipped_count += 1
            
            conn.commit()
    
    return {
        'imported': imported_count,
        'updated': updated_count,
        'skipped': skipped_count
    }

def _save_linetypes(elements):
    """Save linetypes to linetypes table"""
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    
    with get_db() as conn:
        with conn.cursor() as cur:
            for elem in elements:
                action = elem.get('action', 'skip')
                if action == 'skip':
                    skipped_count += 1
                    continue
                
                name = elem.get('name')
                description = elem.get('description', '')
                pattern_definition = elem.get('pattern_definition', '')
                
                cur.execute("""
                    SELECT linetype_id FROM linetypes
                    WHERE LOWER(linetype_name) = LOWER(%s)
                """, (name,))
                
                existing = cur.fetchone()
                
                if action == 'update' and existing:
                    cur.execute("""
                        UPDATE linetypes
                        SET description = %s, pattern_definition = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE linetype_id = %s
                    """, (description, pattern_definition, existing[0]))
                    updated_count += 1
                    
                elif action == 'import' and not existing:
                    cur.execute("""
                        INSERT INTO linetypes 
                        (linetype_name, description, pattern_definition, is_active)
                        VALUES (%s, %s, %s, TRUE)
                    """, (name, description, pattern_definition))
                    imported_count += 1
                else:
                    skipped_count += 1
            
            conn.commit()
    
    return {
        'imported': imported_count,
        'updated': updated_count,
        'skipped': skipped_count
    }

# ============================================================================
# BATCH POINT IMPORT ROUTES
# ============================================================================

def get_batch_import_db_config():
    """
    Get database config for batch point import.
    Uses DATABASE_URL (Neon) instead of DB_* variables (Supabase).
    """
    from urllib.parse import urlparse
    
    database_url = os.getenv('DATABASE_URL')
    if database_url:
        parsed = urlparse(database_url)
        return {
            'host': parsed.hostname,
            'port': parsed.port or 5432,
            'database': parsed.path.lstrip('/').split('?')[0],
            'user': parsed.username,
            'password': parsed.password
        }
    else:
        # Fallback to individual variables
        return {
            'host': os.getenv('DB_HOST', 'localhost'),
            'port': os.getenv('DB_PORT', '5432'),
            'database': os.getenv('DB_NAME', 'acad_gis'),
            'user': os.getenv('DB_USER', 'postgres'),
            'password': os.getenv('DB_PASSWORD', '')
        }

@app.route('/api/batch-point-import/get-projects', methods=['GET'])
def get_projects_for_point_import():
    """Get list of projects for point import selection"""
    try:
        from batch_pnezd_parser import PNEZDParser
        import traceback
        
        db_config = get_batch_import_db_config()
        parser = PNEZDParser(db_config)
        result = parser.get_projects()
        
        return jsonify(result)
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-point-import/get-coordinate-systems', methods=['GET'])
def get_coordinate_systems_for_point_import():
    """Get list of available coordinate systems for point import"""
    try:
        from batch_pnezd_parser import PNEZDParser
        
        db_config = get_batch_import_db_config()
        parser = PNEZDParser(db_config)
        systems = parser.get_coordinate_systems()
        
        return jsonify({'coordinate_systems': systems})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-point-import/parse', methods=['POST'])
def parse_pnezd_file():
    """Parse PNEZD text file and extract point data"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        project_id = request.form.get('project_id')
        epsg_code = request.form.get('epsg_code')
        
        if not project_id or not epsg_code:
            return jsonify({'error': 'Project ID and EPSG code are required'}), 400
        
        from batch_pnezd_parser import PNEZDParser
        
        db_config = get_batch_import_db_config()
        parser = PNEZDParser(db_config)
        
        # Read file content
        file_content = file.read().decode('utf-8')
        
        # Parse file
        result = parser.parse_file(file_content, file.filename)
        
        # Check for existing points in this project
        if result['points']:
            result['points'] = parser.check_existing_points(result['points'], project_id)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-point-import/save', methods=['POST'])
def save_survey_points():
    """Save parsed survey points to database"""
    try:
        data = request.get_json()
        points = data.get('points', [])
        project_id = data.get('project_id')
        source_epsg = data.get('source_epsg')
        
        if not points:
            return jsonify({'error': 'No points provided'}), 400
        
        if not project_id:
            return jsonify({'error': 'Project ID is required'}), 400
        
        if not source_epsg:
            return jsonify({'error': 'Source EPSG code is required'}), 400
        
        from pyproj import Transformer
        
        # Get target EPSG (database uses SRID 2226)
        target_epsg = '2226'
        
        # Normalize EPSG codes: extract numeric code from formats like "EPSG:2226" or "2226"
        # Only strip leading "EPSG:" prefix (case-insensitive), not all occurrences
        import re
        
        def normalize_epsg(code: str) -> str:
            """
            Extract numeric EPSG code from various formats:
            - "2226" → "2226"
            - "EPSG:2226" → "2226"
            - "EPSG:EPSG:2226" → "2226" (handles double-prefix bug)
            - "urn:ogc:def:crs:EPSG::4326" → "4326"
            
            Returns the numeric code or empty string if invalid
            """
            if not code or not code.strip():
                return ''
            
            code = code.strip()
            
            # If it's already just digits, return it
            if code.isdigit():
                return code
            
            # Split by colon and look for numeric segments from right to left
            # This handles: "EPSG:2226", "EPSG:EPSG:2226", "urn:ogc:def:crs:EPSG::4326"
            parts = code.split(':')
            for part in reversed(parts):
                part = part.strip()
                if part.isdigit():
                    return part
            
            # No numeric code found
            return ''
        
        source_epsg_clean = normalize_epsg(source_epsg)
        target_epsg_clean = normalize_epsg(target_epsg)
        
        # Validate EPSG codes are non-empty
        if not source_epsg_clean:
            return jsonify({'error': f'Invalid source coordinate system code: "{source_epsg}"'}), 400
        if not target_epsg_clean:
            return jsonify({'error': f'Invalid target coordinate system code: "{target_epsg}"'}), 400
        
        # Create coordinate transformer if coordinate systems differ
        transformer = None
        if source_epsg_clean != target_epsg_clean:
            try:
                transformer = Transformer.from_crs(
                    f'EPSG:{source_epsg_clean}',
                    f'EPSG:{target_epsg_clean}',
                    always_xy=True
                )
            except Exception as e:
                return jsonify({'error': f'Invalid coordinate system transformation: {str(e)}'}), 400
        
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        with get_db() as conn:
            with conn.cursor() as cur:
                for point in points:
                    action = point.get('action', 'skip')
                    if action == 'skip':
                        skipped_count += 1
                        continue
                    
                    point_number = point['point_number']
                    northing = float(point['northing'])
                    easting = float(point['easting'])
                    elevation = float(point['elevation'])
                    description = point['description']
                    
                    # Transform coordinates if necessary
                    if transformer:
                        easting_transformed, northing_transformed = transformer.transform(easting, northing)
                    else:
                        easting_transformed = easting
                        northing_transformed = northing
                    
                    # Check if point already exists
                    cur.execute("""
                        SELECT point_id FROM survey_points
                        WHERE project_id = %s AND point_number = %s
                    """, (project_id, point_number))
                    
                    existing = cur.fetchone()
                    
                    if action == 'update' and existing:
                        # Update existing point
                        cur.execute("""
                            UPDATE survey_points
                            SET northing = %s,
                                easting = %s,
                                elevation = %s,
                                point_description = %s,
                                geometry = ST_SetSRID(ST_MakePoint(%s, %s, %s), 2226),
                                coordinate_system = %s,
                                epsg_code = %s,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE point_id = %s
                        """, (
                            northing_transformed,
                            easting_transformed,
                            elevation,
                            description,
                            easting_transformed,
                            northing_transformed,
                            elevation,
                            f'EPSG:{target_epsg}',
                            target_epsg,
                            existing[0]
                        ))
                        updated_count += 1
                        
                    elif action == 'import' and not existing:
                        # Insert new point
                        cur.execute("""
                            INSERT INTO survey_points (
                                project_id,
                                point_number,
                                point_description,
                                northing,
                                easting,
                                elevation,
                                geometry,
                                coordinate_system,
                                epsg_code,
                                is_active
                            ) VALUES (%s, %s, %s, %s, %s, %s, 
                                      ST_SetSRID(ST_MakePoint(%s, %s, %s), 2226),
                                      %s, %s, TRUE)
                        """, (
                            project_id,
                            point_number,
                            description,
                            northing_transformed,
                            easting_transformed,
                            elevation,
                            easting_transformed,
                            northing_transformed,
                            elevation,
                            f'EPSG:{target_epsg}',
                            target_epsg
                        ))
                        imported_count += 1
                    else:
                        skipped_count += 1
                
                conn.commit()
        
        cache.clear()
        
        return jsonify({
            'imported': imported_count,
            'updated': updated_count,
            'skipped': skipped_count
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# DETAILS MANAGER ROUTES
# ============================================================================

@app.route('/data-manager/details')
def details_manager():
    """Render the Details Manager page"""
    return render_template('data_manager/details.html')

@app.route('/data-manager/standard-notes')
def standard_notes_manager():
    """Render the Standard Notes Manager page"""
    return render_template('data_manager/standard_notes.html')

@app.route('/data-manager/materials')
def materials_manager():
    """Render the Materials Manager page"""
    return render_template('data_manager/materials.html')

@app.route('/data-manager/projects')
def projects_manager():
    """Redirect to new Projects page location"""
    return redirect(url_for('projects_page'), code=301)

@app.route('/data-manager/hatches')
def hatches_manager():
    """Render the Hatches Manager page"""
    return render_template('data_manager/hatches.html')

@app.route('/data-manager/linetypes')
def linetypes_manager():
    """Render the Linetypes Manager page"""
    return render_template('data_manager/linetypes.html')

@app.route('/data-manager/text-styles')
def text_styles_manager():
    """Render the Text Styles Manager page"""
    return render_template('data_manager/text-styles.html')

@app.route('/data-manager/dimension-styles')
def dimension_styles_manager():
    """Render the Dimension Styles Manager page"""
    return render_template('data_manager/dimension-styles.html')

@app.route('/usage-dashboard')
def usage_dashboard():
    """Render the Usage Tracking Dashboard page"""
    return render_template('usage_dashboard.html')

@app.route('/api/data-manager/details', methods=['GET'])
def get_details():
    """Get all details"""
    try:
        query = """
            SELECT detail_id, detail_number, detail_title, detail_category, 
                   description, usage_context
            FROM detail_standards
            ORDER BY detail_category, detail_number
        """
        details = execute_query(query)
        return jsonify({'details': details})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/details', methods=['POST'])
def create_detail():
    """Create a new detail"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO detail_standards 
                    (detail_number, detail_title, detail_category, description, usage_context)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING detail_id
                """, (
                    data.get('detail_number'),
                    data.get('detail_title'),
                    data.get('detail_category'),
                    data.get('description'),
                    data.get('usage_context')
                ))
                detail_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'detail_id': detail_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/details/<detail_id>', methods=['PUT'])
def update_detail(detail_id):
    """Update an existing detail"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE detail_standards
                    SET detail_number = %s, detail_title = %s, detail_category = %s, 
                        description = %s, usage_context = %s
                    WHERE detail_id = %s
                """, (
                    data.get('detail_number'),
                    data.get('detail_title'),
                    data.get('detail_category'),
                    data.get('description'),
                    data.get('usage_context'),
                    detail_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/details/<detail_id>', methods=['DELETE'])
def delete_detail(detail_id):
    """Delete a detail"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM detail_standards WHERE detail_id = %s", (detail_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/details/import-csv', methods=['POST'])
def import_details_csv():
    """Import details from CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Read CSV file
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported_count = 0
        updated_count = 0
        
        with get_db() as conn:
            with conn.cursor() as cur:
                for row in csv_reader:
                    # Map CSV columns to database columns
                    # CSV: Detail_Number, Detail_Title, Category, Scale, Description
                    # DB: detail_number, detail_title, detail_category, scale, description
                    
                    detail_number = row.get('Detail_Number', '').strip()
                    detail_title = row.get('Detail_Title', '').strip()
                    category = row.get('Category', '').strip()
                    scale = row.get('Scale', '').strip()
                    description = row.get('Description', '').strip()
                    
                    if not detail_number:
                        continue
                    
                    # Check if detail already exists
                    cur.execute("""
                        SELECT detail_id FROM detail_standards 
                        WHERE LOWER(detail_number) = LOWER(%s)
                    """, (detail_number,))
                    
                    existing = cur.fetchone()
                    
                    if existing:
                        # Update existing record
                        cur.execute("""
                            UPDATE detail_standards
                            SET detail_title = %s, detail_category = %s, description = %s
                            WHERE detail_id = %s
                        """, (detail_title, category, description, existing[0]))
                        updated_count += 1
                    else:
                        # Insert new record
                        cur.execute("""
                            INSERT INTO detail_standards (detail_number, detail_title, detail_category, description)
                            VALUES (%s, %s, %s, %s)
                        """, (detail_number, detail_title, category, description))
                        imported_count += 1
                
                conn.commit()
        
        cache.clear()
        return jsonify({
            'imported': imported_count,
            'updated': updated_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/details/export-csv', methods=['GET'])
def export_details_csv():
    """Export details to CSV file"""
    try:
        query = """
            SELECT detail_number, detail_title, detail_category, description
            FROM detail_standards
            ORDER BY detail_category, detail_number
        """
        data = execute_query(query)
        
        # Create CSV in memory
        output = io.StringIO()
        if data:
            fieldnames = ['Detail_Number', 'Detail_Title', 'Category', 'Description']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for row in data:
                writer.writerow({
                    'Detail_Number': row.get('detail_number', ''),
                    'Detail_Title': row.get('detail_title', ''),
                    'Category': row.get('detail_category', ''),
                    'Description': row.get('description', '')
                })
        
        # Convert to bytes
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='details.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# STANDARD NOTES MANAGER API ENDPOINTS
# ============================================================================

@app.route('/api/data-manager/standard-notes', methods=['GET'])
def get_standard_notes_manager():
    """Get all standard notes"""
    try:
        query = """
            SELECT note_id, note_title, note_text, note_category, discipline, 
                   tags, usage_frequency, created_at, updated_at
            FROM standard_notes
            ORDER BY note_category, note_title
        """
        notes = execute_query(query)
        return jsonify({'notes': notes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/standard-notes', methods=['POST'])
def create_standard_note():
    """Create a new standard note"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO standard_notes 
                    (note_title, note_text, note_category, discipline, tags)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING note_id
                """, (
                    data.get('note_title'),
                    data.get('note_text'),
                    data.get('note_category'),
                    data.get('discipline'),
                    data.get('tags', [])
                ))
                note_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'note_id': note_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/standard-notes/<note_id>', methods=['PUT'])
def update_standard_note(note_id):
    """Update an existing standard note"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE standard_notes
                    SET note_title = %s, note_text = %s, note_category = %s, 
                        discipline = %s, tags = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE note_id = %s
                """, (
                    data.get('note_title'),
                    data.get('note_text'),
                    data.get('note_category'),
                    data.get('discipline'),
                    data.get('tags', []),
                    note_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/standard-notes/<note_id>', methods=['DELETE'])
def delete_standard_note(note_id):
    """Delete a standard note"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM standard_notes WHERE note_id = %s", (note_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# MATERIALS MANAGER API ENDPOINTS
# ============================================================================

@app.route('/api/data-manager/materials', methods=['GET'])
def get_materials():
    """Get all materials"""
    try:
        query = """
            SELECT material_id, material_name, material_type, description, 
                   specifications, manufacturer, product_code, cost_per_unit,
                   unit_of_measure, environmental_rating, usage_frequency
            FROM material_standards
            ORDER BY material_type, material_name
        """
        materials = execute_query(query)
        return jsonify({'materials': materials})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/materials', methods=['POST'])
def create_material():
    """Create a new material"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO material_standards 
                    (material_name, material_type, description, specifications, manufacturer,
                     product_code, cost_per_unit, unit_of_measure, environmental_rating)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING material_id
                """, (
                    data.get('material_name'),
                    data.get('material_type'),
                    data.get('description'),
                    data.get('specifications'),
                    data.get('manufacturer'),
                    data.get('product_code'),
                    data.get('cost_per_unit'),
                    data.get('unit_of_measure'),
                    data.get('environmental_rating')
                ))
                material_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'material_id': material_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/materials/<material_id>', methods=['PUT'])
def update_material(material_id):
    """Update an existing material"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE material_standards
                    SET material_name = %s, material_type = %s, description = %s,
                        specifications = %s, manufacturer = %s, product_code = %s,
                        cost_per_unit = %s, unit_of_measure = %s, environmental_rating = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE material_id = %s
                """, (
                    data.get('material_name'),
                    data.get('material_type'),
                    data.get('description'),
                    data.get('specifications'),
                    data.get('manufacturer'),
                    data.get('product_code'),
                    data.get('cost_per_unit'),
                    data.get('unit_of_measure'),
                    data.get('environmental_rating'),
                    material_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/materials/<material_id>', methods=['DELETE'])
def delete_material(material_id):
    """Delete a material"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM material_standards WHERE material_id = %s", (material_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# PROJECT STANDARDS ASSIGNMENT API ENDPOINTS
# ============================================================================

@app.route('/api/project-standards/<project_id>', methods=['GET'])
def get_project_standards(project_id):
    """Get all assigned standards for a project grouped by type"""
    try:
        query = """
            SELECT 
                psa.assignment_id,
                psa.project_id,
                psa.standard_type,
                psa.standard_id,
                psa.is_required,
                psa.notes,
                psa.assigned_date,
                CASE 
                    WHEN psa.standard_type = 'layer' THEN l.layer_name
                    WHEN psa.standard_type = 'block' THEN b.block_name
                    WHEN psa.standard_type = 'hatch' THEN hp.pattern_name
                    WHEN psa.standard_type = 'linetype' THEN lt.linetype_name
                    WHEN psa.standard_type = 'text_style' THEN ts.style_name
                    WHEN psa.standard_type = 'dimension_style' THEN ds.style_name
                    WHEN psa.standard_type = 'material' THEN ms.material_name
                    WHEN psa.standard_type = 'detail' THEN d.detail_title
                    WHEN psa.standard_type = 'standard_note' THEN LEFT(sn.note_text, 50)
                END as standard_name,
                CASE 
                    WHEN psa.standard_type = 'layer' THEN l.description
                    WHEN psa.standard_type = 'block' THEN b.description
                    WHEN psa.standard_type = 'hatch' THEN hp.description
                    WHEN psa.standard_type = 'linetype' THEN lt.description
                    WHEN psa.standard_type = 'text_style' THEN ts.description
                    WHEN psa.standard_type = 'dimension_style' THEN ds.description
                    WHEN psa.standard_type = 'material' THEN ms.description
                    WHEN psa.standard_type = 'detail' THEN d.description
                    WHEN psa.standard_type = 'standard_note' THEN sn.note_category
                END as standard_description
            FROM project_standard_assignments psa
            LEFT JOIN layer_standards l ON psa.standard_type = 'layer' AND psa.standard_id::uuid = l.layer_id
            LEFT JOIN block_definitions b ON psa.standard_type = 'block' AND psa.standard_id::uuid = b.block_id
            LEFT JOIN hatch_patterns hp ON psa.standard_type = 'hatch' AND psa.standard_id::uuid = hp.hatch_id
            LEFT JOIN linetypes lt ON psa.standard_type = 'linetype' AND psa.standard_id::uuid = lt.linetype_id
            LEFT JOIN text_styles ts ON psa.standard_type = 'text_style' AND psa.standard_id::uuid = ts.style_id
            LEFT JOIN dimension_styles ds ON psa.standard_type = 'dimension_style' AND psa.standard_id::uuid = ds.dimension_style_id
            LEFT JOIN material_standards ms ON psa.standard_type = 'material' AND psa.standard_id::uuid = ms.material_id
            LEFT JOIN detail_standards d ON psa.standard_type = 'detail' AND psa.standard_id::uuid = d.detail_id
            LEFT JOIN standard_notes sn ON psa.standard_type = 'standard_note' AND psa.standard_id::uuid = sn.note_id
            WHERE psa.project_id = %s
            ORDER BY psa.standard_type, standard_name
        """
        standards = execute_query(query, (project_id,))
        
        grouped = {}
        for standard in standards:
            std_type = standard['standard_type']
            if std_type not in grouped:
                grouped[std_type] = []
            grouped[std_type].append(standard)
        
        return jsonify({'standards': grouped})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-standards', methods=['POST'])
def add_project_standard():
    """Add a standard to a project"""
    try:
        data = request.get_json()
        
        project_id = data.get('project_id')
        standard_type = data.get('standard_type')
        standard_id = data.get('standard_id')
        is_required = data.get('is_required', False)
        notes = data.get('notes')
        
        if not project_id or not standard_type or not standard_id:
            return jsonify({'error': 'project_id, standard_type, and standard_id are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO project_standard_assignments 
                    (project_id, standard_type, standard_id, is_required, notes)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING assignment_id
                """, (
                    project_id,
                    standard_type,
                    standard_id,
                    is_required,
                    notes
                ))
                assignment_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'assignment_id': assignment_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-standards/<assignment_id>', methods=['DELETE'])
def remove_project_standard(assignment_id):
    """Remove a standard assignment from a project"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_standard_assignments WHERE assignment_id = %s", (assignment_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards-by-type/<standard_type>', methods=['GET'])
def get_standards_by_type(standard_type):
    """Get all available standards of a specific type"""
    try:
        type_mapping = {
            'layer': ('layers', 'layer_id', 'layer_name', 'description'),
            'block': ('blocks', 'block_id', 'block_name', 'description'),
            'hatch': ('hatch_patterns', 'hatch_id', 'pattern_name', 'description'),
            'linetype': ('linetypes', 'linetype_id', 'linetype_name', 'description'),
            'text_style': ('text_styles', 'style_id', 'style_name', 'description'),
            'dimension_style': ('dimension_styles', 'dimension_style_id', 'style_name', 'description'),
            'material': ('material_standards', 'material_id', 'material_name', 'description'),
            'detail': ('details', 'detail_id', 'detail_name', 'description'),
            'standard_note': ('standard_notes', 'note_id', 'note_text', 'category')
        }
        
        if standard_type not in type_mapping:
            return jsonify({'error': 'Invalid standard type'}), 400
        
        table, id_col, name_col, desc_col = type_mapping[standard_type]
        
        if standard_type == 'standard_note':
            query = f"""
                SELECT {id_col} as id, 
                       LEFT({name_col}, 100) as name, 
                       {desc_col} as description
                FROM {table}
                ORDER BY {name_col}
                LIMIT 500
            """
        else:
            query = f"""
                SELECT {id_col} as id, {name_col} as name, {desc_col} as description
                FROM {table}
                ORDER BY {name_col}
                LIMIT 500
            """
        
        standards = execute_query(query)
        return jsonify({'standards': standards})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-compliance/<project_id>', methods=['GET'])
def get_project_compliance(project_id):
    """Get compliance summary for a project"""
    try:
        type_mapping = {
            'layer': ('layers', 'layer_id', 'layer_name'),
            'block': ('blocks', 'block_id', 'block_name'),
            'hatch': ('hatch_patterns', 'hatch_id', 'pattern_name'),
            'linetype': ('linetypes', 'linetype_id', 'linetype_name'),
            'text_style': ('text_styles', 'style_id', 'style_name'),
            'dimension_style': ('dimension_styles', 'dimension_style_id', 'style_name'),
            'material': ('material_standards', 'material_id', 'material_name'),
            'detail': ('details', 'detail_id', 'detail_name'),
            'standard_note': ('standard_notes', 'note_id', 'note_title')
        }
        
        assigned_query = """
            SELECT standard_type, COUNT(*) as count
            FROM project_standard_assignments
            WHERE project_id = %s
            GROUP BY standard_type
        """
        assigned_by_type = execute_query(assigned_query, (project_id,))
        
        deviations_query = """
            SELECT 
                pso.override_id,
                pso.project_id,
                pso.standard_type,
                pso.standard_id,
                pso.override_reason,
                pso.created_date,
                pso.created_by,
                pso.notes,
                'Standard ID: ' || pso.standard_id::text as standard_name
            FROM project_standard_overrides pso
            WHERE pso.project_id = %s
            ORDER BY pso.created_date DESC
        """
        deviations = execute_query(deviations_query, (project_id,))
        
        deviations_by_type_query = """
            SELECT standard_type, COUNT(*) as count
            FROM project_standard_overrides
            WHERE project_id = %s
            GROUP BY standard_type
        """
        deviations_by_type = execute_query(deviations_by_type_query, (project_id,))
        
        deviation_counts = {item['standard_type']: item['count'] for item in deviations_by_type}
        
        total_assigned = sum(item['count'] for item in assigned_by_type)
        total_deviations = len(deviations)
        total_used = total_assigned
        
        compliance_percentage = 0
        if total_assigned > 0:
            compliance_percentage = ((total_assigned - total_deviations) / total_assigned) * 100
        
        by_category = []
        for item in assigned_by_type:
            standard_type = item['standard_type']
            assigned_count = item['count']
            deviations_count = deviation_counts.get(standard_type, 0)
            
            by_category.append({
                'standard_type': standard_type,
                'assigned': assigned_count,
                'used': assigned_count,
                'deviations': deviations_count
            })
        
        return jsonify({
            'total_assigned': total_assigned,
            'total_used': total_used,
            'deviations_count': total_deviations,
            'compliance_percentage': compliance_percentage,
            'by_category': by_category,
            'deviations': deviations
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-compliance/<project_id>/conformance-details', methods=['GET'])
def get_project_conformance_details(project_id):
    """Get detailed conformance tracking data for a project (sheet notes only for now)"""
    try:
        # Get conformance status rollup
        conformance_rollup_query = """
            SELECT 
                cs.status_code,
                cs.status_name,
                cs.color_hex,
                COUNT(psn.conformance_status_id) as count
            FROM conformance_statuses cs
            LEFT JOIN (
                SELECT psn.conformance_status_id
                FROM project_sheet_notes psn
                JOIN sheet_note_sets sns ON psn.set_id = sns.set_id
                WHERE sns.project_id = %s::uuid
            ) psn ON cs.status_id = psn.conformance_status_id
            WHERE cs.is_active = TRUE
            GROUP BY cs.status_id, cs.status_code, cs.status_name, cs.color_hex, cs.sort_order
            ORDER BY cs.sort_order
        """
        conformance_statuses = execute_query(conformance_rollup_query, (project_id,))
        
        # Get deviation categories breakdown
        deviation_categories_query = """
            SELECT 
                dc.category_code,
                dc.category_name,
                COUNT(psn.deviation_category_id) as count,
                ARRAY_AGG(psn.display_code ORDER BY psn.display_code) FILTER (WHERE psn.display_code IS NOT NULL) as sample_codes
            FROM deviation_categories dc
            LEFT JOIN (
                SELECT psn.deviation_category_id, psn.display_code
                FROM project_sheet_notes psn
                JOIN sheet_note_sets sns ON psn.set_id = sns.set_id
                WHERE sns.project_id = %s::uuid
            ) psn ON dc.category_id = psn.deviation_category_id
            WHERE dc.is_active = TRUE
            GROUP BY dc.category_id, dc.category_code, dc.category_name, dc.sort_order
            HAVING COUNT(psn.deviation_category_id) > 0
            ORDER BY dc.sort_order
        """
        deviation_categories = execute_query(deviation_categories_query, (project_id,))
        
        # Get standardization candidates (custom notes used multiple times or nominated)
        standardization_candidates_query = """
            SELECT 
                psn.project_note_id,
                psn.custom_title,
                psn.display_code,
                psn.usage_count,
                sns.set_name,
                ss.status_name as standardization_status
            FROM project_sheet_notes psn
            JOIN sheet_note_sets sns ON psn.set_id = sns.set_id
            LEFT JOIN standardization_statuses ss ON psn.standardization_status_id = ss.status_id
            WHERE sns.project_id = %s::uuid
              AND psn.source_type = 'custom'
              AND (psn.usage_count > 2 OR ss.status_code IN ('NOMINATED', 'UNDER_REVIEW', 'APPROVED'))
            ORDER BY psn.usage_count DESC, psn.custom_title
            LIMIT 20
        """
        standardization_candidates = execute_query(standardization_candidates_query, (project_id,))
        
        # Get source type breakdown
        source_type_breakdown_query = """
            SELECT 
                psn.source_type,
                COUNT(*) as count
            FROM project_sheet_notes psn
            JOIN sheet_note_sets sns ON psn.set_id = sns.set_id
            WHERE sns.project_id = %s::uuid
            GROUP BY psn.source_type
            ORDER BY psn.source_type
        """
        source_types = execute_query(source_type_breakdown_query, (project_id,))
        
        # Calculate summary metrics
        total_notes = sum(status['count'] for status in conformance_statuses)
        total_with_deviations = sum(cat['count'] for cat in deviation_categories)
        
        full_compliance_count = next((s['count'] for s in conformance_statuses if s['status_code'] == 'FULL_COMPLIANCE'), 0)
        conformance_percentage = (full_compliance_count / total_notes * 100) if total_notes > 0 else 0
        
        # Source type counts
        standard_count = next((s['count'] for s in source_types if s['source_type'] == 'standard'), 0)
        modified_count = next((s['count'] for s in source_types if s['source_type'] == 'modified_standard'), 0)
        custom_count = next((s['count'] for s in source_types if s['source_type'] == 'custom'), 0)
        
        return jsonify({
            'summary': {
                'total_notes': total_notes,
                'total_with_deviations': total_with_deviations,
                'conformance_percentage': conformance_percentage,
                'candidates_count': len(standardization_candidates),
                'standard_count': standard_count,
                'modified_count': modified_count,
                'custom_count': custom_count
            },
            'conformance_statuses': conformance_statuses,
            'deviation_categories': deviation_categories,
            'standardization_candidates': standardization_candidates,
            'source_types': source_types
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# HATCH PATTERNS MANAGER API ENDPOINTS
# ============================================================================

@app.route('/api/data-manager/hatches', methods=['GET'])
def get_hatches():
    """Get all hatch patterns"""
    try:
        query = """
            SELECT hatch_id, pattern_name, pattern_type, pattern_definition,
                   description, usage_context
            FROM hatch_patterns
            ORDER BY pattern_type, pattern_name
        """
        hatches = execute_query(query)
        return jsonify({'hatches': hatches})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/hatches', methods=['POST'])
def create_hatch():
    """Create a new hatch pattern"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO hatch_patterns 
                    (pattern_name, pattern_type, pattern_definition, description, usage_context)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING hatch_id
                """, (
                    data.get('pattern_name'),
                    data.get('pattern_type'),
                    data.get('pattern_definition'),
                    data.get('description'),
                    data.get('usage_context')
                ))
                hatch_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'hatch_id': hatch_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/hatches/<hatch_id>', methods=['PUT'])
def update_hatch(hatch_id):
    """Update an existing hatch pattern"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE hatch_patterns
                    SET pattern_name = %s, pattern_type = %s, pattern_definition = %s,
                        description = %s, usage_context = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE hatch_id = %s
                """, (
                    data.get('pattern_name'),
                    data.get('pattern_type'),
                    data.get('pattern_definition'),
                    data.get('description'),
                    data.get('usage_context'),
                    hatch_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/hatches/<hatch_id>', methods=['DELETE'])
def delete_hatch(hatch_id):
    """Delete a hatch pattern"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM hatch_patterns WHERE hatch_id = %s", (hatch_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# LINETYPES MANAGER API ENDPOINTS
# ============================================================================

@app.route('/api/data-manager/linetypes', methods=['GET'])
def get_linetypes():
    """Get all linetypes"""
    try:
        query = """
            SELECT linetype_id, linetype_name, pattern_definition,
                   description, usage_context
            FROM linetypes
            ORDER BY linetype_name
        """
        linetypes = execute_query(query)
        return jsonify({'linetypes': linetypes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/linetypes', methods=['POST'])
def create_linetype():
    """Create a new linetype"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO linetypes 
                    (linetype_name, pattern_definition, description, usage_context)
                    VALUES (%s, %s, %s, %s)
                    RETURNING linetype_id
                """, (
                    data.get('linetype_name'),
                    data.get('pattern_definition'),
                    data.get('description'),
                    data.get('usage_context')
                ))
                linetype_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'linetype_id': linetype_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/linetypes/<linetype_id>', methods=['PUT'])
def update_linetype(linetype_id):
    """Update an existing linetype"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE linetypes
                    SET linetype_name = %s, pattern_definition = %s,
                        description = %s, usage_context = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE linetype_id = %s
                """, (
                    data.get('linetype_name'),
                    data.get('pattern_definition'),
                    data.get('description'),
                    data.get('usage_context'),
                    linetype_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/linetypes/<linetype_id>', methods=['DELETE'])
def delete_linetype(linetype_id):
    """Delete a linetype"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM linetypes WHERE linetype_id = %s", (linetype_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# TEXT STYLES MANAGER API ENDPOINTS
# ============================================================================

@app.route('/api/data-manager/text-styles', methods=['GET'])
def get_text_styles():
    """Get all text styles"""
    try:
        query = """
            SELECT style_id, style_name, font_name, height, width_factor,
                   oblique_angle, description, usage_context, discipline
            FROM text_styles
            ORDER BY discipline, style_name
        """
        text_styles = execute_query(query)
        return jsonify({'text_styles': text_styles})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/text-styles', methods=['POST'])
def create_text_style():
    """Create a new text style"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO text_styles 
                    (style_name, font_name, height, width_factor, oblique_angle,
                     description, usage_context, discipline)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING style_id
                """, (
                    data.get('style_name'),
                    data.get('font_name'),
                    data.get('height'),
                    data.get('width_factor'),
                    data.get('oblique_angle'),
                    data.get('description'),
                    data.get('usage_context'),
                    data.get('discipline')
                ))
                style_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'style_id': style_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/text-styles/<style_id>', methods=['PUT'])
def update_text_style(style_id):
    """Update an existing text style"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE text_styles
                    SET style_name = %s, font_name = %s, height = %s,
                        width_factor = %s, oblique_angle = %s, description = %s,
                        usage_context = %s, discipline = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE style_id = %s
                """, (
                    data.get('style_name'),
                    data.get('font_name'),
                    data.get('height'),
                    data.get('width_factor'),
                    data.get('oblique_angle'),
                    data.get('description'),
                    data.get('usage_context'),
                    data.get('discipline'),
                    style_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/text-styles/<style_id>', methods=['DELETE'])
def delete_text_style(style_id):
    """Delete a text style"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM text_styles WHERE style_id = %s", (style_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# DIMENSION STYLES MANAGER API ENDPOINTS
# ============================================================================

@app.route('/api/data-manager/dimension-styles', methods=['GET'])
def get_dimension_styles():
    """Get all dimension styles"""
    try:
        query = """
            SELECT dimension_style_id, style_name, text_height, arrow_size,
                   extension_line_offset, dimension_line_color, text_color, description
            FROM dimension_styles
            ORDER BY style_name
        """
        dimension_styles = execute_query(query)
        return jsonify({'dimension_styles': dimension_styles})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/dimension-styles', methods=['POST'])
def create_dimension_style():
    """Create a new dimension style"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO dimension_styles 
                    (style_name, text_height, arrow_size, extension_line_offset,
                     dimension_line_color, text_color, description)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING dimension_style_id
                """, (
                    data.get('style_name'),
                    data.get('text_height'),
                    data.get('arrow_size'),
                    data.get('extension_line_offset'),
                    data.get('dimension_line_color'),
                    data.get('text_color'),
                    data.get('description')
                ))
                dimension_style_id = cur.fetchone()[0]
                conn.commit()
        
        cache.clear()
        return jsonify({'dimension_style_id': dimension_style_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/dimension-styles/<dimension_style_id>', methods=['PUT'])
def update_dimension_style(dimension_style_id):
    """Update an existing dimension style"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE dimension_styles
                    SET style_name = %s, text_height = %s, arrow_size = %s,
                        extension_line_offset = %s, dimension_line_color = %s,
                        text_color = %s, description = %s
                    WHERE dimension_style_id = %s
                """, (
                    data.get('style_name'),
                    data.get('text_height'),
                    data.get('arrow_size'),
                    data.get('extension_line_offset'),
                    data.get('dimension_line_color'),
                    data.get('text_color'),
                    data.get('description'),
                    dimension_style_id
                ))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/dimension-styles/<dimension_style_id>', methods=['DELETE'])
def delete_dimension_style(dimension_style_id):
    """Delete a dimension style"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM dimension_styles WHERE dimension_style_id = %s", (dimension_style_id,))
                conn.commit()
        
        cache.clear()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# USAGE TRACKING DASHBOARD API ENDPOINTS
# ============================================================================

@app.route('/api/usage/summary')
def get_usage_summary():
    """Get summary usage statistics"""
    try:
        query = """
            WITH stats AS (
                SELECT 
                    (SELECT COUNT(*) FROM drawings) as total_drawings,
                    (SELECT COUNT(DISTINCT layer_id) FROM drawing_layer_usage) as unique_layers,
                    (SELECT SUM(block_count) FROM drawings) as total_block_instances,
                    (SELECT ROUND(AVG(entity_count)) FROM drawings) as avg_entities_per_drawing
            )
            SELECT * FROM stats
        """
        result = execute_query(query)
        return jsonify(result[0] if result else {})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/usage/top-layers')
def get_top_layers():
    """Get most used layers across all projects"""
    try:
        query = """
            SELECT 
                l.layer_name,
                COUNT(DISTINCT l.project_id) as project_count,
                COUNT(*) as usage_count
            FROM layers l
            WHERE l.layer_name IS NOT NULL
            GROUP BY l.layer_name
            ORDER BY usage_count DESC
            LIMIT 10
        """
        layers = execute_query(query)
        return jsonify({'layers': layers})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/usage/top-notes')
def get_top_notes():
    """Get most used standard notes"""
    try:
        query = """
            SELECT 
                note_id,
                note_title,
                note_category,
                usage_frequency
            FROM standard_notes
            WHERE usage_frequency > 0
            ORDER BY usage_frequency DESC
            LIMIT 10
        """
        notes = execute_query(query)
        return jsonify({'notes': notes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/usage/top-blocks')
def get_top_blocks():
    """Get most used blocks"""
    try:
        query = """
            SELECT 
                block_name,
                category,
                usage_frequency
            FROM block_definitions
            WHERE usage_frequency > 0
            ORDER BY usage_frequency DESC
            LIMIT 10
        """
        blocks = execute_query(query)
        return jsonify({'blocks': blocks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/usage/recent-activity')
def get_usage_recent_activity():
    """Get recent project activity"""
    try:
        query = """
            SELECT 
                project_id,
                project_name,
                project_number,
                created_at,
                updated_at
            FROM projects
            WHERE updated_at IS NOT NULL
            ORDER BY updated_at DESC
            LIMIT 20
        """
        activity = execute_query(query)
        return jsonify({'activity': activity})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# STANDARDS MAPPING DASHBOARD & API ENDPOINTS
# ============================================================================

@app.route('/standards-mapping-dashboard')
def standards_mapping_dashboard():
    """Render the Standards Mapping Dashboard page"""
    return render_template('standards_mapping_dashboard.html')

@app.route('/api/standards-mapping/stats')
def get_mapping_stats():
    """Get counts for all mapping types"""
    try:
        stats = {}
        
        stats['block_mappings'] = execute_query("SELECT COUNT(*) as count FROM block_name_mappings")[0]['count']
        stats['detail_mappings'] = execute_query("SELECT COUNT(*) as count FROM detail_name_mappings")[0]['count']
        stats['hatch_mappings'] = execute_query("SELECT COUNT(*) as count FROM hatch_pattern_name_mappings")[0]['count']
        stats['material_mappings'] = execute_query("SELECT COUNT(*) as count FROM material_name_mappings")[0]['count']
        stats['note_mappings'] = execute_query("SELECT COUNT(*) as count FROM note_name_mappings")[0]['count']
        
        stats['keynote_block_relationships'] = execute_query("SELECT COUNT(*) as count FROM project_keynote_block_mappings")[0]['count']
        stats['keynote_detail_relationships'] = execute_query("SELECT COUNT(*) as count FROM project_keynote_detail_mappings")[0]['count']
        stats['hatch_material_relationships'] = execute_query("SELECT COUNT(*) as count FROM project_hatch_material_mappings")[0]['count']
        stats['detail_material_relationships'] = execute_query("SELECT COUNT(*) as count FROM project_detail_material_mappings")[0]['count']
        stats['block_spec_relationships'] = execute_query("SELECT COUNT(*) as count FROM project_block_specification_mappings")[0]['count']
        stats['cross_references'] = execute_query("SELECT COUNT(*) as count FROM project_element_cross_references")[0]['count']
        
        stats['total_relationships'] = (
            stats['keynote_block_relationships'] + 
            stats['keynote_detail_relationships'] +
            stats['hatch_material_relationships'] +
            stats['detail_material_relationships'] +
            stats['block_spec_relationships'] +
            stats['cross_references']
        )
        
        return jsonify({'stats': stats})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards-mapping/block-mappings')
def get_block_name_mappings():
    """Get all block name mappings"""
    try:
        query = """
            SELECT mapping_id, canonical_name, dxf_alias, description,
                   import_direction, export_direction, client_id,
                   confidence_score, is_active
            FROM block_name_mappings
            WHERE is_active = TRUE
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards-mapping/detail-mappings')
def get_detail_name_mappings():
    """Get all detail name mappings"""
    try:
        query = """
            SELECT mapping_id, canonical_name, dxf_alias, description,
                   import_direction, export_direction, client_id,
                   confidence_score, is_active
            FROM detail_name_mappings
            WHERE is_active = TRUE
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards-mapping/hatch-mappings')
def get_hatch_name_mappings():
    """Get all hatch pattern name mappings"""
    try:
        query = """
            SELECT mapping_id, canonical_name, dxf_alias, description,
                   import_direction, export_direction, client_id,
                   confidence_score, is_active
            FROM hatch_pattern_name_mappings
            WHERE is_active = TRUE
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards-mapping/material-mappings')
def get_material_name_mappings():
    """Get all material name mappings"""
    try:
        query = """
            SELECT mapping_id, canonical_name, dxf_alias, description,
                   import_direction, export_direction, client_id,
                   confidence_score, is_active
            FROM material_name_mappings
            WHERE is_active = TRUE
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/standards-mapping/note-mappings')
def get_note_name_mappings():
    """Get all note name mappings"""
    try:
        query = """
            SELECT mapping_id, canonical_name, dxf_alias, description,
                   import_direction, export_direction, client_id,
                   confidence_score, is_active
            FROM note_name_mappings
            WHERE is_active = TRUE
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# BLOCK NAME MAPPINGS CRUD MANAGER
# ============================================================================

@app.route('/data-manager/block-mappings')
def block_mappings_manager():
    """Render the Block Name Mappings Manager page"""
    return render_template('data_manager/block_mappings.html')

@app.route('/api/data-manager/block-mappings', methods=['GET'])
def get_block_mappings_crud():
    """Get all block mappings for CRUD interface"""
    try:
        query = """
            SELECT * FROM block_name_mappings
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/block-mappings', methods=['POST'])
def create_block_mapping():
    """Create a new block mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO block_name_mappings 
                    (canonical_name, dxf_alias, description, import_direction, 
                     export_direction, client_id, confidence_score)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0)
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        return jsonify({'mapping_id': mapping_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/block-mappings/<mapping_id>', methods=['PUT'])
def update_block_mapping(mapping_id):
    """Update an existing block mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE block_name_mappings
                    SET canonical_name = %s, dxf_alias = %s, description = %s,
                        import_direction = %s, export_direction = %s, 
                        client_id = %s, confidence_score = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0),
                    mapping_id
                ))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/block-mappings/<mapping_id>', methods=['DELETE'])
def delete_block_mapping(mapping_id):
    """Delete a block mapping"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM block_name_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# DETAIL NAME MAPPINGS CRUD MANAGER
# ============================================================================

@app.route('/data-manager/detail-mappings')
def detail_mappings_manager():
    """Render the Detail Name Mappings Manager page"""
    return render_template('data_manager/detail_mappings.html')

@app.route('/api/data-manager/detail-mappings', methods=['GET'])
def get_detail_mappings_crud():
    """Get all detail mappings for CRUD interface"""
    try:
        query = """
            SELECT * FROM detail_name_mappings
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/detail-mappings', methods=['POST'])
def create_detail_mapping():
    """Create a new detail mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO detail_name_mappings 
                    (canonical_name, dxf_alias, description, import_direction, 
                     export_direction, client_id, confidence_score)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0)
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        return jsonify({'mapping_id': mapping_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/detail-mappings/<mapping_id>', methods=['PUT'])
def update_detail_mapping(mapping_id):
    """Update an existing detail mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE detail_name_mappings
                    SET canonical_name = %s, dxf_alias = %s, description = %s,
                        import_direction = %s, export_direction = %s, 
                        client_id = %s, confidence_score = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0),
                    mapping_id
                ))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/detail-mappings/<mapping_id>', methods=['DELETE'])
def delete_detail_mapping(mapping_id):
    """Delete a detail mapping"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM detail_name_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# HATCH PATTERN NAME MAPPINGS CRUD MANAGER
# ============================================================================

@app.route('/data-manager/hatch-mappings')
def hatch_mappings_manager():
    """Render the Hatch Pattern Name Mappings Manager page"""
    return render_template('data_manager/hatch_mappings.html')

@app.route('/api/data-manager/hatch-mappings', methods=['GET'])
def get_hatch_mappings_crud():
    """Get all hatch pattern mappings for CRUD interface"""
    try:
        query = """
            SELECT * FROM hatch_pattern_name_mappings
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/hatch-mappings', methods=['POST'])
def create_hatch_mapping():
    """Create a new hatch pattern mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO hatch_pattern_name_mappings 
                    (canonical_name, dxf_alias, description, import_direction, 
                     export_direction, client_id, confidence_score)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0)
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        return jsonify({'mapping_id': mapping_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/hatch-mappings/<mapping_id>', methods=['PUT'])
def update_hatch_mapping(mapping_id):
    """Update an existing hatch pattern mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE hatch_pattern_name_mappings
                    SET canonical_name = %s, dxf_alias = %s, description = %s,
                        import_direction = %s, export_direction = %s, 
                        client_id = %s, confidence_score = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0),
                    mapping_id
                ))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/hatch-mappings/<mapping_id>', methods=['DELETE'])
def delete_hatch_mapping(mapping_id):
    """Delete a hatch pattern mapping"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM hatch_pattern_name_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# MATERIAL NAME MAPPINGS CRUD MANAGER
# ============================================================================

@app.route('/data-manager/material-mappings')
def material_mappings_manager():
    """Render the Material Name Mappings Manager page"""
    return render_template('data_manager/material_mappings.html')

@app.route('/api/data-manager/material-mappings', methods=['GET'])
def get_material_mappings_crud():
    """Get all material name mappings for CRUD interface"""
    try:
        query = """
            SELECT * FROM material_name_mappings
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/material-mappings', methods=['POST'])
def create_material_mapping():
    """Create a new material name mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO material_name_mappings 
                    (canonical_name, dxf_alias, description, import_direction, 
                     export_direction, client_id, confidence_score)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0)
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        return jsonify({'mapping_id': mapping_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/material-mappings/<mapping_id>', methods=['PUT'])
def update_material_mapping(mapping_id):
    """Update an existing material name mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE material_name_mappings
                    SET canonical_name = %s, dxf_alias = %s, description = %s,
                        import_direction = %s, export_direction = %s, 
                        client_id = %s, confidence_score = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0),
                    mapping_id
                ))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/material-mappings/<mapping_id>', methods=['DELETE'])
def delete_material_mapping(mapping_id):
    """Delete a material name mapping"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM material_name_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# NOTE NAME MAPPINGS CRUD MANAGER
# ============================================================================

@app.route('/data-manager/note-mappings')
def note_mappings_manager():
    """Render the Note Name Mappings Manager page"""
    return render_template('data_manager/note_mappings.html')

@app.route('/api/data-manager/note-mappings', methods=['GET'])
def get_note_mappings_crud():
    """Get all note name mappings for CRUD interface"""
    try:
        query = """
            SELECT * FROM note_name_mappings
            ORDER BY canonical_name
        """
        mappings = execute_query(query)
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/note-mappings', methods=['POST'])
def create_note_mapping():
    """Create a new note name mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO note_name_mappings 
                    (canonical_name, dxf_alias, description, import_direction, 
                     export_direction, client_id, confidence_score)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0)
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        return jsonify({'mapping_id': mapping_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/note-mappings/<mapping_id>', methods=['PUT'])
def update_note_mapping(mapping_id):
    """Update an existing note name mapping"""
    try:
        data = request.get_json()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE note_name_mappings
                    SET canonical_name = %s, dxf_alias = %s, description = %s,
                        import_direction = %s, export_direction = %s, 
                        client_id = %s, confidence_score = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('canonical_name'),
                    data.get('dxf_alias'),
                    data.get('description'),
                    data.get('import_direction', True),
                    data.get('export_direction', True),
                    data.get('client_id'),
                    data.get('confidence_score', 1.0),
                    mapping_id
                ))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data-manager/note-mappings/<mapping_id>', methods=['DELETE'])
def delete_note_mapping(mapping_id):
    """Delete a note name mapping"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM note_name_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# PROJECT CONTEXT MAPPING MANAGER
# ============================================================================

@app.route('/project-context-manager')
def project_context_manager():
    """Project Context Mapping Manager page"""
    return render_template('project_context_manager.html')

@app.route('/api/project-context/keynote-blocks', methods=['GET'])
def get_keynote_blocks():
    """Get keynote-block relationships for a project"""
    try:
        project_id = request.args.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                kbm.mapping_id,
                kbm.project_id,
                kbm.note_id,
                kbm.keynote_number,
                kbm.block_id,
                kbm.relationship_type,
                kbm.usage_context,
                sn.note_text,
                sn.note_title,
                bd.block_name,
                bd.description as block_description
            FROM project_keynote_block_mappings kbm
            LEFT JOIN standard_notes sn ON kbm.note_id = sn.note_id
            LEFT JOIN block_definitions bd ON kbm.block_id = bd.block_id
            WHERE kbm.project_id = %s AND kbm.is_active = TRUE
            ORDER BY kbm.keynote_number, bd.block_name
        """
        mappings = execute_query(query, (project_id,))
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-blocks', methods=['POST'])
def create_keynote_block():
    """Create a new keynote-block relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO project_keynote_block_mappings 
                    (project_id, note_id, keynote_number, block_id, relationship_type, usage_context)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('project_id'),
                    data.get('note_id'),
                    data.get('keynote_number'),
                    data.get('block_id'),
                    data.get('relationship_type', 'references'),
                    data.get('usage_context')
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'mapping_id': str(mapping_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-blocks/<mapping_id>', methods=['GET'])
def get_keynote_block(mapping_id):
    """Get a single keynote-block relationship"""
    try:
        query = """
            SELECT * FROM project_keynote_block_mappings
            WHERE mapping_id = %s
        """
        result = execute_query(query, (mapping_id,))
        if not result:
            return jsonify({'error': 'Mapping not found'}), 404
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-blocks/<mapping_id>', methods=['PUT'])
def update_keynote_block(mapping_id):
    """Update a keynote-block relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE project_keynote_block_mappings
                    SET note_id = %s, keynote_number = %s, block_id = %s,
                        relationship_type = %s, usage_context = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('note_id'),
                    data.get('keynote_number'),
                    data.get('block_id'),
                    data.get('relationship_type'),
                    data.get('usage_context'),
                    mapping_id
                ))
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-blocks/<mapping_id>', methods=['DELETE'])
def delete_keynote_block(mapping_id):
    """Delete a keynote-block relationship"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_keynote_block_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-details', methods=['GET'])
def get_keynote_details():
    """Get keynote-detail relationships for a project"""
    try:
        project_id = request.args.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                kdm.mapping_id,
                kdm.project_id,
                kdm.note_id,
                kdm.keynote_number,
                kdm.detail_id,
                kdm.detail_callout,
                kdm.sheet_number,
                kdm.usage_context,
                sn.note_text,
                sn.note_title,
                ds.detail_title as detail_name,
                ds.description as detail_description
            FROM project_keynote_detail_mappings kdm
            LEFT JOIN standard_notes sn ON kdm.note_id = sn.note_id
            LEFT JOIN detail_standards ds ON kdm.detail_id = ds.detail_id
            WHERE kdm.project_id = %s AND kdm.is_active = TRUE
            ORDER BY kdm.keynote_number, ds.detail_title
        """
        mappings = execute_query(query, (project_id,))
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-details', methods=['POST'])
def create_keynote_detail():
    """Create a new keynote-detail relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO project_keynote_detail_mappings 
                    (project_id, note_id, keynote_number, detail_id, detail_callout, sheet_number, usage_context)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('project_id'),
                    data.get('note_id'),
                    data.get('keynote_number'),
                    data.get('detail_id'),
                    data.get('detail_callout'),
                    data.get('sheet_number'),
                    data.get('usage_context')
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'mapping_id': str(mapping_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-details/<mapping_id>', methods=['GET'])
def get_keynote_detail(mapping_id):
    """Get a single keynote-detail relationship"""
    try:
        query = """
            SELECT * FROM project_keynote_detail_mappings
            WHERE mapping_id = %s
        """
        result = execute_query(query, (mapping_id,))
        if not result:
            return jsonify({'error': 'Mapping not found'}), 404
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-details/<mapping_id>', methods=['PUT'])
def update_keynote_detail(mapping_id):
    """Update a keynote-detail relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE project_keynote_detail_mappings
                    SET note_id = %s, keynote_number = %s, detail_id = %s,
                        detail_callout = %s, sheet_number = %s, usage_context = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('note_id'),
                    data.get('keynote_number'),
                    data.get('detail_id'),
                    data.get('detail_callout'),
                    data.get('sheet_number'),
                    data.get('usage_context'),
                    mapping_id
                ))
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/keynote-details/<mapping_id>', methods=['DELETE'])
def delete_keynote_detail(mapping_id):
    """Delete a keynote-detail relationship"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_keynote_detail_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/hatch-materials', methods=['GET'])
def get_hatch_materials():
    """Get hatch-material relationships for a project"""
    try:
        project_id = request.args.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                hmm.mapping_id,
                hmm.project_id,
                hmm.hatch_id,
                hmm.hatch_name,
                hmm.material_id,
                hmm.material_thickness,
                hmm.material_notes,
                hmm.is_legend_item,
                hp.pattern_name as hatch_pattern_name,
                hp.description as hatch_description,
                ms.material_name,
                ms.description as material_description
            FROM project_hatch_material_mappings hmm
            LEFT JOIN hatch_patterns hp ON hmm.hatch_id = hp.hatch_id
            LEFT JOIN material_standards ms ON hmm.material_id = ms.material_id
            WHERE hmm.project_id = %s AND hmm.is_active = TRUE
            ORDER BY ms.material_name, hp.pattern_name
        """
        mappings = execute_query(query, (project_id,))
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/hatch-materials', methods=['POST'])
def create_hatch_material():
    """Create a new hatch-material relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO project_hatch_material_mappings 
                    (project_id, hatch_id, material_id, material_thickness, material_notes, is_legend_item)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('project_id'),
                    data.get('hatch_id'),
                    data.get('material_id'),
                    data.get('material_thickness'),
                    data.get('material_notes'),
                    data.get('is_legend_item', True)
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'mapping_id': str(mapping_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/hatch-materials/<mapping_id>', methods=['GET'])
def get_hatch_material(mapping_id):
    """Get a single hatch-material relationship"""
    try:
        query = """
            SELECT * FROM project_hatch_material_mappings
            WHERE mapping_id = %s
        """
        result = execute_query(query, (mapping_id,))
        if not result:
            return jsonify({'error': 'Mapping not found'}), 404
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/hatch-materials/<mapping_id>', methods=['PUT'])
def update_hatch_material(mapping_id):
    """Update a hatch-material relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE project_hatch_material_mappings
                    SET hatch_id = %s, material_id = %s, material_thickness = %s,
                        material_notes = %s, is_legend_item = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('hatch_id'),
                    data.get('material_id'),
                    data.get('material_thickness'),
                    data.get('material_notes'),
                    data.get('is_legend_item'),
                    mapping_id
                ))
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/hatch-materials/<mapping_id>', methods=['DELETE'])
def delete_hatch_material(mapping_id):
    """Delete a hatch-material relationship"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_hatch_material_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/detail-materials', methods=['GET'])
def get_detail_materials():
    """Get detail-material relationships for a project"""
    try:
        project_id = request.args.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                dmm.mapping_id,
                dmm.project_id,
                dmm.detail_id,
                dmm.material_id,
                dmm.material_role,
                dmm.material_layer_order,
                dmm.material_thickness,
                dmm.material_notes,
                ds.detail_title as detail_name,
                ds.description as detail_description,
                ms.material_name,
                ms.description as material_description
            FROM project_detail_material_mappings dmm
            LEFT JOIN detail_standards ds ON dmm.detail_id = ds.detail_id
            LEFT JOIN material_standards ms ON dmm.material_id = ms.material_id
            WHERE dmm.project_id = %s AND dmm.is_active = TRUE
            ORDER BY ds.detail_title, dmm.material_layer_order
        """
        mappings = execute_query(query, (project_id,))
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/detail-materials', methods=['POST'])
def create_detail_material():
    """Create a new detail-material relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO project_detail_material_mappings 
                    (project_id, detail_id, material_id, material_role, material_layer_order, 
                     material_thickness, material_notes)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('project_id'),
                    data.get('detail_id'),
                    data.get('material_id'),
                    data.get('material_role', 'primary'),
                    data.get('material_layer_order'),
                    data.get('material_thickness'),
                    data.get('material_notes')
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'mapping_id': str(mapping_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/detail-materials/<mapping_id>', methods=['GET'])
def get_detail_material(mapping_id):
    """Get a single detail-material relationship"""
    try:
        query = """
            SELECT * FROM project_detail_material_mappings
            WHERE mapping_id = %s
        """
        result = execute_query(query, (mapping_id,))
        if not result:
            return jsonify({'error': 'Mapping not found'}), 404
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/detail-materials/<mapping_id>', methods=['PUT'])
def update_detail_material(mapping_id):
    """Update a detail-material relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE project_detail_material_mappings
                    SET detail_id = %s, material_id = %s, material_role = %s,
                        material_layer_order = %s, material_thickness = %s, material_notes = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('detail_id'),
                    data.get('material_id'),
                    data.get('material_role'),
                    data.get('material_layer_order'),
                    data.get('material_thickness'),
                    data.get('material_notes'),
                    mapping_id
                ))
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/detail-materials/<mapping_id>', methods=['DELETE'])
def delete_detail_material(mapping_id):
    """Delete a detail-material relationship"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_detail_material_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/block-specs', methods=['GET'])
def get_block_specs():
    """Get block-specification relationships for a project"""
    try:
        project_id = request.args.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                bsm.mapping_id,
                bsm.project_id,
                bsm.block_id,
                bsm.spec_section,
                bsm.spec_description,
                bsm.manufacturer,
                bsm.model_number,
                bsm.product_url,
                bsm.jurisdiction,
                bd.block_name,
                bd.description as block_description
            FROM project_block_specification_mappings bsm
            LEFT JOIN block_definitions bd ON bsm.block_id = bd.block_id
            WHERE bsm.project_id = %s AND bsm.is_active = TRUE
            ORDER BY bd.block_name, bsm.spec_section
        """
        mappings = execute_query(query, (project_id,))
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/block-specs', methods=['POST'])
def create_block_spec():
    """Create a new block-specification relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO project_block_specification_mappings 
                    (project_id, block_id, spec_section, spec_description, manufacturer, 
                     model_number, product_url, jurisdiction)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING mapping_id
                """, (
                    data.get('project_id'),
                    data.get('block_id'),
                    data.get('spec_section'),
                    data.get('spec_description'),
                    data.get('manufacturer'),
                    data.get('model_number'),
                    data.get('product_url'),
                    data.get('jurisdiction')
                ))
                mapping_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'mapping_id': str(mapping_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/block-specs/<mapping_id>', methods=['GET'])
def get_block_spec(mapping_id):
    """Get a single block-specification relationship"""
    try:
        query = """
            SELECT * FROM project_block_specification_mappings
            WHERE mapping_id = %s
        """
        result = execute_query(query, (mapping_id,))
        if not result:
            return jsonify({'error': 'Mapping not found'}), 404
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/block-specs/<mapping_id>', methods=['PUT'])
def update_block_spec(mapping_id):
    """Update a block-specification relationship"""
    try:
        data = request.get_json()
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE project_block_specification_mappings
                    SET block_id = %s, spec_section = %s, spec_description = %s,
                        manufacturer = %s, model_number = %s, product_url = %s, jurisdiction = %s
                    WHERE mapping_id = %s
                """, (
                    data.get('block_id'),
                    data.get('spec_section'),
                    data.get('spec_description'),
                    data.get('manufacturer'),
                    data.get('model_number'),
                    data.get('product_url'),
                    data.get('jurisdiction'),
                    mapping_id
                ))
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/block-specs/<mapping_id>', methods=['DELETE'])
def delete_block_spec(mapping_id):
    """Delete a block-specification relationship"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_block_specification_mappings WHERE mapping_id = %s", (mapping_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# STANDARDS DOCUMENTATION EXPORT
# ============================================================================

@app.route('/standards-export')
def standards_export_page():
    """Standards Documentation Export page"""
    return render_template('standards_export.html')

@app.route('/api/standards-export/generate', methods=['POST'])
def generate_standards_export():
    """Generate standards documentation export in requested format"""
    try:
        config = request.get_json()
        export_type = config.get('exportType')
        title = config.get('title', 'CAD Standards Documentation')
        description = config.get('description', '')
        name_mappings = config.get('nameMappings', {})
        relationships = config.get('relationships', {})
        
        # Collect data based on configuration
        data = {}
        
        # Get name mappings
        if name_mappings.get('blocks'):
            query = """
                SELECT 
                    bnm.project_id,
                    COALESCE(p.project_name, bnm.client_name, 'Global') as project_name,
                    bnm.dxf_alias as imported_name,
                    bnm.block_id,
                    bd.block_name as standard_name,
                    bd.description
                FROM block_name_mappings bnm
                LEFT JOIN projects p ON bnm.project_id = p.project_id
                LEFT JOIN block_definitions bd ON bnm.block_id = bd.block_id
                WHERE bnm.is_active = TRUE
                ORDER BY project_name, bnm.dxf_alias
            """
            data['block_mappings'] = execute_query(query)
        
        if name_mappings.get('details'):
            query = """
                SELECT 
                    dnm.project_id,
                    COALESCE(p.project_name, dnm.client_name, 'Global') as project_name,
                    dnm.dxf_alias as imported_name,
                    dnm.detail_id,
                    ds.detail_title as standard_name,
                    ds.description
                FROM detail_name_mappings dnm
                LEFT JOIN projects p ON dnm.project_id = p.project_id
                LEFT JOIN detail_standards ds ON dnm.detail_id = ds.detail_id
                WHERE dnm.is_active = TRUE
                ORDER BY project_name, dnm.dxf_alias
            """
            data['detail_mappings'] = execute_query(query)
        
        if name_mappings.get('hatches'):
            query = """
                SELECT 
                    hnm.project_id,
                    COALESCE(p.project_name, hnm.client_name, 'Global') as project_name,
                    hnm.dxf_alias as imported_name,
                    hnm.hatch_id,
                    hp.pattern_name as standard_name,
                    hp.description
                FROM hatch_pattern_name_mappings hnm
                LEFT JOIN projects p ON hnm.project_id = p.project_id
                LEFT JOIN hatch_patterns hp ON hnm.hatch_id = hp.hatch_id
                WHERE hnm.is_active = TRUE
                ORDER BY project_name, hnm.dxf_alias
            """
            data['hatch_mappings'] = execute_query(query)
        
        if name_mappings.get('materials'):
            query = """
                SELECT 
                    mnm.project_id,
                    COALESCE(p.project_name, mnm.client_name, 'Global') as project_name,
                    mnm.dxf_alias as imported_name,
                    mnm.material_id,
                    ms.material_name as standard_name,
                    ms.description
                FROM material_name_mappings mnm
                LEFT JOIN projects p ON mnm.project_id = p.project_id
                LEFT JOIN material_standards ms ON mnm.material_id = ms.material_id
                WHERE mnm.is_active = TRUE
                ORDER BY project_name, mnm.dxf_alias
            """
            data['material_mappings'] = execute_query(query)
        
        if name_mappings.get('notes'):
            query = """
                SELECT 
                    nnm.project_id,
                    COALESCE(p.project_name, nnm.client_name, 'Global') as project_name,
                    nnm.dxf_alias as imported_name,
                    nnm.note_id,
                    sn.note_text as standard_name,
                    sn.note_category as category
                FROM note_name_mappings nnm
                LEFT JOIN projects p ON nnm.project_id = p.project_id
                LEFT JOIN standard_notes sn ON nnm.note_id = sn.note_id
                WHERE nnm.is_active = TRUE
                ORDER BY project_name, nnm.dxf_alias
            """
            data['note_mappings'] = execute_query(query)
        
        # Get relationships
        if relationships.get('keynoteBlocks'):
            query = """
                SELECT 
                    pkbm.project_id,
                    p.project_name,
                    pkbm.keynote_number as keynote,
                    pkbm.block_id,
                    bd.block_name,
                    pkbm.usage_context as context_notes
                FROM project_keynote_block_mappings pkbm
                LEFT JOIN projects p ON pkbm.project_id = p.project_id
                LEFT JOIN block_definitions bd ON pkbm.block_id = bd.block_id
                WHERE pkbm.is_active = TRUE
                ORDER BY p.project_name, pkbm.keynote_number
            """
            data['keynote_block_relationships'] = execute_query(query)
        
        if relationships.get('keynoteDetails'):
            query = """
                SELECT 
                    pkdm.project_id,
                    p.project_name,
                    pkdm.keynote_number as keynote,
                    pkdm.detail_id,
                    ds.detail_title as detail_name,
                    pkdm.usage_context as context_notes
                FROM project_keynote_detail_mappings pkdm
                LEFT JOIN projects p ON pkdm.project_id = p.project_id
                LEFT JOIN detail_standards ds ON pkdm.detail_id = ds.detail_id
                WHERE pkdm.is_active = TRUE
                ORDER BY p.project_name, pkdm.keynote_number
            """
            data['keynote_detail_relationships'] = execute_query(query)
        
        if relationships.get('hatchMaterials'):
            query = """
                SELECT 
                    phmm.project_id,
                    p.project_name,
                    phmm.hatch_id,
                    hp.pattern_name,
                    phmm.material_id,
                    ms.material_name,
                    phmm.material_notes as context_notes
                FROM project_hatch_material_mappings phmm
                LEFT JOIN projects p ON phmm.project_id = p.project_id
                LEFT JOIN hatch_patterns hp ON phmm.hatch_id = hp.hatch_id
                LEFT JOIN material_standards ms ON phmm.material_id = ms.material_id
                WHERE phmm.is_active = TRUE
                ORDER BY p.project_name, hp.pattern_name
            """
            data['hatch_material_relationships'] = execute_query(query)
        
        if relationships.get('detailMaterials'):
            query = """
                SELECT 
                    pdmm.project_id,
                    p.project_name,
                    pdmm.detail_id,
                    ds.detail_title as detail_name,
                    pdmm.material_id,
                    ms.material_name,
                    pdmm.material_notes as context_notes
                FROM project_detail_material_mappings pdmm
                LEFT JOIN projects p ON pdmm.project_id = p.project_id
                LEFT JOIN detail_standards ds ON pdmm.detail_id = ds.detail_id
                LEFT JOIN material_standards ms ON pdmm.material_id = ms.material_id
                WHERE pdmm.is_active = TRUE
                ORDER BY p.project_name, ds.detail_title
            """
            data['detail_material_relationships'] = execute_query(query)
        
        if relationships.get('blockSpecs'):
            query = """
                SELECT 
                    pbsm.project_id,
                    p.project_name,
                    pbsm.block_id,
                    bd.block_name,
                    pbsm.spec_section,
                    pbsm.spec_description,
                    pbsm.manufacturer,
                    pbsm.model_number
                FROM project_block_specification_mappings pbsm
                LEFT JOIN projects p ON pbsm.project_id = p.project_id
                LEFT JOIN block_definitions bd ON pbsm.block_id = bd.block_id
                WHERE pbsm.is_active = TRUE
                ORDER BY p.project_name, bd.block_name
            """
            data['block_spec_relationships'] = execute_query(query)
        
        # Generate export based on type
        if export_type == 'excel':
            return generate_excel_export(data, title, description)
        elif export_type == 'csv':
            return generate_csv_export(data, title, description)
        elif export_type == 'html':
            return generate_html_export(data, title, description)
        elif export_type == 'pdf':
            return generate_pdf_export(data, title, description)
        else:
            return jsonify({'error': 'Invalid export type'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def generate_excel_export(data, title, description):
    """Generate Excel workbook export"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Overview sheet
    ws_overview = wb.create_sheet("Overview")
    ws_overview.append(['Title', title])
    ws_overview.append(['Description', description])
    ws_overview.append(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_overview.append([])
    ws_overview.append(['Statistics'])
    
    for key, items in data.items():
        ws_overview.append([key.replace('_', ' ').title(), len(items)])
    
    # Style overview
    for cell in ws_overview[1]:
        cell.font = Font(bold=True)
    for cell in ws_overview[5]:
        cell.font = Font(bold=True)
    
    # Create sheets for each data type
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00B4D8", end_color="00B4D8", fill_type="solid")
    
    if 'block_mappings' in data:
        ws = wb.create_sheet("Block Mappings")
        ws.append(['Project', 'Imported Name', 'Standard Name', 'Description'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['block_mappings']:
            ws.append([
                item.get('project_name', ''),
                item.get('imported_name', ''),
                item.get('standard_name', ''),
                item.get('description', '')
            ])
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    if 'detail_mappings' in data:
        ws = wb.create_sheet("Detail Mappings")
        ws.append(['Project', 'Imported Name', 'Standard Name', 'Description'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['detail_mappings']:
            ws.append([
                item.get('project_name', ''),
                item.get('imported_name', ''),
                item.get('standard_name', ''),
                item.get('description', '')
            ])
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    if 'hatch_mappings' in data:
        ws = wb.create_sheet("Hatch Mappings")
        ws.append(['Project', 'Imported Name', 'Standard Name', 'Description'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['hatch_mappings']:
            ws.append([
                item.get('project_name', ''),
                item.get('imported_name', ''),
                item.get('standard_name', ''),
                item.get('description', '')
            ])
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    if 'material_mappings' in data:
        ws = wb.create_sheet("Material Mappings")
        ws.append(['Project', 'Imported Name', 'Standard Name', 'Description'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['material_mappings']:
            ws.append([
                item.get('project_name', ''),
                item.get('imported_name', ''),
                item.get('standard_name', ''),
                item.get('description', '')
            ])
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    if 'note_mappings' in data:
        ws = wb.create_sheet("Note Mappings")
        ws.append(['Project', 'Imported Name', 'Standard Note', 'Category'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['note_mappings']:
            ws.append([
                item.get('project_name', ''),
                item.get('imported_name', ''),
                item.get('standard_name', ''),
                item.get('category', '')
            ])
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Add relationship sheets
    if 'keynote_block_relationships' in data:
        ws = wb.create_sheet("Keynote-Block Relations")
        ws.append(['Project', 'Keynote', 'Block Name', 'Notes'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['keynote_block_relationships']:
            ws.append([
                item.get('project_name', ''),
                item.get('keynote', ''),
                item.get('block_name', ''),
                item.get('context_notes', '')
            ])
    
    if 'keynote_detail_relationships' in data:
        ws = wb.create_sheet("Keynote-Detail Relations")
        ws.append(['Project', 'Keynote', 'Detail Name', 'Notes'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['keynote_detail_relationships']:
            ws.append([
                item.get('project_name', ''),
                item.get('keynote', ''),
                item.get('detail_name', ''),
                item.get('context_notes', '')
            ])
    
    if 'hatch_material_relationships' in data:
        ws = wb.create_sheet("Hatch-Material Relations")
        ws.append(['Project', 'Hatch Pattern', 'Material', 'Notes'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['hatch_material_relationships']:
            ws.append([
                item.get('project_name', ''),
                item.get('pattern_name', ''),
                item.get('material_name', ''),
                item.get('context_notes', '')
            ])
    
    if 'detail_material_relationships' in data:
        ws = wb.create_sheet("Detail-Material Relations")
        ws.append(['Project', 'Detail Name', 'Material', 'Notes'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['detail_material_relationships']:
            ws.append([
                item.get('project_name', ''),
                item.get('detail_name', ''),
                item.get('material_name', ''),
                item.get('context_notes', '')
            ])
    
    if 'block_spec_relationships' in data:
        ws = wb.create_sheet("Block-Specification Relations")
        ws.append(['Project', 'Block Name', 'Spec Section', 'Description', 'Manufacturer', 'Model'])
        for row in ws[1]:
            row.font = header_font
            row.fill = header_fill
        for item in data['block_spec_relationships']:
            ws.append([
                item.get('project_name', ''),
                item.get('block_name', ''),
                item.get('spec_section', ''),
                item.get('spec_description', ''),
                item.get('manufacturer', ''),
                item.get('model_number', '')
            ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'CAD_Standards_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

def generate_csv_export(data, title, description):
    """Generate CSV files in ZIP archive"""
    # Create temp directory
    with tempfile.TemporaryDirectory() as tmpdir:
        files_created = []
        
        # Create CSV for each data type
        if 'block_mappings' in data and data['block_mappings']:
            filepath = os.path.join(tmpdir, 'block_mappings.csv')
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Project', 'Imported Name', 'Standard Name', 'Description'])
                for item in data['block_mappings']:
                    writer.writerow([
                        item.get('project_name', ''),
                        item.get('imported_name', ''),
                        item.get('standard_name', ''),
                        item.get('description', '')
                    ])
            files_created.append(filepath)
        
        if 'detail_mappings' in data and data['detail_mappings']:
            filepath = os.path.join(tmpdir, 'detail_mappings.csv')
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Project', 'Imported Name', 'Standard Name', 'Description'])
                for item in data['detail_mappings']:
                    writer.writerow([
                        item.get('project_name', ''),
                        item.get('imported_name', ''),
                        item.get('standard_name', ''),
                        item.get('description', '')
                    ])
            files_created.append(filepath)
        
        if 'hatch_mappings' in data and data['hatch_mappings']:
            filepath = os.path.join(tmpdir, 'hatch_mappings.csv')
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Project', 'Imported Name', 'Standard Name', 'Description'])
                for item in data['hatch_mappings']:
                    writer.writerow([
                        item.get('project_name', ''),
                        item.get('imported_name', ''),
                        item.get('standard_name', ''),
                        item.get('description', '')
                    ])
            files_created.append(filepath)
        
        if 'material_mappings' in data and data['material_mappings']:
            filepath = os.path.join(tmpdir, 'material_mappings.csv')
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Project', 'Imported Name', 'Standard Name', 'Description'])
                for item in data['material_mappings']:
                    writer.writerow([
                        item.get('project_name', ''),
                        item.get('imported_name', ''),
                        item.get('standard_name', ''),
                        item.get('description', '')
                    ])
            files_created.append(filepath)
        
        if 'note_mappings' in data and data['note_mappings']:
            filepath = os.path.join(tmpdir, 'note_mappings.csv')
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Project', 'Imported Name', 'Standard Note', 'Category'])
                for item in data['note_mappings']:
                    writer.writerow([
                        item.get('project_name', ''),
                        item.get('imported_name', ''),
                        item.get('standard_name', ''),
                        item.get('category', '')
                    ])
            files_created.append(filepath)
        
        # Add metadata file
        metadata_path = os.path.join(tmpdir, 'README.txt')
        with open(metadata_path, 'w', encoding='utf-8') as f:
            f.write(f"{title}\n")
            f.write("=" * len(title) + "\n\n")
            f.write(f"{description}\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        files_created.append(metadata_path)
        
        # Create ZIP file
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filepath in files_created:
                zip_file.write(filepath, os.path.basename(filepath))
        
        zip_buffer.seek(0)
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'CAD_Standards_{datetime.now().strftime("%Y%m%d")}.zip'
        )

def generate_html_export(data, title, description):
    """Generate standalone HTML export"""
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #001a33 0%, #003d66 100%);
            color: #e0e0e0;
            padding: 20px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: rgba(0, 40, 80, 0.6);
            border: 1px solid #00b4d8;
            border-radius: 8px;
            padding: 30px;
        }}
        h1 {{
            color: #00b4d8;
            text-align: center;
            margin-bottom: 10px;
            font-size: 2.5em;
        }}
        .subtitle {{
            text-align: center;
            color: #90e0ef;
            margin-bottom: 30px;
        }}
        .section {{
            margin: 30px 0;
        }}
        h2 {{
            color: #00b4d8;
            border-bottom: 2px solid #00b4d8;
            padding-bottom: 10px;
            margin-bottom: 20px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background: rgba(0, 20, 40, 0.6);
        }}
        th {{
            background: #00b4d8;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid rgba(0, 180, 216, 0.2);
        }}
        tr:hover {{
            background: rgba(0, 180, 216, 0.1);
        }}
        .empty {{
            text-align: center;
            padding: 20px;
            color: #90e0ef;
            font-style: italic;
        }}
        .metadata {{
            background: rgba(0, 60, 100, 0.4);
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 30px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <p class="subtitle">{description}</p>
        
        <div class="metadata">
            <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
"""
    
    # Add each data section
    if 'block_mappings' in data:
        html_content += '<div class="section"><h2>Block Name Mappings</h2>'
        if data['block_mappings']:
            html_content += '<table><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
            for item in data['block_mappings']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No block mappings found</p>'
        html_content += '</div>'
    
    if 'detail_mappings' in data:
        html_content += '<div class="section"><h2>Detail Name Mappings</h2>'
        if data['detail_mappings']:
            html_content += '<table><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
            for item in data['detail_mappings']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No detail mappings found</p>'
        html_content += '</div>'
    
    if 'hatch_mappings' in data:
        html_content += '<div class="section"><h2>Hatch Pattern Mappings</h2>'
        if data['hatch_mappings']:
            html_content += '<table><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
            for item in data['hatch_mappings']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No hatch mappings found</p>'
        html_content += '</div>'
    
    if 'material_mappings' in data:
        html_content += '<div class="section"><h2>Material Name Mappings</h2>'
        if data['material_mappings']:
            html_content += '<table><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
            for item in data['material_mappings']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No material mappings found</p>'
        html_content += '</div>'
    
    if 'note_mappings' in data:
        html_content += '<div class="section"><h2>Note/Keynote Mappings</h2>'
        if data['note_mappings']:
            html_content += '<table><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Note</th><th>Category</th></tr></thead><tbody>'
            for item in data['note_mappings']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('category', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No note mappings found</p>'
        html_content += '</div>'
    
    if 'keynote_block_relationships' in data:
        html_content += '<div class="section"><h2>Keynote → Block Relationships</h2>'
        if data['keynote_block_relationships']:
            html_content += '<table><thead><tr><th>Project</th><th>Keynote</th><th>Block Name</th><th>Notes</th></tr></thead><tbody>'
            for item in data['keynote_block_relationships']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('keynote', '')}</td><td>{item.get('block_name', '')}</td><td>{item.get('context_notes', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No keynote-block relationships found</p>'
        html_content += '</div>'
    
    if 'keynote_detail_relationships' in data:
        html_content += '<div class="section"><h2>Keynote → Detail Relationships</h2>'
        if data['keynote_detail_relationships']:
            html_content += '<table><thead><tr><th>Project</th><th>Keynote</th><th>Detail Name</th><th>Notes</th></tr></thead><tbody>'
            for item in data['keynote_detail_relationships']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('keynote', '')}</td><td>{item.get('detail_name', '')}</td><td>{item.get('context_notes', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No keynote-detail relationships found</p>'
        html_content += '</div>'
    
    if 'hatch_material_relationships' in data:
        html_content += '<div class="section"><h2>Hatch → Material Relationships</h2>'
        if data['hatch_material_relationships']:
            html_content += '<table><thead><tr><th>Project</th><th>Hatch Pattern</th><th>Material</th><th>Notes</th></tr></thead><tbody>'
            for item in data['hatch_material_relationships']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('pattern_name', '')}</td><td>{item.get('material_name', '')}</td><td>{item.get('context_notes', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No hatch-material relationships found</p>'
        html_content += '</div>'
    
    if 'detail_material_relationships' in data:
        html_content += '<div class="section"><h2>Detail → Material Relationships</h2>'
        if data['detail_material_relationships']:
            html_content += '<table><thead><tr><th>Project</th><th>Detail Name</th><th>Material</th><th>Notes</th></tr></thead><tbody>'
            for item in data['detail_material_relationships']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('detail_name', '')}</td><td>{item.get('material_name', '')}</td><td>{item.get('context_notes', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No detail-material relationships found</p>'
        html_content += '</div>'
    
    if 'block_spec_relationships' in data:
        html_content += '<div class="section"><h2>Block → Specification Relationships</h2>'
        if data['block_spec_relationships']:
            html_content += '<table><thead><tr><th>Project</th><th>Block Name</th><th>Spec Section</th><th>Description</th><th>Manufacturer</th><th>Model</th></tr></thead><tbody>'
            for item in data['block_spec_relationships']:
                html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('block_name', '')}</td><td>{item.get('spec_section', '')}</td><td>{item.get('spec_description', '')}</td><td>{item.get('manufacturer', '')}</td><td>{item.get('model_number', '')}</td></tr>"
            html_content += '</tbody></table>'
        else:
            html_content += '<p class="empty">No block-specification relationships found</p>'
        html_content += '</div>'
    
    html_content += """
    </div>
</body>
</html>
"""
    
    # Return as file download
    return send_file(
        io.BytesIO(html_content.encode('utf-8')),
        mimetype='text/html',
        as_attachment=True,
        download_name=f'CAD_Standards_{datetime.now().strftime("%Y%m%d")}.html'
    )

def generate_pdf_export(data, title, description):
    """Generate PDF export from HTML"""
    # Generate HTML content first
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
</head>
<body>
    <h1>{title}</h1>
    <p>{description}</p>
    <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
"""
    
    # Add each data section
    if 'block_mappings' in data and data['block_mappings']:
        html_content += '<h2>Block Name Mappings</h2><table border="1" cellpadding="5" cellspacing="0" style="width:100%; border-collapse:collapse;"><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
        for item in data['block_mappings']:
            html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
        html_content += '</tbody></table>'
    
    if 'detail_mappings' in data and data['detail_mappings']:
        html_content += '<h2>Detail Name Mappings</h2><table border="1" cellpadding="5" cellspacing="0" style="width:100%; border-collapse:collapse;"><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
        for item in data['detail_mappings']:
            html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
        html_content += '</tbody></table>'
    
    if 'hatch_mappings' in data and data['hatch_mappings']:
        html_content += '<h2>Hatch Pattern Mappings</h2><table border="1" cellpadding="5" cellspacing="0" style="width:100%; border-collapse:collapse;"><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
        for item in data['hatch_mappings']:
            html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
        html_content += '</tbody></table>'
    
    if 'material_mappings' in data and data['material_mappings']:
        html_content += '<h2>Material Name Mappings</h2><table border="1" cellpadding="5" cellspacing="0" style="width:100%; border-collapse:collapse;"><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Name</th><th>Description</th></tr></thead><tbody>'
        for item in data['material_mappings']:
            html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('description', '')}</td></tr>"
        html_content += '</tbody></table>'
    
    if 'note_mappings' in data and data['note_mappings']:
        html_content += '<h2>Note/Keynote Mappings</h2><table border="1" cellpadding="5" cellspacing="0" style="width:100%; border-collapse:collapse;"><thead><tr><th>Project</th><th>Imported Name</th><th>Standard Note</th><th>Category</th></tr></thead><tbody>'
        for item in data['note_mappings']:
            html_content += f"<tr><td>{item.get('project_name', '')}</td><td>{item.get('imported_name', '')}</td><td>{item.get('standard_name', '')}</td><td>{item.get('category', '')}</td></tr>"
        html_content += '</tbody></table>'
    
    html_content += '</body></html>'
    
    # Convert to PDF using WeasyPrint
    pdf_bytes = HTML(string=html_content).write_pdf()
    
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'CAD_Standards_{datetime.now().strftime("%Y%m%d")}.pdf'
    )

# ============================================================================
# PROJECT ELEMENT CROSS-REFERENCES API ENDPOINTS
# ============================================================================

def validate_element_exists(element_type, element_id):
    """Validate that an element ID exists in the appropriate table for its type"""
    element_tables = {
        'keynote': ('standard_notes', 'note_id'),
        'block': ('block_definitions', 'block_id'),
        'detail': ('detail_standards', 'detail_id'),
        'hatch': ('hatch_patterns', 'hatch_id'),
        'material': ('material_standards', 'material_id')
    }
    
    if element_type not in element_tables:
        return False, f"Invalid element type: {element_type}"
    
    table_name, id_column = element_tables[element_type]
    
    try:
        query = f"SELECT 1 FROM {table_name} WHERE {id_column} = %s LIMIT 1"
        result = execute_query(query, (element_id,))
        if not result:
            return False, f"{element_type} with ID {element_id} not found in {table_name}"
        return True, None
    except Exception as e:
        return False, f"Validation error: {str(e)}"

@app.route('/api/project-context/cross-references', methods=['GET'])
def get_cross_references():
    """Get cross-reference relationships for a project"""
    try:
        project_id = request.args.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                xref_id,
                project_id,
                source_element_type,
                source_element_id,
                source_element_name,
                target_element_type,
                target_element_id,
                target_element_name,
                relationship_type,
                relationship_strength,
                context_description,
                sheet_references,
                tags,
                attributes,
                is_active,
                created_at,
                updated_at
            FROM project_element_cross_references
            WHERE project_id = %s AND is_active = TRUE
            ORDER BY created_at DESC
        """
        mappings = execute_query(query, (project_id,))
        return jsonify({'mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/cross-references', methods=['POST'])
def create_cross_reference():
    """Create a new cross-reference relationship"""
    try:
        data = request.get_json()
        
        # Validate source element
        source_valid, source_error = validate_element_exists(
            data.get('source_element_type'),
            data.get('source_element_id')
        )
        if not source_valid:
            return jsonify({'error': source_error}), 400
        
        # Validate target element
        target_valid, target_error = validate_element_exists(
            data.get('target_element_type'),
            data.get('target_element_id')
        )
        if not target_valid:
            return jsonify({'error': target_error}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO project_element_cross_references 
                    (project_id, source_element_type, source_element_id, source_element_name,
                     target_element_type, target_element_id, target_element_name,
                     relationship_type, relationship_strength, context_description,
                     sheet_references, tags)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING xref_id
                """, (
                    data.get('project_id'),
                    data.get('source_element_type'),
                    data.get('source_element_id'),
                    data.get('source_element_name'),
                    data.get('target_element_type'),
                    data.get('target_element_id'),
                    data.get('target_element_name'),
                    data.get('relationship_type'),
                    data.get('relationship_strength', 'normal'),
                    data.get('context_description'),
                    data.get('sheet_references', []),
                    data.get('tags', [])
                ))
                xref_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'xref_id': str(xref_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/cross-references/<xref_id>', methods=['GET'])
def get_cross_reference(xref_id):
    """Get a single cross-reference relationship"""
    try:
        query = """
            SELECT * FROM project_element_cross_references
            WHERE xref_id = %s
        """
        result = execute_query(query, (xref_id,))
        if not result:
            return jsonify({'error': 'Cross-reference not found'}), 404
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/cross-references/<xref_id>', methods=['PUT'])
def update_cross_reference(xref_id):
    """Update a cross-reference relationship"""
    try:
        data = request.get_json()
        
        # Validate source element
        source_valid, source_error = validate_element_exists(
            data.get('source_element_type'),
            data.get('source_element_id')
        )
        if not source_valid:
            return jsonify({'error': source_error}), 400
        
        # Validate target element
        target_valid, target_error = validate_element_exists(
            data.get('target_element_type'),
            data.get('target_element_id')
        )
        if not target_valid:
            return jsonify({'error': target_error}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE project_element_cross_references
                    SET source_element_type = %s, source_element_id = %s, source_element_name = %s,
                        target_element_type = %s, target_element_id = %s, target_element_name = %s,
                        relationship_type = %s, relationship_strength = %s, context_description = %s,
                        sheet_references = %s, tags = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE xref_id = %s
                """, (
                    data.get('source_element_type'),
                    data.get('source_element_id'),
                    data.get('source_element_name'),
                    data.get('target_element_type'),
                    data.get('target_element_id'),
                    data.get('target_element_name'),
                    data.get('relationship_type'),
                    data.get('relationship_strength'),
                    data.get('context_description'),
                    data.get('sheet_references', []),
                    data.get('tags', []),
                    xref_id
                ))
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/cross-references/<xref_id>', methods=['DELETE'])
def delete_cross_reference(xref_id):
    """Delete a cross-reference relationship"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM project_element_cross_references WHERE xref_id = %s", (xref_id,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-context/search-elements', methods=['GET'])
def search_elements():
    """Search for elements by type and query string (for typeahead)"""
    try:
        element_type = request.args.get('type')
        query_str = request.args.get('query', '')
        
        if not element_type:
            return jsonify({'error': 'type is required'}), 400
        
        results = []
        
        if element_type == 'keynote':
            query = """
                SELECT note_id as id, 
                       CONCAT(COALESCE(note_title, ''), ' - ', LEFT(note_text, 100)) as name,
                       note_title,
                       note_text
                FROM standard_notes
                WHERE (note_title ILIKE %s OR note_text ILIKE %s)
                ORDER BY note_title, note_text
                LIMIT 20
            """
            search_pattern = f'%{query_str}%'
            results = execute_query(query, (search_pattern, search_pattern))
            
        elif element_type == 'block':
            query = """
                SELECT block_id as id,
                       CONCAT(block_name, ' - ', COALESCE(description, '')) as name,
                       block_name,
                       description
                FROM block_definitions
                WHERE (block_name ILIKE %s OR description ILIKE %s)
                  AND is_active = TRUE
                ORDER BY block_name
                LIMIT 20
            """
            search_pattern = f'%{query_str}%'
            results = execute_query(query, (search_pattern, search_pattern))
            
        elif element_type == 'detail':
            query = """
                SELECT detail_id as id,
                       CONCAT(COALESCE(detail_number, ''), ' - ', COALESCE(detail_title, '')) as name,
                       detail_number,
                       detail_title,
                       description
                FROM detail_standards
                WHERE (detail_number ILIKE %s OR detail_title ILIKE %s OR description ILIKE %s)
                ORDER BY detail_number, detail_title
                LIMIT 20
            """
            search_pattern = f'%{query_str}%'
            results = execute_query(query, (search_pattern, search_pattern, search_pattern))
            
        elif element_type == 'hatch':
            query = """
                SELECT hatch_id as id,
                       CONCAT(pattern_name, ' - ', COALESCE(description, '')) as name,
                       pattern_name,
                       description
                FROM hatch_patterns
                WHERE (pattern_name ILIKE %s OR description ILIKE %s)
                  AND (is_active IS NULL OR is_active = TRUE)
                ORDER BY pattern_name
                LIMIT 20
            """
            search_pattern = f'%{query_str}%'
            results = execute_query(query, (search_pattern, search_pattern))
            
        elif element_type == 'material':
            query = """
                SELECT material_id as id,
                       CONCAT(material_name, ' - ', COALESCE(description, '')) as name,
                       material_name,
                       description
                FROM material_standards
                WHERE (material_name ILIKE %s OR description ILIKE %s)
                ORDER BY material_name
                LIMIT 20
            """
            search_pattern = f'%{query_str}%'
            results = execute_query(query, (search_pattern, search_pattern))
        else:
            return jsonify({'error': 'Invalid element type'}), 400
        
        return jsonify({'results': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# PIPE NETWORK EDITOR API ENDPOINTS
# ============================================================================

@app.route('/api/pipe-networks')
def get_pipe_networks():
    """Get list of all pipe networks, optionally filtered by project or mode"""
    try:
        project_id = request.args.get('project_id')
        network_mode = request.args.get('mode')
        
        query = """
            SELECT 
                pn.network_id,
                pn.project_id,
                pn.network_name,
                pn.utility_system,
                pn.network_mode,
                pn.network_status,
                pn.description,
                p.project_name,
                COUNT(DISTINCT unm_lines.line_id) as line_count,
                COUNT(DISTINCT unm_struct.structure_id) as structure_count
            FROM pipe_networks pn
            LEFT JOIN projects p ON pn.project_id = p.project_id
            LEFT JOIN utility_network_memberships unm_lines ON pn.network_id = unm_lines.network_id AND unm_lines.line_id IS NOT NULL
            LEFT JOIN utility_network_memberships unm_struct ON pn.network_id = unm_struct.network_id AND unm_struct.structure_id IS NOT NULL
            WHERE 1=1
        """
        params = []
        
        if project_id:
            query += " AND pn.project_id = %s"
            params.append(project_id)
        
        if network_mode:
            query += " AND pn.network_mode = %s"
            params.append(network_mode)
        
        query += """
            GROUP BY pn.network_id, pn.project_id, pn.network_name, pn.utility_system, 
                     pn.network_mode, pn.network_status, pn.description, p.project_name
            ORDER BY pn.created_at DESC
        """
        
        networks = execute_query(query, tuple(params) if params else None)
        return jsonify({'networks': networks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pipe-networks/<network_id>')
def get_network_details(network_id):
    """Get detailed information about a specific pipe network"""
    try:
        query = """
            SELECT 
                pn.network_id,
                pn.project_id,
                pn.network_name,
                pn.utility_system,
                pn.network_mode,
                pn.network_status,
                pn.description,
                pn.attributes,
                p.project_name,
                p.client_name
            FROM pipe_networks pn
            LEFT JOIN projects p ON pn.project_id = p.project_id
            WHERE pn.network_id = %s
        """
        network = execute_query(query, (network_id,))
        
        if not network or len(network) == 0:
            return jsonify({'error': 'Network not found'}), 404
        
        return jsonify({'network': network[0]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pipe-networks/<network_id>/pipes')
def get_network_pipes(network_id):
    """Get all pipes in a network with from/to structure information"""
    try:
        query = """
            SELECT 
                ul.line_id,
                ul.line_number,
                ul.utility_system,
                ul.material,
                ul.diameter_mm,
                ul.invert_elevation_start,
                ul.invert_elevation_end,
                ul.slope,
                ul.length,
                ul.from_structure_id,
                ul.to_structure_id,
                from_struct.structure_number as from_structure_number,
                from_struct.structure_type as from_structure_type,
                to_struct.structure_number as to_structure_number,
                to_struct.structure_type as to_structure_type,
                ST_AsGeoJSON(ul.geometry) as geometry,
                ul.attributes
            FROM utility_network_memberships unm
            JOIN utility_lines ul ON unm.line_id = ul.line_id
            LEFT JOIN utility_structures from_struct ON ul.from_structure_id = from_struct.structure_id
            LEFT JOIN utility_structures to_struct ON ul.to_structure_id = to_struct.structure_id
            WHERE unm.network_id = %s
            ORDER BY ul.line_number, ul.created_at
        """
        pipes = execute_query(query, (network_id,))
        return jsonify({'pipes': pipes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pipe-networks/<network_id>/structures')
def get_network_structures(network_id):
    """Get all structures in a network"""
    try:
        query = """
            SELECT 
                us.structure_id,
                us.structure_number,
                us.structure_type,
                us.utility_system,
                us.rim_elevation,
                us.invert_elevation,
                us.size_mm,
                us.material,
                us.manhole_depth_ft,
                us.condition,
                ST_AsGeoJSON(us.rim_geometry) as geometry,
                us.attributes
            FROM utility_network_memberships unm
            JOIN utility_structures us ON unm.structure_id = us.structure_id
            WHERE unm.network_id = %s
            ORDER BY us.structure_number, us.created_at
        """
        structures = execute_query(query, (network_id,))
        return jsonify({'structures': structures})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pipe-networks/<network_id>/pipes/<pipe_id>', methods=['PUT'])
def update_network_pipe(network_id, pipe_id):
    """Update pipe attributes (preserves geometry)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
        
        allowed_fields = [
            'line_number', 'material', 'diameter_mm', 
            'invert_elevation_start', 'invert_elevation_end', 
            'slope', 'from_structure_id', 'to_structure_id', 'notes'
        ]
        
        updates = []
        params = []
        for field in allowed_fields:
            if field in data:
                updates.append(f"{field} = %s")
                params.append(data[field])
        
        if not updates:
            return jsonify({'error': 'No valid fields to update'}), 400
        
        params.append(pipe_id)
        
        with get_db() as conn:
            with conn.cursor() as cur:
                query = f"""
                    UPDATE utility_lines 
                    SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP
                    WHERE line_id = %s
                    RETURNING line_id
                """
                cur.execute(query, tuple(params))
                result = cur.fetchone()
                conn.commit()
                
                if not result:
                    return jsonify({'error': 'Pipe not found'}), 404
                
                return jsonify({'success': True, 'line_id': str(result[0])})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pipe-networks/<network_id>/structures/<structure_id>', methods=['PUT'])
def update_network_structure(network_id, structure_id):
    """Update structure attributes (preserves geometry)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
        
        allowed_fields = [
            'structure_number', 'structure_type', 'rim_elevation', 
            'invert_elevation', 'size_mm', 'material', 
            'manhole_depth_ft', 'condition', 'notes'
        ]
        
        updates = []
        params = []
        for field in allowed_fields:
            if field in data:
                updates.append(f"{field} = %s")
                params.append(data[field])
        
        if not updates:
            return jsonify({'error': 'No valid fields to update'}), 400
        
        params.append(structure_id)
        
        with get_db() as conn:
            with conn.cursor() as cur:
                query = f"""
                    UPDATE utility_structures 
                    SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP
                    WHERE structure_id = %s
                    RETURNING structure_id
                """
                cur.execute(query, tuple(params))
                result = cur.fetchone()
                conn.commit()
                
                if not result:
                    return jsonify({'error': 'Structure not found'}), 404
                
                return jsonify({'success': True, 'structure_id': str(result[0])})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pipe-networks/<network_id>/auto-connect', methods=['POST'])
def auto_connect_pipes(network_id):
    """Automatically connect pipes to nearby structures using spatial snapping"""
    try:
        from pipe_structure_connector import PipeStructureConnector
        
        data = request.get_json() or {}
        tolerance_feet = data.get('tolerance_feet', 5.0)
        
        connector = PipeStructureConnector(tolerance_feet=tolerance_feet)
        results = connector.connect_network_pipes(network_id)
        
        if results.get('success'):
            return jsonify(results), 200
        else:
            return jsonify(results), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/pipe-networks/<network_id>/viewer-entities')
def get_network_viewer_entities(network_id):
    """Get network entities (pipes + structures) in Entity Viewer format with transformed geometries"""
    try:
        all_entities = []
        
        # Fetch pipes
        pipes_query = """
            SELECT 
                ul.line_id::text as entity_id,
                ul.line_number as label,
                'pipe' as entity_type,
                ul.utility_system,
                ul.material,
                ul.diameter_mm,
                ul.slope,
                ul.length,
                ul.from_structure_id::text,
                ul.to_structure_id::text,
                ST_AsGeoJSON(ST_Transform(ul.geometry, 4326))::json as geometry
            FROM utility_network_memberships unm
            JOIN utility_lines ul ON unm.line_id = ul.line_id
            WHERE unm.network_id = %s AND ul.geometry IS NOT NULL
            ORDER BY ul.line_number
        """
        pipes = execute_query(pipes_query, (network_id,))
        
        for pipe in pipes:
            all_entities.append({
                'entity_id': pipe['entity_id'],
                'entity_type': 'pipe',
                'label': pipe['label'] or 'Unnamed Pipe',
                'layer_name': pipe['utility_system'] or 'Unknown',
                'category': 'Utilities',
                'geometry_type': 'line',
                'color': '#f7b801' if pipe['utility_system'] == 'Storm' else '#0096ff',
                'geometry': pipe['geometry'],
                'properties': {
                    'material': pipe['material'],
                    'diameter_mm': pipe['diameter_mm'],
                    'slope': pipe['slope'],
                    'length': pipe['length'],
                    'from_structure': pipe['from_structure_id'],
                    'to_structure': pipe['to_structure_id']
                }
            })
        
        # Fetch structures
        structures_query = """
            SELECT 
                us.structure_id::text as entity_id,
                us.structure_number as label,
                'structure' as entity_type,
                us.structure_type,
                us.utility_system,
                us.rim_elevation,
                us.invert_elevation,
                us.manhole_depth_ft,
                us.condition,
                ST_AsGeoJSON(ST_Transform(us.rim_geometry, 4326))::json as geometry
            FROM utility_network_memberships unm
            JOIN utility_structures us ON unm.structure_id = us.structure_id
            WHERE unm.network_id = %s AND us.rim_geometry IS NOT NULL
            ORDER BY us.structure_number
        """
        structures = execute_query(structures_query, (network_id,))
        
        for struct in structures:
            all_entities.append({
                'entity_id': struct['entity_id'],
                'entity_type': 'structure',
                'label': struct['label'] or 'Unnamed Structure',
                'layer_name': struct['utility_system'] or 'Unknown',
                'category': 'Utilities',
                'geometry_type': 'point',
                'color': '#6a994e',
                'geometry': struct['geometry'],
                'properties': {
                    'type': struct['structure_type'],
                    'rim_elevation': struct['rim_elevation'],
                    'invert_elevation': struct['invert_elevation'],
                    'depth_ft': struct['manhole_depth_ft'],
                    'condition': struct['condition']
                }
            })
        
        # Calculate combined bounding box
        bbox = None
        if all_entities:
            bbox_query = """
                SELECT ST_Extent(
                    ST_Transform(
                        ST_Union(
                            ARRAY[
                                (SELECT ST_Collect(ul.geometry) 
                                 FROM utility_network_memberships unm
                                 JOIN utility_lines ul ON unm.line_id = ul.line_id
                                 WHERE unm.network_id = %s AND ul.geometry IS NOT NULL),
                                (SELECT ST_Collect(us.rim_geometry) 
                                 FROM utility_network_memberships unm
                                 JOIN utility_structures us ON unm.structure_id = us.structure_id
                                 WHERE unm.network_id = %s AND us.rim_geometry IS NOT NULL)
                            ]
                        ),
                        4326
                    )
                ) as bbox
            """
            bbox_result = execute_query(bbox_query, (network_id, network_id))
            
            if bbox_result and bbox_result[0]['bbox']:
                bbox_str = bbox_result[0]['bbox']
                coords = bbox_str.replace('BOX(', '').replace(')', '').split(',')
                min_coords = coords[0].strip().split()
                max_coords = coords[1].strip().split()
                bbox = {
                    'minX': float(min_coords[0]),
                    'minY': float(min_coords[1]),
                    'maxX': float(max_coords[0]),
                    'maxY': float(max_coords[1])
                }
        
        # Calculate counts
        type_counts = {}
        layer_counts = {}
        for entity in all_entities:
            etype = entity.get('entity_type', 'Unknown')
            type_counts[etype] = type_counts.get(etype, 0) + 1
            
            layer = entity.get('layer_name', 'Unknown')
            layer_counts[layer] = layer_counts.get(layer, 0) + 1
        
        return jsonify({
            'entities': all_entities,
            'bbox': bbox,
            'type_counts': type_counts,
            'layer_counts': layer_counts,
            'total_count': len(all_entities)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# SCHEMA RELATIONSHIPS ROUTES
# ============================================================================

@app.route('/schema/relationships')
def schema_relationships():
    """Render the Schema Relationships visualization page"""
    return render_template('schema_relationships.html')

@app.route('/why-this-matters')
def why_this_matters():
    """Render the Why This Matters comparison page"""
    return render_template('why_this_matters.html')

@app.route('/architecture')
def architecture():
    """Render the System Architecture diagram page"""
    return render_template('architecture.html')

@app.route('/digital-twin')
def digital_twin():
    """Render the Digital Twin concept page"""
    return render_template('digital_twin.html')

@app.route('/evolution')
def evolution():
    """Render the Evolution timeline page"""
    return render_template('evolution.html')

@app.route('/standards/layer-generator')
def layer_generator():
    """Render the CAD Layer Generator page"""
    return render_template('layer_generator.html')

@app.route('/tools/dxf-test-generator')
def dxf_test_generator():
    """Render the DXF Test Generator page"""
    return render_template('tools/dxf_test_generator.html')

@app.route('/api/z-stress-test/run', methods=['POST'])
def run_z_stress_test():
    """Run Z-value elevation preservation stress test"""
    import sys
    import os
    import tempfile
    from werkzeug.utils import secure_filename
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    
    from scripts.z_stress_harness import ZStressHarness
    
    try:
        user_dxf_path = None
        
        # Check if this is a file upload or JSON request
        if 'file' in request.files:
            # File upload mode
            file = request.files['file']
            
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            if not file.filename.lower().endswith('.dxf'):
                return jsonify({'error': 'File must be a DXF file'}), 400
            
            # Save uploaded file to temp directory
            filename = secure_filename(file.filename)
            user_dxf_path = os.path.join(tempfile.gettempdir(), f'upload_{filename}')
            file.save(user_dxf_path)
            
            # Get parameters from form data
            num_cycles = int(request.form.get('num_cycles', 20))
            srid = int(request.form.get('srid', 0))
            tolerance = float(request.form.get('tolerance', 0.001))
        else:
            # JSON mode (test fixtures)
            data = request.get_json(silent=True)
            if not data:
                return jsonify({'error': 'Invalid request'}), 400
            
            num_cycles = data.get('num_cycles', 20)
            srid = data.get('srid', 0)
            tolerance = data.get('tolerance', 0.001)
        
        # Validate parameters
        if num_cycles < 1 or num_cycles > 100:
            return jsonify({'error': 'Number of cycles must be between 1 and 100'}), 400
        
        if srid not in [0, 2226, 4326]:
            return jsonify({'error': 'SRID must be 0 (local CAD), 2226 (CA State Plane), or 4326 (WGS84)'}), 400
        
        if tolerance < 0 or tolerance > 1:
            return jsonify({'error': 'Tolerance must be between 0 and 1 feet'}), 400
        
        # Run the stress test
        harness = ZStressHarness()
        results = harness.run_stress_test(
            num_cycles=num_cycles,
            srid=srid,
            tolerance_ft=tolerance,
            user_dxf_path=user_dxf_path
        )
        
        # Sanitize results to prevent NaN/Infinity JSON errors
        sanitized_results = harness._sanitize_for_json(results)
        
        # DON'T clean up uploaded file - let harness manage it for multi-cycle testing
        
        return jsonify(sanitized_results)
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print(f"Z-Stress Test Error: {error_msg}")
        print(error_trace)
        
        # Clean up uploaded file on error
        if 'user_dxf_path' in locals() and user_dxf_path and os.path.exists(user_dxf_path):
            try:
                os.remove(user_dxf_path)
            except:
                pass
        
        return jsonify({'error': error_msg}), 500

@app.route('/api/tools/generate-test-dxf', methods=['POST'])
def generate_test_dxf():
    """Generate a test DXF file with sample geometries and network objects"""
    try:
        import ezdxf
        import random
        
        data = request.get_json(silent=True)
        if not data:
            return jsonify({'error': 'Invalid JSON payload'}), 400
            
        geometries = data.get('geometries', [])
        networks = data.get('networks', [])
        bmps = data.get('bmps', [])
        layer_mode = data.get('layerMode', 'canonical')
        complexity = data.get('complexity', 'basic')
        
        # Determine feature count based on complexity
        feature_counts = {
            'minimal': 2,
            'basic': 5,
            'complex': 10
        }
        count = feature_counts.get(complexity, 5)
        
        # Santa Rosa, California coordinates in SRID 2226 (State Plane CA Zone 2, US Survey Feet)
        # 2286 Chanate Road, Santa Rosa, CA (Lat: 38.4468, Lon: -122.7047)
        base_x = 6359843  # feet
        base_y = 1925319  # feet
        
        # Create new DXF document (AutoCAD 2013)
        doc = ezdxf.new('AC1027')
        msp = doc.modelspace()
        
        # Helper function for random offset
        def random_offset():
            return random.uniform(-500, 500)
        
        # Generate basic geometries
        if 'AXIS' in geometries:  # Lines/Centerlines
            layer_name = 'CIV-ROAD-CNTR-EXST-AXIS' if layer_mode == 'canonical' else 'Centerlines'
            doc.layers.new(name=layer_name, dxfattribs={'color': 3})  # Green
            for i in range(count):
                x1 = base_x + random_offset()
                y1 = base_y + random_offset()
                x2 = x1 + random.uniform(50, 200)
                y2 = y1 + random.uniform(-50, 50)
                msp.add_line((x1, y1), (x2, y2), dxfattribs={'layer': layer_name})
        
        if 'PT' in geometries:  # Points
            layer_name = 'SURV-CTRL-MONUMENT-EXST-PT' if layer_mode == 'canonical' else 'Points'
            doc.layers.new(name=layer_name, dxfattribs={'color': 1})  # Red
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                msp.add_point((x, y), dxfattribs={'layer': layer_name})
        
        if 'AREA' in geometries:  # Areas/Polygons
            layer_name = 'SITE-GRAD-PAD-EXST-AREA' if layer_mode == 'canonical' else 'Areas'
            doc.layers.new(name=layer_name, dxfattribs={'color': 5})  # Blue
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                size = random.uniform(30, 80)
                points = [
                    (x, y),
                    (x + size, y),
                    (x + size, y + size),
                    (x, y + size),
                    (x, y)
                ]
                msp.add_lwpolyline(points, close=True, dxfattribs={'layer': layer_name})
        
        if 'TEXT' in geometries:  # Text Labels
            layer_name = 'SITE-ANNO-LABL-EXST-TEXT' if layer_mode == 'canonical' else 'Labels'
            doc.layers.new(name=layer_name, dxfattribs={'color': 7})  # White
            labels = ['TEST POINT', 'SAMPLE AREA', 'REFERENCE', 'BENCHMARK', 'CONTROL POINT']
            for i in range(min(count, len(labels))):
                x = base_x + random_offset()
                y = base_y + random_offset()
                msp.add_text(labels[i], dxfattribs={
                    'layer': layer_name,
                    'height': 5.0,
                    'insert': (x, y)
                })
        
        if 'SYMB' in geometries:  # Blocks/Symbols
            layer_name = 'SITE-UTIL-VALVE-EXST-SYMB' if layer_mode == 'canonical' else 'Symbols'
            doc.layers.new(name=layer_name, dxfattribs={'color': 6})  # Magenta
            # Create a simple block (circle symbol)
            if 'VALVE_SYMBOL' not in doc.blocks:
                blk = doc.blocks.new(name='VALVE_SYMBOL')
                blk.add_circle((0, 0), radius=3)
                blk.add_line((-3, 0), (3, 0))
                blk.add_line((0, -3), (0, 3))
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                msp.add_blockref('VALVE_SYMBOL', (x, y), dxfattribs={'layer': layer_name})
        
        if 'HATCH' in geometries:  # Hatches
            layer_name = 'SITE-LAND-GRASS-EXST-HATCH' if layer_mode == 'canonical' else 'Hatches'
            doc.layers.new(name=layer_name, dxfattribs={'color': 92})  # Light green
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                size = random.uniform(25, 60)
                hatch = msp.add_hatch(color=92, dxfattribs={'layer': layer_name})
                hatch.paths.add_polyline_path([
                    (x, y), (x + size, y), (x + size, y + size), (x, y + size)
                ], is_closed=True)
                hatch.set_pattern_fill('ANSI31', scale=0.5)
        
        # Generate network objects
        if 'GRAV' in networks:  # Gravity Pipe System
            pipe_layer = 'CIV-PIPE-GRAV-EXST-AXIS' if layer_mode == 'canonical' else 'Gravity_Pipes'
            mh_layer = 'CIV-STRC-MH-EXST-SYMB' if layer_mode == 'canonical' else 'Manholes'
            doc.layers.new(name=pipe_layer, dxfattribs={'color': 2})  # Yellow
            doc.layers.new(name=mh_layer, dxfattribs={'color': 1})  # Red
            
            # Create simple manhole block
            if 'MH_SYMBOL' not in doc.blocks:
                blk = doc.blocks.new(name='MH_SYMBOL')
                blk.add_circle((0, 0), radius=2.5)
                blk.add_text('MH', dxfattribs={'height': 1.5, 'insert': (-1, -0.75)})
            
            # Generate pipe network
            x_start = base_x - 100
            y_start = base_y + 100
            for i in range(count):
                x1 = x_start + (i * 50)
                y1 = y_start - (i * 30)
                x2 = x1 + 50
                y2 = y1 - 30
                # Add pipe
                msp.add_line((x1, y1), (x2, y2), dxfattribs={'layer': pipe_layer})
                # Add manholes
                msp.add_blockref('MH_SYMBOL', (x1, y1), dxfattribs={'layer': mh_layer})
                if i == count - 1:
                    msp.add_blockref('MH_SYMBOL', (x2, y2), dxfattribs={'layer': mh_layer})
        
        if 'PRES' in networks:  # Pressure Pipe System
            pipe_layer = 'CIV-PIPE-PRES-EXST-AXIS' if layer_mode == 'canonical' else 'Pressure_Pipes'
            valve_layer = 'CIV-STRC-VALVE-EXST-SYMB' if layer_mode == 'canonical' else 'Valves'
            doc.layers.new(name=pipe_layer, dxfattribs={'color': 4})  # Cyan
            doc.layers.new(name=valve_layer, dxfattribs={'color': 6})  # Magenta
            
            # Create valve symbol
            if 'VALVE_SYM' not in doc.blocks:
                blk = doc.blocks.new(name='VALVE_SYM')
                blk.add_circle((0, 0), radius=2)
                blk.add_line((-2, 0), (2, 0))
            
            # Generate pressure network
            x_start = base_x + 100
            y_start = base_y - 100
            for i in range(count):
                x1 = x_start + (i * 40)
                y1 = y_start + (i * 20)
                x2 = x1 + 40
                y2 = y1 + 20
                msp.add_line((x1, y1), (x2, y2), dxfattribs={'layer': pipe_layer})
                msp.add_blockref('VALVE_SYM', (x1, y1), dxfattribs={'layer': valve_layer})
        
        if 'STORM' in networks:  # Storm Drain Network
            pipe_layer = 'CIV-PIPE-STORM-EXST-AXIS' if layer_mode == 'canonical' else 'Storm_Pipes'
            inlet_layer = 'CIV-STRC-INLET-EXST-SYMB' if layer_mode == 'canonical' else 'Inlets'
            doc.layers.new(name=pipe_layer, dxfattribs={'color': 3})  # Green
            doc.layers.new(name=inlet_layer, dxfattribs={'color': 5})  # Blue
            
            # Create inlet symbol
            if 'INLET_SYM' not in doc.blocks:
                blk = doc.blocks.new(name='INLET_SYM')
                blk.add_lwpolyline([(0, 2), (1.5, 0), (0, -2), (-1.5, 0), (0, 2)])
                blk.add_text('I', dxfattribs={'height': 1, 'insert': (-0.3, -0.5)})
            
            # Generate storm network
            for i in range(count):
                x1 = base_x + random_offset()
                y1 = base_y + random_offset()
                x2 = x1 + random.uniform(30, 70)
                y2 = y1 - random.uniform(20, 40)
                msp.add_line((x1, y1), (x2, y2), dxfattribs={'layer': pipe_layer})
                msp.add_blockref('INLET_SYM', (x1, y1), dxfattribs={'layer': inlet_layer})
        
        if 'WATER' in networks:  # Water Distribution
            pipe_layer = 'CIV-PIPE-WATER-EXST-AXIS' if layer_mode == 'canonical' else 'Water_Pipes'
            hydrant_layer = 'CIV-STRC-HYDRANT-EXST-SYMB' if layer_mode == 'canonical' else 'Hydrants'
            doc.layers.new(name=pipe_layer, dxfattribs={'color': 4})  # Cyan
            doc.layers.new(name=hydrant_layer, dxfattribs={'color': 1})  # Red
            
            # Create hydrant symbol
            if 'HYDRANT_SYM' not in doc.blocks:
                blk = doc.blocks.new(name='HYDRANT_SYM')
                blk.add_circle((0, 0), radius=2.5)
                blk.add_text('H', dxfattribs={'height': 1.5, 'insert': (-0.6, -0.75)})
            
            # Generate water network in grid pattern
            grid_size = 60
            for i in range(int(count / 2) + 1):
                for j in range(2):
                    x = base_x + (i * grid_size) - 150
                    y = base_y + (j * grid_size) - 150
                    if i < count / 2:
                        msp.add_line((x, y), (x + grid_size, y), dxfattribs={'layer': pipe_layer})
                    if j == 0:
                        msp.add_line((x, y), (x, y + grid_size), dxfattribs={'layer': pipe_layer})
                    if i % 2 == 0:
                        msp.add_blockref('HYDRANT_SYM', (x, y), dxfattribs={'layer': hydrant_layer})
        
        # Generate BMP features
        if 'BIOR' in bmps:  # Bioretention
            layer_name = 'SITE-BMP-BIOR-EXST-AREA' if layer_mode == 'canonical' else 'Bioretention'
            doc.layers.new(name=layer_name, dxfattribs={'color': 82})  # Light green
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                size_x = random.uniform(20, 40)
                size_y = random.uniform(15, 30)
                points = [
                    (x, y), (x + size_x, y), (x + size_x, y + size_y), (x, y + size_y), (x, y)
                ]
                msp.add_lwpolyline(points, close=True, dxfattribs={'layer': layer_name})
                msp.add_text('BIORETENTION', dxfattribs={
                    'layer': layer_name, 'height': 3.0, 'insert': (x + 2, y + size_y/2)
                })
        
        if 'SWAL' in bmps:  # Bioswale
            layer_name = 'SITE-BMP-SWAL-EXST-AXIS' if layer_mode == 'canonical' else 'Bioswale'
            doc.layers.new(name=layer_name, dxfattribs={'color': 93})  # Light cyan
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                length = random.uniform(40, 80)
                width = random.uniform(8, 15)
                # Create swale outline
                points = [
                    (x, y), (x + length, y), (x + length, y + width), (x, y + width), (x, y)
                ]
                msp.add_lwpolyline(points, close=True, dxfattribs={'layer': layer_name})
                # Add centerline
                msp.add_line((x, y + width/2), (x + length, y + width/2), 
                           dxfattribs={'layer': layer_name, 'linetype': 'DASHED'})
        
        if 'POND' in bmps:  # Detention Pond
            layer_name = 'SITE-BMP-POND-EXST-AREA' if layer_mode == 'canonical' else 'Detention_Pond'
            doc.layers.new(name=layer_name, dxfattribs={'color': 150})  # Blue
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                radius = random.uniform(25, 50)
                # Create irregular pond shape using multiple points
                angle_step = 360 / 12
                points = []
                for angle in range(0, 360, int(angle_step)):
                    rad = radius + random.uniform(-5, 5)
                    import math
                    px = x + rad * math.cos(math.radians(angle))
                    py = y + rad * math.sin(math.radians(angle))
                    points.append((px, py))
                points.append(points[0])
                msp.add_lwpolyline(points, close=True, dxfattribs={'layer': layer_name})
                msp.add_text('DETENTION POND', dxfattribs={
                    'layer': layer_name, 'height': 4.0, 'insert': (x - 15, y)
                })
        
        if 'INFIL' in bmps:  # Infiltration Basin
            layer_name = 'SITE-BMP-INFIL-EXST-AREA' if layer_mode == 'canonical' else 'Infiltration'
            doc.layers.new(name=layer_name, dxfattribs={'color': 52})  # Dark green
            for i in range(count):
                x = base_x + random_offset()
                y = base_y + random_offset()
                size_x = random.uniform(30, 50)
                size_y = random.uniform(20, 35)
                points = [
                    (x, y), (x + size_x, y), (x + size_x, y + size_y), (x, y + size_y), (x, y)
                ]
                msp.add_lwpolyline(points, close=True, dxfattribs={'layer': layer_name})
                # Add hatch pattern
                hatch = msp.add_hatch(color=52, dxfattribs={'layer': layer_name})
                hatch.paths.add_polyline_path(points, is_closed=True)
                hatch.set_pattern_fill('EARTH', scale=0.3)
        
        # Save to temporary file
        temp_path = f'/tmp/test-data-{datetime.now().strftime("%Y%m%d-%H%M%S")}.dxf'
        doc.saveas(temp_path)
        
        # Return file for download
        return send_file(
            temp_path,
            mimetype='application/dxf',
            as_attachment=True,
            download_name=f'test-data-{datetime.now().strftime("%Y%m%d")}.dxf'
        )
        
    except Exception as e:
        print(f"Error generating DXF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/schema/relationships', methods=['GET'])
def get_schema_relationships():
    """Get table relationships and metadata for visualization"""
    try:
        # Get all tables with estimated row counts (much faster than COUNT(*))
        tables_query = """
            SELECT 
                t.table_name,
                obj_description((quote_ident(t.table_schema)||'.'||quote_ident(t.table_name))::regclass, 'pg_class') as table_description,
                (SELECT count(*) FROM information_schema.columns c 
                 WHERE c.table_name = t.table_name AND c.table_schema = t.table_schema) as column_count,
                COALESCE(
                    (SELECT n_live_tup 
                     FROM pg_stat_user_tables 
                     WHERE schemaname = 'public' AND relname = t.table_name),
                    0
                ) as row_count
            FROM information_schema.tables t
            WHERE t.table_schema = 'public' 
            AND t.table_type = 'BASE TABLE'
            ORDER BY t.table_name
        """
        tables = execute_query(tables_query)
        
        # Get foreign key relationships
        fk_query = """
            SELECT
                tc.table_name as source_table,
                kcu.column_name as source_column,
                ccu.table_name AS target_table,
                ccu.column_name AS target_column,
                tc.constraint_name
            FROM information_schema.table_constraints AS tc
            JOIN information_schema.key_column_usage AS kcu
                ON tc.constraint_name = kcu.constraint_name
                AND tc.table_schema = kcu.table_schema
            JOIN information_schema.constraint_column_usage AS ccu
                ON ccu.constraint_name = tc.constraint_name
                AND ccu.table_schema = tc.table_schema
            WHERE tc.constraint_type = 'FOREIGN KEY'
            AND tc.table_schema = 'public'
            ORDER BY tc.table_name
        """
        relationships = execute_query(fk_query)
        
        # Get column details for each table
        columns_query = """
            SELECT 
                c.table_name,
                c.column_name,
                c.data_type,
                c.is_nullable,
                CASE 
                    WHEN pk.column_name IS NOT NULL THEN true 
                    ELSE false 
                END as is_primary_key
            FROM information_schema.columns c
            LEFT JOIN (
                SELECT ku.table_name, ku.column_name
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage ku
                    ON tc.constraint_name = ku.constraint_name
                WHERE tc.constraint_type = 'PRIMARY KEY'
                AND tc.table_schema = 'public'
            ) pk ON c.table_name = pk.table_name AND c.column_name = pk.column_name
            WHERE c.table_schema = 'public'
            ORDER BY c.table_name, c.ordinal_position
        """
        all_columns = execute_query(columns_query)
        
        # Group columns by table
        columns_by_table = {}
        for col in all_columns:
            table_name = col['table_name']
            if table_name not in columns_by_table:
                columns_by_table[table_name] = []
            columns_by_table[table_name].append(col)
        
        return jsonify({
            'tables': tables,
            'relationships': relationships,
            'columns': columns_by_table
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# SHEET NOTE MANAGER
# ============================================

@app.route('/sheet-notes')
def sheet_notes_page():
    """Sheet Note Manager page"""
    return render_template('sheet_notes.html')

# ============================================
# GRAVITY PIPE NETWORK MANAGER
# ============================================

@app.route('/gravity-network-manager')
def gravity_network_manager():
    """Gravity Pipe Network Manager page"""
    return render_template('gravity_network_manager.html')

# ============================================
# PRESSURE PIPE NETWORK MANAGER
# ============================================

@app.route('/pressure-network-manager')
def pressure_network_manager():
    """Pressure Pipe Network Manager page"""
    return render_template('pressure_network_manager.html')

# ============================================
# BMP MANAGER
# ============================================

@app.route('/bmp-manager')
def bmp_manager():
    """BMP Manager page"""
    return render_template('bmp_manager.html')

@app.route('/api/pressure-networks')
def get_pressure_networks():
    """Get list of pressure pipe networks (potable, reclaimed, fire)"""
    try:
        project_id = request.args.get('project_id')
        
        query = """
            SELECT 
                pn.network_id,
                pn.project_id,
                pn.network_name,
                pn.utility_system,
                pn.network_mode,
                pn.network_status,
                pn.description,
                p.project_name,
                COUNT(DISTINCT unm_lines.line_id) as line_count,
                COUNT(DISTINCT unm_struct.structure_id) as structure_count
            FROM pipe_networks pn
            LEFT JOIN projects p ON pn.project_id = p.project_id
            LEFT JOIN utility_network_memberships unm_lines ON pn.network_id = unm_lines.network_id AND unm_lines.line_id IS NOT NULL
            LEFT JOIN utility_network_memberships unm_struct ON pn.network_id = unm_struct.network_id AND unm_struct.structure_id IS NOT NULL
            WHERE pn.utility_system IN ('potable', 'Potable', 'reclaimed', 'Reclaimed', 'fire', 'Fire')
        """
        params = []
        
        if project_id:
            query += " AND pn.project_id = %s"
            params.append(project_id)
        
        query += """
            GROUP BY pn.network_id, pn.project_id, pn.network_name, 
                     pn.utility_system, pn.network_mode, pn.network_status, 
                     pn.description, p.project_name
            ORDER BY p.project_name, pn.network_name
        """
        
        networks = execute_query(query, tuple(params) if params else None)
        return jsonify({'networks': networks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pressure-networks/<network_id>')
def get_pressure_network_details(network_id):
    """Get detailed information about a specific pressure network"""
    try:
        query = """
            SELECT 
                pn.network_id,
                pn.project_id,
                pn.network_name,
                pn.utility_system,
                pn.network_mode,
                pn.network_status,
                pn.description,
                pn.attributes,
                p.project_name,
                p.client_name
            FROM pipe_networks pn
            LEFT JOIN projects p ON pn.project_id = p.project_id
            WHERE pn.network_id = %s
        """
        network = execute_query(query, (network_id,))
        
        if not network or len(network) == 0:
            return jsonify({'error': 'Network not found'}), 404
        
        return jsonify({'network': network[0]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pressure-networks/<network_id>/pipes')
def get_pressure_network_pipes(network_id):
    """Get all pipes in a pressure network with pressure-specific data"""
    try:
        query = """
            SELECT 
                ul.line_id,
                ul.line_number,
                ul.utility_system,
                ul.material,
                ul.diameter_mm,
                ul.length,
                ul.from_structure_id,
                ul.to_structure_id,
                from_struct.structure_number as from_structure_number,
                from_struct.structure_type as from_structure_type,
                to_struct.structure_number as to_structure_number,
                to_struct.structure_type as to_structure_type,
                ST_AsGeoJSON(ul.geometry) as geometry,
                ul.attributes,
                json_build_object(
                    'pressure_rating_psi', ulpd.pressure_rating_psi,
                    'pipe_class', ulpd.pipe_class,
                    'operating_pressure_psi', ulpd.operating_pressure_psi,
                    'flow_direction', ulpd.flow_direction
                ) as pressure_data
            FROM utility_network_memberships unm
            JOIN utility_lines ul ON unm.line_id = ul.line_id
            LEFT JOIN utility_structures from_struct ON ul.from_structure_id = from_struct.structure_id
            LEFT JOIN utility_structures to_struct ON ul.to_structure_id = to_struct.structure_id
            LEFT JOIN utility_line_pressure_data ulpd ON ul.line_id = ulpd.line_id
            WHERE unm.network_id = %s
            ORDER BY ul.line_number, ul.created_at
        """
        pipes = execute_query(query, (network_id,))
        return jsonify({'pipes': pipes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pressure-networks/<network_id>/structures')
def get_pressure_network_structures(network_id):
    """Get all structures in a pressure network with pressure-specific data"""
    try:
        query = """
            SELECT 
                us.structure_id,
                us.structure_number,
                us.structure_type,
                us.utility_system,
                us.rim_elevation,
                us.condition,
                ST_AsGeoJSON(us.rim_geometry) as geometry,
                us.attributes,
                json_build_object(
                    'valve_type', uspd.valve_type,
                    'valve_status', uspd.valve_status,
                    'hydrant_flow_gpm', uspd.hydrant_flow_gpm,
                    'operating_pressure_psi', uspd.operating_pressure_psi,
                    'pressure_class', uspd.pressure_class
                ) as pressure_data
            FROM utility_network_memberships unm
            JOIN utility_structures us ON unm.structure_id = us.structure_id
            LEFT JOIN utility_structure_pressure_data uspd ON us.structure_id = uspd.structure_id
            WHERE unm.network_id = %s
            ORDER BY us.structure_number, us.created_at
        """
        structures = execute_query(query, (network_id,))
        return jsonify({'structures': structures})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pressure-networks/<network_id>/pipes/<pipe_id>', methods=['PUT'])
def update_pressure_network_pipe(network_id, pipe_id):
    """Update pressure pipe attributes (preserves geometry)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                base_fields = ['line_number', 'material', 'diameter_mm', 'from_structure_id', 'to_structure_id', 'notes']
                pressure_fields = ['pressure_rating_psi', 'pipe_class', 'operating_pressure_psi', 'flow_direction']
                
                base_updates = []
                base_params = []
                for field in base_fields:
                    if field in data:
                        base_updates.append(f"{field} = %s")
                        base_params.append(data[field])
                
                if base_updates:
                    base_params.append(pipe_id)
                    base_query = f"""
                        UPDATE utility_lines 
                        SET {', '.join(base_updates)}, updated_at = CURRENT_TIMESTAMP
                        WHERE line_id = %s
                        RETURNING line_id
                    """
                    cur.execute(base_query, tuple(base_params))
                    result = cur.fetchone()
                    if not result:
                        return jsonify({'error': 'Pipe not found'}), 404
                
                pressure_updates = []
                pressure_params = []
                for field in pressure_fields:
                    if field in data:
                        pressure_updates.append(f"{field} = %s")
                        pressure_params.append(data[field])
                
                if pressure_updates:
                    pressure_params.extend([pipe_id, pipe_id])
                    pressure_query = f"""
                        INSERT INTO utility_line_pressure_data (line_id, {', '.join([f.split('=')[0].strip() for f in pressure_updates])})
                        VALUES (%s, {', '.join(['%s'] * len(pressure_updates))})
                        ON CONFLICT (line_id) 
                        DO UPDATE SET {', '.join(pressure_updates)}, updated_at = CURRENT_TIMESTAMP
                    """
                    all_pressure_params = [pipe_id] + [data[f] for f in pressure_fields if f in data]
                    cur.execute(f"""
                        INSERT INTO utility_line_pressure_data (line_id, {', '.join(pressure_fields)})
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (line_id) 
                        DO UPDATE SET {', '.join(pressure_updates)}, updated_at = CURRENT_TIMESTAMP
                    """, tuple(all_pressure_params))
                
                conn.commit()
                return jsonify({'success': True, 'line_id': str(pipe_id)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pressure-networks/<network_id>/structures/<structure_id>', methods=['PUT'])
def update_pressure_network_structure(network_id, structure_id):
    """Update pressure structure attributes (preserves geometry)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request body'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                base_fields = ['structure_number', 'structure_type', 'condition', 'notes']
                pressure_fields = ['valve_type', 'valve_status', 'hydrant_flow_gpm', 'operating_pressure_psi', 'pressure_class']
                
                base_updates = []
                base_params = []
                for field in base_fields:
                    if field in data:
                        base_updates.append(f"{field} = %s")
                        base_params.append(data[field])
                
                if base_updates:
                    base_params.append(structure_id)
                    base_query = f"""
                        UPDATE utility_structures 
                        SET {', '.join(base_updates)}, updated_at = CURRENT_TIMESTAMP
                        WHERE structure_id = %s
                        RETURNING structure_id
                    """
                    cur.execute(base_query, tuple(base_params))
                    result = cur.fetchone()
                    if not result:
                        return jsonify({'error': 'Structure not found'}), 404
                
                pressure_updates = []
                pressure_params = []
                for field in pressure_fields:
                    if field in data:
                        pressure_updates.append(f"{field} = %s")
                        pressure_params.append(data[field])
                
                if pressure_updates:
                    all_pressure_params = [structure_id] + [data[f] for f in pressure_fields if f in data]
                    cur.execute(f"""
                        INSERT INTO utility_structure_pressure_data (structure_id, {', '.join(pressure_fields)})
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT (structure_id) 
                        DO UPDATE SET {', '.join(pressure_updates)}, updated_at = CURRENT_TIMESTAMP
                    """, tuple(all_pressure_params))
                
                conn.commit()
                return jsonify({'success': True, 'structure_id': str(structure_id)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pressure-networks/<network_id>/viewer-entities')
def get_pressure_network_viewer_entities(network_id):
    """Get pressure network entities (pipes + structures) in Entity Viewer format with transformed geometries"""
    try:
        all_entities = []
        
        pipes_query = """
            SELECT 
                ul.line_id::text as entity_id,
                ul.line_number as label,
                LOWER(ul.utility_system) || '_pipe' as entity_type,
                ul.utility_system,
                ul.material,
                ul.diameter_mm,
                ul.length,
                ul.from_structure_id::text,
                ul.to_structure_id::text,
                ulpd.pressure_rating_psi,
                ulpd.pipe_class,
                ulpd.operating_pressure_psi,
                ulpd.flow_direction,
                ST_AsGeoJSON(ST_Transform(ul.geometry, 4326))::json as geometry
            FROM utility_network_memberships unm
            JOIN utility_lines ul ON unm.line_id = ul.line_id
            LEFT JOIN utility_line_pressure_data ulpd ON ul.line_id = ulpd.line_id
            WHERE unm.network_id = %s AND ul.geometry IS NOT NULL
            ORDER BY ul.line_number
        """
        pipes = execute_query(pipes_query, (network_id,))
        
        for pipe in pipes:
            utility_system = (pipe['utility_system'] or 'unknown').lower()
            if utility_system == 'potable':
                color = '#0096ff'
            elif utility_system == 'reclaimed':
                color = '#9b59b6'
            elif utility_system == 'fire':
                color = '#e74c3c'
            else:
                color = '#0096ff'
            
            all_entities.append({
                'entity_id': pipe['entity_id'],
                'entity_type': pipe['entity_type'],
                'label': pipe['label'] or 'Unnamed Pipe',
                'layer_name': pipe['utility_system'] or 'Unknown',
                'category': 'Pressure Systems',
                'geometry_type': 'line',
                'color': color,
                'geometry': pipe['geometry'],
                'properties': {
                    'material': pipe['material'],
                    'diameter_mm': pipe['diameter_mm'],
                    'length': pipe['length'],
                    'pressure_rating_psi': pipe['pressure_rating_psi'],
                    'pipe_class': pipe['pipe_class'],
                    'operating_pressure_psi': pipe['operating_pressure_psi'],
                    'flow_direction': pipe['flow_direction'],
                    'from_structure': pipe['from_structure_id'],
                    'to_structure': pipe['to_structure_id']
                }
            })
        
        structures_query = """
            SELECT 
                us.structure_id::text as entity_id,
                us.structure_number as label,
                CASE 
                    WHEN us.structure_type ILIKE '%valve%' THEN 'valve'
                    WHEN us.structure_type ILIKE '%hydrant%' THEN 'hydrant'
                    ELSE 'structure'
                END as entity_type,
                us.structure_type,
                us.utility_system,
                us.rim_elevation,
                us.condition,
                uspd.valve_type,
                uspd.valve_status,
                uspd.hydrant_flow_gpm,
                uspd.operating_pressure_psi,
                uspd.pressure_class,
                ST_AsGeoJSON(ST_Transform(us.rim_geometry, 4326))::json as geometry
            FROM utility_network_memberships unm
            JOIN utility_structures us ON unm.structure_id = us.structure_id
            LEFT JOIN utility_structure_pressure_data uspd ON us.structure_id = uspd.structure_id
            WHERE unm.network_id = %s AND us.rim_geometry IS NOT NULL
            ORDER BY us.structure_number
        """
        structures = execute_query(structures_query, (network_id,))
        
        for struct in structures:
            entity_type = struct['entity_type']
            if entity_type == 'valve':
                color = '#f39c12'
            elif entity_type == 'hydrant':
                color = '#e74c3c'
            else:
                color = '#6a994e'
            
            all_entities.append({
                'entity_id': struct['entity_id'],
                'entity_type': entity_type,
                'label': struct['label'] or 'Unnamed Structure',
                'layer_name': struct['utility_system'] or 'Unknown',
                'category': 'Pressure Systems',
                'geometry_type': 'point',
                'color': color,
                'geometry': struct['geometry'],
                'properties': {
                    'type': struct['structure_type'],
                    'rim_elevation': struct['rim_elevation'],
                    'condition': struct['condition'],
                    'valve_type': struct['valve_type'],
                    'valve_status': struct['valve_status'],
                    'hydrant_flow_gpm': struct['hydrant_flow_gpm'],
                    'operating_pressure_psi': struct['operating_pressure_psi'],
                    'pressure_class': struct['pressure_class']
                }
            })
        
        bbox = None
        if all_entities:
            bbox_query = """
                SELECT ST_Extent(
                    ST_Transform(
                        ST_Union(
                            ARRAY[
                                (SELECT ST_Collect(ul.geometry) 
                                 FROM utility_network_memberships unm
                                 JOIN utility_lines ul ON unm.line_id = ul.line_id
                                 WHERE unm.network_id = %s AND ul.geometry IS NOT NULL),
                                (SELECT ST_Collect(us.rim_geometry) 
                                 FROM utility_network_memberships unm
                                 JOIN utility_structures us ON unm.structure_id = us.structure_id
                                 WHERE unm.network_id = %s AND us.rim_geometry IS NOT NULL)
                            ]
                        ),
                        4326
                    )
                ) as bbox
            """
            bbox_result = execute_query(bbox_query, (network_id, network_id))
            
            if bbox_result and bbox_result[0]['bbox']:
                bbox_str = bbox_result[0]['bbox']
                coords = bbox_str.replace('BOX(', '').replace(')', '').split(',')
                min_coords = coords[0].strip().split()
                max_coords = coords[1].strip().split()
                bbox = {
                    'minX': float(min_coords[0]),
                    'minY': float(min_coords[1]),
                    'maxX': float(max_coords[0]),
                    'maxY': float(max_coords[1])
                }
        
        type_counts = {}
        layer_counts = {}
        for entity in all_entities:
            etype = entity.get('entity_type', 'Unknown')
            type_counts[etype] = type_counts.get(etype, 0) + 1
            
            layer = entity.get('layer_name', 'Unknown')
            layer_counts[layer] = layer_counts.get(layer, 0) + 1
        
        return jsonify({
            'entities': all_entities,
            'bbox': bbox,
            'type_counts': type_counts,
            'layer_counts': layer_counts,
            'total_count': len(all_entities)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# BMP API ENDPOINTS
# ============================================

@app.route('/api/bmps')
def get_bmps():
    """Get list of all BMPs"""
    try:
        query = """
            SELECT 
                bmp_id,
                bmp_name,
                bmp_type,
                bmp_status,
                treatment_volume_cf,
                storage_volume_cf,
                infiltration_rate_in_hr,
                bypass_elevation_ft,
                maintenance_schedule,
                drainage_area_acres,
                design_storm,
                created_at,
                updated_at
            FROM storm_bmps
            ORDER BY bmp_name
        """
        bmps = execute_query(query)
        return jsonify({'bmps': bmps})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps/<bmp_id>')
def get_bmp_details(bmp_id):
    """Get detailed information about a specific BMP"""
    try:
        query = """
            SELECT 
                bmp_id,
                bmp_name,
                bmp_type,
                bmp_status,
                treatment_volume_cf,
                storage_volume_cf,
                infiltration_rate_in_hr,
                bypass_elevation_ft,
                overflow_structure_id,
                maintenance_schedule,
                drainage_area_acres,
                design_storm,
                attributes,
                created_at,
                updated_at
            FROM storm_bmps
            WHERE bmp_id = %s
        """
        bmp = execute_query(query, (bmp_id,))
        
        if not bmp or len(bmp) == 0:
            return jsonify({'error': 'BMP not found'}), 404
        
        return jsonify({'bmp': bmp[0]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps/<bmp_id>', methods=['PUT'])
def update_bmp(bmp_id):
    """Update BMP attributes"""
    try:
        data = request.get_json()
        
        allowed_fields = [
            'bmp_name', 'bmp_type', 'bmp_status', 'treatment_volume_cf',
            'storage_volume_cf', 'infiltration_rate_in_hr', 'bypass_elevation_ft',
            'maintenance_schedule', 'drainage_area_acres', 'design_storm'
        ]
        
        updates = []
        params = []
        
        for field in allowed_fields:
            if field in data:
                updates.append(f"{field} = %s")
                params.append(data[field])
        
        if not updates:
            return jsonify({'error': 'No valid fields to update'}), 400
        
        params.append(bmp_id)
        
        with get_db() as conn:
            with conn.cursor() as cur:
                query = f"""
                    UPDATE storm_bmps 
                    SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP
                    WHERE bmp_id = %s
                    RETURNING bmp_id
                """
                cur.execute(query, tuple(params))
                result = cur.fetchone()
                
                if not result:
                    return jsonify({'error': 'BMP not found'}), 404
                
                conn.commit()
        
        return jsonify({'success': True, 'bmp_id': str(result[0])})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps', methods=['POST'])
def create_bmp():
    """Create a new BMP"""
    try:
        data = request.get_json()
        
        required_fields = ['bmp_name', 'bmp_type']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO storm_bmps (
                        bmp_name, bmp_type, bmp_status, treatment_volume_cf,
                        storage_volume_cf, infiltration_rate_in_hr, bypass_elevation_ft,
                        maintenance_schedule, drainage_area_acres, design_storm, geometry
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            ST_GeomFromText('MULTIPOLYGON EMPTY', 3857))
                    RETURNING bmp_id
                """, (
                    data.get('bmp_name'),
                    data.get('bmp_type'),
                    data.get('bmp_status', 'active'),
                    data.get('treatment_volume_cf'),
                    data.get('storage_volume_cf'),
                    data.get('infiltration_rate_in_hr'),
                    data.get('bypass_elevation_ft'),
                    data.get('maintenance_schedule'),
                    data.get('drainage_area_acres'),
                    data.get('design_storm')
                ))
                bmp_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'bmp_id': str(bmp_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps/<bmp_id>', methods=['DELETE'])
def delete_bmp(bmp_id):
    """Delete a BMP"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM storm_bmps WHERE bmp_id = %s RETURNING bmp_id", (bmp_id,))
                result = cur.fetchone()
                
                if not result:
                    return jsonify({'error': 'BMP not found'}), 404
                
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps/<bmp_id>/connections')
def get_bmp_connections(bmp_id):
    """Get all inflow/outflow connections for a BMP"""
    try:
        query = """
            SELECT 
                bio.bmp_io_id,
                bio.bmp_id,
                bio.connection_type,
                bio.connected_structure_id,
                bio.connected_line_id,
                bio.invert_elevation_ft,
                bio.notes,
                us.structure_number,
                us.structure_type,
                ul.line_number,
                ul.material,
                ul.diameter_mm
            FROM bmp_inflow_outflow bio
            LEFT JOIN utility_structures us ON bio.connected_structure_id = us.structure_id
            LEFT JOIN utility_lines ul ON bio.connected_line_id = ul.line_id
            WHERE bio.bmp_id = %s
            ORDER BY bio.connection_type, bio.created_at
        """
        connections = execute_query(query, (bmp_id,))
        return jsonify({'connections': connections})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps/<bmp_id>/connections', methods=['POST'])
def add_bmp_connection(bmp_id):
    """Add a new connection to a BMP"""
    try:
        data = request.get_json()
        
        if 'connection_type' not in data:
            return jsonify({'error': 'Missing required field: connection_type'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO bmp_inflow_outflow (
                        bmp_id, connection_type, connected_structure_id,
                        connected_line_id, invert_elevation_ft, notes
                    )
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING bmp_io_id
                """, (
                    bmp_id,
                    data.get('connection_type'),
                    data.get('connected_structure_id'),
                    data.get('connected_line_id'),
                    data.get('invert_elevation_ft'),
                    data.get('notes')
                ))
                connection_id = cur.fetchone()[0]
                conn.commit()
        
        return jsonify({'bmp_io_id': str(connection_id)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps/<bmp_id>/connections/<connection_id>', methods=['DELETE'])
def delete_bmp_connection(bmp_id, connection_id):
    """Delete a BMP connection"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM bmp_inflow_outflow 
                    WHERE bmp_io_id = %s AND bmp_id = %s
                    RETURNING bmp_io_id
                """, (connection_id, bmp_id))
                result = cur.fetchone()
                
                if not result:
                    return jsonify({'error': 'Connection not found'}), 404
                
                conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bmps/<bmp_id>/viewer-entities')
def get_bmp_viewer_entities(bmp_id):
    """Get BMP geometry and connections in Entity Viewer format with EPSG:4326 transformation"""
    try:
        all_entities = []
        
        bmp_query = """
            SELECT 
                bmp_id::text as entity_id,
                bmp_name as label,
                'bmp_polygon' as entity_type,
                bmp_type,
                bmp_status,
                treatment_volume_cf,
                storage_volume_cf,
                infiltration_rate_in_hr,
                bypass_elevation_ft,
                maintenance_schedule,
                drainage_area_acres,
                design_storm,
                ST_AsGeoJSON(ST_Transform(geometry, 4326))::json as geometry
            FROM storm_bmps
            WHERE bmp_id = %s AND geometry IS NOT NULL
        """
        bmp = execute_query(bmp_query, (bmp_id,))
        
        if bmp and len(bmp) > 0:
            bmp_data = bmp[0]
            
            bmp_type_colors = {
                'bioswale': '#2ecc71',
                'detention_basin': '#3498db',
                'retention_basin': '#1abc9c',
                'infiltration_trench': '#95a5a6',
                'permeable_pavement': '#34495e',
                'rain_garden': '#27ae60'
            }
            color = bmp_type_colors.get(bmp_data['bmp_type'], '#00ffff')
            
            all_entities.append({
                'entity_id': bmp_data['entity_id'],
                'entity_type': bmp_data['bmp_type'] or 'bmp_polygon',
                'label': bmp_data['label'] or 'Unnamed BMP',
                'layer_name': 'BMP',
                'category': 'Stormwater',
                'geometry_type': 'polygon',
                'color': color,
                'geometry': bmp_data['geometry'],
                'properties': {
                    'bmp_name': bmp_data['label'],
                    'bmp_type': bmp_data['bmp_type'],
                    'bmp_status': bmp_data['bmp_status'],
                    'treatment_volume_cf': bmp_data['treatment_volume_cf'],
                    'storage_volume_cf': bmp_data['storage_volume_cf'],
                    'infiltration_rate_in_hr': bmp_data['infiltration_rate_in_hr'],
                    'bypass_elevation_ft': bmp_data['bypass_elevation_ft'],
                    'drainage_area_acres': bmp_data['drainage_area_acres'],
                    'design_storm': bmp_data['design_storm']
                }
            })
        
        connections_query = """
            SELECT 
                bio.bmp_io_id::text,
                bio.connection_type,
                bio.connected_structure_id::text,
                bio.connected_line_id::text,
                us.structure_number,
                us.structure_type,
                ST_AsGeoJSON(ST_Transform(us.rim_geometry, 4326))::json as structure_geometry,
                ul.line_number,
                ul.material,
                ul.diameter_mm,
                ST_AsGeoJSON(ST_Transform(ul.geometry, 4326))::json as line_geometry
            FROM bmp_inflow_outflow bio
            LEFT JOIN utility_structures us ON bio.connected_structure_id = us.structure_id
            LEFT JOIN utility_lines ul ON bio.connected_line_id = ul.line_id
            WHERE bio.bmp_id = %s 
              AND (us.rim_geometry IS NOT NULL OR ul.geometry IS NOT NULL)
        """
        connections = execute_query(connections_query, (bmp_id,))
        
        for conn in connections:
            if conn['structure_geometry']:
                all_entities.append({
                    'entity_id': conn['connected_structure_id'],
                    'entity_type': 'connected_structure',
                    'label': conn['structure_number'] or 'Structure',
                    'layer_name': 'Connected Structure',
                    'category': 'Connection',
                    'geometry_type': 'point',
                    'color': '#ff9900',
                    'geometry': conn['structure_geometry'],
                    'properties': {
                        'structure_number': conn['structure_number'],
                        'structure_type': conn['structure_type'],
                        'connection_type': conn['connection_type']
                    }
                })
            
            if conn['line_geometry']:
                all_entities.append({
                    'entity_id': conn['connected_line_id'],
                    'entity_type': 'connected_line',
                    'label': conn['line_number'] or 'Line',
                    'layer_name': 'Connected Line',
                    'category': 'Connection',
                    'geometry_type': 'line',
                    'color': '#9966ff',
                    'geometry': conn['line_geometry'],
                    'properties': {
                        'line_number': conn['line_number'],
                        'material': conn['material'],
                        'diameter_mm': conn['diameter_mm'],
                        'connection_type': conn['connection_type']
                    }
                })
        
        bbox = None
        if all_entities:
            bbox_query = """
                SELECT ST_Extent(ST_Transform(geometry, 4326)) as bbox
                FROM storm_bmps
                WHERE bmp_id = %s AND geometry IS NOT NULL
            """
            bbox_result = execute_query(bbox_query, (bmp_id,))
            
            if bbox_result and bbox_result[0]['bbox']:
                bbox_str = bbox_result[0]['bbox']
                coords = bbox_str.replace('BOX(', '').replace(')', '').split(',')
                min_coords = coords[0].strip().split()
                max_coords = coords[1].strip().split()
                bbox = {
                    'minX': float(min_coords[0]),
                    'minY': float(min_coords[1]),
                    'maxX': float(max_coords[0]),
                    'maxY': float(max_coords[1])
                }
        
        type_counts = {}
        for entity in all_entities:
            etype = entity.get('entity_type', 'Unknown')
            type_counts[etype] = type_counts.get(etype, 0) + 1
        
        return jsonify({
            'entities': all_entities,
            'bbox': bbox,
            'type_counts': type_counts,
            'total_count': len(all_entities)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# DXF IMPORT/EXPORT
# ============================================

@app.route('/dxf-tools')
def dxf_tools_page():
    """DXF Import/Export tools page"""
    return render_template('dxf_tools.html')

@app.route('/api/dxf/import', methods=['POST'])
def import_dxf():
    """Import DXF file into database - with optional pattern-based classification"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith('.dxf'):
            return jsonify({'error': 'File must be a DXF file'}), 400
        
        # Get parameters
        project_id = request.form.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        import_modelspace = request.form.get('import_modelspace', 'true') == 'true'
        import_paperspace = request.form.get('import_paperspace', 'true') == 'true'
        pattern_id = request.form.get('pattern_id')  # Optional import pattern
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join('/tmp', f'{uuid.uuid4()}_{filename}')
        file.save(temp_path)
        
        try:
            # If pattern_id is provided, use intelligent object creation
            # NOTE: Currently the DXFImporter uses its own pattern matching logic
            # The selected pattern_id serves as a hint but automatic detection is used
            use_intelligent = pattern_id is not None
            
            # Import DXF with intelligent object creation enabled
            importer = DXFImporter(DB_CONFIG, create_intelligent_objects=use_intelligent)
            stats = importer.import_dxf(
                temp_path,
                project_id,
                import_modelspace=import_modelspace,
                import_paperspace=import_paperspace
            )
            
            # Add pattern matching info if pattern was selected
            pattern_info = None
            if pattern_id and use_intelligent:
                # Get selected pattern details
                pattern_query = "SELECT client_name, source_pattern, confidence_score FROM import_mapping_patterns WHERE mapping_id = %s"
                pattern_result = execute_query(pattern_query, (pattern_id,))
                
                if pattern_result:
                    pattern = pattern_result[0]
                    # Get actual stats from import
                    total_layers = len(stats.get('layers', set()))
                    matched = stats.get('intelligent_objects_created', 0)
                    
                    pattern_info = {
                        'pattern_name': f"{pattern['client_name']} - {pattern['source_pattern']}",
                        'matched_count': matched,
                        'unmatched_count': max(0, total_layers - (1 if matched > 0 else 0)),
                        'avg_confidence': pattern['confidence_score'],
                        'note': 'Using automatic pattern detection with intelligent object creation'
                    }
            
            response_data = {
                'success': len(stats['errors']) == 0,
                'stats': stats
            }
            
            if pattern_info:
                response_data['pattern_info'] = pattern_info
            
            return jsonify(response_data)
        
        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dxf/export', methods=['POST'])
def export_dxf():
    """Export project to DXF file"""
    try:
        data = request.get_json()
        
        project_id = data.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        dxf_version = data.get('dxf_version', 'AC1027')
        include_modelspace = data.get('include_modelspace', True)
        include_paperspace = data.get('include_paperspace', True)
        layer_filter = data.get('layer_filter')
        
        # Generate output file
        output_filename = f'project_{project_id}_{uuid.uuid4().hex[:8]}.dxf'
        output_path = os.path.join('/tmp', output_filename)
        
        # Export DXF
        exporter = DXFExporter(DB_CONFIG)
        stats = exporter.export_dxf(
            project_id,
            output_path,
            dxf_version=dxf_version,
            include_modelspace=include_modelspace,
            include_paperspace=include_paperspace,
            layer_filter=layer_filter
        )
        
        if not os.path.exists(output_path):
            return jsonify({'error': 'Export failed - file not created'}), 500
        
        # Send file for download
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/dxf'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dxf/export-jobs', methods=['GET'])
def get_export_jobs():
    """Get export job history from map export jobs table"""
    try:
        query = """
            SELECT 
                ej.id as export_job_id,
                ej.params->>'project_id' as project_id,
                p.project_name,
                ej.params->>'export_format' as export_format,
                ej.params->>'dxf_version' as dxf_version,
                ej.status,
                ej.params->>'metrics' as metrics,
                ej.created_at as started_at,
                ej.expires_at as completed_at,
                ej.download_url,
                ej.file_size_mb,
                ej.error_message
            FROM export_jobs ej
            LEFT JOIN projects p ON (ej.params->>'project_id')::uuid = p.project_id
            ORDER BY ej.created_at DESC
            LIMIT 50
        """
        
        jobs = execute_query(query)
        return jsonify({'jobs': jobs})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# INTELLIGENT DXF WORKFLOW API
# ============================================

@app.route('/api/dxf/import-intelligent', methods=['POST'])
def import_intelligent_dxf():
    """Import DXF file with intelligent object creation from layer patterns"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith('.dxf'):
            return jsonify({'error': 'File must be a DXF file'}), 400
        
        # Get parameters
        drawing_id = request.form.get('drawing_id')
        if not drawing_id:
            return jsonify({'error': 'drawing_id is required'}), 400
        
        import_modelspace = request.form.get('import_modelspace', 'true') == 'true'
        import_paperspace = request.form.get('import_paperspace', 'true') == 'true'
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join('/tmp', f'{uuid.uuid4()}_{filename}')
        file.save(temp_path)
        
        try:
            # Import DXF with intelligent object creation enabled
            importer = DXFImporter(DB_CONFIG, create_intelligent_objects=True)
            stats = importer.import_dxf(
                temp_path,
                drawing_id,
                import_modelspace=import_modelspace,
                import_paperspace=import_paperspace
            )
            
            return jsonify({
                'success': len(stats['errors']) == 0,
                'stats': stats,
                'message': f"Imported {stats.get('entities', 0)} entities and created {stats.get('intelligent_objects_created', 0)} intelligent objects"
            })
        
        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dxf/export-intelligent', methods=['POST'])
def export_intelligent_dxf():
    """Export intelligent objects from a project to DXF file with proper layer names"""
    try:
        data = request.get_json()
        
        project_id = data.get('project_id')
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        include_types = data.get('include_types')  # Optional filter
        
        # Generate output file
        output_filename = f'project_{project_id}_{uuid.uuid4().hex[:8]}.dxf'
        output_path = os.path.join('/tmp', output_filename)
        
        # Export intelligent objects
        exporter = DXFExporter(DB_CONFIG)
        stats = exporter.export_intelligent_objects_to_dxf(
            project_id,
            output_path,
            include_types=include_types
        )
        
        if not os.path.exists(output_path):
            return jsonify({'error': 'Export failed - file not created'}), 500
        
        # Send file for download
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/dxf'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dxf/reimport', methods=['POST'])
def reimport_dxf_with_changes():
    """Re-import a DXF file and detect/merge changes to intelligent objects"""
    try:
        from dxf_change_detector import DXFChangeDetector
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith('.dxf'):
            return jsonify({'error': 'File must be a DXF file'}), 400
        
        # Get parameters
        drawing_id = request.form.get('drawing_id')
        if not drawing_id:
            return jsonify({'error': 'drawing_id is required'}), 400
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join('/tmp', f'{uuid.uuid4()}_{filename}')
        file.save(temp_path)
        
        try:
            # Step 1: Import DXF entities (without creating new intelligent objects yet)
            importer = DXFImporter(DB_CONFIG, create_intelligent_objects=False)
            import_stats = importer.import_dxf(
                temp_path,
                drawing_id,
                import_modelspace=True,
                import_paperspace=True
            )
            
            # Step 2: Get the reimported entities from database
            with get_db() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("""
                        SELECT 
                            entity_id,
                            entity_type,
                            layer_name,
                            ST_AsText(geometry) as geometry_wkt,
                            ST_GeometryType(geometry) as geometry_type,
                            dxf_handle
                        FROM drawing_entities
                        WHERE drawing_id = %s
                    """, (drawing_id,))
                    
                    reimported_entities = [dict(row) for row in cur.fetchall()]
            
            # Step 3: Detect changes
            detector = DXFChangeDetector(DB_CONFIG)
            change_stats = detector.detect_changes(drawing_id, reimported_entities)
            
            return jsonify({
                'success': len(import_stats['errors']) == 0 and len(change_stats['errors']) == 0,
                'import_stats': import_stats,
                'change_stats': change_stats,
                'message': f"Reimported {import_stats.get('entities', 0)} entities, detected {change_stats['geometry_changes']} geometry changes, {change_stats['layer_changes']} layer changes, {change_stats['new_entities']} new entities ({change_stats['new_objects_created']} intelligent objects created), {change_stats['deleted_entities']} deletions"
            })
        
        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# PROJECT SHEET NOTES API (project-scoped, not drawing-scoped)
# ============================================

@app.route('/api/project-sheet-notes', methods=['GET'])
def get_project_sheet_notes():
    """Get all project sheet notes for a set"""
    try:
        set_id = request.args.get('set_id')
        
        if not set_id:
            return jsonify({'error': 'set_id is required'}), 400
        
        query = """
            SELECT 
                psn.project_note_id,
                psn.set_id,
                psn.standard_note_id,
                psn.display_code,
                psn.custom_title,
                psn.custom_text,
                psn.is_modified,
                psn.sort_order,
                psn.usage_count,
                psn.source_type,
                psn.deviation_reason,
                psn.standardization_note,
                psn.first_used_at,
                psn.last_used_at,
                sn.note_title as standard_title,
                sn.note_text as standard_text,
                sn.note_category,
                sn.discipline,
                dc.category_code as deviation_category_code,
                dc.category_name as deviation_category_name,
                cs.status_code as conformance_status_code,
                cs.status_name as conformance_status_name,
                cs.color_hex as conformance_color,
                ss.status_code as standardization_status_code,
                ss.status_name as standardization_status_name
            FROM project_sheet_notes psn
            LEFT JOIN standard_notes sn ON psn.standard_note_id = sn.note_id
            LEFT JOIN deviation_categories dc ON psn.deviation_category_id = dc.category_id
            LEFT JOIN conformance_statuses cs ON psn.conformance_status_id = cs.status_id
            LEFT JOIN standardization_statuses ss ON psn.standardization_status_id = ss.status_id
            WHERE psn.set_id = %s::uuid
            ORDER BY psn.sort_order, psn.display_code
        """
        
        notes = execute_query(query, (set_id,))
        return jsonify({'notes': notes})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes/<project_note_id>', methods=['GET'])
def get_project_sheet_note(project_note_id):
    """Get a single project sheet note by ID"""
    try:
        query = """
            SELECT 
                psn.*,
                sn.note_title as standard_title,
                sn.note_text as standard_text,
                sn.note_category,
                sn.discipline,
                dc.category_code as deviation_category_code,
                dc.category_name as deviation_category_name,
                cs.status_code as conformance_status_code,
                cs.status_name as conformance_status_name,
                cs.color_hex as conformance_color,
                ss.status_code as standardization_status_code,
                ss.status_name as standardization_status_name
            FROM project_sheet_notes psn
            LEFT JOIN standard_notes sn ON psn.standard_note_id = sn.note_id
            LEFT JOIN deviation_categories dc ON psn.deviation_category_id = dc.category_id
            LEFT JOIN conformance_statuses cs ON psn.conformance_status_id = cs.status_id
            LEFT JOIN standardization_statuses ss ON psn.standardization_status_id = ss.status_id
            WHERE psn.project_note_id = %s::uuid
        """
        
        notes = execute_query(query, (project_note_id,))
        
        if not notes:
            return jsonify({'error': 'Note not found'}), 404
        
        return jsonify({'note': notes[0]})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes', methods=['POST'])
def create_project_sheet_note():
    """Create a new project sheet note"""
    try:
        data = request.get_json()
        
        set_id = data.get('set_id')
        standard_note_id = data.get('standard_note_id')
        display_code = data.get('display_code')
        custom_title = data.get('custom_title')
        custom_text = data.get('custom_text')
        
        if not set_id or not display_code:
            return jsonify({'error': 'set_id and display_code are required'}), 400
        
        is_custom = standard_note_id is None
        
        if is_custom and (not custom_title or not custom_text):
            return jsonify({'error': 'custom_title and custom_text required for custom notes'}), 400
        
        if not is_custom and not standard_note_id:
            return jsonify({'error': 'standard_note_id required for standard notes'}), 400
        
        project_note_id = str(uuid.uuid4())
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Get next sort order
                cur.execute('SELECT COALESCE(MAX(sort_order), 0) + 1 as next_order FROM project_sheet_notes WHERE set_id = %s::uuid', (set_id,))
                next_sort_order = cur.fetchone()['next_order']
                
                # Get conformance status based on whether it's custom or standard
                if is_custom:
                    cur.execute("SELECT status_id FROM conformance_statuses WHERE status_code = 'NON_STANDARD'")
                    conformance_status = cur.fetchone()
                else:
                    cur.execute("SELECT status_id FROM conformance_statuses WHERE status_code = 'FULL_COMPLIANCE'")
                    conformance_status = cur.fetchone()
                
                # Get standardization status for not nominated
                cur.execute("SELECT status_id FROM standardization_statuses WHERE status_code = 'NOT_NOMINATED'")
                standardization_status = cur.fetchone()
                
                cur.execute("""
                    INSERT INTO project_sheet_notes 
                    (project_note_id, set_id, standard_note_id, standard_reference_id, display_code, custom_title, custom_text, 
                     source_type, sort_order, conformance_status_id, standardization_status_id,
                     first_used_at, last_used_at, usage_count)
                    VALUES (%s::uuid, %s::uuid, %s, %s, %s, %s, %s, %s, %s, %s::uuid, %s::uuid, 
                            CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, 0)
                    RETURNING project_note_id, set_id, standard_note_id, display_code, custom_title, custom_text, 
                              is_modified, sort_order, usage_count, source_type
                """, (project_note_id, set_id, standard_note_id, standard_note_id, display_code, custom_title, custom_text,
                      'custom' if is_custom else 'standard', next_sort_order, conformance_status['status_id'], 
                      standardization_status['status_id']))
                
                new_note = dict(cur.fetchone())
                conn.commit()
                
                cache.clear()
                return jsonify({'note': new_note}), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes/<project_note_id>', methods=['PUT'])
def update_project_sheet_note(project_note_id):
    """Update an existing project sheet note"""
    try:
        data = request.get_json()
        
        display_code = data.get('display_code')
        custom_title = data.get('custom_title')
        custom_text = data.get('custom_text')
        is_modified = data.get('is_modified')
        
        updates = []
        params = []
        
        if display_code is not None:
            updates.append('display_code = %s')
            params.append(display_code)
        if custom_title is not None:
            updates.append('custom_title = %s')
            params.append(custom_title)
        if custom_text is not None:
            updates.append('custom_text = %s')
            params.append(custom_text)
        if is_modified is not None:
            updates.append('is_modified = %s')
            params.append(is_modified)
        
        if not updates:
            return jsonify({'error': 'No fields to update'}), 400
        
        params.append(project_note_id)
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(f"""
                    UPDATE project_sheet_notes 
                    SET {', '.join(updates)}
                    WHERE project_note_id = %s::uuid
                    RETURNING project_note_id, set_id, standard_note_id, display_code, custom_title, custom_text, is_modified, sort_order, usage_count
                """, params)
                
                updated_note = cur.fetchone()
                if not updated_note:
                    return jsonify({'error': 'Note not found'}), 404
                
                conn.commit()
                cache.clear()
                return jsonify({'note': dict(updated_note)})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes/<project_note_id>', methods=['DELETE'])
def delete_project_sheet_note(project_note_id):
    """Delete a project sheet note"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute('DELETE FROM project_sheet_notes WHERE project_note_id = %s::uuid', (project_note_id,))
                
                if cur.rowcount == 0:
                    return jsonify({'error': 'Note not found'}), 404
                
                conn.commit()
                cache.clear()
                return jsonify({'message': 'Note deleted successfully'})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes/<project_note_id>/reorder', methods=['PATCH'])
def reorder_project_sheet_note(project_note_id):
    """Update sort order for a project sheet note"""
    try:
        data = request.get_json()
        new_sort_order = data.get('sort_order')
        
        if new_sort_order is None:
            return jsonify({'error': 'sort_order is required'}), 400
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    UPDATE project_sheet_notes 
                    SET sort_order = %s
                    WHERE project_note_id = %s::uuid
                    RETURNING project_note_id, sort_order
                """, (new_sort_order, project_note_id))
                
                updated_note = cur.fetchone()
                if not updated_note:
                    return jsonify({'error': 'Note not found'}), 404
                
                conn.commit()
                cache.clear()
                return jsonify({'note': dict(updated_note)})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-note-assignments', methods=['GET'])
def get_sheet_note_assignments():
    """Get sheet note assignments by project or by project note"""
    try:
        project_id = request.args.get('project_id')
        project_note_id = request.args.get('project_note_id')
        
        if project_note_id:
            query = """
                SELECT 
                    sna.assignment_id,
                    sna.project_note_id,
                    sna.project_id,
                    sna.legend_sequence,
                    p.project_name
                FROM sheet_note_assignments sna
                LEFT JOIN projects p ON sna.project_id = p.project_id
                WHERE sna.project_note_id = %s::uuid
                ORDER BY p.project_name, sna.legend_sequence
            """
            assignments = execute_query(query, (project_note_id,))
        
        elif project_id:
            query = """
                SELECT 
                    sna.assignment_id,
                    sna.project_note_id,
                    sna.project_id,
                    sna.legend_sequence,
                    psn.display_code,
                    COALESCE(psn.custom_title, sn.note_title) as note_title,
                    COALESCE(psn.custom_text, sn.note_text) as note_text
                FROM sheet_note_assignments sna
                LEFT JOIN project_sheet_notes psn ON sna.project_note_id = psn.project_note_id
                LEFT JOIN standard_notes sn ON psn.standard_note_id = sn.note_id
                WHERE sna.project_id = %s::uuid
                ORDER BY sna.legend_sequence
            """
            assignments = execute_query(query, (project_id,))
        
        else:
            return jsonify({'error': 'Either project_id or project_note_id required'}), 400
        
        return jsonify({'assignments': assignments})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-note-assignments', methods=['POST'])
def create_sheet_note_assignment():
    """Create a new sheet note assignment"""
    try:
        data = request.get_json()
        
        project_note_id = data.get('project_note_id')
        project_id = data.get('project_id')
        
        if not project_note_id or not project_id:
            return jsonify({'error': 'project_note_id and project_id are required'}), 400
        
        assignment_id = str(uuid.uuid4())
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT COALESCE(MAX(legend_sequence), 0) + 1 
                    FROM sheet_note_assignments 
                    WHERE project_id = %s::uuid
                """, (project_id,))
                next_sequence = cur.fetchone()[0]
                
                cur.execute("""
                    INSERT INTO sheet_note_assignments 
                    (assignment_id, project_note_id, project_id, legend_sequence)
                    VALUES (%s::uuid, %s::uuid, %s::uuid, %s)
                    RETURNING assignment_id, project_note_id, project_id, legend_sequence
                """, (assignment_id, project_note_id, project_id, next_sequence))
                
                new_assignment = dict(cur.fetchone())
                conn.commit()
                
                cache.clear()
                return jsonify({'assignment': new_assignment}), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-note-assignments/<assignment_id>', methods=['DELETE'])
def delete_sheet_note_assignment(assignment_id):
    """Delete a sheet note assignment"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute('DELETE FROM sheet_note_assignments WHERE assignment_id = %s::uuid', (assignment_id,))
                
                if cur.rowcount == 0:
                    return jsonify({'error': 'Assignment not found'}), 404
                
                conn.commit()
                cache.clear()
                return jsonify({'message': 'Assignment deleted successfully'})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-note-legend', methods=['GET'])
def get_sheet_note_legend():
    """Generate note legend for a project"""
    try:
        project_id = request.args.get('project_id')
        
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                sna.legend_sequence,
                psn.display_code,
                COALESCE(psn.custom_title, sn.note_title) as note_title,
                COALESCE(psn.custom_text, sn.note_text) as note_text,
                psn.standard_note_id,
                psn.is_modified
            FROM sheet_note_assignments sna
            LEFT JOIN project_sheet_notes psn ON sna.project_note_id = psn.project_note_id
            LEFT JOIN standard_notes sn ON psn.standard_note_id = sn.note_id
            WHERE sna.project_id = %s::uuid
            ORDER BY sna.legend_sequence
        """
        
        legend_items = execute_query(query, (project_id,))
        
        return jsonify({
            'project_id': project_id,
            'legend': legend_items,
            'total_notes': len(legend_items)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==========================
# CONFORMANCE TRACKING API
# ==========================

def validate_project_note_membership(cur, project_note_id, set_id=None, project_id=None):
    """
    Validate that a project note exists and optionally belongs to a specific set and/or project.
    Returns dict with note details if valid, raises ValueError if invalid.
    
    Args:
        cur: Database cursor
        project_note_id: UUID of the project note
        set_id: Optional UUID of the set to validate membership
        project_id: Optional UUID of the project to validate membership
    
    Returns:
        dict: {'set_id': UUID, 'project_id': UUID, 'source_type': str, 'standard_note_id': UUID}
    
    Raises:
        ValueError: If note doesn't exist or doesn't match set/project
    """
    query = """
        SELECT psn.set_id, sns.project_id, psn.source_type, psn.standard_note_id, psn.display_code
        FROM project_sheet_notes psn
        JOIN sheet_note_sets sns ON psn.set_id = sns.set_id
        WHERE psn.project_note_id = %s::uuid
    """
    cur.execute(query, (project_note_id,))
    result = cur.fetchone()
    
    if not result:
        raise ValueError(f'Project note {project_note_id} not found')
    
    note_set_id = result['set_id']
    note_project_id = result['project_id']
    
    if set_id and note_set_id != set_id:
        raise ValueError(f'Project note {project_note_id} does not belong to set {set_id}')
    
    if project_id and note_project_id != project_id:
        raise ValueError(f'Project note {project_note_id} does not belong to project {project_id}')
    
    return {
        'set_id': note_set_id, 
        'project_id': note_project_id,
        'source_type': result['source_type'],
        'standard_note_id': result['standard_note_id'],
        'display_code': result['display_code']
    }

def validate_set_membership(cur, set_id, project_id=None):
    """
    Validate that a sheet note set exists and optionally belongs to a specific project.
    Returns dict with {set_id, project_id} if valid, raises ValueError if invalid.
    
    Args:
        cur: Database cursor
        set_id: UUID of the sheet note set
        project_id: Optional UUID of the project to validate membership
    
    Returns:
        dict: {'set_id': UUID, 'project_id': UUID}
    
    Raises:
        ValueError: If set doesn't exist or doesn't match project
    """
    query = "SELECT set_id, project_id FROM sheet_note_sets WHERE set_id = %s::uuid"
    cur.execute(query, (set_id,))
    result = cur.fetchone()
    
    if not result:
        raise ValueError(f'Sheet note set {set_id} not found')
    
    set_project_id = result['project_id']
    
    if project_id and set_project_id != project_id:
        raise ValueError(f'Sheet note set {set_id} does not belong to project {project_id}')
    
    return {'set_id': result['set_id'], 'project_id': set_project_id}

@app.route('/api/conformance/deviation-categories', methods=['GET'])
def get_deviation_categories():
    """Get all deviation categories"""
    try:
        query = """
            SELECT category_id, category_code, category_name, description, sort_order
            FROM deviation_categories
            WHERE is_active = TRUE
            ORDER BY sort_order, category_name
        """
        categories = execute_query(query)
        return jsonify({'categories': categories})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/conformance/statuses', methods=['GET'])
def get_conformance_statuses():
    """Get all conformance and standardization statuses"""
    try:
        conformance_query = """
            SELECT status_id, status_code, status_name, description, color_hex, sort_order
            FROM conformance_statuses
            WHERE is_active = TRUE
            ORDER BY sort_order, status_name
        """
        standardization_query = """
            SELECT status_id, status_code, status_name, description, workflow_order
            FROM standardization_statuses
            WHERE is_active = TRUE
            ORDER BY workflow_order, status_name
        """
        conformance_statuses = execute_query(conformance_query)
        standardization_statuses = execute_query(standardization_query)
        
        return jsonify({
            'conformance_statuses': conformance_statuses,
            'standardization_statuses': standardization_statuses
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes/<project_note_id>/conformance', methods=['PATCH'])
def update_note_conformance(project_note_id):
    """Update conformance tracking for a project sheet note"""
    try:
        data = request.get_json()
        
        deviation_category_id = data.get('deviation_category_id')
        deviation_reason = data.get('deviation_reason')
        conformance_status_id = data.get('conformance_status_id')
        standardization_status_id = data.get('standardization_status_id')
        standardization_note = data.get('standardization_note')
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Validate project note exists and get its project context
                try:
                    note_info = validate_project_note_membership(cur, project_note_id)
                except ValueError as e:
                    return jsonify({'error': str(e)}), 404
                
                # Validate foreign key references
                if deviation_category_id:
                    cur.execute('SELECT 1 FROM deviation_categories WHERE category_id = %s::uuid AND is_active = TRUE', (deviation_category_id,))
                    if not cur.fetchone():
                        return jsonify({'error': 'Invalid or inactive deviation_category_id'}), 400
                
                if conformance_status_id:
                    cur.execute('SELECT 1 FROM conformance_statuses WHERE status_id = %s::uuid AND is_active = TRUE', (conformance_status_id,))
                    if not cur.fetchone():
                        return jsonify({'error': 'Invalid or inactive conformance_status_id'}), 400
                
                if standardization_status_id:
                    cur.execute('SELECT 1 FROM standardization_statuses WHERE status_id = %s::uuid AND is_active = TRUE', (standardization_status_id,))
                    if not cur.fetchone():
                        return jsonify({'error': 'Invalid or inactive standardization_status_id'}), 400
                
                update_parts = []
                params = []
                
                if deviation_category_id is not None:
                    update_parts.append('deviation_category_id = %s::uuid')
                    params.append(deviation_category_id)
                
                if deviation_reason is not None:
                    update_parts.append('deviation_reason = %s')
                    params.append(deviation_reason)
                
                if conformance_status_id is not None:
                    update_parts.append('conformance_status_id = %s::uuid')
                    params.append(conformance_status_id)
                
                if standardization_status_id is not None:
                    update_parts.append('standardization_status_id = %s::uuid')
                    params.append(standardization_status_id)
                
                if standardization_note is not None:
                    update_parts.append('standardization_note = %s')
                    params.append(standardization_note)
                
                if not update_parts:
                    return jsonify({'error': 'No fields to update'}), 400
                
                update_parts.append('updated_at = CURRENT_TIMESTAMP')
                params.append(project_note_id)
                
                query = f"""
                    UPDATE project_sheet_notes
                    SET {', '.join(update_parts)}
                    WHERE project_note_id = %s::uuid
                    RETURNING *
                """
                
                cur.execute(query, params)
                updated_note = cur.fetchone()
                
                conn.commit()
                cache.clear()
                return jsonify({'note': dict(updated_note)})
    
    except ValueError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes/assign-standard', methods=['POST'])
def assign_standard_note():
    """Assign a standard note to a project note set"""
    try:
        data = request.get_json()
        
        set_id = data.get('set_id')
        standard_note_id = data.get('standard_note_id')
        display_code = data.get('display_code')
        
        if not set_id or not standard_note_id or not display_code:
            return jsonify({'error': 'set_id, standard_note_id, and display_code are required'}), 400
        
        project_note_id = str(uuid.uuid4())
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Validate set exists and get its project context
                try:
                    set_info = validate_set_membership(cur, set_id)
                except ValueError as e:
                    return jsonify({'error': str(e)}), 404
                
                # Validate standard note exists
                cur.execute('SELECT 1 FROM standard_notes WHERE note_id = %s::uuid', (standard_note_id,))
                if not cur.fetchone():
                    return jsonify({'error': 'Invalid standard_note_id'}), 400
                
                # Check for duplicate display code in this set
                cur.execute('SELECT 1 FROM project_sheet_notes WHERE set_id = %s::uuid AND display_code = %s', (set_id, display_code))
                if cur.fetchone():
                    return jsonify({'error': f'Display code {display_code} already exists in this set'}), 400
                
                # Get next sort order
                cur.execute('SELECT COALESCE(MAX(sort_order), 0) + 1 as next_order FROM project_sheet_notes WHERE set_id = %s::uuid', (set_id,))
                next_sort_order = cur.fetchone()['next_order']
                
                # Get conformance status for full compliance
                cur.execute("SELECT status_id FROM conformance_statuses WHERE status_code = 'FULL_COMPLIANCE'")
                conformance_status = cur.fetchone()
                
                # Get standardization status for not nominated
                cur.execute("SELECT status_id FROM standardization_statuses WHERE status_code = 'NOT_NOMINATED'")
                standardization_status = cur.fetchone()
                
                # Create project note from standard
                query = """
                    INSERT INTO project_sheet_notes 
                    (project_note_id, set_id, standard_note_id, standard_reference_id, display_code, 
                     source_type, is_modified, sort_order, conformance_status_id, standardization_status_id,
                     first_used_at, last_used_at, usage_count)
                    VALUES (%s::uuid, %s::uuid, %s::uuid, %s::uuid, %s, 'standard', FALSE, %s, %s::uuid, %s::uuid,
                            CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, 0)
                    RETURNING *
                """
                
                cur.execute(query, (
                    project_note_id, set_id, standard_note_id, standard_note_id, display_code,
                    next_sort_order, conformance_status['status_id'], standardization_status['status_id']
                ))
                
                new_note = cur.fetchone()
                conn.commit()
                cache.clear()
                return jsonify({'note': dict(new_note)}), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-sheet-notes/<project_note_id>/create-modified-copy', methods=['POST'])
def create_modified_copy(project_note_id):
    """Create a modified copy of a standard note with deviation tracking"""
    try:
        data = request.get_json()
        
        deviation_category = data.get('deviation_category')
        deviation_reason = data.get('deviation_reason')
        conformance_status = data.get('conformance_status')
        standardization_status = data.get('standardization_status')
        standardization_note = data.get('standardization_note')
        custom_title = data.get('custom_title')
        custom_text = data.get('custom_text')
        
        if not deviation_category or not deviation_reason or not conformance_status:
            return jsonify({'error': 'deviation_category, deviation_reason, and conformance_status are required'}), 400
        
        new_note_id = str(uuid.uuid4())
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Validate original note exists and get its data
                try:
                    note_info = validate_project_note_membership(cur, project_note_id)
                except ValueError as e:
                    return jsonify({'error': str(e)}), 404
                
                # Verify it's a standard note (source_type = 'standard')
                if note_info['source_type'] != 'standard':
                    return jsonify({'error': 'Can only create modified copies from standard notes'}), 400
                
                # Get the standard note data
                cur.execute("""
                    SELECT note_title, note_text 
                    FROM standard_notes 
                    WHERE note_id = %s::uuid
                """, (note_info['standard_note_id'],))
                standard_data = cur.fetchone()
                
                if not standard_data:
                    return jsonify({'error': 'Standard note not found'}), 404
                
                # Get deviation category ID
                cur.execute("SELECT category_id FROM deviation_categories WHERE category_code = %s", (deviation_category,))
                category = cur.fetchone()
                if not category:
                    return jsonify({'error': 'Invalid deviation_category'}), 400
                
                # Get conformance status ID
                cur.execute("SELECT status_id FROM conformance_statuses WHERE status_code = %s", (conformance_status,))
                status = cur.fetchone()
                if not status:
                    return jsonify({'error': 'Invalid conformance_status'}), 400
                
                # Get standardization status ID
                if standardization_status:
                    cur.execute("SELECT status_id FROM standardization_statuses WHERE status_code = %s", (standardization_status,))
                    std_status = cur.fetchone()
                    std_status_id = std_status['status_id'] if std_status else None
                else:
                    # Default to NOT_NOMINATED
                    cur.execute("SELECT status_id FROM standardization_statuses WHERE status_code = 'NOT_NOMINATED'")
                    std_status_id = cur.fetchone()['status_id']
                
                # Get next sort order in the set
                cur.execute('SELECT COALESCE(MAX(sort_order), 0) + 1 as next_order FROM project_sheet_notes WHERE set_id = %s::uuid', (note_info['set_id'],))
                next_sort_order = cur.fetchone()['next_order']
                
                # Generate new display code (original code + -M suffix)
                base_code = note_info['display_code']
                new_display_code = f"{base_code}-M"
                
                # Check if display code already exists, increment if needed
                counter = 1
                while True:
                    cur.execute('SELECT 1 FROM project_sheet_notes WHERE set_id = %s::uuid AND display_code = %s', 
                               (note_info['set_id'], new_display_code))
                    if not cur.fetchone():
                        break
                    counter += 1
                    new_display_code = f"{base_code}-M{counter}"
                
                # Use provided custom title/text or copy from standard
                final_title = custom_title if custom_title else standard_data['note_title']
                final_text = custom_text if custom_text else standard_data['note_text']
                
                # Create the modified copy
                query = """
                    INSERT INTO project_sheet_notes 
                    (project_note_id, set_id, standard_note_id, standard_reference_id, display_code,
                     custom_title, custom_text, source_type, is_modified, sort_order,
                     deviation_category_id, deviation_reason, conformance_status_id,
                     standardization_status_id, standardization_note,
                     first_used_at, last_used_at, usage_count)
                    VALUES (%s::uuid, %s::uuid, %s::uuid, %s::uuid, %s, %s, %s, 'modified_standard', TRUE, %s,
                            %s::uuid, %s, %s::uuid, %s::uuid, %s,
                            CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, 0)
                    RETURNING *
                """
                
                cur.execute(query, (
                    new_note_id, note_info['set_id'], note_info['standard_note_id'], 
                    note_info['standard_note_id'], new_display_code,
                    final_title, final_text, next_sort_order,
                    category['category_id'], deviation_reason, status['status_id'],
                    std_status_id, standardization_note
                ))
                
                # Fetch the created note with all related data (codes, names, etc.)
                cur.execute("""
                    SELECT 
                        psn.*,
                        sn.note_title as standard_title,
                        dc.category_code as deviation_category,
                        dc.category_name as deviation_category_name,
                        cs.status_code as conformance_status,
                        cs.status_name as conformance_status_name,
                        ss.status_code as standardization_status,
                        ss.status_name as standardization_status_name
                    FROM project_sheet_notes psn
                    LEFT JOIN standard_notes sn ON psn.standard_note_id = sn.note_id
                    LEFT JOIN deviation_categories dc ON psn.deviation_category_id = dc.category_id
                    LEFT JOIN conformance_statuses cs ON psn.conformance_status_id = cs.status_id
                    LEFT JOIN standardization_statuses ss ON psn.standardization_status_id = ss.status_id
                    WHERE psn.project_note_id = %s::uuid
                """, (new_note_id,))
                
                new_note = cur.fetchone()
                conn.commit()
                cache.clear()
                return jsonify({'note': dict(new_note), 'message': 'Modified copy created successfully'}), 201
    
    except ValueError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==========================
# SHEET NOTE SETS API
# ==========================

@app.route('/api/sheet-note-sets', methods=['GET'])
def get_sheet_note_sets():
    """Get all sheet note sets for a project"""
    try:
        project_id = request.args.get('project_id')
        
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        query = """
            SELECT 
                set_id,
                project_id,
                set_name,
                description,
                discipline,
                phase,
                status,
                issued_date,
                issued_to,
                created_at,
                updated_at
            FROM sheet_note_sets
            WHERE project_id = %s::uuid
            ORDER BY created_at DESC
        """
        
        sets = execute_query(query, (project_id,))
        return jsonify({'sets': sets})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-note-sets', methods=['POST'])
def create_sheet_note_set():
    """Create a new sheet note set"""
    try:
        data = request.get_json()
        
        project_id = data.get('project_id')
        set_name = data.get('set_name')
        description = data.get('description')
        discipline = data.get('discipline')
        phase = data.get('phase')
        
        if not project_id or not set_name:
            return jsonify({'error': 'project_id and set_name are required'}), 400
        
        set_id = str(uuid.uuid4())
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    INSERT INTO sheet_note_sets 
                    (set_id, project_id, set_name, description, discipline, phase, status, created_at, updated_at)
                    VALUES (%s::uuid, %s::uuid, %s, %s, %s, %s, 'Draft', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    RETURNING *
                """, (set_id, project_id, set_name, description, discipline, phase))
                
                new_set = dict(cur.fetchone())
                conn.commit()
                cache.clear()
                return jsonify({'set': new_set}), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-note-sets/<set_id>', methods=['PUT'])
def update_sheet_note_set(set_id):
    """Update an existing sheet note set"""
    try:
        data = request.get_json()
        
        update_parts = []
        params = []
        
        if 'set_name' in data:
            update_parts.append('set_name = %s')
            params.append(data['set_name'])
        if 'description' in data:
            update_parts.append('description = %s')
            params.append(data['description'])
        if 'discipline' in data:
            update_parts.append('discipline = %s')
            params.append(data['discipline'])
        if 'phase' in data:
            update_parts.append('phase = %s')
            params.append(data['phase'])
        if 'status' in data:
            update_parts.append('status = %s')
            params.append(data['status'])
        if 'issued_date' in data:
            update_parts.append('issued_date = %s')
            params.append(data['issued_date'])
        if 'issued_to' in data:
            update_parts.append('issued_to = %s')
            params.append(data['issued_to'])
        
        if not update_parts:
            return jsonify({'error': 'No fields to update'}), 400
        
        update_parts.append('updated_at = CURRENT_TIMESTAMP')
        params.append(set_id)
        
        query = f"""
            UPDATE sheet_note_sets
            SET {', '.join(update_parts)}
            WHERE set_id = %s::uuid
            RETURNING *
        """
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query, params)
                updated_set = cur.fetchone()
                
                if not updated_set:
                    return jsonify({'error': 'Sheet note set not found'}), 404
                
                conn.commit()
                cache.clear()
                return jsonify({'set': dict(updated_set)})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-note-sets/<set_id>', methods=['DELETE'])
def delete_sheet_note_set(set_id):
    """Delete a sheet note set"""
    try:
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Check if set exists
                cur.execute('SELECT 1 FROM sheet_note_sets WHERE set_id = %s::uuid', (set_id,))
                if not cur.fetchone():
                    return jsonify({'error': 'Sheet note set not found'}), 404
                
                # Delete (cascades to project_sheet_notes and sheet_note_assignments)
                cur.execute('DELETE FROM sheet_note_sets WHERE set_id = %s::uuid', (set_id,))
                conn.commit()
                cache.clear()
                return jsonify({'message': 'Sheet note set deleted successfully'})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==========================
# SHEET SET MANAGER API
# ==========================

@app.route('/api/project-details', methods=['GET'])
def get_all_project_details():
    """Get project details for all projects"""
    try:
        query = """
            SELECT pd.*, p.project_name, p.project_number
            FROM project_details pd
            JOIN projects p ON pd.project_id = p.project_id
            ORDER BY p.project_name
        """
        result = execute_query(query)
        return jsonify({'project_details': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-details/<project_id>', methods=['GET'])
def get_project_details(project_id):
    """Get project details for a specific project"""
    try:
        query = """
            SELECT pd.*, p.project_name, p.project_number
            FROM project_details pd
            JOIN projects p ON pd.project_id = p.project_id
            WHERE pd.project_id = %s::uuid
        """
        result = execute_query(query, (project_id,))
        
        if not result:
            return jsonify({'project_details': None})
        
        return jsonify({'project_details': result[0]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-details', methods=['POST'])
def create_project_details():
    """Create project details"""
    try:
        data = request.json
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    INSERT INTO project_details 
                    (project_id, project_address, project_city, project_state, project_zip,
                     engineer_name, engineer_license, jurisdiction, permit_number,
                     contact_name, contact_phone, contact_email, notes)
                    VALUES (%s::uuid, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING *
                """, (
                    data['project_id'], data.get('project_address'), data.get('project_city'),
                    data.get('project_state'), data.get('project_zip'), data.get('engineer_name'),
                    data.get('engineer_license'), data.get('jurisdiction'), data.get('permit_number'),
                    data.get('contact_name'), data.get('contact_phone'), data.get('contact_email'),
                    data.get('notes')
                ))
                
                details = dict(cur.fetchone())
                conn.commit()
                return jsonify({'project_details': details}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/project-details/<project_id>', methods=['PUT'])
def update_project_details(project_id):
    """Update project details"""
    try:
        data = request.json
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    UPDATE project_details
                    SET project_address = %s, project_city = %s, project_state = %s, project_zip = %s,
                        engineer_name = %s, engineer_license = %s, jurisdiction = %s, permit_number = %s,
                        contact_name = %s, contact_phone = %s, contact_email = %s, notes = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE project_id = %s::uuid
                    RETURNING *
                """, (
                    data.get('project_address'), data.get('project_city'), data.get('project_state'),
                    data.get('project_zip'), data.get('engineer_name'), data.get('engineer_license'),
                    data.get('jurisdiction'), data.get('permit_number'), data.get('contact_name'),
                    data.get('contact_phone'), data.get('contact_email'), data.get('notes'),
                    project_id
                ))
                
                if cur.rowcount == 0:
                    return jsonify({'error': 'Project details not found'}), 404
                
                details = dict(cur.fetchone())
                conn.commit()
                return jsonify({'project_details': details})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Sheet Category Standards
@app.route('/api/sheet-category-standards', methods=['GET'])
def get_sheet_category_standards():
    """Get all sheet category standards"""
    try:
        query = """
            SELECT * FROM sheet_category_standards
            WHERE is_active = TRUE
            ORDER BY default_hierarchy
        """
        result = execute_query(query)
        return jsonify({'categories': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sheet-category-standards', methods=['POST'])
def create_sheet_category_standard():
    """Create new sheet category standard"""
    try:
        data = request.json
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    INSERT INTO sheet_category_standards
                    (category_code, category_name, default_hierarchy, description)
                    VALUES (%s, %s, %s, %s)
                    RETURNING *
                """, (
                    data['category_code'], data['category_name'],
                    data['default_hierarchy'], data.get('description')
                ))
                
                category = dict(cur.fetchone())
                conn.commit()
                return jsonify({'category': category}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# AI TOOLKIT ROUTES
# ============================================

# Import and register toolkit blueprint
try:
    import sys
    sys.path.append('tools')
    from api.toolkit_routes import toolkit_bp
    app.register_blueprint(toolkit_bp)
    print("✓ AI Toolkit API routes registered at /api/toolkit")
except Exception as e:
    print(f"✗ Failed to load AI Toolkit routes: {e}")

@app.route('/toolkit')
def toolkit_page():
    """AI Toolkit Management Page"""
    return render_template('toolkit.html')

@app.route('/graph')
def graph_page():
    """Knowledge Graph Visualization Page"""
    return render_template('graph.html')

@app.route('/quality-dashboard')
def quality_dashboard_page():
    """Quality and Health Dashboard Page"""
    return render_template('quality-dashboard.html')

# ============================================
# MAP VIEWER & EXPORT ROUTES
# ============================================

# Initialize map export service
map_export = MapExportService()

@app.route('/map-viewer')
def map_viewer_page():
    """Redirect to uncached map viewer"""
    return redirect('/map-viewer-v2', code=302)

@app.route('/map-viewer-v2')
def map_viewer_v2_page():
    """Map Viewer Page V2 - Uncached"""
    import time
    response = make_response(render_template('map_viewer_simple.html', cache_bust=int(time.time())))
    response.headers['Cache-Control'] = 'private, no-cache, no-store, must-revalidate, max-age=0, s-maxage=0'
    response.headers['Surrogate-Control'] = 'no-store'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['X-Accel-Expires'] = '0'
    return response

@app.route('/map-test')
def map_test_page():
    """Simple Map Test Page (no base template)"""
    return render_template('map_viewer_simple.html')

@app.route('/api/map-viewer/layers')
def get_gis_layers():
    """Get available GIS layers (external Sonoma County layers)"""
    try:
        query = "SELECT * FROM gis_layers WHERE enabled = true ORDER BY name"
        layers = execute_query(query)
        return jsonify({'layers': layers})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/database-layers')
def get_database_layers():
    """Get available database layers (user's PostGIS data) - optionally filtered by CAD standards"""
    try:
        # Get filter parameters from query string
        disciplines = request.args.getlist('disciplines')
        categories = request.args.getlist('categories')
        phases = request.args.getlist('phases')
        
        has_filters = len(disciplines) > 0 or len(categories) > 0 or len(phases) > 0
        
        # Define the database layers that contain geospatial data
        db_layers = [
            {
                'id': 'survey_points',
                'name': 'Survey Points',
                'table': 'survey_points',
                'geom_column': 'geometry',
                'geom_type': 'Point',
                'enabled': True,
                'description': 'Survey control points and topo shots',
                'has_layer_name': True
            },
            {
                'id': 'utility_lines',
                'name': 'Utility Lines',
                'table': 'utility_lines',
                'geom_column': 'geometry',
                'geom_type': 'LineString',
                'enabled': True,
                'description': 'Water, sewer, storm, electric, gas lines',
                'has_layer_name': True
            },
            {
                'id': 'utility_structures',
                'name': 'Utility Structures',
                'table': 'utility_structures',
                'geom_column': 'geometry',
                'geom_type': 'Point',
                'enabled': True,
                'description': 'Manholes, valves, meters, structures',
                'has_layer_name': True
            },
            {
                'id': 'parcels',
                'name': 'Parcels',
                'table': 'parcels',
                'geom_column': 'geometry',
                'geom_type': 'Polygon',
                'enabled': True,
                'description': 'Property parcel boundaries',
                'has_layer_name': False
            },
            {
                'id': 'horizontal_alignments',
                'name': 'Alignments',
                'table': 'horizontal_alignments',
                'geom_column': 'geometry',
                'geom_type': 'LineString',
                'enabled': True,
                'description': 'Horizontal centerline alignments',
                'has_layer_name': True
            },
            {
                'id': 'surface_features',
                'name': 'Surface Features',
                'table': 'surface_features',
                'geom_column': 'geometry',
                'geom_type': 'Mixed',
                'enabled': True,
                'description': 'General site features (points, lines, polygons)',
                'has_layer_name': True
            },
            {
                'id': 'drawing_entities',
                'name': 'CAD Entities',
                'table': 'drawing_entities',
                'geom_column': 'geometry',
                'geom_type': 'Mixed',
                'enabled': True,
                'description': 'Lines, polylines, arcs, circles from CAD drawings',
                'has_layer_name': True
            }
        ]
        
        # Check which tables actually exist and have data
        available_layers = []
        for layer in db_layers:
            try:
                # Build count query with optional standards filter
                if has_filters and layer.get('has_layer_name'):
                    # Check if table has layer_name column and filter by standards
                    filter_conditions = []
                    if disciplines:
                        disc_pattern = '|'.join([f'^{d}-' for d in disciplines])
                        filter_conditions.append(f"layer_name ~ '{disc_pattern}'")
                    if categories:
                        cat_pattern = '|'.join([f'-{c}-' for c in categories])
                        filter_conditions.append(f"layer_name ~ '{cat_pattern}'")
                    if phases:
                        phase_pattern = '|'.join([f'-{p}-' for p in phases])
                        filter_conditions.append(f"layer_name ~ '{phase_pattern}'")
                    
                    where_clause = " OR ".join(filter_conditions) if filter_conditions else "1=1"
                    count_query = f"SELECT COUNT(*) as count FROM {layer['table']} WHERE {where_clause}"
                else:
                    count_query = f"SELECT COUNT(*) as count FROM {layer['table']}"
                
                result = execute_query(count_query)
                if result and result[0]['count'] > 0:
                    layer['feature_count'] = result[0]['count']
                    available_layers.append(layer)
            except Exception as e:
                # Table doesn't exist or error checking - skip it
                print(f"Skipping layer {layer['id']}: {e}")
                continue
        
        return jsonify({'layers': available_layers})
    except Exception as e:
        print(f"Error getting database layers: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/layer-data/<layer_id>')
def get_layer_data(layer_id):
    """Fetch GIS layer data from FeatureServer and convert to GeoJSON"""
    try:
        import requests
        from arcgis2geojson import arcgis2geojson
        
        # Get layer config
        query = "SELECT * FROM gis_layers WHERE id = %s AND enabled = true"
        layers = execute_query(query, (layer_id,))
        
        if not layers:
            return jsonify({'error': 'Layer not found'}), 404
        
        layer_config = layers[0]
        
        # Get bbox from request params
        minx = request.args.get('minx', type=float)
        miny = request.args.get('miny', type=float)
        maxx = request.args.get('maxx', type=float)
        maxy = request.args.get('maxy', type=float)
        
        if not all([minx, miny, maxx, maxy]):
            return jsonify({'error': 'Missing bbox parameters'}), 400
        
        # Query FeatureServer with ESRI JSON format
        query_url = f"{layer_config['url']}/query"
        query_params = {
            'where': '1=1',
            'geometry': f'{minx},{miny},{maxx},{maxy}',
            'geometryType': 'esriGeometryEnvelope',
            'inSR': '4326',
            'outSR': '4326',
            'spatialRel': 'esriSpatialRelIntersects',
            'outFields': '*',
            'returnGeometry': 'true',
            'resultRecordCount': '1000',
            'f': 'json'  # ESRI JSON format
        }
        
        response = requests.get(query_url, params=query_params, timeout=30)
        response.raise_for_status()
        esri_data = response.json()
        
        # Convert ESRI JSON to GeoJSON
        features = []
        for esri_feature in esri_data.get('features', []):
            try:
                geojson_feature = arcgis2geojson(esri_feature)
                if geojson_feature.get('geometry') is not None:
                    features.append(geojson_feature)
            except Exception as e:
                print(f"Failed to convert feature: {e}")
                continue
        
        return jsonify({
            'type': 'FeatureCollection',
            'features': features
        })
        
    except Exception as e:
        print(f"Error fetching layer data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/database-layer-data/<layer_id>')
def get_database_layer_data(layer_id):
    """Fetch database layer data from PostGIS tables"""
    try:
        # Map layer IDs to table and geometry column
        layer_map = {
            'survey_points': ('survey_points', 'geometry', 'Point'),
            'utility_lines': ('utility_lines', 'geometry', 'LineString'),
            'utility_structures': ('utility_structures', 'geometry', 'Point'),
            'parcels': ('parcels', 'geometry', 'Polygon'),
            'horizontal_alignments': ('horizontal_alignments', 'geometry', 'LineString'),
            'surface_features': ('surface_features', 'geometry', 'Mixed'),
            'drawing_entities': ('drawing_entities', 'geometry', 'Mixed')
        }
        
        if layer_id not in layer_map:
            return jsonify({'error': 'Invalid layer ID'}), 400
        
        table_name, geom_column, geom_type = layer_map[layer_id]
        
        # Get bbox from request params (WGS84)
        minx = request.args.get('minx', type=float)
        miny = request.args.get('miny', type=float)
        maxx = request.args.get('maxx', type=float)
        maxy = request.args.get('maxy', type=float)
        
        if not all([minx, miny, maxx, maxy]):
            return jsonify({'error': 'Missing bbox parameters'}), 400
        
        # Transform bbox from WGS84 to EPSG:2226 for PostGIS query
        from pyproj import Transformer
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:2226", always_xy=True)
        min_x_2226, min_y_2226 = transformer.transform(minx, miny)
        max_x_2226, max_y_2226 = transformer.transform(maxx, maxy)
        
        # Build query to fetch features within bbox (data in EPSG:2226, transformed to WGS84 for output)
        # Get all columns except the geometry column to avoid conflicts
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(f"""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = '{table_name}' 
                    AND column_name != '{geom_column}'
                """)
                columns = [row[0] for row in cur.fetchall()]
        
        columns_str = ', '.join(columns)
        query = f"""
            SELECT 
                ST_AsGeoJSON(ST_Transform({geom_column}, 4326)) as geometry_json,
                {columns_str}
            FROM {table_name}
            WHERE {geom_column} && ST_MakeEnvelope(%s, %s, %s, %s, 2226)
            LIMIT 1000
        """
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query, (min_x_2226, min_y_2226, max_x_2226, max_y_2226))
                rows = cur.fetchall()
        
        # Convert to GeoJSON features
        features = []
        for row in rows:
            import json
            geom_json = json.loads(row['geometry_json'])
            
            # Build properties (exclude geometry_json column)
            properties = {}
            for key, value in row.items():
                if key != 'geometry_json':
                    # Convert datetime/UUID/etc to string
                    if hasattr(value, 'isoformat'):
                        properties[key] = value.isoformat()
                    elif value is None:
                        continue
                    else:
                        properties[key] = str(value)
            
            features.append({
                'type': 'Feature',
                'geometry': geom_json,
                'properties': properties
            })
        
        return jsonify({
            'type': 'FeatureCollection',
            'features': features
        })
        
    except Exception as e:
        print(f"Error fetching database layer data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/projects')
def get_map_projects():
    """Get all projects with spatial data for map display"""
    try:
        from pyproj import Transformer
        
        # Query all drawings with their projects and bbox
        query = """
            SELECT 
                d.drawing_id,
                d.drawing_name,
                d.drawing_number,
                d.bbox_min_x,
                d.bbox_min_y,
                d.bbox_max_x,
                d.bbox_max_y,
                d.created_at,
                p.project_id,
                p.project_name,
                p.client_name
            FROM drawings d
            JOIN projects p ON d.project_id = p.project_id
            WHERE d.bbox_min_x IS NOT NULL 
              AND d.bbox_min_y IS NOT NULL 
              AND d.bbox_max_x IS NOT NULL 
              AND d.bbox_max_y IS NOT NULL
            ORDER BY d.created_at DESC
        """
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query)
                drawings = cur.fetchall()
        
        # Transformer to convert EPSG:2226 to WGS84 for map display
        transformer = Transformer.from_crs("EPSG:2226", "EPSG:4326", always_xy=True)
        
        features = []
        for drawing in drawings:
            # Get bbox in EPSG:2226 (State Plane)
            min_x_2226 = drawing['bbox_min_x']
            min_y_2226 = drawing['bbox_min_y']
            max_x_2226 = drawing['bbox_max_x']
            max_y_2226 = drawing['bbox_max_y']
            
            # Transform all 4 corners to WGS84
            min_lon, min_lat = transformer.transform(min_x_2226, min_y_2226)
            max_lon, max_lat = transformer.transform(max_x_2226, max_y_2226)
            
            # Also transform the other two corners for accurate bbox
            top_left_lon, top_left_lat = transformer.transform(min_x_2226, max_y_2226)
            bottom_right_lon, bottom_right_lat = transformer.transform(max_x_2226, min_y_2226)
            
            # Create GeoJSON polygon in WGS84 (lon, lat order for GeoJSON)
            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Polygon',
                    'coordinates': [[
                        [min_lon, min_lat],              # Bottom-left
                        [bottom_right_lon, bottom_right_lat],  # Bottom-right
                        [max_lon, max_lat],              # Top-right
                        [top_left_lon, top_left_lat],    # Top-left
                        [min_lon, min_lat]               # Close the ring
                    ]]
                },
                'properties': {
                    'drawing_id': str(drawing['drawing_id']),
                    'drawing_name': drawing['drawing_name'],
                    'drawing_number': drawing['drawing_number'],
                    'project_id': str(drawing['project_id']),
                    'project_name': drawing['project_name'],
                    'client_name': drawing['client_name'],
                    'epsg_code': 'EPSG:2226',
                    'created_at': drawing['created_at'].isoformat() if drawing['created_at'] else None
                }
            }
            features.append(feature)
        
        return jsonify({
            'type': 'FeatureCollection',
            'features': features
        })
    except Exception as e:
        print(f"Error fetching projects: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/project-structure')
def get_project_structure():
    """Get all projects with nested drawings and entity type counts"""
    try:
        # Query all projects with their drawings
        # Include projects even without bbox - will calculate from entities if needed
        query = """
            SELECT 
                p.project_id,
                p.project_name,
                p.client_name,
                p.description,
                d.drawing_id,
                d.drawing_name,
                d.drawing_number,
                d.bbox_min_x,
                d.bbox_min_y,
                d.bbox_max_x,
                d.bbox_max_y,
                -- Calculate bbox from entities if drawing bbox is null
                -- Only calculate from entities if they exist (COUNT > 0)
                COALESCE(
                    d.bbox_min_x, 
                    CASE WHEN (SELECT COUNT(*) FROM drawing_entities WHERE drawing_id = d.drawing_id) > 0 
                        THEN (SELECT ST_XMin(ST_Extent(geometry)) FROM drawing_entities WHERE drawing_id = d.drawing_id AND geometry IS NOT NULL)
                        ELSE NULL 
                    END
                ) as calc_min_x,
                COALESCE(
                    d.bbox_min_y, 
                    CASE WHEN (SELECT COUNT(*) FROM drawing_entities WHERE drawing_id = d.drawing_id) > 0 
                        THEN (SELECT ST_YMin(ST_Extent(geometry)) FROM drawing_entities WHERE drawing_id = d.drawing_id AND geometry IS NOT NULL)
                        ELSE NULL 
                    END
                ) as calc_min_y,
                COALESCE(
                    d.bbox_max_x, 
                    CASE WHEN (SELECT COUNT(*) FROM drawing_entities WHERE drawing_id = d.drawing_id) > 0 
                        THEN (SELECT ST_XMax(ST_Extent(geometry)) FROM drawing_entities WHERE drawing_id = d.drawing_id AND geometry IS NOT NULL)
                        ELSE NULL 
                    END
                ) as calc_max_x,
                COALESCE(
                    d.bbox_max_y, 
                    CASE WHEN (SELECT COUNT(*) FROM drawing_entities WHERE drawing_id = d.drawing_id) > 0 
                        THEN (SELECT ST_YMax(ST_Extent(geometry)) FROM drawing_entities WHERE drawing_id = d.drawing_id AND geometry IS NOT NULL)
                        ELSE NULL 
                    END
                ) as calc_max_y
            FROM projects p
            LEFT JOIN drawings d ON p.project_id = d.project_id
            WHERE d.drawing_id IS NOT NULL
            ORDER BY p.created_at DESC, p.project_name, d.drawing_name
        """
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query)
                rows = cur.fetchall()
        
        # Group by project
        projects = {}
        for row in rows:
            project_id = str(row['project_id'])
            
            # Use calculated bbox (from entities) if drawing bbox is null
            bbox_min_x = row['calc_min_x']
            bbox_min_y = row['calc_min_y']
            bbox_max_x = row['calc_max_x']
            bbox_max_y = row['calc_max_y']
            
            if project_id not in projects:
                projects[project_id] = {
                    'project_id': project_id,
                    'project_name': row['project_name'],
                    'client_name': row['client_name'],
                    'description': row['description'],
                    'drawings': [],
                    'bbox': {
                        'min_x': bbox_min_x,
                        'min_y': bbox_min_y,
                        'max_x': bbox_max_x,
                        'max_y': bbox_max_y
                    } if bbox_min_x is not None else None
                }
            
            # Update project bbox to include all drawings
            if bbox_min_x is not None:
                if projects[project_id]['bbox'] is None:
                    projects[project_id]['bbox'] = {
                        'min_x': bbox_min_x,
                        'min_y': bbox_min_y,
                        'max_x': bbox_max_x,
                        'max_y': bbox_max_y
                    }
                else:
                    projects[project_id]['bbox']['min_x'] = min(projects[project_id]['bbox']['min_x'], bbox_min_x)
                    projects[project_id]['bbox']['min_y'] = min(projects[project_id]['bbox']['min_y'], bbox_min_y)
                    projects[project_id]['bbox']['max_x'] = max(projects[project_id]['bbox']['max_x'], bbox_max_x)
                    projects[project_id]['bbox']['max_y'] = max(projects[project_id]['bbox']['max_y'], bbox_max_y)
            
            if row['drawing_id']:
                drawing_id = str(row['drawing_id'])
                
                # Get entity type counts for this drawing (combine LINE/ARC/LWPOLYLINE into Linework)
                entity_query = """
                    SELECT 
                        CASE 
                            WHEN entity_type IN ('LINE', 'ARC', 'LWPOLYLINE', 'POLYLINE', 'CIRCLE', 'ELLIPSE', 'SPLINE') 
                            THEN 'Linework'
                            ELSE entity_type
                        END as entity_type,
                        COUNT(*) as count
                    FROM drawing_entities
                    WHERE drawing_id = %s
                    AND entity_type NOT IN ('TEXT', 'MTEXT', 'HATCH', 'ATTDEF', 'ATTRIB')
                    GROUP BY CASE 
                        WHEN entity_type IN ('LINE', 'ARC', 'LWPOLYLINE', 'POLYLINE', 'CIRCLE', 'ELLIPSE', 'SPLINE') 
                        THEN 'Linework'
                        ELSE entity_type
                    END
                    ORDER BY count DESC
                """
                
                with get_db() as conn2:
                    with conn2.cursor(cursor_factory=RealDictCursor) as cur2:
                        cur2.execute(entity_query, (drawing_id,))
                        entity_types = cur2.fetchall()
                
                projects[project_id]['drawings'].append({
                    'drawing_id': drawing_id,
                    'drawing_name': row['drawing_name'],
                    'drawing_number': row['drawing_number'],
                    'entity_types': [
                        {
                            'type': et['entity_type'],
                            'count': et['count']
                        }
                        for et in entity_types
                    ],
                    'total_entities': sum(et['count'] for et in entity_types)
                })
        
        return jsonify({
            'projects': list(projects.values())
        })
    except Exception as e:
        print(f"Error fetching project structure: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/project-entities/<project_id>')
def get_project_entities(project_id):
    """Get all entities for a specific project, optionally filtered by entity type or layer"""
    try:
        import json
        
        # Get optional filters from query params
        entity_type = request.args.get('entity_type', None)
        layer_name = request.args.get('layer', None)
        
        # Query project entities with layer information
        query = """
            SELECT 
                e.entity_id,
                e.entity_type,
                e.color_aci,
                e.linetype,
                e.lineweight,
                e.transparency,
                e.attributes,
                l.layer_name,
                l.color as layer_color,
                CASE 
                    WHEN ST_SRID(e.geometry) = 0 THEN ST_AsGeoJSON(e.geometry)
                    ELSE ST_AsGeoJSON(ST_Transform(e.geometry, 4326))
                END as geometry,
                ST_SRID(e.geometry) as srid
            FROM drawing_entities e
            LEFT JOIN layers l ON e.layer_id = l.layer_id
            WHERE e.project_id = %s
            AND e.entity_type NOT IN ('TEXT', 'MTEXT', 'HATCH', 'ATTDEF', 'ATTRIB')
        """
        
        params = [project_id]
        
        # Add layer filter if specified
        if layer_name:
            query += " AND l.layer_name = %s"
            params.append(layer_name)
        
        # Add entity type filter if specified (translate "Linework" to all line-based types)
        if entity_type:
            if entity_type == 'Linework':
                query += " AND e.entity_type IN ('LINE', 'ARC', 'LWPOLYLINE', 'POLYLINE', 'CIRCLE', 'ELLIPSE', 'SPLINE')"
            else:
                query += " AND e.entity_type = %s"
                params.append(entity_type)
        
        query += " LIMIT 5000"
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query, tuple(params))
                entities = cur.fetchall()
        
        # Convert to GeoJSON features
        features = []
        for entity in entities:
            try:
                geom_json = json.loads(entity['geometry'])
                
                # Build properties
                properties = {
                    'entity_id': str(entity['entity_id']),
                    'entity_type': entity['entity_type'],
                    'layer_name': entity['layer_name'],
                    'color_aci': entity['color_aci'],
                    'layer_color': entity['layer_color'],
                    'linetype': entity['linetype'],
                    'lineweight': entity['lineweight'],
                    'transparency': entity['transparency'],
                    'srid': entity['srid']
                }
                
                # Add attributes if present
                if entity['attributes']:
                    properties['attributes'] = entity['attributes']
                
                features.append({
                    'type': 'Feature',
                    'geometry': geom_json,
                    'properties': properties
                })
            except Exception as e:
                print(f"Error processing entity {entity.get('entity_id')}: {e}")
                continue
        
        return jsonify({
            'type': 'FeatureCollection',
            'features': features,
            'count': len(features)
        })
        
    except Exception as e:
        print(f"Error fetching project entities: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/project-layers/<project_id>')
def get_project_layers(project_id):
    """Get all layers for a specific project with entity counts"""
    try:
        query = """
            SELECT 
                l.layer_name,
                COUNT(e.entity_id) as entity_count
            FROM layers l
            LEFT JOIN drawing_entities e ON l.layer_id = e.layer_id
            WHERE l.project_id = %s
            AND (e.entity_type IS NULL OR e.entity_type NOT IN ('TEXT', 'MTEXT', 'HATCH', 'ATTDEF', 'ATTRIB'))
            GROUP BY l.layer_name
            HAVING COUNT(e.entity_id) > 0
            ORDER BY l.layer_name
        """
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query, (project_id,))
                layers = cur.fetchall()
        
        return jsonify({
            'layers': [
                {
                    'layer_name': layer['layer_name'],
                    'entity_count': layer['entity_count']
                }
                for layer in layers
            ]
        })
        
    except Exception as e:
        print(f"Error fetching project layers: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-viewer/project-extent/<project_id>')
def get_project_extent(project_id):
    """Get bounding box extent for a project's entities"""
    try:
        from pyproj import Transformer
        
        query = """
            SELECT 
                ST_XMin(ST_Extent(geometry)) as xmin,
                ST_YMin(ST_Extent(geometry)) as ymin,
                ST_XMax(ST_Extent(geometry)) as xmax,
                ST_YMax(ST_Extent(geometry)) as ymax,
                COUNT(*) as entity_count,
                (SELECT ST_SRID(geometry) FROM drawing_entities WHERE project_id = %s LIMIT 1) as srid
            FROM drawing_entities
            WHERE project_id = %s
        """
        
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query, (project_id, project_id))
                result = cur.fetchone()
        
        if not result or result['entity_count'] == 0:
            return jsonify({'error': 'No entities found for this project'}), 404
        
        srid = result['srid'] or 0
        
        # For SRID 0 (local CAD coordinates), return native bounds without transformation
        if srid == 0:
            return jsonify({
                'bounds': None,  # Can't display on geographic map
                'entity_count': result['entity_count'],
                'native_bounds': {
                    'xmin': float(result['xmin']),
                    'ymin': float(result['ymin']),
                    'xmax': float(result['xmax']),
                    'ymax': float(result['ymax']),
                    'srid': 0
                },
                'is_local_coordinates': True
            })
        
        # Transform from geographic SRID to WGS84 for map display
        transformer = Transformer.from_crs(f"EPSG:{srid}", "EPSG:4326", always_xy=True)
        
        # Transform the bounding box corners
        sw_lon, sw_lat = transformer.transform(result['xmin'], result['ymin'])
        ne_lon, ne_lat = transformer.transform(result['xmax'], result['ymax'])
        
        return jsonify({
            'bounds': [[sw_lat, sw_lon], [ne_lat, ne_lon]],
            'entity_count': result['entity_count'],
            'native_bounds': {
                'xmin': float(result['xmin']),
                'ymin': float(result['ymin']),
                'xmax': float(result['xmax']),
                'ymax': float(result['ymax']),
                'srid': srid
            },
            'is_local_coordinates': False
        })
        
    except Exception as e:
        print(f"Error getting drawing extent: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-export/create-simple', methods=['POST'])
def create_simple_export():
    """Create shapefile with real GIS data from Sonoma County"""
    try:
        import requests
        from shapely.geometry import shape, box
        from shapely import geometry as geom_lib
        from pyproj import Transformer
        import fiona
        from fiona.crs import from_epsg
        
        params = request.json
        bbox = params.get('bbox', {})
        requested_layers = params.get('layers', ['parcels', 'buildings', 'roads'])
        
        print(f"Creating export for bbox: {bbox}")
        print(f"Requested layers: {requested_layers}")
        
        # Create unique ID for this export
        export_id = str(uuid.uuid4())
        export_dir = os.path.join('/tmp/exports', export_id)
        os.makedirs(export_dir, exist_ok=True)
        
        # Get bounding box in WGS84
        minx, miny, maxx, maxy = bbox['minx'], bbox['miny'], bbox['maxx'], bbox['maxy']
        
        # Coordinate transformer
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:2226", always_xy=True)
        
        # Get GIS layer configurations from database
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM gis_layers WHERE enabled = true AND id = ANY(%s)", (requested_layers,))
                layers = cur.fetchall()
        
        print(f"Found {len(layers)} layers in database matching requested layers")
        if len(layers) == 0:
            print(f"WARNING: No layers found! Requested: {requested_layers}")
            print("Checking all available layers...")
            with get_db() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT id, name, enabled FROM gis_layers")
                    all_layers = cur.fetchall()
                    print(f"All layers in DB: {all_layers}")
        
        feature_counts = {}
        
        # Fetch and export each layer
        for layer_config in layers:
            layer_id = layer_config['id']
            layer_url = layer_config['url']
            layer_name = layer_config['name']
            
            print(f"Fetching {layer_name} from {layer_url}")
            
            try:
                # Query FeatureServer with bounding box
                query_url = f"{layer_url}/query"
                query_params = {
                    'where': '1=1',
                    'geometry': f'{minx},{miny},{maxx},{maxy}',
                    'geometryType': 'esriGeometryEnvelope',
                    'inSR': '4326',
                    'spatialRel': 'esriSpatialRelIntersects',
                    'outFields': '*',
                    'returnGeometry': 'true',
                    'outSR': '4326',  # Return in WGS84
                    'f': 'json'  # ESRI JSON format (geojson not supported)
                }
                
                response = requests.get(query_url, params=query_params, timeout=30)
                response.raise_for_status()
                esri_data = response.json()
                
                # Convert ESRI JSON features to GeoJSON using arcgis2geojson library
                from arcgis2geojson import arcgis2geojson
                
                features = []
                skipped = 0
                total_features = len(esri_data.get('features', []))
                
                for esri_feature in esri_data.get('features', []):
                    try:
                        # arcgis2geojson handles all geometry types, MultiPolygons, null geoms
                        geojson_feature = arcgis2geojson(esri_feature)
                        
                        # Skip features with null geometries
                        if geojson_feature.get('geometry') is not None:
                            features.append(geojson_feature)
                        else:
                            skipped += 1
                    except Exception as e:
                        print(f"  WARNING: Failed to convert feature: {e}")
                        skipped += 1
                        continue
                
                feature_counts[layer_id] = len(features)
                if skipped > 0:
                    print(f"  Skipped {skipped}/{total_features} features (null/invalid geometries)")
                
                print(f"  Found {len(features)} features in {layer_name}")
                
                if len(features) == 0:
                    continue
                
                # Determine geometry type from first feature
                first_geom = shape(features[0]['geometry'])
                geom_type = first_geom.geom_type
                
                # Create schema based on first feature properties
                sample_props = features[0]['properties']
                schema_props = {}
                for key, value in sample_props.items():
                    if value is None:
                        schema_props[key] = 'str'
                    elif isinstance(value, bool):
                        schema_props[key] = 'bool'
                    elif isinstance(value, int):
                        schema_props[key] = 'int'
                    elif isinstance(value, float):
                        schema_props[key] = 'float'
                    else:
                        schema_props[key] = 'str'
                
                schema = {
                    'geometry': geom_type,
                    'properties': schema_props
                }
                
                # Create shapefile for this layer
                shp_path = os.path.join(export_dir, f'{layer_id}.shp')
                
                with fiona.open(shp_path, 'w', driver='ESRI Shapefile',
                               crs=from_epsg(2226), schema=schema) as output:
                    for feature in features:
                        # Transform geometry to EPSG:2226
                        geom_wgs84 = shape(feature['geometry'])
                        
                        # Transform coordinates
                        if geom_wgs84.geom_type == 'Point':
                            x, y = transformer.transform(geom_wgs84.x, geom_wgs84.y)
                            geom_2226 = geom_lib.Point(x, y)
                        elif geom_wgs84.geom_type in ['LineString', 'MultiLineString']:
                            coords_2226 = [transformer.transform(x, y) for x, y in geom_wgs84.coords]
                            geom_2226 = geom_lib.LineString(coords_2226)
                        elif geom_wgs84.geom_type == 'Polygon':
                            exterior_2226 = [transformer.transform(x, y) for x, y in geom_wgs84.exterior.coords]
                            geom_2226 = geom_lib.Polygon(exterior_2226)
                        elif geom_wgs84.geom_type == 'MultiPolygon':
                            polys_2226 = []
                            for poly in geom_wgs84.geoms:
                                exterior_2226 = [transformer.transform(x, y) for x, y in poly.exterior.coords]
                                polys_2226.append(geom_lib.Polygon(exterior_2226))
                            geom_2226 = geom_lib.MultiPolygon(polys_2226)
                        else:
                            continue  # Skip unsupported geometry types
                        
                        output.write({
                            'geometry': geom_2226.__geo_interface__,
                            'properties': feature['properties']
                        })
                
                print(f"  Wrote {len(features)} features to {layer_id}.shp")
                
            except Exception as e:
                print(f"  ERROR fetching {layer_name}: {e}")
                feature_counts[layer_id] = 0
        
        # Create download URL
        download_url = f"/api/map-export/download-simple/{export_id}"
        
        return jsonify({
            'status': 'complete',
            'download_url': download_url,
            'export_id': export_id,
            'feature_counts': feature_counts,
            'message': f'Exported {sum(feature_counts.values())} total features'
        })
        
    except Exception as e:
        print(f"ERROR in create_simple_export: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-export/create', methods=['POST'])
def create_multi_format_export():
    """Create export with multiple formats (PNG, DXF, SHP, KML)"""
    try:
        import requests
        from shapely.geometry import shape
        from shapely import geometry as geom_lib
        from pyproj import Transformer
        
        params = request.json
        bbox = params.get('bbox', {})
        requested_layers = params.get('layers', [])
        formats = params.get('formats', ['shp'])  # Default to shapefile only
        png_options = params.get('png_options', {'north_arrow': True, 'scale_bar': True})
        
        print(f"Creating multi-format export:")
        print(f"  Formats: {formats}")
        print(f"  Layers: {requested_layers}")
        print(f"  PNG options: {png_options}")
        
        # Create unique ID for this export
        export_id = str(uuid.uuid4())
        export_dir = os.path.join('/tmp/exports', export_id)
        os.makedirs(export_dir, exist_ok=True)
        
        # Get bounding box in WGS84
        minx, miny, maxx, maxy = bbox['minx'], bbox['miny'], bbox['maxx'], bbox['maxy']
        
        # Coordinate transformer
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:2226", always_xy=True)
        
        # Get GIS layer configurations from database
        with get_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM gis_layers WHERE enabled = true AND id = ANY(%s)", (requested_layers,))
                layers = cur.fetchall()
        
        print(f"Found {len(layers)} layers in database")
        
        # PRIORITY 1: Fetch drawing entities from database (automatic, ALL layers)
        print("Fetching drawing entities from database...")
        all_layers_data = {}
        feature_counts = {}
        
        # Transform bbox to EPSG:2226 for PostGIS query
        min_x_2226, min_y_2226 = transformer.transform(minx, miny)
        max_x_2226, max_y_2226 = transformer.transform(maxx, maxy)
        bbox_2226 = (min_x_2226, min_y_2226, max_x_2226, max_y_2226)
        
        # Use MapExportService with database connection
        from map_export_service import MapExportService
        with get_db() as conn:
            export_service = MapExportService(db_conn=conn)
            drawing_layers = export_service.fetch_drawing_entities_by_layer(bbox_2226)
            
            if drawing_layers:
                print(f"Found {len(drawing_layers)} DXF-imported layers")
                for layer_name, features in drawing_layers.items():
                    # Features are already in EPSG:2226, need to transform to WGS84 for consistency
                    features_wgs84 = []
                    back_transformer = Transformer.from_crs("EPSG:2226", "EPSG:4326", always_xy=True)
                    
                    for feature in features:
                        try:
                            geom_2226 = shape(feature['geometry'])
                            
                            # Transform to WGS84 (handle both 2D and 3D geometries)
                            if geom_2226.geom_type == 'LineString':
                                # Handle 3D coordinates (x, y, z) by ignoring Z
                                coords_wgs84 = []
                                for coord in geom_2226.coords:
                                    x, y = coord[0], coord[1]  # Ignore Z if present
                                    lon, lat = back_transformer.transform(x, y)
                                    coords_wgs84.append((lon, lat))
                                geom_wgs84 = geom_lib.LineString(coords_wgs84)
                            elif geom_2226.geom_type == 'Polygon':
                                exterior_wgs84 = []
                                for coord in geom_2226.exterior.coords:
                                    x, y = coord[0], coord[1]  # Ignore Z if present
                                    lon, lat = back_transformer.transform(x, y)
                                    exterior_wgs84.append((lon, lat))
                                geom_wgs84 = geom_lib.Polygon(exterior_wgs84)
                            elif geom_2226.geom_type == 'Point':
                                x, y = geom_2226.x, geom_2226.y
                                lon, lat = back_transformer.transform(x, y)
                                geom_wgs84 = geom_lib.Point(lon, lat)
                            else:
                                geom_wgs84 = geom_2226  # Keep as-is for unsupported types
                            
                            features_wgs84.append({
                                'type': 'Feature',
                                'geometry': geom_wgs84.__geo_interface__,
                                'properties': feature['properties']
                            })
                        except Exception as e:
                            print(f"Error transforming drawing entity: {e}")
                            import traceback
                            traceback.print_exc()
                            continue
                    
                    all_layers_data[layer_name] = features_wgs84
                    feature_counts[layer_name] = len(features_wgs84)
                    print(f"  {layer_name}: {len(features_wgs84)} features")
        
        # PRIORITY 2: Fetch external WFS layers
        for layer_config in layers:
            layer_id = layer_config['id']
            layer_url = layer_config['url']
            layer_name = layer_config['name']
            
            print(f"Fetching external WFS layer {layer_name}...")
            
            try:
                # Query FeatureServer with bounding box
                query_url = f"{layer_url}/query"
                query_params = {
                    'where': '1=1',
                    'geometry': f'{minx},{miny},{maxx},{maxy}',
                    'geometryType': 'esriGeometryEnvelope',
                    'inSR': '4326',
                    'spatialRel': 'esriSpatialRelIntersects',
                    'outFields': '*',
                    'returnGeometry': 'true',
                    'outSR': '4326',
                    'f': 'json'
                }
                
                response = requests.get(query_url, params=query_params, timeout=30)
                response.raise_for_status()
                esri_data = response.json()
                
                # Convert ESRI JSON to GeoJSON
                from arcgis2geojson import arcgis2geojson
                
                features = []
                for esri_feature in esri_data.get('features', []):
                    try:
                        geojson_feature = arcgis2geojson(esri_feature)
                        if geojson_feature.get('geometry') is not None:
                            features.append(geojson_feature)
                    except Exception as e:
                        continue
                
                feature_counts[layer_id] = len(features)
                all_layers_data[layer_id] = features
                print(f"  Found {len(features)} features")
                
            except Exception as e:
                print(f"  ERROR fetching {layer_name}: {e}")
                feature_counts[layer_id] = 0
        
        # Legacy: Fetch CAD entity layers if explicitly requested (backwards compatibility)
        entity_layers = params.get('entity_layers', [])
        if entity_layers:
            print(f"Fetching {len(entity_layers)} CAD entity layer(s)...")
            
            # Transform bbox to EPSG:2226 for PostGIS query
            min_x_2226, min_y_2226 = transformer.transform(minx, miny)
            max_x_2226, max_y_2226 = transformer.transform(maxx, maxy)
            
            for entity_layer in entity_layers:
                drawing_id = entity_layer.get('drawingId')
                entity_type = entity_layer.get('entityType')
                layer_key = f"entities_{drawing_id}_{entity_type}"
                
                print(f"  Fetching {entity_type} from drawing {drawing_id}...")
                
                try:
                    with get_db() as conn:
                        with conn.cursor(cursor_factory=RealDictCursor) as cur:
                            # Query drawing_entities within bbox
                            cur.execute("""
                                SELECT 
                                    entity_id,
                                    entity_type,
                                    color_aci,
                                    linetype,
                                    lineweight,
                                    transparency,
                                    ST_AsGeoJSON(ST_Transform(geometry, 4326)) as geometry_json,
                                    attributes
                                FROM drawing_entities
                                WHERE drawing_id = %s 
                                AND entity_type = %s
                                AND geometry && ST_MakeEnvelope(%s, %s, %s, %s, 2226)
                                AND geometry IS NOT NULL
                                LIMIT 5000
                            """, (drawing_id, entity_type, min_x_2226, min_y_2226, max_x_2226, max_y_2226))
                            
                            rows = cur.fetchall()
                            
                            # Convert to GeoJSON features
                            features = []
                            for row in rows:
                                feature = {
                                    'type': 'Feature',
                                    'geometry': json.loads(row['geometry_json']),
                                    'properties': {
                                        'entity_id': str(row['entity_id']),
                                        'entity_type': row['entity_type'],
                                        'color_aci': row['color_aci'],
                                        'linetype': row['linetype'] or 'Continuous',
                                        'lineweight': row['lineweight'] or 0,
                                        'transparency': row['transparency'] or 0
                                    }
                                }
                                features.append(feature)
                            
                            feature_counts[layer_key] = len(features)
                            all_layers_data[layer_key] = features
                            print(f"    Found {len(features)} {entity_type} entities")
                            
                except Exception as e:
                    print(f"    ERROR fetching {entity_type}: {e}")
                    import traceback
                    traceback.print_exc()
                    feature_counts[layer_key] = 0
        
        # Transform all features to EPSG:2226 first (needed for DXF, SHP, KML, PNG)
        from shapely import geometry as geom_lib
        all_layers_2226 = {}
        
        for layer_id, features_wgs84 in all_layers_data.items():
            if not features_wgs84:
                continue
            
            transformed_features = []
            for feature in features_wgs84:
                try:
                    geom_wgs84 = shape(feature['geometry'])
                    
                    # Transform to EPSG:2226
                    if geom_wgs84.geom_type == 'Polygon':
                        exterior_2226 = [transformer.transform(x, y) for x, y in geom_wgs84.exterior.coords]
                        geom_2226 = geom_lib.Polygon(exterior_2226)
                    elif geom_wgs84.geom_type == 'MultiPolygon':
                        polys_2226 = []
                        for poly in geom_wgs84.geoms:
                            exterior_2226 = [transformer.transform(x, y) for x, y in poly.exterior.coords]
                            polys_2226.append(geom_lib.Polygon(exterior_2226))
                        geom_2226 = geom_lib.MultiPolygon(polys_2226)
                    elif geom_wgs84.geom_type == 'LineString':
                        coords_2226 = [transformer.transform(x, y) for x, y in geom_wgs84.coords]
                        geom_2226 = geom_lib.LineString(coords_2226)
                    elif geom_wgs84.geom_type == 'MultiLineString':
                        lines_2226 = []
                        for line in geom_wgs84.geoms:
                            coords_2226 = [transformer.transform(x, y) for x, y in line.coords]
                            lines_2226.append(geom_lib.LineString(coords_2226))
                        geom_2226 = geom_lib.MultiLineString(lines_2226)
                    elif geom_wgs84.geom_type == 'Point':
                        x, y = transformer.transform(geom_wgs84.x, geom_wgs84.y)
                        geom_2226 = geom_lib.Point(x, y)
                    elif geom_wgs84.geom_type == 'MultiPoint':
                        points_2226 = []
                        for point in geom_wgs84.geoms:
                            x, y = transformer.transform(point.x, point.y)
                            points_2226.append(geom_lib.Point(x, y))
                        geom_2226 = geom_lib.MultiPoint(points_2226)
                    else:
                        continue  # Skip unsupported geometry types
                    
                    transformed_features.append({
                        'type': 'Feature',
                        'geometry': geom_2226.__geo_interface__,
                        'properties': feature['properties']
                    })
                except Exception as e:
                    print(f"Error transforming feature: {e}")
                    continue
            
            all_layers_2226[layer_id] = transformed_features
        
        # Now create each requested format
        created_files = []
        
        # 1. Shapefile (SHP)
        if 'shp' in formats:
            print("Creating Shapefile...")
            from map_export_service import MapExportService
            export_service = MapExportService()
            
            for layer_id, features_2226 in all_layers_2226.items():
                if not features_2226:
                    continue
                
                try:
                    shp_path = os.path.join(export_dir, f'{layer_id}.shp')
                    if export_service.export_to_shapefile(features_2226, layer_id, shp_path):
                        created_files.append(f'{layer_id}.shp')
                except Exception as e:
                    print(f"Error creating shapefile for {layer_id}: {e}")
        
        # 2. DXF
        if 'dxf' in formats:
            print("Creating DXF...")
            try:
                from map_export_service import MapExportService
                export_service = MapExportService()
                
                dxf_path = os.path.join(export_dir, 'export.dxf')
                if export_service.export_to_dxf(all_layers_2226, dxf_path):
                    created_files.append('export.dxf')
            except Exception as e:
                print(f"Error creating DXF: {e}")
        
        # 3. KML
        if 'kml' in formats:
            print("Creating KML...")
            try:
                from map_export_service import MapExportService
                export_service = MapExportService()
                
                # KML export expects EPSG:2226 data and transforms to WGS84 internally
                kml_path = os.path.join(export_dir, 'export.kml')
                if export_service.export_to_kml(all_layers_2226, kml_path):
                    created_files.append('export.kml')
            except Exception as e:
                print(f"Error creating KML: {e}")
        
        # 4. PNG
        if 'png' in formats:
            print("Creating PNG...")
            try:
                from map_export_service import MapExportService
                export_service = MapExportService()
                
                # Transform bbox to EPSG:2226 for scale calculations
                bbox_2226 = export_service.transform_bbox(bbox, 'EPSG:4326')
                bbox_dict = {
                    'minx': bbox_2226[0],
                    'miny': bbox_2226[1],
                    'maxx': bbox_2226[2],
                    'maxy': bbox_2226[3]
                }
                
                png_path = export_service.create_map_image(
                    bbox_dict,
                    north_arrow=png_options.get('north_arrow', True),
                    scale_bar=png_options.get('scale_bar', True)
                )
                
                if png_path and os.path.exists(png_path):
                    # Move to export dir
                    import shutil
                    dest_path = os.path.join(export_dir, 'map.png')
                    shutil.move(png_path, dest_path)
                    created_files.append('map.png')
            except Exception as e:
                print(f"Error creating PNG: {e}")
        
        print(f"Created files: {created_files}")
        
        # Create download URL
        download_url = f"/api/map-export/download-simple/{export_id}"
        
        return jsonify({
            'status': 'complete',
            'download_url': download_url,
            'export_id': export_id,
            'feature_counts': feature_counts,
            'formats': formats,
            'created_files': created_files,
            'message': f'Exported {sum(feature_counts.values())} total features in {len(formats)} format(s)'
        })
        
    except Exception as e:
        print(f"ERROR in create_multi_format_export: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-export/download-simple/<export_id>')
def download_simple_export(export_id):
    """Download simple export as ZIP file"""
    try:
        export_dir = os.path.join('/tmp/exports', export_id)
        
        if not os.path.exists(export_dir):
            return jsonify({'error': 'Export not found or expired'}), 404
        
        # Create ZIP file with all shapefile components
        import zipfile
        zip_path = os.path.join('/tmp', f'{export_id}.zip')
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for filename in os.listdir(export_dir):
                file_path = os.path.join(export_dir, filename)
                zipf.write(file_path, filename)
        
        return send_file(zip_path, 
                        as_attachment=True,
                        download_name=f'sonoma_export_{export_id[:8]}.zip',
                        mimetype='application/zip')
        
    except Exception as e:
        print(f"Download error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-export/status/<job_id>')
def get_export_status(job_id):
    """Get export job status"""
    try:
        query = """
            SELECT id, status, download_url, file_size_mb, 
                   error_message, created_at, expires_at
            FROM map_export_jobs
            WHERE id = %s::uuid
        """
        result = execute_query(query, (job_id,))
        
        if not result:
            return jsonify({'error': 'Job not found'}), 404
        
        job = result[0]
        response = {
            'job_id': str(job['id']),
            'status': job['status']
        }
        
        if job['status'] == 'complete':
            response.update({
                'download_url': job['download_url'],
                'file_size_mb': job['file_size_mb'],
                'expires_at': job['expires_at'].isoformat() if job['expires_at'] else None
            })
        elif job['status'] == 'failed':
            response['error'] = job['error_message']
        
        return jsonify(response)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/map-export/download/<job_id>/<filename>')
def download_export(job_id, filename):
    """Download export file"""
    try:
        file_path = os.path.join('/tmp/exports', job_id, filename)
        
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found or expired'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# CAD STANDARDS VOCABULARY API
# ============================================

@app.route('/api/vocabulary/disciplines')
def get_vocabulary_disciplines():
    """Get all discipline codes (legacy vocabulary table)"""
    try:
        query = """
            SELECT discipline_id, code, full_name, description, sort_order
            FROM discipline_codes
            WHERE is_active = TRUE
            ORDER BY sort_order NULLS LAST, code
        """
        disciplines = execute_query(query)
        return jsonify({'disciplines': disciplines})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/categories')
def get_vocabulary_categories():
    """Get all category codes with discipline info (legacy vocabulary table)"""
    try:
        discipline_id = request.args.get('discipline_id')
        
        query = """
            SELECT c.category_id, c.code, c.full_name, c.description, c.sort_order,
                   d.code as discipline_code, d.full_name as discipline_name
            FROM category_codes c
            JOIN discipline_codes d ON c.discipline_id = d.discipline_id
            WHERE c.is_active = TRUE
        """
        
        params = None
        if discipline_id:
            query += " AND c.discipline_id = %s"
            params = (discipline_id,)
        
        query += " ORDER BY d.code, c.sort_order NULLS LAST, c.code"
        
        categories = execute_query(query, params)
        return jsonify({'categories': categories})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/object-types')
def get_object_types():
    """Get all object type codes"""
    try:
        category_id = request.args.get('category_id')
        
        query = """
            SELECT t.type_id, t.code, t.full_name, t.description, t.database_table,
                   c.code as category_code, c.full_name as category_name,
                   d.code as discipline_code
            FROM object_type_codes t
            JOIN category_codes c ON t.category_id = c.category_id
            JOIN discipline_codes d ON c.discipline_id = d.discipline_id
            WHERE t.is_active = TRUE
        """
        
        params = None
        if category_id:
            query += " AND t.category_id = %s"
            params = (category_id,)
        
        query += " ORDER BY d.code, c.code, t.sort_order NULLS LAST, t.code"
        
        types = execute_query(query, params)
        return jsonify({'object_types': types})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/phases')
def get_phases():
    """Get all phase codes"""
    try:
        query = """
            SELECT phase_id, code, full_name, description, color_rgb, sort_order
            FROM phase_codes
            WHERE is_active = TRUE
            ORDER BY sort_order NULLS LAST, code
        """
        phases = execute_query(query)
        return jsonify({'phases': phases})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/geometries')
def get_geometries():
    """Get all geometry codes"""
    try:
        query = """
            SELECT geometry_id, code, full_name, description, dxf_entity_types, sort_order
            FROM geometry_codes
            WHERE is_active = TRUE
            ORDER BY sort_order NULLS LAST, code
        """
        geometries = execute_query(query)
        return jsonify({'geometries': geometries})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/layer-patterns')
def get_layer_patterns():
    """Get standard layer patterns"""
    try:
        query = """
            SELECT p.pattern_id, p.full_layer_name, p.description,
                   p.database_table, p.example_attributes,
                   d.code as discipline_code, d.full_name as discipline_name,
                   c.code as category_code, c.full_name as category_name,
                   t.code as type_code, t.full_name as type_name,
                   ph.code as phase_code, ph.full_name as phase_name,
                   g.code as geometry_code, g.full_name as geometry_name
            FROM standard_layer_patterns p
            JOIN discipline_codes d ON p.discipline_id = d.discipline_id
            JOIN category_codes c ON p.category_id = c.category_id
            JOIN object_type_codes t ON p.type_id = t.type_id
            JOIN phase_codes ph ON p.phase_id = ph.phase_id
            JOIN geometry_codes g ON p.geometry_id = g.geometry_id
            WHERE p.is_active = TRUE
            ORDER BY p.full_layer_name
        """
        patterns = execute_query(query)
        return jsonify({'layer_patterns': patterns})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/import-mappings')
def get_import_mappings():
    """Get import mapping patterns"""
    try:
        query = """
            SELECT m.mapping_id, m.client_name, m.source_pattern,
                   m.regex_pattern, m.extraction_rules, m.confidence_score,
                   d.code as discipline_code,
                   c.code as category_code,
                   t.code as type_code
            FROM import_mapping_patterns m
            LEFT JOIN discipline_codes d ON m.target_discipline_id = d.discipline_id
            LEFT JOIN category_codes c ON m.target_category_id = c.category_id
            LEFT JOIN object_type_codes t ON m.target_type_id = t.type_id
            WHERE m.is_active = TRUE
            ORDER BY m.confidence_score DESC, m.client_name, m.source_pattern
        """
        mappings = execute_query(query)
        return jsonify({'import_mappings': mappings})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/layer-examples')
def get_layer_examples():
    """Generate curated layer name examples covering all geometry types"""
    try:
        # Curated catalog of examples ensuring all 8 geometry types are represented
        # Format: (discipline, category, object_type, phase, geometry, description, db_table)
        catalog = [
            # === LN (Lines) - Pipes, Alignments, Contours, Boundaries ===
            ('CIV', 'STOR', 'STORM', 'NEW', 'LN', 'New storm drain pipe', 'utility_lines'),
            ('CIV', 'STOR', 'SANIT', 'EXIST', 'LN', 'Existing sanitary sewer', 'utility_lines'),
            ('CIV', 'ROAD', 'WATER', 'PROP', 'LN', 'Proposed water main', 'utility_lines'),
            ('CIV', 'ROAD', 'RECYC', 'NEW', 'LN', 'New recycled water pipe', 'utility_lines'),
            ('CIV', 'ROAD', 'CNTR', 'EXIST', 'LN', 'Existing road centerline', 'alignments'),
            ('CIV', 'GRAD', 'CNTR', 'EXIST', 'LN', 'Existing contour line', 'contours'),
            ('SURV', 'BNDY', 'BNDY', 'EXIST', 'LN', 'Existing property boundary', 'boundaries'),
            ('SURV', 'TOPO', 'BREAK', 'EXIST', 'LN', 'Breakline', 'breaklines'),
            
            # === PT (Points) - Structures, Monuments, Trees ===
            ('CIV', 'STOR', 'MH', 'NEW', 'PT', 'New manhole', 'utility_structures'),
            ('CIV', 'STOR', 'CB', 'EXIST', 'PT', 'Existing catch basin', 'utility_structures'),
            ('CIV', 'STOR', 'INLET', 'PROP', 'PT', 'Proposed inlet', 'utility_structures'),
            ('CIV', 'ROAD', 'HYDRA', 'NEW', 'PT', 'New fire hydrant', 'utility_structures'),
            ('CIV', 'ROAD', 'VALVE', 'EXIST', 'PT', 'Existing water valve', 'utility_structures'),
            ('SURV', 'CTRL', 'BENCH', 'EXIST', 'PT', 'Survey benchmark', 'survey_points'),
            ('SURV', 'CTRL', 'MONUMENT', 'EXIST', 'PT', 'Survey monument', 'survey_points'),
            ('SURV', 'TOPO', 'SHOT', 'EXIST', 'PT', 'Topographic shot', 'survey_points'),
            ('LAND', 'TREE', 'OAK', 'EXIST', 'PT', 'Existing oak tree', 'site_trees'),
            ('LAND', 'TREE', 'PINE', 'PROP', 'PT', 'Proposed pine tree', 'site_trees'),
            
            # === PG (Polygons) - Areas, Basins, Buildings ===
            ('CIV', 'STOR', 'BASIN', 'NEW', 'PG', 'New detention basin', 'bmps'),
            ('CIV', 'STOR', 'BIORT', 'PROP', 'PG', 'Proposed bioretention area', 'bmps'),
            ('CIV', 'STOR', 'SWALE', 'NEW', 'PG', 'New bioswale', 'bmps'),
            ('CIV', 'STOR', 'POND', 'EXIST', 'PG', 'Existing retention pond', 'ponds'),
            ('CIV', 'GRAD', 'EG', 'EXIST', 'PG', 'Existing grade surface', 'surfaces'),
            ('CIV', 'GRAD', 'FG', 'PROP', 'PG', 'Proposed finished grade', 'surfaces'),
            ('CIV', 'ADA', 'RAMP', 'NEW', 'PG', 'New ADA ramp', 'ada_features'),
            ('SURV', 'BLDG', 'FOOT', 'EXIST', 'PG', 'Building footprint', 'buildings'),
            ('LAND', 'PLNT', 'SHRUB', 'PROP', 'PG', 'Proposed shrub bed', 'planting_areas'),
            
            # === TX (Text) - Labels, Callouts, Notes ===
            ('ANNO', 'NOTE', 'GEN', 'NEW', 'TX', 'General note', 'drawing_entities'),
            ('ANNO', 'LABL', 'SPOT', 'EXIST', 'TX', 'Spot elevation label', 'drawing_entities'),
            ('ANNO', 'LABL', 'PIPE', 'NEW', 'TX', 'Pipe size label', 'drawing_entities'),
            ('ANNO', 'CALL', 'DET', 'NEW', 'TX', 'Detail callout', 'drawing_entities'),
            ('SURV', 'TOPO', 'ELEV', 'EXIST', 'TX', 'Elevation text', 'drawing_entities'),
            
            # === 3D (3D Objects) - Surfaces, Terrain Models ===
            ('CIV', 'GRAD', 'DTM', 'EXIST', '3D', 'Digital terrain model', 'surfaces'),
            ('CIV', 'GRAD', 'TIN', 'PROP', '3D', 'Proposed TIN surface', 'surfaces'),
            ('SURV', 'TOPO', '3DSURF', 'EXIST', '3D', '3D survey surface', 'surfaces'),
            
            # === BK (Blocks) - Symbols, Fixtures ===
            ('CIV', 'UTIL', 'SYM', 'NEW', 'BK', 'Utility symbol block', 'blocks'),
            ('LAND', 'PLNT', 'SYM', 'PROP', 'BK', 'Plant symbol block', 'blocks'),
            ('ANNO', 'SYM', 'NORTH', 'NEW', 'BK', 'North arrow symbol', 'blocks'),
            ('ANNO', 'SYM', 'TITLE', 'NEW', 'BK', 'Title block', 'blocks'),
            
            # === HT (Hatches) - Fill Patterns ===
            ('CIV', 'UTIL', 'FILL', 'NEW', 'HT', 'Utility trench hatch', 'drawing_entities'),
            ('CIV', 'GRAD', 'FILL', 'PROP', 'HT', 'Grading fill pattern', 'drawing_entities'),
            ('LAND', 'PLNT', 'FILL', 'PROP', 'HT', 'Planting area hatch', 'drawing_entities'),
            
            # === DIM (Dimensions) - Measurements ===
            ('ANNO', 'DIM', 'LIN', 'NEW', 'DIM', 'Linear dimension', 'drawing_entities'),
            ('ANNO', 'DIM', 'RAD', 'NEW', 'DIM', 'Radius dimension', 'drawing_entities'),
            ('ANNO', 'DIM', 'ANGLE', 'NEW', 'DIM', 'Angular dimension', 'drawing_entities'),
        ]
        
        # Get metadata from database for enrichment
        phases_query = """
            SELECT code, full_name, color_rgb FROM phase_codes 
            WHERE is_active = TRUE
        """
        phases_data = {p['code']: p for p in execute_query(phases_query)}
        
        geometries_query = """
            SELECT code, full_name, dxf_entity_types FROM geometry_codes 
            WHERE is_active = TRUE
        """
        geometries_data = {g['code']: g for g in execute_query(geometries_query)}
        
        disciplines_query = """
            SELECT code, full_name FROM discipline_codes 
            WHERE is_active = TRUE
        """
        disciplines_data = {d['code']: d for d in execute_query(disciplines_query)}
        
        categories_query = """
            SELECT c.code, c.full_name, d.code as discipline_code
            FROM category_codes c
            JOIN discipline_codes d ON c.discipline_id = d.discipline_id
            WHERE c.is_active = TRUE
        """
        categories_data = {}
        for c in execute_query(categories_query):
            key = f"{c['discipline_code']}-{c['code']}"
            categories_data[key] = c
        
        object_types_query = """
            SELECT t.code, t.full_name, c.code as category_code, d.code as discipline_code
            FROM object_type_codes t
            JOIN category_codes c ON t.category_id = c.category_id
            JOIN discipline_codes d ON c.discipline_id = d.discipline_id
            WHERE t.is_active = TRUE
        """
        object_types_data = {}
        for t in execute_query(object_types_query):
            key = f"{t['discipline_code']}-{t['category_code']}-{t['code']}"
            object_types_data[key] = t
        
        # Build enriched examples from catalog
        examples = []
        for disc, cat, obj, phase, geom, description, db_table in catalog:
            # Build layer name using standard pattern
            layer_name = f"{disc}-{cat}-{obj}-{phase}-{geom}"
            
            # Enrich with metadata
            phase_info = phases_data.get(phase, {})
            geom_info = geometries_data.get(geom, {})
            disc_info = disciplines_data.get(disc, {})
            cat_key = f"{disc}-{cat}"
            cat_info = categories_data.get(cat_key, {})
            obj_key = f"{disc}-{cat}-{obj}"
            obj_info = object_types_data.get(obj_key, {})
            
            examples.append({
                'layer_name': layer_name,
                'description': description,
                'object_type': obj_info.get('full_name', obj),
                'object_code': obj,
                'discipline': disc_info.get('full_name', disc),
                'discipline_code': disc,
                'category': cat_info.get('full_name', cat),
                'category_code': cat,
                'phase': phase_info.get('full_name', phase),
                'phase_code': phase,
                'phase_color': phase_info.get('color_rgb'),
                'geometry': geom_info.get('full_name', geom),
                'geometry_code': geom,
                'dxf_types': geom_info.get('dxf_entity_types', ''),
                'database_table': db_table
            })
        
        return jsonify({
            'examples': examples,
            'count': len(examples)
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

# ============================================
# IMPORT TEMPLATE MANAGER API ENDPOINTS
# ============================================

@app.route('/api/import-templates', methods=['GET'])
def get_import_templates():
    """Get all import mapping patterns with full details"""
    try:
        query = """
            SELECT 
                m.mapping_id, m.client_name, m.source_pattern,
                m.regex_pattern, m.extraction_rules, m.confidence_score,
                m.is_active, m.created_at, m.updated_at,
                d.code as discipline_code, d.full_name as discipline_name,
                c.code as category_code, c.full_name as category_name,
                t.code as type_code, t.full_name as type_name
            FROM import_mapping_patterns m
            LEFT JOIN discipline_codes d ON m.target_discipline_id = d.discipline_id
            LEFT JOIN category_codes c ON m.target_category_id = c.category_id
            LEFT JOIN object_type_codes t ON m.target_type_id = t.type_id
            ORDER BY m.confidence_score DESC, m.client_name, m.source_pattern
        """
        patterns = execute_query(query)
        return jsonify({'patterns': patterns})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-templates/<int:mapping_id>', methods=['GET'])
def get_import_template(mapping_id):
    """Get a single import mapping pattern by ID"""
    try:
        query = """
            SELECT 
                m.mapping_id, m.client_name, m.source_pattern,
                m.regex_pattern, m.extraction_rules, m.confidence_score,
                m.is_active, m.created_at, m.updated_at,
                m.target_discipline_id, m.target_category_id, m.target_type_id,
                d.code as discipline_code, d.full_name as discipline_name,
                c.code as category_code, c.full_name as category_name,
                t.code as type_code, t.full_name as type_name
            FROM import_mapping_patterns m
            LEFT JOIN discipline_codes d ON m.target_discipline_id = d.discipline_id
            LEFT JOIN category_codes c ON m.target_category_id = c.category_id
            LEFT JOIN object_type_codes t ON m.target_type_id = t.type_id
            WHERE m.mapping_id = %s
        """
        result = execute_query(query, (mapping_id,))
        if result:
            return jsonify({'pattern': result[0]})
        else:
            return jsonify({'error': 'Pattern not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-templates', methods=['POST'])
def create_import_template():
    """Create a new import mapping pattern"""
    try:
        data = request.get_json()
        
        # Get IDs for vocabulary codes if provided
        discipline_id = None
        category_id = None
        type_id = None
        
        if data.get('discipline_code'):
            result = execute_query(
                "SELECT discipline_id FROM discipline_codes WHERE code = %s",
                (data['discipline_code'],)
            )
            if result:
                discipline_id = result[0]['discipline_id']
        
        if data.get('category_code') and discipline_id:
            result = execute_query(
                "SELECT category_id FROM category_codes WHERE code = %s AND discipline_id = %s",
                (data['category_code'], discipline_id)
            )
            if result:
                category_id = result[0]['category_id']
        
        if data.get('type_code') and category_id:
            result = execute_query(
                "SELECT type_id FROM object_type_codes WHERE code = %s AND category_id = %s",
                (data['type_code'], category_id)
            )
            if result:
                type_id = result[0]['type_id']
        
        query = """
            INSERT INTO import_mapping_patterns
            (client_name, source_pattern, regex_pattern, extraction_rules,
             target_discipline_id, target_category_id, target_type_id, confidence_score)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING mapping_id
        """
        
        params = (
            data['client_name'],
            data['source_pattern'],
            data['regex_pattern'],
            json.dumps(data.get('extraction_rules', {})),
            discipline_id,
            category_id,
            type_id,
            data.get('confidence_score', 80)
        )
        
        result = execute_query(query, params)
        if result:
            return jsonify({'mapping_id': result[0]['mapping_id'], 'message': 'Pattern created successfully'}), 201
        else:
            return jsonify({'error': 'Failed to create pattern'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-templates/<int:mapping_id>', methods=['PUT'])
def update_import_template(mapping_id):
    """Update an existing import mapping pattern"""
    try:
        data = request.get_json()
        
        # Get IDs for vocabulary codes if provided
        discipline_id = None
        category_id = None
        type_id = None
        
        if data.get('discipline_code'):
            result = execute_query(
                "SELECT discipline_id FROM discipline_codes WHERE code = %s",
                (data['discipline_code'],)
            )
            if result:
                discipline_id = result[0]['discipline_id']
        
        if data.get('category_code') and discipline_id:
            result = execute_query(
                "SELECT category_id FROM category_codes WHERE code = %s AND discipline_id = %s",
                (data['category_code'], discipline_id)
            )
            if result:
                category_id = result[0]['category_id']
        
        if data.get('type_code') and category_id:
            result = execute_query(
                "SELECT type_id FROM object_type_codes WHERE code = %s AND category_id = %s",
                (data['type_code'], category_id)
            )
            if result:
                type_id = result[0]['type_id']
        
        query = """
            UPDATE import_mapping_patterns
            SET client_name = %s,
                source_pattern = %s,
                regex_pattern = %s,
                extraction_rules = %s,
                target_discipline_id = %s,
                target_category_id = %s,
                target_type_id = %s,
                confidence_score = %s,
                is_active = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE mapping_id = %s
        """
        
        params = (
            data['client_name'],
            data['source_pattern'],
            data['regex_pattern'],
            json.dumps(data.get('extraction_rules', {})),
            discipline_id,
            category_id,
            type_id,
            data.get('confidence_score', 80),
            data.get('is_active', True),
            mapping_id
        )
        
        execute_query(query, params, fetch=False)
        return jsonify({'message': 'Pattern updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-templates/<int:mapping_id>', methods=['DELETE'])
def delete_import_template(mapping_id):
    """Delete an import mapping pattern"""
    try:
        query = "DELETE FROM import_mapping_patterns WHERE mapping_id = %s"
        execute_query(query, (mapping_id,), fetch=False)
        return jsonify({'message': 'Pattern deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-templates/test', methods=['POST'])
def test_import_template():
    """Test a regex pattern against layer names"""
    try:
        data = request.get_json()
        regex_pattern = data.get('regex_pattern')
        test_names = data.get('test_names', [])
        extraction_rules = data.get('extraction_rules', {})
        
        import re
        results = []
        
        for layer_name in test_names:
            try:
                match = re.match(regex_pattern, layer_name, re.IGNORECASE)
                if match:
                    groups = match.groupdict()
                    results.append({
                        'layer_name': layer_name,
                        'matched': True,
                        'groups': groups,
                        'full_match': match.group(0)
                    })
                else:
                    results.append({
                        'layer_name': layer_name,
                        'matched': False
                    })
            except re.error as e:
                return jsonify({'error': f'Invalid regex: {str(e)}'}), 400
        
        return jsonify({'results': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# VOCABULARY CRUD API ENDPOINTS (PRODUCTION-READY)
# ============================================

def invalidate_classifier_cache():
    """Reload LayerClassifierV3 cache after vocabulary changes"""
    try:
        from standards.layer_classifier_v3 import LayerClassifierV3
        classifier = LayerClassifierV3()
        classifier.reload_codes()
    except Exception as e:
        print(f"Warning: Failed to reload classifier cache: {e}")

def validate_table_exists(table_name):
    """
    Validate that a table exists in the database by checking information_schema.
    Returns (is_valid, message) tuple.
    """
    if not table_name or not isinstance(table_name, str):
        return False, "Table name must be a non-empty string"
    
    table_name = table_name.strip().lower()
    
    if not table_name.replace('_', '').isalnum():
        return False, "Table name contains invalid characters"
    
    try:
        query = """
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            AND table_name = %s
        """
        result = execute_query(query, (table_name,))
        
        if result:
            return True, f"Table '{table_name}' exists"
        else:
            return False, f"Table '{table_name}' does not exist in the database"
    except Exception as e:
        return False, f"Error validating table: {str(e)}"

# ===== DISCIPLINES =====

@app.route('/api/vocabulary/disciplines', methods=['POST'])
def create_vocabulary_discipline():
    """Create a new discipline code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        check_query = "SELECT discipline_id FROM discipline_codes WHERE UPPER(code) = %s"
        existing = execute_query(check_query, (code,))
        if existing:
            return jsonify({'error': f'Discipline code {code} already exists'}), 409
        
        query = """
            INSERT INTO discipline_codes (code, full_name, description, sort_order, is_active)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING discipline_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True)
        ))
        
        invalidate_classifier_cache()
        
        return jsonify({
            'discipline_id': result[0]['discipline_id'],
            'message': f'Discipline {code} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/disciplines/<int:discipline_id>', methods=['PUT'])
def update_vocabulary_discipline(discipline_id):
    """Update a discipline code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        check_query = """
            SELECT discipline_id FROM discipline_codes 
            WHERE UPPER(code) = %s AND discipline_id != %s
        """
        existing = execute_query(check_query, (code, discipline_id))
        if existing:
            return jsonify({'error': f'Discipline code {code} already exists'}), 409
        
        query = """
            UPDATE discipline_codes 
            SET code = %s, full_name = %s, description = %s, sort_order = %s, is_active = %s
            WHERE discipline_id = %s
            RETURNING discipline_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True),
            discipline_id
        ))
        
        if not result:
            return jsonify({'error': 'Discipline not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': f'Discipline {code} updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/disciplines/<int:discipline_id>', methods=['DELETE'])
def delete_vocabulary_discipline(discipline_id):
    """Delete a discipline code with dependency checking"""
    try:
        exists_query = "SELECT discipline_id FROM discipline_codes WHERE discipline_id = %s"
        exists = execute_query(exists_query, (discipline_id,))
        if not exists:
            return jsonify({'error': 'Discipline not found'}), 404
        
        dep_query = """
            SELECT COUNT(*) as count FROM category_codes WHERE discipline_id = %s
        """
        deps = execute_query(dep_query, (discipline_id,))
        if deps[0]['count'] > 0:
            return jsonify({'error': f'Cannot delete discipline: {deps[0]["count"]} categories depend on it'}), 409
        
        execute_query("DELETE FROM discipline_codes WHERE discipline_id = %s", (discipline_id,), fetch=False)
        
        invalidate_classifier_cache()
        
        return jsonify({'message': 'Discipline deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== CATEGORIES =====

@app.route('/api/vocabulary/categories', methods=['POST'])
def create_vocabulary_category():
    """Create a new category code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name') or not data.get('discipline_id'):
            return jsonify({'error': 'Code, full_name, and discipline_id are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        disc_query = "SELECT discipline_id FROM discipline_codes WHERE discipline_id = %s"
        disc_exists = execute_query(disc_query, (data['discipline_id'],))
        if not disc_exists:
            return jsonify({'error': 'Invalid discipline_id'}), 400
        
        check_query = """
            SELECT category_id FROM category_codes 
            WHERE UPPER(code) = %s AND discipline_id = %s
        """
        existing = execute_query(check_query, (code, data['discipline_id']))
        if existing:
            return jsonify({'error': f'Category code {code} already exists for this discipline'}), 409
        
        query = """
            INSERT INTO category_codes (discipline_id, code, full_name, description, sort_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING category_id
        """
        result = execute_query(query, (
            data['discipline_id'],
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True)
        ))
        
        invalidate_classifier_cache()
        
        return jsonify({
            'category_id': result[0]['category_id'],
            'message': f'Category {code} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/categories/<int:category_id>', methods=['PUT'])
def update_vocabulary_category(category_id):
    """Update a category code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        current_query = "SELECT discipline_id FROM category_codes WHERE category_id = %s"
        current = execute_query(current_query, (category_id,))
        if not current:
            return jsonify({'error': 'Category not found'}), 404
        
        check_query = """
            SELECT category_id FROM category_codes 
            WHERE UPPER(code) = %s AND discipline_id = %s AND category_id != %s
        """
        existing = execute_query(check_query, (code, current[0]['discipline_id'], category_id))
        if existing:
            return jsonify({'error': f'Category code {code} already exists for this discipline'}), 409
        
        query = """
            UPDATE category_codes 
            SET code = %s, full_name = %s, description = %s, sort_order = %s, is_active = %s
            WHERE category_id = %s
            RETURNING category_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True),
            category_id
        ))
        
        if not result:
            return jsonify({'error': 'Category not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': f'Category {code} updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/categories/<int:category_id>', methods=['DELETE'])
def delete_vocabulary_category(category_id):
    """Delete a category code with dependency checking"""
    try:
        exists_query = "SELECT category_id FROM category_codes WHERE category_id = %s"
        exists = execute_query(exists_query, (category_id,))
        if not exists:
            return jsonify({'error': 'Category not found'}), 404
        
        dep_query = """
            SELECT COUNT(*) as count FROM object_type_codes WHERE category_id = %s
        """
        deps = execute_query(dep_query, (category_id,))
        if deps[0]['count'] > 0:
            return jsonify({'error': f'Cannot delete category: {deps[0]["count"]} object types depend on it'}), 409
        
        execute_query("DELETE FROM category_codes WHERE category_id = %s", (category_id,), fetch=False)
        
        invalidate_classifier_cache()
        
        return jsonify({'message': 'Category deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== OBJECT TYPES =====

@app.route('/api/vocabulary/object-types', methods=['POST'])
def create_object_type():
    """Create a new object type code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name') or not data.get('category_id'):
            return jsonify({'error': 'Code, full_name, and category_id are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        cat_query = "SELECT category_id FROM category_codes WHERE category_id = %s"
        cat_exists = execute_query(cat_query, (data['category_id'],))
        if not cat_exists:
            return jsonify({'error': 'Invalid category_id'}), 400
        
        check_query = """
            SELECT type_id FROM object_type_codes 
            WHERE UPPER(code) = %s AND category_id = %s
        """
        existing = execute_query(check_query, (code, data['category_id']))
        if existing:
            return jsonify({'error': f'Object type code {code} already exists for this category'}), 409
        
        database_table = data.get('database_table', '').strip() or None
        if database_table:
            registry_check = "SELECT registry_id FROM entity_registry WHERE table_name = %s AND is_active = TRUE"
            registry_entry = execute_query(registry_check, (database_table,))
            if not registry_entry:
                return jsonify({'error': f'Database table {database_table} is not registered in entity registry. Please register it first.'}), 400
        
        query = """
            INSERT INTO object_type_codes (category_id, code, full_name, description, database_table, sort_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING type_id
        """
        result = execute_query(query, (
            data['category_id'],
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            database_table,
            data.get('sort_order', 100),
            data.get('is_active', True)
        ))
        
        invalidate_classifier_cache()
        
        return jsonify({
            'type_id': result[0]['type_id'],
            'message': f'Object type {code} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/object-types/<int:type_id>', methods=['PUT'])
def update_object_type(type_id):
    """Update an object type code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        current_query = "SELECT category_id FROM object_type_codes WHERE type_id = %s"
        current = execute_query(current_query, (type_id,))
        if not current:
            return jsonify({'error': 'Object type not found'}), 404
        
        check_query = """
            SELECT type_id FROM object_type_codes 
            WHERE UPPER(code) = %s AND category_id = %s AND type_id != %s
        """
        existing = execute_query(check_query, (code, current[0]['category_id'], type_id))
        if existing:
            return jsonify({'error': f'Object type code {code} already exists for this category'}), 409
        
        database_table = data.get('database_table', '').strip() or None
        if database_table:
            registry_check = "SELECT registry_id FROM entity_registry WHERE table_name = %s AND is_active = TRUE"
            registry_entry = execute_query(registry_check, (database_table,))
            if not registry_entry:
                return jsonify({'error': f'Database table {database_table} is not registered in entity registry. Please register it first.'}), 400
        
        query = """
            UPDATE object_type_codes 
            SET code = %s, full_name = %s, description = %s, database_table = %s, sort_order = %s, is_active = %s
            WHERE type_id = %s
            RETURNING type_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            database_table,
            data.get('sort_order', 100),
            data.get('is_active', True),
            type_id
        ))
        
        if not result:
            return jsonify({'error': 'Object type not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': f'Object type {code} updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/object-types/<int:type_id>', methods=['DELETE'])
def delete_object_type(type_id):
    """Delete an object type code (soft delete - mark inactive)"""
    try:
        query = "UPDATE object_type_codes SET is_active = FALSE WHERE type_id = %s RETURNING type_id"
        result = execute_query(query, (type_id,))
        
        if not result:
            return jsonify({'error': 'Object type not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': 'Object type deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== PHASES =====

@app.route('/api/vocabulary/phases', methods=['POST'])
def create_phase():
    """Create a new phase code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        check_query = "SELECT phase_id FROM phase_codes WHERE UPPER(code) = %s"
        existing = execute_query(check_query, (code,))
        if existing:
            return jsonify({'error': f'Phase code {code} already exists'}), 409
        
        query = """
            INSERT INTO phase_codes (code, full_name, description, color_rgb, sort_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING phase_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('color_rgb', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True)
        ))
        
        invalidate_classifier_cache()
        
        return jsonify({
            'phase_id': result[0]['phase_id'],
            'message': f'Phase {code} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/phases/<int:phase_id>', methods=['PUT'])
def update_phase(phase_id):
    """Update a phase code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        check_query = """
            SELECT phase_id FROM phase_codes 
            WHERE UPPER(code) = %s AND phase_id != %s
        """
        existing = execute_query(check_query, (code, phase_id))
        if existing:
            return jsonify({'error': f'Phase code {code} already exists'}), 409
        
        query = """
            UPDATE phase_codes 
            SET code = %s, full_name = %s, description = %s, color_rgb = %s, sort_order = %s, is_active = %s
            WHERE phase_id = %s
            RETURNING phase_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('color_rgb', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True),
            phase_id
        ))
        
        if not result:
            return jsonify({'error': 'Phase not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': f'Phase {code} updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/phases/<int:phase_id>', methods=['DELETE'])
def delete_phase(phase_id):
    """Delete a phase code (soft delete - mark inactive)"""
    try:
        query = "UPDATE phase_codes SET is_active = FALSE WHERE phase_id = %s RETURNING phase_id"
        result = execute_query(query, (phase_id,))
        
        if not result:
            return jsonify({'error': 'Phase not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': 'Phase deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== GEOMETRIES =====

@app.route('/api/vocabulary/geometries', methods=['POST'])
def create_geometry():
    """Create a new geometry code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        check_query = "SELECT geometry_id FROM geometry_codes WHERE UPPER(code) = %s"
        existing = execute_query(check_query, (code,))
        if existing:
            return jsonify({'error': f'Geometry code {code} already exists'}), 409
        
        query = """
            INSERT INTO geometry_codes (code, full_name, description, dxf_entity_types, sort_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING geometry_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('dxf_entity_types', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True)
        ))
        
        invalidate_classifier_cache()
        
        return jsonify({
            'geometry_id': result[0]['geometry_id'],
            'message': f'Geometry {code} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/geometries/<int:geometry_id>', methods=['PUT'])
def update_geometry(geometry_id):
    """Update a geometry code with validation"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('full_name'):
            return jsonify({'error': 'Code and full_name are required'}), 400
        
        code = data['code'].upper().strip()
        if len(code) > 20:
            return jsonify({'error': 'Code must be 20 characters or less'}), 400
        
        check_query = """
            SELECT geometry_id FROM geometry_codes 
            WHERE UPPER(code) = %s AND geometry_id != %s
        """
        existing = execute_query(check_query, (code, geometry_id))
        if existing:
            return jsonify({'error': f'Geometry code {code} already exists'}), 409
        
        query = """
            UPDATE geometry_codes 
            SET code = %s, full_name = %s, description = %s, dxf_entity_types = %s, sort_order = %s, is_active = %s
            WHERE geometry_id = %s
            RETURNING geometry_id
        """
        result = execute_query(query, (
            code,
            data['full_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('dxf_entity_types', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True),
            geometry_id
        ))
        
        if not result:
            return jsonify({'error': 'Geometry not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': f'Geometry {code} updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/geometries/<int:geometry_id>', methods=['DELETE'])
def delete_geometry(geometry_id):
    """Delete a geometry code (soft delete - mark inactive)"""
    try:
        query = "UPDATE geometry_codes SET is_active = FALSE WHERE geometry_id = %s RETURNING geometry_id"
        result = execute_query(query, (geometry_id,))
        
        if not result:
            return jsonify({'error': 'Geometry not found'}), 404
        
        invalidate_classifier_cache()
        
        return jsonify({'message': 'Geometry deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== ENTITY REGISTRY =====

@app.route('/api/entity-registry')
def get_entity_registry():
    """Get all entity registry entries"""
    try:
        query = """
            SELECT registry_id, table_name, display_name, description, icon, category, is_active, sort_order
            FROM entity_registry
            WHERE is_active = TRUE
            ORDER BY category, sort_order, display_name
        """
        entries = execute_query(query)
        return jsonify({'entities': entries})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity-registry/<int:registry_id>')
def get_entity_registry_detail(registry_id):
    """Get a specific entity registry entry"""
    try:
        query = """
            SELECT registry_id, table_name, display_name, description, icon, category, is_active, sort_order
            FROM entity_registry
            WHERE registry_id = %s
        """
        result = execute_query(query, (registry_id,))
        
        if not result:
            return jsonify({'error': 'Entity registry entry not found'}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity-registry', methods=['POST'])
def create_entity_registry():
    """Create a new entity registry entry with table validation"""
    try:
        data = request.get_json()
        
        if not data.get('table_name') or not data.get('display_name'):
            return jsonify({'error': 'table_name and display_name are required'}), 400
        
        table_name = data['table_name'].strip().lower()
        
        is_valid, message = validate_table_exists(table_name)
        if not is_valid:
            return jsonify({'error': f'Table validation failed: {message}'}), 400
        
        check_query = "SELECT registry_id FROM entity_registry WHERE table_name = %s"
        existing = execute_query(check_query, (table_name,))
        if existing:
            return jsonify({'error': f'Entity registry entry for table {table_name} already exists'}), 409
        
        query = """
            INSERT INTO entity_registry (table_name, display_name, description, icon, category, sort_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING registry_id
        """
        result = execute_query(query, (
            table_name,
            data['display_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('icon', '').strip() or None,
            data.get('category', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True)
        ))
        
        return jsonify({
            'registry_id': result[0]['registry_id'],
            'message': f'Entity registry entry for {table_name} created successfully',
            'table_validation': message
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity-registry/<int:registry_id>', methods=['PUT'])
def update_entity_registry(registry_id):
    """Update an entity registry entry with table validation"""
    try:
        data = request.get_json()
        
        if not data.get('table_name') or not data.get('display_name'):
            return jsonify({'error': 'table_name and display_name are required'}), 400
        
        table_name = data['table_name'].strip().lower()
        
        is_valid, message = validate_table_exists(table_name)
        if not is_valid:
            return jsonify({'error': f'Table validation failed: {message}'}), 400
        
        check_query = """
            SELECT registry_id FROM entity_registry 
            WHERE table_name = %s AND registry_id != %s
        """
        existing = execute_query(check_query, (table_name, registry_id))
        if existing:
            return jsonify({'error': f'Entity registry entry for table {table_name} already exists'}), 409
        
        query = """
            UPDATE entity_registry 
            SET table_name = %s, display_name = %s, description = %s, icon = %s, category = %s, sort_order = %s, is_active = %s
            WHERE registry_id = %s
            RETURNING registry_id
        """
        result = execute_query(query, (
            table_name,
            data['display_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('icon', '').strip() or None,
            data.get('category', '').strip() or None,
            data.get('sort_order', 100),
            data.get('is_active', True),
            registry_id
        ))
        
        if not result:
            return jsonify({'error': 'Entity registry entry not found'}), 404
        
        return jsonify({
            'message': f'Entity registry entry for {table_name} updated successfully',
            'table_validation': message
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity-registry/<int:registry_id>', methods=['DELETE'])
def delete_entity_registry(registry_id):
    """Delete an entity registry entry (soft delete - mark inactive)"""
    try:
        query = "UPDATE entity_registry SET is_active = FALSE WHERE registry_id = %s RETURNING registry_id"
        result = execute_query(query, (registry_id,))
        
        if not result:
            return jsonify({'error': 'Entity registry entry not found'}), 404
        
        return jsonify({'message': 'Entity registry entry deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# CAD STANDARDS / LAYER GENERATOR API
# ============================================================================

@app.route('/api/cad-standards/disciplines')
def get_layer_disciplines():
    """Get all active layer disciplines"""
    try:
        query = """
            SELECT discipline_id, discipline_code, discipline_name, description, display_order, is_active
            FROM layer_disciplines
            WHERE is_active = TRUE
            ORDER BY display_order, discipline_name
        """
        disciplines = execute_query(query)
        return jsonify(disciplines)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/disciplines', methods=['POST'])
def create_layer_discipline():
    """Create a new discipline"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                # Check for duplicate code
                cur.execute("SELECT 1 FROM layer_disciplines WHERE discipline_code = %s", (code,))
                if cur.fetchone():
                    return jsonify({'error': f'Discipline code "{code}" already exists'}), 409
                
                # Insert new discipline
                cur.execute("""
                    INSERT INTO layer_disciplines (discipline_code, discipline_name, description, display_order, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING discipline_id
                """, (code, name, description, display_order, is_active))
                
                result = cur.fetchone()
                conn.commit()
                
                return jsonify({
                    'message': 'Discipline created successfully',
                    'discipline_id': result[0]
                }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/disciplines/<int:discipline_id>', methods=['PUT'])
def update_layer_discipline(discipline_id):
    """Update an existing discipline"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                # Check for duplicate code (excluding current record)
                cur.execute("""
                    SELECT 1 FROM layer_disciplines 
                    WHERE discipline_code = %s AND discipline_id != %s
                """, (code, discipline_id))
                if cur.fetchone():
                    return jsonify({'error': f'Discipline code "{code}" already exists'}), 409
                
                # Update discipline
                cur.execute("""
                    UPDATE layer_disciplines
                    SET discipline_code = %s, discipline_name = %s, description = %s,
                        display_order = %s, is_active = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE discipline_id = %s
                    RETURNING discipline_id
                """, (code, name, description, display_order, is_active, discipline_id))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Discipline not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Discipline updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/disciplines/<int:discipline_id>', methods=['DELETE'])
def delete_layer_discipline(discipline_id):
    """Delete a discipline"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                # Check if discipline is in use (you can add more checks here for layer_name_config, etc.)
                # For now, just delete it
                cur.execute("""
                    DELETE FROM layer_disciplines
                    WHERE discipline_id = %s
                    RETURNING discipline_id
                """, (discipline_id,))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Discipline not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Discipline deleted successfully'})
    except Exception as e:
        # Catch foreign key constraint violations
        if 'foreign key' in str(e).lower() or 'violates' in str(e).lower():
            return jsonify({'error': 'Cannot delete: discipline is in use'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/categories')
def get_layer_categories():
    """Get layer categories, optionally filtered by discipline"""
    try:
        discipline = request.args.get('discipline')
        
        if discipline:
            query = """
                SELECT category_id, category_code, category_name, description, 
                       valid_for_disciplines, display_order
                FROM layer_categories
                WHERE is_active = TRUE
                AND (valid_for_disciplines = '{}' OR %s = ANY(valid_for_disciplines))
                ORDER BY display_order, category_name
            """
            categories = execute_query(query, (discipline,))
        else:
            query = """
                SELECT category_id, category_code, category_name, description, 
                       valid_for_disciplines, display_order
                FROM layer_categories
                WHERE is_active = TRUE
                ORDER BY display_order, category_name
            """
            categories = execute_query(query)
        
        return jsonify(categories)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/categories', methods=['POST'])
def create_layer_category():
    """Create a new category"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM layer_categories WHERE category_code = %s", (code,))
                if cur.fetchone():
                    return jsonify({'error': f'Category code "{code}" already exists'}), 409
                
                cur.execute("""
                    INSERT INTO layer_categories (category_code, category_name, description, display_order, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING category_id
                """, (code, name, description, display_order, is_active))
                
                result = cur.fetchone()
                conn.commit()
                
                return jsonify({
                    'message': 'Category created successfully',
                    'category_id': result[0]
                }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/categories/<int:category_id>', methods=['PUT'])
def update_layer_category(category_id):
    """Update an existing category"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT 1 FROM layer_categories 
                    WHERE category_code = %s AND category_id != %s
                """, (code, category_id))
                if cur.fetchone():
                    return jsonify({'error': f'Category code "{code}" already exists'}), 409
                
                cur.execute("""
                    UPDATE layer_categories
                    SET category_code = %s, category_name = %s, description = %s,
                        display_order = %s, is_active = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE category_id = %s
                    RETURNING category_id
                """, (code, name, description, display_order, is_active, category_id))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Category not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Category updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/categories/<int:category_id>', methods=['DELETE'])
def delete_layer_category(category_id):
    """Delete a category"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM layer_categories
                    WHERE category_id = %s
                    RETURNING category_id
                """, (category_id,))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Category not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Category deleted successfully'})
    except Exception as e:
        if 'foreign key' in str(e).lower() or 'violates' in str(e).lower():
            return jsonify({'error': 'Cannot delete: category is in use'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/objects')
def get_layer_objects():
    """Get layer objects, optionally filtered by category"""
    try:
        category = request.args.get('category')
        
        if category:
            query = """
                SELECT object_id, object_code, object_name, description, 
                       valid_for_categories, network_mode, utility_type, display_order
                FROM layer_objects
                WHERE is_active = TRUE
                AND (valid_for_categories = '{}' OR %s = ANY(valid_for_categories))
                ORDER BY display_order, object_name
            """
            objects = execute_query(query, (category,))
        else:
            query = """
                SELECT object_id, object_code, object_name, description, 
                       valid_for_categories, network_mode, utility_type, display_order
                FROM layer_objects
                WHERE is_active = TRUE
                ORDER BY display_order, object_name
            """
            objects = execute_query(query)
        
        return jsonify(objects)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/objects', methods=['POST'])
def create_layer_object():
    """Create a new object"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM layer_objects WHERE object_code = %s", (code,))
                if cur.fetchone():
                    return jsonify({'error': f'Object code "{code}" already exists'}), 409
                
                cur.execute("""
                    INSERT INTO layer_objects (object_code, object_name, description, display_order, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING object_id
                """, (code, name, description, display_order, is_active))
                
                result = cur.fetchone()
                conn.commit()
                
                return jsonify({
                    'message': 'Object created successfully',
                    'object_id': result[0]
                }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/objects/<int:object_id>', methods=['PUT'])
def update_layer_object(object_id):
    """Update an existing object"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT 1 FROM layer_objects 
                    WHERE object_code = %s AND object_id != %s
                """, (code, object_id))
                if cur.fetchone():
                    return jsonify({'error': f'Object code "{code}" already exists'}), 409
                
                cur.execute("""
                    UPDATE layer_objects
                    SET object_code = %s, object_name = %s, description = %s,
                        display_order = %s, is_active = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE object_id = %s
                    RETURNING object_id
                """, (code, name, description, display_order, is_active, object_id))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Object not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Object updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/objects/<int:object_id>', methods=['DELETE'])
def delete_layer_object(object_id):
    """Delete an object"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM layer_objects
                    WHERE object_id = %s
                    RETURNING object_id
                """, (object_id,))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Object not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Object deleted successfully'})
    except Exception as e:
        if 'foreign key' in str(e).lower() or 'violates' in str(e).lower():
            return jsonify({'error': 'Cannot delete: object is in use'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/object-tools')
def get_layer_object_tools():
    """Get specialized tools available for layer objects"""
    try:
        object_code = request.args.get('object_code')
        
        if object_code:
            query = """
                SELECT 
                    lot.id,
                    lot.object_code,
                    lot.tool_name,
                    lot.tool_url,
                    lot.tool_icon,
                    lot.description,
                    lot.display_order,
                    lo.object_name
                FROM layer_object_tools lot
                LEFT JOIN layer_objects lo ON lot.object_code = lo.object_code
                WHERE lot.is_active = TRUE 
                AND lot.object_code = %s
                ORDER BY lot.display_order, lot.tool_name
            """
            tools = execute_query(query, (object_code.upper(),))
        else:
            query = """
                SELECT 
                    lot.id,
                    lot.object_code,
                    lot.tool_name,
                    lot.tool_url,
                    lot.tool_icon,
                    lot.description,
                    lot.display_order,
                    lo.object_name
                FROM layer_object_tools lot
                LEFT JOIN layer_objects lo ON lot.object_code = lo.object_code
                WHERE lot.is_active = TRUE
                ORDER BY lot.object_code, lot.display_order, lot.tool_name
            """
            tools = execute_query(query)
        
        return jsonify(tools)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/specialized-tools-directory')
@cache.cached(timeout=300, query_string=True)
def get_specialized_tools_directory():
    """Get all specialized tools with curated layer examples for discovery/directory views"""
    try:
        tools_query = """
            SELECT 
                lot.id,
                lot.object_code,
                lot.tool_name,
                lot.tool_url,
                lot.tool_icon,
                lot.description,
                lot.display_order,
                lo.object_name,
                lo.valid_for_categories
            FROM layer_object_tools lot
            LEFT JOIN layer_objects lo ON lot.object_code = lo.object_code
            WHERE lot.is_active = TRUE
            ORDER BY lot.display_order, lot.tool_name
        """
        tools = execute_query(tools_query)
        
        phases_query = """
            SELECT phase_code, phase_name 
            FROM layer_phases 
            WHERE is_active = TRUE 
            ORDER BY display_order 
            LIMIT 5
        """
        phases = execute_query(phases_query)
        
        geometries_query = """
            SELECT geom_code as geometry_code, geom_name as geometry_name 
            FROM layer_geometries 
            WHERE is_active = TRUE 
            ORDER BY display_order 
            LIMIT 5
        """
        geometries = execute_query(geometries_query)
        
        categories_query = """
            SELECT category_code, category_name, valid_for_disciplines
            FROM layer_categories
            WHERE is_active = TRUE
            ORDER BY display_order
        """
        categories = execute_query(categories_query)
        
        disciplines_query = """
            SELECT discipline_code, discipline_name
            FROM layer_disciplines
            WHERE is_active = TRUE
            ORDER BY discipline_name
        """
        disciplines = execute_query(disciplines_query)
        
        default_phases = [p['phase_code'] for p in phases[:3]] if phases else ['EXST', 'PROP', 'DEMO']
        default_geometries = [g['geometry_code'] for g in geometries[:3]] if geometries else ['LN', 'PT', 'PG']
        
        category_lookup = {c['category_code']: c for c in categories}
        discipline_lookup = {d['discipline_code']: d for d in disciplines}
        
        result_tools = []
        for tool in tools:
            object_code = tool['object_code']
            valid_categories = tool.get('valid_for_categories', []) or []
            
            layer_examples = []
            examples_count = 0
            max_examples = 3
            
            for category_code in valid_categories[:2]:
                if examples_count >= max_examples:
                    break
                    
                category = category_lookup.get(category_code)
                if not category:
                    continue
                    
                valid_disciplines = category.get('valid_for_disciplines', []) or []
                
                for discipline_code in valid_disciplines[:2]:
                    if examples_count >= max_examples:
                        break
                    
                    phase_code = default_phases[examples_count % len(default_phases)]
                    geometry_code = default_geometries[examples_count % len(default_geometries)]
                    
                    layer_name = f"{discipline_code}-{category_code}-{object_code}-{phase_code}-{geometry_code}"
                    
                    discipline_name = discipline_lookup.get(discipline_code, {}).get('discipline_name', discipline_code)
                    category_name = category.get('category_name', category_code)
                    phase_obj = next((p for p in phases if p['phase_code'] == phase_code), None)
                    geom_obj = next((g for g in geometries if g['geometry_code'] == geometry_code), None)
                    
                    phase_name = phase_obj['phase_name'] if phase_obj else phase_code
                    geometry_name = geom_obj['geometry_name'] if geom_obj else geometry_code
                    
                    layer_examples.append({
                        'layer': layer_name,
                        'description': f"{phase_name} {tool['object_name']} - {discipline_name}/{category_name}",
                        'discipline_code': discipline_code,
                        'category_code': category_code,
                        'object_code': object_code,
                        'phase_code': phase_code,
                        'geometry_code': geometry_code
                    })
                    
                    examples_count += 1
            
            if not layer_examples:
                discipline_code = 'CIV'
                category_code = valid_categories[0] if valid_categories else 'MISC'
                phase_code = default_phases[0]
                geometry_code = default_geometries[0]
                layer_name = f"{discipline_code}-{category_code}-{object_code}-{phase_code}-{geometry_code}"
                
                layer_examples.append({
                    'layer': layer_name,
                    'description': f"{tool['object_name']} layer example",
                    'discipline_code': discipline_code,
                    'category_code': category_code,
                    'object_code': object_code,
                    'phase_code': phase_code,
                    'geometry_code': geometry_code
                })
            
            result_tools.append({
                'id': tool['id'],
                'tool_name': tool['tool_name'],
                'tool_url': tool['tool_url'],
                'tool_icon': tool['tool_icon'],
                'description': tool['description'],
                'object_code': tool['object_code'],
                'object_name': tool['object_name'],
                'layer_examples': layer_examples
            })
        
        return jsonify({
            'tools': result_tools,
            'phases': {p['phase_code']: p['phase_name'] for p in phases},
            'geometries': {g['geometry_code']: g['geometry_name'] for g in geometries}
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/phases')
def get_layer_phases():
    """Get all active layer phases"""
    try:
        query = """
            SELECT phase_id, phase_code, phase_name, description, display_order, is_active
            FROM layer_phases
            WHERE is_active = TRUE
            ORDER BY display_order, phase_name
        """
        phases = execute_query(query)
        return jsonify(phases)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/phases', methods=['POST'])
def create_layer_phase():
    """Create a new phase"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM layer_phases WHERE phase_code = %s", (code,))
                if cur.fetchone():
                    return jsonify({'error': f'Phase code "{code}" already exists'}), 409
                
                cur.execute("""
                    INSERT INTO layer_phases (phase_code, phase_name, description, display_order, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING phase_id
                """, (code, name, description, display_order, is_active))
                
                result = cur.fetchone()
                conn.commit()
                
                return jsonify({
                    'message': 'Phase created successfully',
                    'phase_id': result[0]
                }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/phases/<int:phase_id>', methods=['PUT'])
def update_layer_phase(phase_id):
    """Update an existing phase"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT 1 FROM layer_phases 
                    WHERE phase_code = %s AND phase_id != %s
                """, (code, phase_id))
                if cur.fetchone():
                    return jsonify({'error': f'Phase code "{code}" already exists'}), 409
                
                cur.execute("""
                    UPDATE layer_phases
                    SET phase_code = %s, phase_name = %s, description = %s,
                        display_order = %s, is_active = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE phase_id = %s
                    RETURNING phase_id
                """, (code, name, description, display_order, is_active, phase_id))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Phase not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Phase updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/phases/<int:phase_id>', methods=['DELETE'])
def delete_layer_phase(phase_id):
    """Delete a phase"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM layer_phases
                    WHERE phase_id = %s
                    RETURNING phase_id
                """, (phase_id,))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Phase not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Phase deleted successfully'})
    except Exception as e:
        if 'foreign key' in str(e).lower() or 'violates' in str(e).lower():
            return jsonify({'error': 'Cannot delete: phase is in use'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/geometries')
def get_layer_geometries():
    """Get all active layer geometry types"""
    try:
        query = """
            SELECT geometry_id, geom_code AS geometry_code, geom_name AS geometry_name, 
                   description, expected_entity_types, display_order, is_active
            FROM layer_geometries
            WHERE is_active = TRUE
            ORDER BY display_order, geom_name
        """
        geometries = execute_query(query)
        return jsonify(geometries)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/geometries', methods=['POST'])
def create_layer_geometry():
    """Create a new geometry"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM layer_geometries WHERE geom_code = %s", (code,))
                if cur.fetchone():
                    return jsonify({'error': f'Geometry code "{code}" already exists'}), 409
                
                cur.execute("""
                    INSERT INTO layer_geometries (geom_code, geom_name, description, display_order, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING geometry_id
                """, (code, name, description, display_order, is_active))
                
                result = cur.fetchone()
                conn.commit()
                
                return jsonify({
                    'message': 'Geometry created successfully',
                    'geometry_id': result[0]
                }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/geometries/<int:geometry_id>', methods=['PUT'])
def update_layer_geometry(geometry_id):
    """Update an existing geometry"""
    try:
        data = request.get_json()
        code = data.get('code', '').strip().upper()
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        display_order = data.get('display_order', 0)
        is_active = data.get('is_active', True)
        
        if not code or not name:
            return jsonify({'error': 'Code and name are required'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT 1 FROM layer_geometries 
                    WHERE geom_code = %s AND geometry_id != %s
                """, (code, geometry_id))
                if cur.fetchone():
                    return jsonify({'error': f'Geometry code "{code}" already exists'}), 409
                
                cur.execute("""
                    UPDATE layer_geometries
                    SET geom_code = %s, geom_name = %s, description = %s,
                        display_order = %s, is_active = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE geometry_id = %s
                    RETURNING geometry_id
                """, (code, name, description, display_order, is_active, geometry_id))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Geometry not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Geometry updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/geometries/<int:geometry_id>', methods=['DELETE'])
def delete_layer_geometry(geometry_id):
    """Delete a geometry"""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM layer_geometries
                    WHERE geometry_id = %s
                    RETURNING geometry_id
                """, (geometry_id,))
                
                result = cur.fetchone()
                if not result:
                    return jsonify({'error': 'Geometry not found'}), 404
                
                conn.commit()
                return jsonify({'message': 'Geometry deleted successfully'})
    except Exception as e:
        if 'foreign key' in str(e).lower() or 'violates' in str(e).lower():
            return jsonify({'error': 'Cannot delete: geometry is in use'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/generate-layer-name', methods=['POST'])
def generate_layer_name():
    """Generate a canonical layer name from components"""
    try:
        data = request.get_json()
        
        discipline = data.get('discipline', '').strip()
        category = data.get('category', '').strip()
        object_code = data.get('object', '').strip()
        phase = data.get('phase', '').strip()
        geometry = data.get('geometry', '').strip()
        attributes = data.get('attributes', [])
        
        if not all([discipline, category, object_code, phase, geometry]):
            return jsonify({
                'error': 'All components required: discipline, category, object, phase, geometry'
            }), 400
        
        # Build layer name with optional attributes
        # Format: DISCIPLINE-CATEGORY-TYPE-[ATTRIBUTES]-PHASE-GEOMETRY
        parts = [discipline, category, object_code]
        
        # Add attributes if provided (inserted between type and phase)
        if attributes:
            if isinstance(attributes, str):
                # Split comma-separated string
                attributes = [attr.strip().upper() for attr in attributes.split(',') if attr.strip()]
            parts.extend(attributes)
        
        parts.extend([phase, geometry])
        canonical_name = '-'.join(parts)
        
        alias_query = """
            SELECT alias, notes FROM layer_aliases
            WHERE canonical_pattern = %s AND is_active = TRUE
            LIMIT 1
        """
        alias_result = execute_query(alias_query, (canonical_name,))
        
        result = {
            'canonical_name': canonical_name,
            'alias': alias_result[0]['alias'] if alias_result else None,
            'alias_notes': alias_result[0]['notes'] if alias_result else None,
            'components': {
                'discipline': discipline,
                'category': category,
                'object': object_code,
                'attributes': attributes if attributes else [],
                'phase': phase,
                'geometry': geometry
            }
        }
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/export-all-layers')
def export_all_layer_combinations():
    """Export all valid layer combinations to CSV"""
    try:
        disciplines_query = """
            SELECT discipline_code, discipline_name 
            FROM layer_disciplines 
            WHERE is_active = TRUE 
            ORDER BY display_order, discipline_name
        """
        disciplines = execute_query(disciplines_query)
        
        categories_query = """
            SELECT category_code, category_name 
            FROM layer_categories 
            WHERE is_active = TRUE 
            ORDER BY display_order, category_name
        """
        categories = execute_query(categories_query)
        
        objects_query = """
            SELECT object_code, object_name, valid_for_categories 
            FROM layer_objects 
            WHERE is_active = TRUE 
            ORDER BY display_order, object_name
        """
        objects = execute_query(objects_query)
        
        phases_query = """
            SELECT phase_code, phase_name 
            FROM layer_phases 
            WHERE is_active = TRUE 
            ORDER BY display_order, phase_name
        """
        phases = execute_query(phases_query)
        
        geometries_query = """
            SELECT geom_code, geom_name 
            FROM layer_geometries 
            WHERE is_active = TRUE 
            ORDER BY display_order, geom_name
        """
        geometries = execute_query(geometries_query)
        
        aliases_query = """
            SELECT canonical_pattern, alias, notes 
            FROM layer_aliases 
            WHERE is_active = TRUE
        """
        aliases_data = execute_query(aliases_query)
        aliases_map = {row['canonical_pattern']: row for row in aliases_data}
        
        output = io.StringIO()
        fieldnames = [
            'Canonical_Name', 'Alias', 'Alias_Notes',
            'Discipline_Code', 'Discipline_Name',
            'Category_Code', 'Category_Name',
            'Object_Code', 'Object_Name',
            'Phase_Code', 'Phase_Name',
            'Geometry_Code', 'Geometry_Name'
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for discipline in disciplines:
            for category in categories:
                for obj in objects:
                    valid_categories = obj.get('valid_for_categories', [])
                    if valid_categories and category['category_code'] not in valid_categories:
                        continue
                    
                    for phase in phases:
                        for geometry in geometries:
                            canonical_name = f"{discipline['discipline_code']}-{category['category_code']}-{obj['object_code']}-{phase['phase_code']}-{geometry['geom_code']}"
                            
                            alias_info = aliases_map.get(canonical_name, {})
                            
                            writer.writerow({
                                'Canonical_Name': canonical_name,
                                'Alias': alias_info.get('alias', ''),
                                'Alias_Notes': alias_info.get('notes', ''),
                                'Discipline_Code': discipline['discipline_code'],
                                'Discipline_Name': discipline['discipline_name'],
                                'Category_Code': category['category_code'],
                                'Category_Name': category['category_name'],
                                'Object_Code': obj['object_code'],
                                'Object_Name': obj['object_name'],
                                'Phase_Code': phase['phase_code'],
                                'Phase_Name': phase['phase_name'],
                                'Geometry_Code': geometry['geom_code'],
                                'Geometry_Name': geometry['geom_name']
                            })
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='layer_combinations.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/aliases')
def get_layer_aliases():
    """Get all active layer aliases"""
    try:
        query = """
            SELECT alias_id, alias, canonical_pattern, discipline_code, 
                   category_code, object_code, phase_code, geom_code, 
                   notes, usage_count
            FROM layer_aliases
            WHERE is_active = TRUE
            ORDER BY alias
        """
        aliases = execute_query(query)
        return jsonify(aliases)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/translate-alias', methods=['POST'])
def translate_alias():
    """Translate a CAD-friendly alias to canonical layer name"""
    try:
        data = request.get_json()
        alias = data.get('alias', '').strip()
        
        if not alias:
            return jsonify({'error': 'Alias is required'}), 400
        
        query = """
            SELECT alias_id, alias, canonical_pattern, discipline_code, 
                   category_code, object_code, phase_code, geom_code, notes
            FROM layer_aliases
            WHERE UPPER(alias) = UPPER(%s) AND is_active = TRUE
            LIMIT 1
        """
        result = execute_query(query, (alias,))
        
        if not result:
            return jsonify({'error': 'Alias not found', 'alias': alias}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cad-standards/block-schemas')
def get_block_schemas():
    """Get block attribute schemas, optionally filtered by family"""
    try:
        family = request.args.get('family')
        
        if family:
            query = """
                SELECT schema_id, block_family, attribute_name, data_type, 
                       is_required, default_value, validation_rule, 
                       maps_to_db_table, maps_to_db_column, description, display_order
                FROM block_attribute_schemas
                WHERE block_family = %s AND is_active = TRUE
                ORDER BY display_order, attribute_name
            """
            schemas = execute_query(query, (family,))
        else:
            query = """
                SELECT DISTINCT block_family
                FROM block_attribute_schemas
                WHERE is_active = TRUE
                ORDER BY block_family
            """
            families = execute_query(query)
            return jsonify(families)
        
        return jsonify(schemas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== CLIENTS API =====

@app.route('/api/clients')
def get_clients():
    """Get all clients"""
    try:
        query = """
            SELECT client_id, client_name, short_name, contact_name, contact_email, contact_phone,
                   address, city, state, zip_code, country, website, notes, is_active, tags, attributes
            FROM clients
            WHERE is_active = TRUE
            ORDER BY client_name
        """
        clients = execute_query(query)
        return jsonify({'clients': clients})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clients/<int:client_id>')
def get_client_detail(client_id):
    """Get a specific client"""
    try:
        query = """
            SELECT client_id, client_name, short_name, contact_name, contact_email, contact_phone,
                   address, city, state, zip_code, country, website, notes, is_active, tags, attributes
            FROM clients
            WHERE client_id = %s
        """
        result = execute_query(query, (client_id,))
        
        if not result:
            return jsonify({'error': 'Client not found'}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clients', methods=['POST'])
def create_client():
    """Create a new client"""
    try:
        data = request.get_json()
        
        if not data.get('client_name'):
            return jsonify({'error': 'client_name is required'}), 400
        
        query = """
            INSERT INTO clients (client_name, short_name, contact_name, contact_email, contact_phone,
                                address, city, state, zip_code, country, website, notes, tags, attributes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING client_id
        """
        result = execute_query(query, (
            data['client_name'].strip(),
            data.get('short_name', '').strip() or None,
            data.get('contact_name', '').strip() or None,
            data.get('contact_email', '').strip() or None,
            data.get('contact_phone', '').strip() or None,
            data.get('address', '').strip() or None,
            data.get('city', '').strip() or None,
            data.get('state', '').strip() or None,
            data.get('zip_code', '').strip() or None,
            data.get('country', 'USA'),
            data.get('website', '').strip() or None,
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes')
        ))
        
        return jsonify({
            'client_id': result[0]['client_id'],
            'message': f'Client {data["client_name"]} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clients/<int:client_id>', methods=['PUT'])
def update_client(client_id):
    """Update a client"""
    try:
        data = request.get_json()
        
        if not data.get('client_name'):
            return jsonify({'error': 'client_name is required'}), 400
        
        query = """
            UPDATE clients 
            SET client_name = %s, short_name = %s, contact_name = %s, contact_email = %s, contact_phone = %s,
                address = %s, city = %s, state = %s, zip_code = %s, country = %s, website = %s, notes = %s,
                tags = %s, attributes = %s, updated_at = CURRENT_TIMESTAMP
            WHERE client_id = %s
            RETURNING client_id
        """
        result = execute_query(query, (
            data['client_name'].strip(),
            data.get('short_name', '').strip() or None,
            data.get('contact_name', '').strip() or None,
            data.get('contact_email', '').strip() or None,
            data.get('contact_phone', '').strip() or None,
            data.get('address', '').strip() or None,
            data.get('city', '').strip() or None,
            data.get('state', '').strip() or None,
            data.get('zip_code', '').strip() or None,
            data.get('country', 'USA'),
            data.get('website', '').strip() or None,
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes'),
            client_id
        ))
        
        if not result:
            return jsonify({'error': 'Client not found'}), 404
        
        return jsonify({'message': f'Client updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clients/<int:client_id>', methods=['DELETE'])
def delete_client(client_id):
    """Delete a client (soft delete)"""
    try:
        query = "UPDATE clients SET is_active = FALSE WHERE client_id = %s RETURNING client_id"
        result = execute_query(query, (client_id,))
        
        if not result:
            return jsonify({'error': 'Client not found'}), 404
        
        return jsonify({'message': 'Client deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== VENDORS API =====

@app.route('/api/vendors')
def get_vendors():
    """Get all vendors"""
    try:
        query = """
            SELECT vendor_id, vendor_name, vendor_type, contact_name, contact_email, contact_phone,
                   address, city, state, zip_code, country, website, specialty, certifications, notes,
                   is_active, tags, attributes
            FROM vendors
            WHERE is_active = TRUE
            ORDER BY vendor_name
        """
        vendors = execute_query(query)
        return jsonify({'vendors': vendors})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendors/<int:vendor_id>')
def get_vendor_detail(vendor_id):
    """Get a specific vendor"""
    try:
        query = """
            SELECT vendor_id, vendor_name, vendor_type, contact_name, contact_email, contact_phone,
                   address, city, state, zip_code, country, website, specialty, certifications, notes,
                   is_active, tags, attributes
            FROM vendors
            WHERE vendor_id = %s
        """
        result = execute_query(query, (vendor_id,))
        
        if not result:
            return jsonify({'error': 'Vendor not found'}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendors', methods=['POST'])
def create_vendor():
    """Create a new vendor"""
    try:
        data = request.get_json()
        
        if not data.get('vendor_name'):
            return jsonify({'error': 'vendor_name is required'}), 400
        
        query = """
            INSERT INTO vendors (vendor_name, vendor_type, contact_name, contact_email, contact_phone,
                                address, city, state, zip_code, country, website, specialty, certifications, notes,
                                tags, attributes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING vendor_id
        """
        result = execute_query(query, (
            data['vendor_name'].strip(),
            data.get('vendor_type', '').strip() or None,
            data.get('contact_name', '').strip() or None,
            data.get('contact_email', '').strip() or None,
            data.get('contact_phone', '').strip() or None,
            data.get('address', '').strip() or None,
            data.get('city', '').strip() or None,
            data.get('state', '').strip() or None,
            data.get('zip_code', '').strip() or None,
            data.get('country', 'USA'),
            data.get('website', '').strip() or None,
            data.get('specialty', '').strip() or None,
            data.get('certifications', []),
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes')
        ))
        
        return jsonify({
            'vendor_id': result[0]['vendor_id'],
            'message': f'Vendor {data["vendor_name"]} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendors/<int:vendor_id>', methods=['PUT'])
def update_vendor(vendor_id):
    """Update a vendor"""
    try:
        data = request.get_json()
        
        if not data.get('vendor_name'):
            return jsonify({'error': 'vendor_name is required'}), 400
        
        query = """
            UPDATE vendors 
            SET vendor_name = %s, vendor_type = %s, contact_name = %s, contact_email = %s, contact_phone = %s,
                address = %s, city = %s, state = %s, zip_code = %s, country = %s, website = %s,
                specialty = %s, certifications = %s, notes = %s, tags = %s, attributes = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE vendor_id = %s
            RETURNING vendor_id
        """
        result = execute_query(query, (
            data['vendor_name'].strip(),
            data.get('vendor_type', '').strip() or None,
            data.get('contact_name', '').strip() or None,
            data.get('contact_email', '').strip() or None,
            data.get('contact_phone', '').strip() or None,
            data.get('address', '').strip() or None,
            data.get('city', '').strip() or None,
            data.get('state', '').strip() or None,
            data.get('zip_code', '').strip() or None,
            data.get('country', 'USA'),
            data.get('website', '').strip() or None,
            data.get('specialty', '').strip() or None,
            data.get('certifications', []),
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes'),
            vendor_id
        ))
        
        if not result:
            return jsonify({'error': 'Vendor not found'}), 404
        
        return jsonify({'message': f'Vendor updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendors/<int:vendor_id>', methods=['DELETE'])
def delete_vendor(vendor_id):
    """Delete a vendor (soft delete)"""
    try:
        query = "UPDATE vendors SET is_active = FALSE WHERE vendor_id = %s RETURNING vendor_id"
        result = execute_query(query, (vendor_id,))
        
        if not result:
            return jsonify({'error': 'Vendor not found'}), 404
        
        return jsonify({'message': 'Vendor deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== MUNICIPALITIES API =====

@app.route('/api/municipalities')
def get_municipalities():
    """Get all municipalities"""
    try:
        query = """
            SELECT municipality_id, municipality_name, municipality_type, county, state, country, fips_code,
                   contact_department, contact_name, contact_email, contact_phone, address, website,
                   permit_portal_url, cad_standards_url, notes, is_active, tags, attributes
            FROM municipalities
            WHERE is_active = TRUE
            ORDER BY state, municipality_name
        """
        municipalities = execute_query(query)
        return jsonify({'municipalities': municipalities})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/municipalities/<int:municipality_id>')
def get_municipality_detail(municipality_id):
    """Get a specific municipality"""
    try:
        query = """
            SELECT municipality_id, municipality_name, municipality_type, county, state, country, fips_code,
                   contact_department, contact_name, contact_email, contact_phone, address, website,
                   permit_portal_url, cad_standards_url, notes, is_active, tags, attributes
            FROM municipalities
            WHERE municipality_id = %s
        """
        result = execute_query(query, (municipality_id,))
        
        if not result:
            return jsonify({'error': 'Municipality not found'}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/municipalities', methods=['POST'])
def create_municipality():
    """Create a new municipality"""
    try:
        data = request.get_json()
        
        if not data.get('municipality_name') or not data.get('state'):
            return jsonify({'error': 'municipality_name and state are required'}), 400
        
        query = """
            INSERT INTO municipalities (municipality_name, municipality_type, county, state, country, fips_code,
                                       contact_department, contact_name, contact_email, contact_phone, address,
                                       website, permit_portal_url, cad_standards_url, notes, tags, attributes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING municipality_id
        """
        result = execute_query(query, (
            data['municipality_name'].strip(),
            data.get('municipality_type', '').strip() or None,
            data.get('county', '').strip() or None,
            data['state'].strip(),
            data.get('country', 'USA'),
            data.get('fips_code', '').strip() or None,
            data.get('contact_department', '').strip() or None,
            data.get('contact_name', '').strip() or None,
            data.get('contact_email', '').strip() or None,
            data.get('contact_phone', '').strip() or None,
            data.get('address', '').strip() or None,
            data.get('website', '').strip() or None,
            data.get('permit_portal_url', '').strip() or None,
            data.get('cad_standards_url', '').strip() or None,
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes')
        ))
        
        return jsonify({
            'municipality_id': result[0]['municipality_id'],
            'message': f'Municipality {data["municipality_name"]} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/municipalities/<int:municipality_id>', methods=['PUT'])
def update_municipality(municipality_id):
    """Update a municipality"""
    try:
        data = request.get_json()
        
        if not data.get('municipality_name') or not data.get('state'):
            return jsonify({'error': 'municipality_name and state are required'}), 400
        
        query = """
            UPDATE municipalities 
            SET municipality_name = %s, municipality_type = %s, county = %s, state = %s, country = %s, fips_code = %s,
                contact_department = %s, contact_name = %s, contact_email = %s, contact_phone = %s, address = %s,
                website = %s, permit_portal_url = %s, cad_standards_url = %s, notes = %s, tags = %s, attributes = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE municipality_id = %s
            RETURNING municipality_id
        """
        result = execute_query(query, (
            data['municipality_name'].strip(),
            data.get('municipality_type', '').strip() or None,
            data.get('county', '').strip() or None,
            data['state'].strip(),
            data.get('country', 'USA'),
            data.get('fips_code', '').strip() or None,
            data.get('contact_department', '').strip() or None,
            data.get('contact_name', '').strip() or None,
            data.get('contact_email', '').strip() or None,
            data.get('contact_phone', '').strip() or None,
            data.get('address', '').strip() or None,
            data.get('website', '').strip() or None,
            data.get('permit_portal_url', '').strip() or None,
            data.get('cad_standards_url', '').strip() or None,
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes'),
            municipality_id
        ))
        
        if not result:
            return jsonify({'error': 'Municipality not found'}), 404
        
        return jsonify({'message': f'Municipality updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/municipalities/<int:municipality_id>', methods=['DELETE'])
def delete_municipality(municipality_id):
    """Delete a municipality (soft delete)"""
    try:
        query = "UPDATE municipalities SET is_active = FALSE WHERE municipality_id = %s RETURNING municipality_id"
        result = execute_query(query, (municipality_id,))
        
        if not result:
            return jsonify({'error': 'Municipality not found'}), 404
        
        return jsonify({'message': 'Municipality deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== COORDINATE SYSTEMS API =====

@app.route('/api/coordinate-systems')
def get_coordinate_systems():
    """Get all coordinate systems"""
    try:
        query = """
            SELECT system_id, epsg_code, system_name, region, datum, units, zone_number, notes,
                   is_active, tags, attributes
            FROM coordinate_systems
            WHERE is_active = TRUE
            ORDER BY region, system_name
        """
        systems = execute_query(query)
        return jsonify({'coordinate_systems': systems})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/coordinate-systems/<uuid:system_id>')
def get_coordinate_system_detail(system_id):
    """Get a specific coordinate system"""
    try:
        query = """
            SELECT system_id, epsg_code, system_name, region, datum, units, zone_number, notes,
                   is_active, tags, attributes
            FROM coordinate_systems
            WHERE system_id = %s
        """
        result = execute_query(query, (str(system_id),))
        
        if not result:
            return jsonify({'error': 'Coordinate system not found'}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/coordinate-systems', methods=['POST'])
def create_coordinate_system():
    """Create a new coordinate system"""
    try:
        data = request.get_json()
        
        if not data.get('system_name') or not data.get('epsg_code'):
            return jsonify({'error': 'system_name and epsg_code are required'}), 400
        
        query = """
            INSERT INTO coordinate_systems (epsg_code, system_name, region, datum, units, zone_number, notes, tags, attributes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING system_id
        """
        result = execute_query(query, (
            data['epsg_code'].strip(),
            data['system_name'].strip(),
            data.get('region', '').strip() or None,
            data.get('datum', '').strip() or None,
            data.get('units', '').strip() or None,
            data.get('zone_number'),
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes')
        ))
        
        return jsonify({
            'system_id': result[0]['system_id'],
            'message': f'Coordinate system {data["system_name"]} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/coordinate-systems/<uuid:system_id>', methods=['PUT'])
def update_coordinate_system(system_id):
    """Update a coordinate system"""
    try:
        data = request.get_json()
        
        if not data.get('system_name') or not data.get('epsg_code'):
            return jsonify({'error': 'system_name and epsg_code are required'}), 400
        
        query = """
            UPDATE coordinate_systems 
            SET epsg_code = %s, system_name = %s, region = %s, datum = %s, units = %s, zone_number = %s,
                notes = %s, tags = %s, attributes = %s, updated_at = CURRENT_TIMESTAMP
            WHERE system_id = %s
            RETURNING system_id
        """
        result = execute_query(query, (
            data['epsg_code'].strip(),
            data['system_name'].strip(),
            data.get('region', '').strip() or None,
            data.get('datum', '').strip() or None,
            data.get('units', '').strip() or None,
            data.get('zone_number'),
            data.get('notes', '').strip() or None,
            data.get('tags', []),
            data.get('attributes'),
            str(system_id)
        ))
        
        if not result:
            return jsonify({'error': 'Coordinate system not found'}), 404
        
        return jsonify({'message': f'Coordinate system updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/coordinate-systems/<uuid:system_id>', methods=['DELETE'])
def delete_coordinate_system(system_id):
    """Delete a coordinate system (soft delete)"""
    try:
        query = "UPDATE coordinate_systems SET is_active = FALSE WHERE system_id = %s RETURNING system_id"
        result = execute_query(query, (str(system_id),))
        
        if not result:
            return jsonify({'error': 'Coordinate system not found'}), 404
        
        return jsonify({'message': 'Coordinate system deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== SURVEY CODE LIBRARY API =====

@app.route('/api/survey-codes')
def get_survey_codes():
    """Get all survey codes with optional filtering"""
    try:
        discipline = request.args.get('discipline')
        category = request.args.get('category')
        connectivity = request.args.get('connectivity')
        category_group = request.args.get('category_group')
        favorites_only = request.args.get('favorites_only') == 'true'
        search = request.args.get('search', '').strip()
        
        where_clauses = ['is_active = TRUE']
        params = []
        
        if discipline:
            where_clauses.append('discipline_code = %s')
            params.append(discipline)
        
        if category:
            where_clauses.append('category_code = %s')
            params.append(category)
        
        if connectivity:
            where_clauses.append('connectivity_type = %s')
            params.append(connectivity)
        
        if category_group:
            where_clauses.append('category_group = %s')
            params.append(category_group)
        
        if favorites_only:
            where_clauses.append('is_favorite = TRUE')
        
        if search:
            where_clauses.append('(code ILIKE %s OR display_name ILIKE %s OR description ILIKE %s)')
            search_pattern = f'%{search}%'
            params.extend([search_pattern, search_pattern, search_pattern])
        
        where_sql = ' AND '.join(where_clauses)
        
        query = f"""
            SELECT code_id, code, display_name, description, discipline_code, category_code,
                   feature_type, icon_name, connectivity_type, geometry_output, auto_connect,
                   create_block, block_name, available_attributes, default_attributes,
                   required_attributes, layer_template, default_phase, category_group,
                   sort_order, is_favorite, usage_count, last_used_at, quality_score,
                   tags, attributes, created_at, updated_at
            FROM survey_code_library
            WHERE {where_sql}
            ORDER BY is_favorite DESC, sort_order, display_name
        """
        
        codes = execute_query(query, tuple(params))
        return jsonify({'survey_codes': codes, 'count': len(codes)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/<uuid:code_id>')
def get_survey_code_detail(code_id):
    """Get a specific survey code"""
    try:
        query = """
            SELECT code_id, code, display_name, description, discipline_code, category_code,
                   feature_type, icon_name, connectivity_type, geometry_output, auto_connect,
                   create_block, block_name, available_attributes, default_attributes,
                   required_attributes, layer_template, default_phase, category_group,
                   sort_order, is_favorite, usage_count, last_used_at, quality_score,
                   tags, attributes, created_at, updated_at
            FROM survey_code_library
            WHERE code_id = %s
        """
        result = execute_query(query, (str(code_id),))
        
        if not result:
            return jsonify({'error': 'Survey code not found'}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes', methods=['POST'])
def create_survey_code():
    """Create a new survey code"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('display_name'):
            return jsonify({'error': 'code and display_name are required'}), 400
        
        if not data.get('feature_type'):
            return jsonify({'error': 'feature_type is required'}), 400
        
        check_query = "SELECT code_id FROM survey_code_library WHERE code = %s"
        existing = execute_query(check_query, (data['code'].strip(),))
        if existing:
            return jsonify({'error': f'Code {data["code"]} already exists'}), 409
        
        query = """
            INSERT INTO survey_code_library (
                code, display_name, description, discipline_code, category_code, feature_type,
                icon_name, connectivity_type, geometry_output, auto_connect, create_block,
                block_name, available_attributes, default_attributes, required_attributes,
                layer_template, default_phase, category_group, sort_order, is_favorite,
                tags, attributes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING code_id, code, display_name, connectivity_type, is_favorite, created_at
        """
        
        result = execute_query(query, (
            data['code'].strip().upper(),
            data['display_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('discipline_code', '').strip() or None,
            data.get('category_code', '').strip() or None,
            data['feature_type'].strip(),
            data.get('icon_name', '').strip() or None,
            data.get('connectivity_type', 'POINT'),
            data.get('geometry_output', '').strip() or None,
            data.get('auto_connect', False),
            data.get('create_block', False),
            data.get('block_name', '').strip() or None,
            data.get('available_attributes', []),
            data.get('default_attributes', {}),
            data.get('required_attributes', []),
            data.get('layer_template', '').strip() or None,
            data.get('default_phase', 'EXIST'),
            data.get('category_group', '').strip() or None,
            data.get('sort_order', 0),
            data.get('is_favorite', False),
            data.get('tags', []),
            data.get('attributes', {})
        ))
        
        return jsonify({
            'survey_code': result[0],
            'message': f'Survey code {data["code"]} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/<uuid:code_id>', methods=['PUT'])
def update_survey_code(code_id):
    """Update a survey code"""
    try:
        data = request.get_json()
        
        if not data.get('code') or not data.get('display_name'):
            return jsonify({'error': 'code and display_name are required'}), 400
        
        check_query = """
            SELECT code_id FROM survey_code_library 
            WHERE code = %s AND code_id != %s
        """
        existing = execute_query(check_query, (data['code'].strip(), str(code_id)))
        if existing:
            return jsonify({'error': f'Code {data["code"]} already exists'}), 409
        
        query = """
            UPDATE survey_code_library 
            SET code = %s, display_name = %s, description = %s, discipline_code = %s,
                category_code = %s, feature_type = %s, icon_name = %s, connectivity_type = %s,
                geometry_output = %s, auto_connect = %s, create_block = %s, block_name = %s,
                available_attributes = %s, default_attributes = %s, required_attributes = %s,
                layer_template = %s, default_phase = %s, category_group = %s, sort_order = %s,
                is_favorite = %s, tags = %s, attributes = %s, updated_at = CURRENT_TIMESTAMP
            WHERE code_id = %s
            RETURNING code_id, code, display_name, connectivity_type, is_favorite, updated_at
        """
        
        result = execute_query(query, (
            data['code'].strip().upper(),
            data['display_name'].strip(),
            data.get('description', '').strip() or None,
            data.get('discipline_code', '').strip() or None,
            data.get('category_code', '').strip() or None,
            data['feature_type'].strip(),
            data.get('icon_name', '').strip() or None,
            data.get('connectivity_type', 'POINT'),
            data.get('geometry_output', '').strip() or None,
            data.get('auto_connect', False),
            data.get('create_block', False),
            data.get('block_name', '').strip() or None,
            data.get('available_attributes', []),
            data.get('default_attributes', {}),
            data.get('required_attributes', []),
            data.get('layer_template', '').strip() or None,
            data.get('default_phase', 'EXIST'),
            data.get('category_group', '').strip() or None,
            data.get('sort_order', 0),
            data.get('is_favorite', False),
            data.get('tags', []),
            data.get('attributes', {}),
            str(code_id)
        ))
        
        if not result:
            return jsonify({'error': 'Survey code not found'}), 404
        
        return jsonify({
            'survey_code': result[0],
            'message': 'Survey code updated successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/<uuid:code_id>', methods=['DELETE'])
def delete_survey_code(code_id):
    """Delete a survey code (soft delete)"""
    try:
        query = "UPDATE survey_code_library SET is_active = FALSE, updated_at = CURRENT_TIMESTAMP WHERE code_id = %s RETURNING code_id, code"
        result = execute_query(query, (str(code_id),))
        
        if not result:
            return jsonify({'error': 'Survey code not found'}), 404
        
        return jsonify({
            'message': f'Survey code {result[0]["code"]} deactivated successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/<uuid:code_id>/favorite', methods=['POST'])
def toggle_survey_code_favorite(code_id):
    """Toggle favorite status of a survey code"""
    try:
        data = request.get_json()
        is_favorite = data.get('is_favorite', False)
        
        query = """
            UPDATE survey_code_library 
            SET is_favorite = %s, updated_at = CURRENT_TIMESTAMP
            WHERE code_id = %s
            RETURNING code_id, code, display_name, is_favorite
        """
        
        result = execute_query(query, (is_favorite, str(code_id)))
        
        if not result:
            return jsonify({'error': 'Survey code not found'}), 404
        
        return jsonify({
            'survey_code': result[0],
            'message': f'Survey code favorite status updated'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/export')
def export_survey_codes_csv():
    """Export survey codes to CSV for data collector import"""
    import csv
    import io
    from flask import make_response
    
    try:
        query = """
            SELECT code, display_name, description, discipline_code, category_code,
                   feature_type, connectivity_type, geometry_output, category_group,
                   auto_connect, create_block, block_name, layer_template, default_phase,
                   is_favorite, usage_count
            FROM survey_code_library
            WHERE is_active = TRUE
            ORDER BY is_favorite DESC, category_group, sort_order, display_name
        """
        
        codes = execute_query(query)
        
        # Create CSV in memory
        si = io.StringIO()
        writer = csv.writer(si)
        
        # Write header
        writer.writerow([
            'Code', 'Display Name', 'Description', 'Discipline', 'Category',
            'Feature Type', 'Connectivity', 'Geometry', 'Group',
            'Auto-Connect', 'Create Block', 'Block Name', 'Layer Template', 'Phase',
            'Favorite', 'Usage Count'
        ])
        
        # Write data rows
        for code in codes:
            writer.writerow([
                code.get('code', ''),
                code.get('display_name', ''),
                code.get('description', ''),
                code.get('discipline_code', ''),
                code.get('category_code', ''),
                code.get('feature_type', ''),
                code.get('connectivity_type', ''),
                code.get('geometry_output', ''),
                code.get('category_group', ''),
                'Yes' if code.get('auto_connect') else 'No',
                'Yes' if code.get('create_block') else 'No',
                code.get('block_name', ''),
                code.get('layer_template', ''),
                code.get('default_phase', ''),
                'Yes' if code.get('is_favorite') else 'No',
                code.get('usage_count', 0)
            ])
        
        # Create response
        output = make_response(si.getvalue())
        output.headers['Content-Type'] = 'text/csv'
        output.headers['Content-Disposition'] = 'attachment; filename=survey_codes_library.csv'
        
        return output
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== SURVEY CODE TESTING/PARSING API =====

@app.route('/api/survey-codes/parse', methods=['POST'])
def parse_survey_code():
    """Parse a single survey code and return all properties"""
    from survey_code_parser import SurveyCodeParser
    
    try:
        data = request.get_json()
        code = data.get('code', '').strip()
        phase = data.get('phase')
        
        if not code:
            return jsonify({'error': 'code is required'}), 400
        
        parser = SurveyCodeParser(DB_CONFIG)
        parsed = parser.parse_code(code)
        
        if parsed.get('valid'):
            layer_name = parser.resolve_layer_name(parsed, phase)
            parsed['layer_name'] = layer_name
        
        return jsonify(parsed)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/simulate', methods=['POST'])
def simulate_field_sequence():
    """Simulate a sequence of field shots and determine connectivity"""
    from survey_code_parser import SurveyCodeParser
    
    try:
        data = request.get_json()
        shots = data.get('shots', [])
        
        if not shots:
            return jsonify({'error': 'shots array is required'}), 400
        
        parser = SurveyCodeParser(DB_CONFIG)
        result = parser.simulate_field_sequence(shots)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

batch_results_cache = {}

def process_batch_validation(codes):
    """Shared helper for batch validation logic"""
    from survey_code_parser import SurveyCodeParser
    
    parser = SurveyCodeParser(DB_CONFIG)
    result = parser.batch_validate(codes)
    
    results_token = str(uuid.uuid4())
    batch_results_cache[results_token] = result
    
    result['results_token'] = results_token
    return result

@app.route('/api/survey-codes/batch-validate', methods=['POST'])
def batch_validate_codes():
    """Validate multiple codes at once (manual entry)"""
    try:
        data = request.get_json()
        codes = data.get('codes', [])
        
        if not codes:
            return jsonify({'error': 'codes array is required'}), 400
        
        if len(codes) > 5000:
            return jsonify({'error': 'Maximum 5,000 codes allowed per batch'}), 400
        
        result = process_batch_validation(codes)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/batch-validate-upload', methods=['POST'])
def batch_validate_upload():
    """Validate codes from uploaded CSV file"""
    import csv
    import io
    
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.csv', '.txt')):
            return jsonify({'error': 'File must be CSV or TXT format'}), 400
        
        content = file.read()
        
        if len(content) > 2 * 1024 * 1024:
            return jsonify({'error': 'File size exceeds 2 MB limit'}), 400
        
        content_str = content.decode('utf-8-sig')
        
        reader = csv.reader(io.StringIO(content_str))
        codes = []
        
        for row_num, row in enumerate(reader, start=1):
            if row and row[0].strip():
                code = row[0].strip()
                if not code.startswith('#'):
                    codes.append(code)
            
            if len(codes) > 5000:
                return jsonify({'error': 'File contains more than 5,000 codes'}), 400
        
        if not codes:
            return jsonify({'error': 'No valid codes found in file'}), 400
        
        result = process_batch_validation(codes)
        result['filename'] = secure_filename(file.filename)
        result['codes_extracted'] = len(codes)
        
        return jsonify(result)
    except UnicodeDecodeError:
        return jsonify({'error': 'File encoding error. Please use UTF-8 encoded CSV'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/batch-validate-download/<token>')
def batch_validate_download(token):
    """Download batch validation results as CSV"""
    import csv
    import io
    
    try:
        if token not in batch_results_cache:
            return jsonify({'error': 'Results not found or expired'}), 404
        
        data = batch_results_cache[token]
        
        si = io.StringIO()
        writer = csv.writer(si, quoting=csv.QUOTE_MINIMAL)
        
        writer.writerow(['Code', 'Valid', 'Display Name', 'Error Message'])
        
        for result in data.get('results', []):
            if result.get('valid'):
                writer.writerow([
                    result.get('code', ''),
                    'TRUE',
                    result.get('display_name', '').replace('\n', ' ').replace('\r', ''),
                    ''
                ])
            else:
                error_msg = result.get('error', '').replace('\n', ' ').replace('\r', '')
                code_from_error = error_msg.split('"')[1] if '"' in error_msg else 'Unknown'
                writer.writerow([
                    code_from_error,
                    'FALSE',
                    '',
                    error_msg
                ])
        
        output = make_response(si.getvalue())
        output.headers['Content-Type'] = 'text/csv'
        output.headers['Content-Disposition'] = 'attachment; filename=batch_validation_results.csv'
        
        return output
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-codes/connectivity-rules')
def get_connectivity_rules():
    """Get documentation for connectivity types"""
    from survey_code_parser import SurveyCodeParser
    
    try:
        parser = SurveyCodeParser(DB_CONFIG)
        rules = parser.get_connectivity_rules()
        return jsonify(rules)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-import/preview', methods=['POST'])
def survey_import_preview():
    """
    Preview survey data import with code parsing and connectivity analysis
    
    Accepts: multipart/form-data with file upload
    Returns: JSON with parsed points, sequences, line segments, warnings
    """
    from survey_import_service import SurveyImportService
    
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        coordinate_system = request.form.get('coordinate_system', 'SRID_2226')
        drawing_id = request.form.get('drawing_id')
        
        content = file.read().decode('utf-8-sig')
        
        service = SurveyImportService(DB_CONFIG)
        
        points, parse_errors = service.parse_pnezd_with_codes(content, coordinate_system)
        
        if parse_errors and not points:
            return jsonify({
                'error': 'Failed to parse file',
                'details': parse_errors
            }), 400
        
        preview_data = service.generate_preview(points, drawing_id)
        
        preview_data['parse_errors'] = parse_errors
        preview_data['filename'] = secure_filename(file.filename)
        preview_data['coordinate_system'] = coordinate_system
        
        return jsonify(preview_data)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-import/commit', methods=['POST'])
def survey_import_commit():
    """
    Commit survey import to database (transactional)
    
    Accepts: JSON with preview_data, project_id, drawing_id
    Returns: Import summary with counts and IDs
    """
    from survey_import_service import SurveyImportService
    
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        preview_data = data.get('preview_data')
        project_id = data.get('project_id')
        drawing_id = data.get('drawing_id')
        
        if not preview_data:
            return jsonify({'error': 'preview_data is required'}), 400
        
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        service = SurveyImportService(DB_CONFIG)
        
        result = service.commit_import(preview_data, project_id, drawing_id)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== SURVEY POINT MANAGER API =====

@app.route('/api/survey-points')
def get_survey_points():
    """
    Get survey points with filtering
    Query params: project_id, drawing_id, discipline, connectivity, sequence_status, is_active, search
    """
    try:
        # Build dynamic WHERE clause based on query parameters
        where_conditions = []
        params = []
        
        project_id = request.args.get('project_id')
        drawing_id = request.args.get('drawing_id')
        discipline = request.args.get('discipline')
        connectivity = request.args.get('connectivity')
        sequence_status = request.args.get('sequence_status')
        is_active = request.args.get('is_active')
        search = request.args.get('search')
        
        if project_id:
            where_conditions.append("sp.project_id = %s")
            params.append(project_id)
        
        if drawing_id:
            where_conditions.append("sp.drawing_id = %s")
            params.append(drawing_id)
        
        if discipline:
            where_conditions.append("sp.discipline_code = %s")
            params.append(discipline)
        
        if connectivity:
            where_conditions.append("sp.connectivity_type = %s")
            params.append(connectivity)
        
        if sequence_status == 'connected':
            where_conditions.append("sp.sequence_id IS NOT NULL")
        elif sequence_status == 'standalone':
            where_conditions.append("sp.sequence_id IS NULL")
        
        if is_active:
            where_conditions.append("sp.is_active = %s")
            params.append(is_active == 'true')
        
        if search:
            where_conditions.append("(sp.point_number ILIKE %s OR sp.code ILIKE %s)")
            search_pattern = f"%{search}%"
            params.extend([search_pattern, search_pattern])
        
        where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"
        
        query = f"""
            SELECT 
                sp.point_id,
                sp.project_id,
                sp.drawing_id,
                sp.point_number,
                sp.point_description,
                sp.code,
                sp.code_id,
                sp.discipline_code,
                sp.category_code,
                sp.feature_type,
                sp.connectivity_type,
                sp.layer_name,
                sp.phase,
                sp.auto_connected,
                sp.sequence_id,
                sp.northing,
                sp.easting,
                sp.elevation,
                sp.coordinate_system,
                sp.is_active,
                sp.created_at,
                sp.updated_at,
                p.project_number,
                p.project_name,
                d.drawing_number,
                d.drawing_name
            FROM survey_points sp
            LEFT JOIN projects p ON sp.project_id = p.project_id
            LEFT JOIN drawings d ON sp.drawing_id = d.drawing_id
            WHERE {where_clause}
            ORDER BY sp.created_at DESC, sp.point_number
            LIMIT 1000
        """
        
        points = execute_query(query, params)
        
        return jsonify({
            'points': points,
            'count': len(points)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-points/soft-delete', methods=['POST'])
def soft_delete_survey_points():
    """
    Soft-delete survey points (set is_active = false)
    Accepts: JSON with point_ids array
    """
    try:
        data = request.get_json()
        
        if not data or 'point_ids' not in data:
            return jsonify({'error': 'point_ids array is required'}), 400
        
        point_ids = data['point_ids']
        
        if not isinstance(point_ids, list) or len(point_ids) == 0:
            return jsonify({'error': 'point_ids must be a non-empty array'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE survey_points 
                    SET is_active = false, updated_at = CURRENT_TIMESTAMP
                    WHERE point_id = ANY(%s::uuid[])
                    RETURNING point_id
                    """,
                    (point_ids,)
                )
                
                deleted_ids = [row[0] for row in cur.fetchall()]
                conn.commit()
                
                return jsonify({
                    'success': True,
                    'deleted_count': len(deleted_ids),
                    'deleted_ids': deleted_ids
                })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/survey-points/restore', methods=['POST'])
def restore_survey_points():
    """
    Restore soft-deleted survey points (set is_active = true)
    Accepts: JSON with point_ids array
    """
    try:
        data = request.get_json()
        
        if not data or 'point_ids' not in data:
            return jsonify({'error': 'point_ids array is required'}), 400
        
        point_ids = data['point_ids']
        
        if not isinstance(point_ids, list) or len(point_ids) == 0:
            return jsonify({'error': 'point_ids must be a non-empty array'}), 400
        
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE survey_points 
                    SET is_active = true, updated_at = CURRENT_TIMESTAMP
                    WHERE point_id = ANY(%s::uuid[])
                    RETURNING point_id
                    """,
                    (point_ids,)
                )
                
                restored_ids = [row[0] for row in cur.fetchall()]
                conn.commit()
                
                return jsonify({
                    'success': True,
                    'restored_count': len(restored_ids),
                    'restored_ids': restored_ids
                })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== GIS DATA LAYERS API =====

@app.route('/api/gis-data-layers')
def get_gis_data_layers():
    """Get all GIS data layers"""
    try:
        query = """
            SELECT layer_id, service_id, name, jurisdiction, category, rest_url, layer, behavior,
                   cs_wkid, query_where, extent_mode, label_field, label_format, sym_preset, block_name,
                   min_scale, max_scale, status, last_verified, notes, is_active, tags, attributes
            FROM gis_data_layers
            WHERE is_active = TRUE
            ORDER BY jurisdiction, category, name
        """
        layers = execute_query(query)
        return jsonify({'gis_data_layers': layers})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gis-data-layers/<int:layer_id>')
def get_gis_data_layer_detail(layer_id):
    """Get a specific GIS data layer"""
    try:
        query = """
            SELECT layer_id, service_id, name, jurisdiction, category, rest_url, layer, behavior,
                   cs_wkid, query_where, extent_mode, label_field, label_format, sym_preset, block_name,
                   min_scale, max_scale, status, last_verified, notes, is_active, tags, attributes
            FROM gis_data_layers
            WHERE layer_id = %s
        """
        result = execute_query(query, (layer_id,))
        
        if not result:
            return jsonify({'error': 'GIS data layer not found'}), 404
        
        return jsonify(result[0])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gis-data-layers', methods=['POST'])
def create_gis_data_layer():
    """Create a new GIS data layer"""
    try:
        data = request.get_json()
        
        if not data.get('service_id') or not data.get('name') or not data.get('rest_url'):
            return jsonify({'error': 'service_id, name, and rest_url are required'}), 400
        
        check_query = "SELECT layer_id FROM gis_data_layers WHERE service_id = %s"
        existing = execute_query(check_query, (data['service_id'].strip(),))
        if existing:
            return jsonify({'error': f'Service ID {data["service_id"]} already exists'}), 409
        
        query = """
            INSERT INTO gis_data_layers (service_id, name, jurisdiction, category, rest_url, layer, behavior,
                                        cs_wkid, query_where, extent_mode, label_field, label_format, sym_preset,
                                        block_name, min_scale, max_scale, status, last_verified, notes, tags, attributes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING layer_id
        """
        result = execute_query(query, (
            data['service_id'].strip(),
            data['name'].strip(),
            (data.get('jurisdiction') or '').strip() or None,
            (data.get('category') or '').strip() or None,
            data['rest_url'].strip(),
            data.get('layer'),
            (data.get('behavior') or '').strip() or None,
            data.get('cs_wkid'),
            (data.get('query_where') or '').strip() or None,
            (data.get('extent_mode') or '').strip() or None,
            (data.get('label_field') or '').strip() or None,
            (data.get('label_format') or '').strip() or None,
            (data.get('sym_preset') or '').strip() or None,
            (data.get('block_name') or '').strip() or None,
            data.get('min_scale', 0),
            data.get('max_scale', 0),
            (data.get('status') or 'active').strip(),
            data.get('last_verified'),
            (data.get('notes') or '').strip() or None,
            data.get('tags', []),
            data.get('attributes')
        ))
        
        return jsonify({
            'layer_id': result[0]['layer_id'],
            'message': f'GIS data layer {data["name"]} created successfully'
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gis-data-layers/<int:layer_id>', methods=['PUT'])
def update_gis_data_layer(layer_id):
    """Update a GIS data layer"""
    try:
        data = request.get_json()
        
        if not data.get('service_id') or not data.get('name') or not data.get('rest_url'):
            return jsonify({'error': 'service_id, name, and rest_url are required'}), 400
        
        check_query = """
            SELECT layer_id FROM gis_data_layers 
            WHERE service_id = %s AND layer_id != %s
        """
        existing = execute_query(check_query, (data['service_id'].strip(), layer_id))
        if existing:
            return jsonify({'error': f'Service ID {data["service_id"]} already exists'}), 409
        
        query = """
            UPDATE gis_data_layers 
            SET service_id = %s, name = %s, jurisdiction = %s, category = %s, rest_url = %s, layer = %s,
                behavior = %s, cs_wkid = %s, query_where = %s, extent_mode = %s, label_field = %s,
                label_format = %s, sym_preset = %s, block_name = %s, min_scale = %s, max_scale = %s,
                status = %s, last_verified = %s, notes = %s, tags = %s, attributes = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE layer_id = %s
            RETURNING layer_id
        """
        result = execute_query(query, (
            data['service_id'].strip(),
            data['name'].strip(),
            (data.get('jurisdiction') or '').strip() or None,
            (data.get('category') or '').strip() or None,
            data['rest_url'].strip(),
            data.get('layer'),
            (data.get('behavior') or '').strip() or None,
            data.get('cs_wkid'),
            (data.get('query_where') or '').strip() or None,
            (data.get('extent_mode') or '').strip() or None,
            (data.get('label_field') or '').strip() or None,
            (data.get('label_format') or '').strip() or None,
            (data.get('sym_preset') or '').strip() or None,
            (data.get('block_name') or '').strip() or None,
            data.get('min_scale', 0),
            data.get('max_scale', 0),
            (data.get('status') or 'active').strip(),
            data.get('last_verified'),
            (data.get('notes') or '').strip() or None,
            data.get('tags', []),
            data.get('attributes'),
            layer_id
        ))
        
        if not result:
            return jsonify({'error': 'GIS data layer not found'}), 404
        
        return jsonify({'message': f'GIS data layer updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gis-data-layers/<int:layer_id>', methods=['DELETE'])
def delete_gis_data_layer(layer_id):
    """Delete a GIS data layer (soft delete)"""
    try:
        query = "UPDATE gis_data_layers SET is_active = FALSE WHERE layer_id = %s RETURNING layer_id"
        result = execute_query(query, (layer_id,))
        
        if not result:
            return jsonify({'error': 'GIS data layer not found'}), 404
        
        return jsonify({'message': 'GIS data layer deactivated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/attributes')
def get_attributes():
    """Get all attribute codes"""
    try:
        query = """
            SELECT attribute_id, code, full_name, attribute_category, description, pattern, is_active
            FROM attribute_codes
            WHERE is_active = TRUE
            ORDER BY attribute_category, code
        """
        attributes = execute_query(query)
        return jsonify({'attributes': attributes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/attributes', methods=['POST'])
def create_attribute():
    """Create a new attribute code"""
    try:
        data = request.get_json()
        query = """
            INSERT INTO attribute_codes (code, full_name, attribute_category, description, pattern)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING attribute_id
        """
        result = execute_query(query, (data['code'], data['full_name'], data.get('attribute_category'), data.get('description'), data.get('pattern')))
        return jsonify({'attribute_id': result[0]['attribute_id'], 'message': 'Attribute created'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/attributes/<int:attribute_id>', methods=['PUT'])
def update_attribute(attribute_id):
    """Update an attribute code"""
    try:
        data = request.get_json()
        query = """
            UPDATE attribute_codes 
            SET code = %s, full_name = %s, attribute_category = %s, description = %s, pattern = %s, is_active = %s
            WHERE attribute_id = %s
        """
        execute_query(query, (data['code'], data['full_name'], data.get('attribute_category'), data.get('description'), data.get('pattern'), data.get('is_active', True), attribute_id), fetch=False)
        return jsonify({'message': 'Attribute updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/vocabulary/attributes/<int:attribute_id>', methods=['DELETE'])
def delete_attribute(attribute_id):
    """Delete an attribute code"""
    try:
        execute_query("DELETE FROM attribute_codes WHERE attribute_id = %s", (attribute_id,), fetch=False)
        return jsonify({'message': 'Attribute deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# ENTITY VIEWER API
# ============================================

ENTITY_VIEWER_REGISTRY = {
    'parcels': {
        'table': 'parcels',
        'geometry_column': 'boundary_geometry',
        'srid': 2226,
        'id_column': 'parcel_id',
        'label_column': 'parcel_name',
        'type': 'polygon',
        'category': 'Property',
        'color': '#ff6b35'
    },
    'utility_lines': {
        'table': 'utility_lines',
        'geometry_column': 'geometry',
        'srid': 2226,
        'id_column': 'line_id',
        'label_column': 'line_number',
        'type': 'line',
        'category': 'Utilities',
        'color': '#f7b801'
    },
    'utility_structures': {
        'table': 'utility_structures',
        'geometry_column': 'rim_geometry',
        'srid': 2226,
        'id_column': 'structure_id',
        'label_column': 'structure_number',
        'type': 'point',
        'category': 'Utilities',
        'color': '#6a994e'
    },
    'survey_points': {
        'table': 'survey_points',
        'geometry_column': 'geometry',
        'srid': 2226,
        'id_column': 'point_id',
        'label_column': 'point_number',
        'type': 'point',
        'category': 'Survey',
        'color': '#00d9ff'
    },
    'horizontal_alignments': {
        'table': 'horizontal_alignments',
        'geometry_column': 'geometry',
        'srid': 2226,
        'id_column': 'alignment_id',
        'label_column': 'alignment_name',
        'type': 'line',
        'category': 'Civil',
        'color': '#bc4b51'
    },
    'easements': {
        'table': 'easements',
        'geometry_column': 'geometry',
        'srid': 2226,
        'id_column': 'easement_id',
        'label_column': 'easement_type',
        'type': 'polygon',
        'category': 'Property',
        'color': '#8e7dbe'
    },
    'right_of_way': {
        'table': 'right_of_way',
        'geometry_column': 'geometry',
        'srid': 2226,
        'id_column': 'row_id',
        'label_column': 'row_type',
        'type': 'polygon',
        'category': 'Property',
        'color': '#5f0f40'
    },
    'drawing_entities': {
        'table': 'drawing_entities',
        'geometry_column': 'geometry',
        'srid': 0,
        'id_column': 'entity_id',
        'label_column': 'entity_type',
        'type': 'mixed',
        'category': 'Drawing',
        'color': '#06ffa5'
    }
}

@app.route('/entity-viewer')
def entity_viewer_page():
    """Entity Viewer page"""
    return render_template('entity_viewer.html')

@app.route('/api/entity-viewer/catalog')
def get_entity_viewer_catalog():
    """Get available entity types with counts per project"""
    try:
        project_id = request.args.get('project_id')
        
        catalog = []
        for entity_key, config in ENTITY_VIEWER_REGISTRY.items():
            try:
                col_check = execute_query(f"""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name='{config['table']}' 
                    AND column_name IN ('project_id', 'drawing_id')
                """)
                col_names = [c['column_name'] for c in col_check]
                has_project_id = 'project_id' in col_names
                has_drawing_id = 'drawing_id' in col_names
                
                if has_project_id:
                    query = f"""
                        SELECT 
                            '{entity_key}' as entity_type,
                            '{config['category']}' as category,
                            '{config['type']}' as geometry_type,
                            '{config['color']}' as color,
                            COUNT(*) as count
                        FROM {config['table']}
                        WHERE project_id = %s
                    """
                    result = execute_query(query, (project_id,))
                else:
                    continue
                
                if result and result[0]['count'] > 0:
                    catalog.append(result[0])
            except Exception as e:
                print(f"Error loading catalog for {entity_key}: {e}")
                continue
        
        return jsonify({'catalog': catalog})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity-viewer/layers')
def get_entity_viewer_layers():
    """Get available layers for a project"""
    try:
        project_id = request.args.get('project_id')
        
        query = """
            SELECT DISTINCT l.layer_name
            FROM layers l
            WHERE l.project_id = %s
            ORDER BY l.layer_name
        """
        
        layers = execute_query(query, (project_id,))
        return jsonify({'layers': [l['layer_name'] for l in layers]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity-viewer/entities')
def get_entity_viewer_entities():
    """Get entities with geometries, transformed to EPSG:4326"""
    try:
        project_id = request.args.get('project_id')
        entity_types = request.args.getlist('entity_type')
        layer_names = request.args.getlist('layer')
        
        if not project_id:
            return jsonify({'error': 'project_id is required'}), 400
        
        if not entity_types:
            entity_types = list(ENTITY_VIEWER_REGISTRY.keys())
        
        all_entities = []
        union_queries = []
        bbox_queries = []
        
        for entity_type in entity_types:
            if entity_type not in ENTITY_VIEWER_REGISTRY:
                continue
            
            config = ENTITY_VIEWER_REGISTRY[entity_type]
            geom_col = config['geometry_column']
            srid = config['srid']
            
            transform_func = f"ST_Transform({geom_col}, 4326)" if srid != 4326 else geom_col
            
            has_project_id = False
            has_drawing_id = False
            has_layer_name = False
            
            try:
                col_check = execute_query(f"""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name='{config['table']}' 
                    AND column_name IN ('project_id', 'drawing_id', 'layer_name')
                """)
                col_names = [c['column_name'] for c in col_check]
                has_project_id = 'project_id' in col_names
                has_drawing_id = 'drawing_id' in col_names
                has_layer_name = 'layer_name' in col_names
            except:
                pass
            
            if has_project_id:
                query = f"""
                    SELECT 
                        '{entity_type}' as entity_type,
                        t.{config['id_column']}::text as entity_id,
                        t.{config['label_column']} as label,
                        {'t.layer_name' if has_layer_name else "'' as layer_name"},
                        '{config['category']}' as category,
                        '{config['type']}' as geometry_type,
                        '{config['color']}' as color,
                        ST_AsGeoJSON({transform_func})::json as geometry,
                        ST_GeometryType(t.{geom_col}) as geom_type
                    FROM {config['table']} t
                    WHERE t.project_id = %s
                """
                params = [project_id]
                
                if layer_names and has_layer_name:
                    placeholders = ','.join(['%s'] * len(layer_names))
                    query += f" AND t.layer_name IN ({placeholders})"
                    params.extend(layer_names)
            else:
                continue
            
            query += f" AND t.{geom_col} IS NOT NULL"
            
            entities = execute_query(query, tuple(params))
            all_entities.extend(entities)
            
            if entities:
                if has_project_id:
                    bbox_query = f"""
                        SELECT ST_Extent(ST_Transform(t.{geom_col}, 4326)) as bbox
                        FROM {config['table']} t
                        WHERE t.project_id = %s AND t.{geom_col} IS NOT NULL
                    """
                    bbox_params = [project_id]
                    
                    if layer_names and has_layer_name:
                        placeholders = ','.join(['%s'] * len(layer_names))
                        bbox_query += f" AND t.layer_name IN ({placeholders})"
                        bbox_params.extend(layer_names)
                    
                    bbox_queries.append((bbox_query, tuple(bbox_params)))
        
        bbox = None
        if bbox_queries:
            bbox_union = " UNION ALL ".join([q[0] for q in bbox_queries])
            all_bbox_params = []
            for _, params in bbox_queries:
                all_bbox_params.extend(params)
            
            bbox_result = execute_query(f"SELECT ST_Extent(bbox::geometry) as combined_bbox FROM ({bbox_union}) t", tuple(all_bbox_params))
            if bbox_result and bbox_result[0]['combined_bbox']:
                bbox_str = bbox_result[0]['combined_bbox']
                coords = bbox_str.replace('BOX(', '').replace(')', '').split(',')
                min_coords = coords[0].strip().split()
                max_coords = coords[1].strip().split()
                bbox = {
                    'minX': float(min_coords[0]),
                    'minY': float(min_coords[1]),
                    'maxX': float(max_coords[0]),
                    'maxY': float(max_coords[1])
                }
        
        layer_counts = {}
        type_counts = {}
        for entity in all_entities:
            layer = entity.get('layer_name', 'Unknown')
            layer_counts[layer] = layer_counts.get(layer, 0) + 1
            
            etype = entity.get('entity_type', 'Unknown')
            type_counts[etype] = type_counts.get(etype, 0) + 1
        
        return jsonify({
            'entities': all_entities,
            'bbox': bbox,
            'layer_counts': layer_counts,
            'type_counts': type_counts,
            'total_count': len(all_entities)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
